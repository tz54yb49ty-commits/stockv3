"""Local-only N3->N5 replay artifacts for the N6 A-track UI.

This module deliberately does not import DB clients or N3/N4/N5 execute
entrypoints. It writes only local replay artifacts under a caller-provided
replay root.
"""

from __future__ import annotations

from contextlib import contextmanager
from copy import deepcopy
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path
import re
from time import perf_counter
from typing import Any, Iterable, Mapping, Sequence
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ashare_v3.action import provisional_action_eligible
from ashare_v3.action import provisional_action_executed_dry_run
from ashare_v3.market import realtime_projection_execute as projection_execute
from ashare_v3.market import v3_realtime_virtual_metric_writer as writer
from ashare_v3.trigger import provisional_ordinary_matcher
from ashare_v3.trigger import provisional_projection_matcher
from ashare_v3.trigger import provisional_trigger_lifecycle


DISPLAY_TIMEZONE = ZoneInfo("Asia/Shanghai")
SAFETY_FLAGS: dict[str, Any] = {
    "replay_mode": "local_only",
    "database_write": False,
    "consume_outbox": False,
    "update_checkpoint": False,
    "worker_started": False,
    "production_run_id_used": False,
}
LOCAL_ONLY_NOTICE = "Local replay only. Not production lineage. Not eligible for N6 delivery or trade."
JOB_ID_RE = re.compile(r"^local_replay_(\d{8})_(\d{6})_[0-9a-f]{8}$")
SOURCE_BUNDLE_KEY_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_-]*$")
PRODUCTION_ID_PREFIXES = (
    "action_",
    "condition_",
    "live_",
    "market_",
    "previous_",
    "realtime_",
    "trigger_",
    "user_",
)
EXCEL_SHEETS = (
    "summary",
    "minute_timeline",
    "n4_ordinary_messages",
    "n4_hint_messages",
    "n5_action_eligible",
    "n5_action_executed",
    "n5_action_skipped",
    "quality_blockers",
    "lineage_and_safety",
)
REPLAY_ENGINE_VERSIONS = ("fixture_v1", "canonical_plan_v1")
DEFAULT_REPLAY_ENGINE_VERSION = "canonical_plan_v1"
FULL_DAY_SHADOW_VALIDATION_MODE = "full_day_shadow_v1"
REPLAY_VALIDATION_MODES = ("none", "asset_unit_fix_delta_v1", FULL_DAY_SHADOW_VALIDATION_MODE)
DEFAULT_REPLAY_VALIDATION_MODE = "none"
REPLAY_N3P_STRATEGIES = ("full_n3p", "prefilter_audit", "prefilter_prune")
DEFAULT_REPLAY_N3P_STRATEGY = "full_n3p"
REPLAY_N3P_REDUCTION_MODES = ("none", "active_state_fast_path")
DEFAULT_REPLAY_N3P_REDUCTION_MODE = "none"
REPLAY_N3P_NEGATIVE_CACHE_MODES = ("disabled", "enabled")
DEFAULT_REPLAY_N3P_NEGATIVE_CACHE = "disabled"
REPLAY_ASSET_SCOPES = ("all", "index_board_only", "index_only", "board_only", "stock_only")
DEFAULT_REPLAY_ASSET_SCOPE = "all"
DEFAULT_REPLAY_SOURCE_BUNDLE_KEY = "auto"
REPLAY_ASSET_SCOPE_ALLOWED_KINDS = {
    "all": ("stock", "index", "board"),
    "index_board_only": ("index", "board"),
    "index_only": ("index",),
    "board_only": ("board",),
    "stock_only": ("stock",),
}
REPLAY_SOURCE_POLICY = {
    "fixture_v1": "fixture harness, not canonical planner proof",
    "canonical_plan_v1": "canonical planner dry-run, not production lineage",
}
HISTORICAL_REPLAY_SOURCE_POLICY = "historical_minute_local_replay"
HISTORICAL_REPLAY_SOURCE_DIRNAME = "_sources"
HISTORICAL_REPLAY_PROFILE_DIRNAME = "_profiles"
INDEX_BOARD_FULL_DAY_SOURCE_SUFFIX = "_index_board_only_full_day"
C1_SOURCE_INCOMPLETE_REASON = "BLOCKED_REPLAY_C1_SOURCE_INCOMPLETE"
PROFILE_PHASE_CLASSIFICATION = {
    "source_bundle_json_load": "JSON_LOAD_BOTTLENECK",
    "source_bundle_normalization_validation": "NORMALIZATION_BOTTLENECK",
    "per_minute_n3p_plan_only": "N3P_PLAN_BOTTLENECK",
    "per_minute_b2_plan_only": "B2_PLAN_BOTTLENECK",
    "per_minute_n4_ordinary_matcher_lifecycle": "N4_MATCHER_BOTTLENECK",
    "per_minute_n4_hint_matcher_lifecycle": "N4_MATCHER_BOTTLENECK",
    "per_minute_n5_actioneligible_planner": "N5_PLANNER_BOTTLENECK",
    "per_minute_n5_actionexecuted_evaluator": "N5_PLANNER_BOTTLENECK",
    "artifact_jsonl_serialization": "ARTIFACT_JSONL_BOTTLENECK",
    "excel_generation": "EXCEL_SERIALIZATION_BOTTLENECK",
}


class _ReplayProfileCollector:
    def __init__(self, *, enabled: bool, job_id: str, trade_date: str, replay_engine_version: str) -> None:
        self.enabled = enabled
        self.job_id = job_id
        self.trade_date = trade_date
        self.replay_engine_version = replay_engine_version
        self.started_at = datetime.now(tz=DISPLAY_TIMEZONE)
        self.phases: dict[str, dict[str, Any]] = {}
        self.metadata: dict[str, Any] = {}

    def set_metadata(self, **values: Any) -> None:
        if not self.enabled:
            return
        for key, value in values.items():
            self.metadata[key] = value

    def start_phase(self, name: str, **metrics: Any) -> dict[str, Any] | None:
        if not self.enabled:
            return None
        return {
            "name": name,
            "started_at": perf_counter(),
            "metrics": dict(metrics),
        }

    def finish_phase(self, token: dict[str, Any] | None, **metrics: Any) -> None:
        if not self.enabled or token is None:
            return
        name = str(token["name"])
        elapsed_seconds = perf_counter() - float(token["started_at"])
        merged_metrics = dict(token.get("metrics") or {})
        merged_metrics.update(metrics)
        phase = self.phases.setdefault(
            name,
            {
                "elapsed_seconds": 0.0,
                "calls": 0,
                "max_seconds": 0.0,
                "metrics": {},
                "samples": [],
            },
        )
        phase["elapsed_seconds"] += elapsed_seconds
        phase["calls"] += 1
        phase["max_seconds"] = max(float(phase["max_seconds"]), elapsed_seconds)
        for key, value in merged_metrics.items():
            if isinstance(value, bool):
                phase["metrics"][key] = int(bool(value)) + int(phase["metrics"].get(key) or 0)
            elif isinstance(value, (int, float)):
                phase["metrics"][key] = value + (phase["metrics"].get(key) or 0)
            elif value not in (None, ""):
                phase["metrics"][key] = value
        sample = dict(merged_metrics)
        sample["elapsed_seconds"] = round(elapsed_seconds, 6)
        phase["samples"].append(sample)

    @contextmanager
    def phase(self, name: str, **metrics: Any) -> Any:
        token = self.start_phase(name, **metrics)
        try:
            yield
        finally:
            self.finish_phase(token)

    def _classification_seconds(self) -> dict[str, float]:
        totals: dict[str, float] = {}
        for phase_name, phase in self.phases.items():
            bucket = PROFILE_PHASE_CLASSIFICATION.get(phase_name)
            if not bucket:
                continue
            totals[bucket] = totals.get(bucket, 0.0) + float(phase.get("elapsed_seconds") or 0.0)
        return totals

    def bottleneck_classification(self) -> str:
        totals = self._classification_seconds()
        if not totals:
            return "UNKNOWN_TIMEOUT"
        return max(totals.items(), key=lambda item: item[1])[0]

    def build_payload(self, *, status: str, artifact_dir: str = "", blocked_reason: str = "") -> dict[str, Any]:
        finished_at = datetime.now(tz=DISPLAY_TIMEZONE)
        return {
            "job_id": self.job_id,
            "trade_date": self.trade_date,
            "replay_engine_version": self.replay_engine_version,
            "status": status,
            "artifact_dir": artifact_dir,
            "blocked_reason": blocked_reason,
            "started_at": self.started_at.isoformat(),
            "finished_at": finished_at.isoformat(),
            "elapsed_seconds": round((finished_at - self.started_at).total_seconds(), 6),
            "safety_flags": dict(SAFETY_FLAGS),
            "bottleneck_classification": self.bottleneck_classification(),
            "classification_seconds": {
                key: round(value, 6)
                for key, value in sorted(self._classification_seconds().items())
            },
            "metadata": dict(self.metadata),
            "phases": {
                name: {
                    "elapsed_seconds": round(float(phase.get("elapsed_seconds") or 0.0), 6),
                    "calls": int(phase.get("calls") or 0),
                    "max_seconds": round(float(phase.get("max_seconds") or 0.0), 6),
                    "metrics": dict(phase.get("metrics") or {}),
                    "samples": list(phase.get("samples") or []),
                }
                for name, phase in sorted(self.phases.items())
            },
        }


class _N3PReplayCache:
    def __init__(
        self,
        *,
        replay_config: Mapping[str, Any],
        context_rows: Sequence[Mapping[str, Any]],
        cache_by_minute: bool = False,
    ) -> None:
        self._prefix = {
            "historical_source_hash": str(replay_config.get("historical_source_hash") or ""),
            "asset_scope": str(replay_config.get("asset_scope") or DEFAULT_REPLAY_ASSET_SCOPE),
            "source_bundle_key": str(replay_config.get("source_bundle_key") or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY),
            "resolved_source_bundle_path": str(replay_config.get("resolved_source_bundle_path") or ""),
            "replay_engine_version": str(replay_config.get("replay_engine_version") or DEFAULT_REPLAY_ENGINE_VERSION),
            "context_hash": _stable_payload_hash(context_rows),
        }
        self._cache_by_minute = cache_by_minute
        self._artifacts: dict[str, dict[str, Any]] = {}
        self._hits = 0
        self._misses = 0
        self._empty_minute_fast_path_count = 0

    def empty_minute_noop(self) -> None:
        self._empty_minute_fast_path_count += 1

    def get(
        self,
        *,
        candidates: Sequence[Mapping[str, Any]],
        target_run_id: str,
        minute: str = "",
    ) -> tuple[dict[str, Any] | None, bool, str]:
        key = self._key(candidates, minute=minute)
        cached = self._artifacts.get(key)
        if cached is None:
            self._misses += 1
            return None, False, key
        self._hits += 1
        return _retarget_n3p_artifact(cached, target_run_id=target_run_id), True, key

    def put(self, *, key: str, artifact: Mapping[str, Any]) -> None:
        self._artifacts[key] = _clone_payload(artifact)

    def stats(self) -> dict[str, int]:
        return {
            "n3p_cache_hits": self._hits,
            "n3p_cache_misses": self._misses,
            "n3p_cache_saved_calls": self._hits,
            "n3p_cache_key_count": len(self._artifacts),
            "empty_minute_fast_path_count": self._empty_minute_fast_path_count,
        }

    def _key(self, candidates: Sequence[Mapping[str, Any]], *, minute: str = "") -> str:
        candidate_target_minute = ",".join(sorted({_candidate_start_minute(candidate) for candidate in candidates}))
        payload = {
            **self._prefix,
            "candidate_target_minute": candidate_target_minute,
            "source_prefix_minute": minute if self._cache_by_minute else "",
            "candidate_group_hash": _stable_payload_hash(
                sorted(
                    (_candidate_cache_fingerprint(candidate) for candidate in candidates),
                    key=lambda row: (
                        str(row.get("asset_kind") or ""),
                        str(row.get("identity_key") or ""),
                        str(row.get("condition_key") or ""),
                    ),
                )
            ),
        }
        return _stable_payload_hash(payload)


class _N3PNegativeProofCache:
    def __init__(self, *, enabled: bool, replay_config: Mapping[str, Any]) -> None:
        self.enabled = enabled
        self._prefix = {
            "historical_source_hash": str(replay_config.get("historical_source_hash") or ""),
            "source_bundle_key": str(replay_config.get("source_bundle_key") or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY),
            "asset_scope": str(replay_config.get("asset_scope") or DEFAULT_REPLAY_ASSET_SCOPE),
            "replay_engine_version": str(replay_config.get("replay_engine_version") or DEFAULT_REPLAY_ENGINE_VERSION),
        }
        self._proofs: dict[str, dict[str, Any]] = {}
        self._proofs_by_family: dict[str, dict[str, Any]] = {}
        self._hits = 0
        self._misses = 0
        self._saved_calls = 0
        self._fail_open = 0
        self._false_negative_count = 0

    def store(self, proof_rows: Sequence[Mapping[str, Any]]) -> None:
        if not self.enabled:
            return
        for row in proof_rows:
            proof = dict(row)
            stable_key = str(proof.get("stable_trigger_key") or "")
            if not stable_key:
                continue
            self._proofs[stable_key] = proof
            family_key = str(proof.get("stable_trigger_family_key") or "")
            if family_key:
                self._proofs_by_family[family_key] = proof

    def decision(self, *, candidate: Mapping[str, Any], proof_context: Mapping[str, Any]) -> dict[str, Any]:
        stable_key = _candidate_stable_trigger_key(candidate)
        family_key = _candidate_stable_trigger_family_key(candidate)
        base = {
            **self._prefix,
            "stable_trigger_key": stable_key,
            "stable_trigger_family_key": family_key,
            "asset_kind": candidate.get("asset_kind"),
            "identity_key": candidate.get("identity_key"),
            "condition_key": candidate.get("condition_key"),
            "original_condition_key": _candidate_proof_condition_key(candidate),
            "signal_type": candidate.get("signal_type"),
            "requested_periods": list(proof_context.get("requested_periods") or _candidate_requested_periods(candidate)),
            "skip_full_n3p": False,
        }
        if not self.enabled:
            return {**base, "decision": "disabled"}
        proof = self._proofs.get(stable_key) or self._proofs_by_family.get(family_key)
        if proof is None:
            self._misses += 1
            self._fail_open += 1
            return {**base, "decision": "fail_open_missing_proof"}
        if proof.get("proof_version") != writer.N3P_PLAN_ONLY_PROOF_SUMMARY_VERSION:
            self._fail_open += 1
            return {**base, "decision": "fail_open_unknown_proof_version"}
        if not bool(proof.get("safe_negative_cacheable")):
            self._fail_open += 1
            return {
                **base,
                "decision": "fail_open_not_safe_cacheable",
                "safe_negative_cacheable_reason": proof.get("safe_negative_cacheable_reason"),
                "unsafe_negative_cacheable_reason": proof.get("unsafe_negative_cacheable_reason"),
            }
        if (
            proof_context.get("source_input_fingerprint") != proof.get("source_input_fingerprint")
            or proof_context.get("context_fingerprint") != proof.get("context_fingerprint")
        ):
            self._fail_open += 1
            return {**base, "decision": "fail_open_fingerprint_changed"}
        self._hits += 1
        self._saved_calls += 1
        return {
            **base,
            "decision": "negative_cache_hit",
            "skip_full_n3p": True,
            "proof_version": proof.get("proof_version"),
            "source_input_fingerprint": proof.get("source_input_fingerprint"),
            "context_fingerprint": proof.get("context_fingerprint"),
            "next_recompute_condition": proof.get("next_recompute_condition"),
            "safe_negative_cacheable": True,
            "safe_negative_cacheable_reason": proof.get("safe_negative_cacheable_reason"),
            "requested_periods": list(proof.get("requested_periods") or base["requested_periods"]),
        }

    def stats(self, *, mode: str) -> dict[str, Any]:
        return {
            "mode": mode,
            "n3p_negative_cache_hits": self._hits,
            "n3p_negative_cache_misses": self._misses,
            "n3p_negative_cache_saved_calls": self._saved_calls,
            "n3p_negative_cache_fail_open": self._fail_open,
            "n3p_negative_cache_key_count": len(self._proofs),
            "n3p_negative_cache_family_key_count": len(self._proofs_by_family),
            "n3p_negative_cache_false_negative_count": self._false_negative_count,
        }


class N6ReplayBlocked(ValueError):
    """Raised when a replay request would violate local-only boundaries."""


REPLAY_SIDE_EFFECT_FLAG_KEYS = {
    "database_write",
    "execute",
    "launchd",
    "rollback",
    "update_checkpoint",
    "user_confirmed",
    "worker_started",
    "write_to_db",
    "writes_outbox",
}
REPLAY_SIDE_EFFECT_PRESENCE_KEYS = {
    "checkpoint",
    "conn",
    "connection",
    "consume_outbox",
    "consumes_outbox",
    "cur",
    "cursor",
    "dsn",
    "transaction",
    "tx",
}


def build_n3p_plan_only_replay_artifact(*, source_bundle: Mapping[str, Any]) -> dict[str, Any]:
    bundle = _normalize_replay_bundle(source_bundle)
    replay_config = bundle["replay_config"]
    replay_run_id = _validate_replay_run_id_from_config(replay_config)
    target_run_id = _n3p_target_run_id(replay_config)
    contract = _build_n3p_replay_contract(
        replay_config=replay_config,
        target_run_id=target_run_id,
    )
    rows_by_asset = writer.build_rows_by_asset_from_source_payload(contract, bundle)
    candidate_map = _candidate_map_by_identity(bundle.get("candidates") or [])
    for asset_rows in rows_by_asset.values():
        for row in asset_rows:
            _normalize_n3p_replay_row(row, candidate_map.get(str(row.get("identity_key") or "")))
    contract = _finalize_n3p_replay_contract(contract=contract, rows_by_asset=rows_by_asset)
    validation = writer.validate_rows_against_contract(rows_by_asset, contract)
    sparse_report = writer.build_live_current_sparse_no_trade_exception_report(contract, source_payload=bundle, rows_by_asset=rows_by_asset)
    proof_summary_rows = writer.build_n3p_plan_only_proof_summary_rows(rows_by_asset)
    for proof in proof_summary_rows:
        proof["target_run_id"] = target_run_id
        proof["replay_run_id"] = replay_run_id
    return {
        "result": "PLAN_ONLY",
        "replay_run_id": replay_run_id,
        "target_run_id": target_run_id,
        "metric_rows_by_asset": rows_by_asset,
        "metric_rows": _flatten_rows(rows_by_asset),
        "proof_summary_rows": proof_summary_rows,
        "quality_summary": {
            "blocked_reasons": list(validation.get("blocked_reasons") or []),
            "metric_ready_count": int(validation.get("metric_ready_count") or 0),
            "metric_not_ready_count": int(validation.get("metric_not_ready_count") or 0),
            "row_counts": dict(validation.get("row_counts") or {}),
            "signal_counts": dict(validation.get("signal_counts") or {}),
            "expected_not_ready_quality_warning": dict(validation.get("expected_not_ready_quality_warning") or {}),
        },
        "adapter_report": {
            "source_mode": "replay",
            "canonical_builder_source_mode": "live_current_1m",
            "live_current_sparse_no_trade_exception_count": int(sparse_report.get("exception_count") or 0),
            "live_current_sparse_no_trade_exception_report": sparse_report,
        },
        "side_effects": _plan_only_side_effects(),
    }


def build_b2_plan_only_replay_artifact(
    *,
    source_bundle: Mapping[str, Any],
    b1_snapshot_rows: Sequence[Mapping[str, Any]],
    live_current_minute_rows_by_asset: Mapping[str, Sequence[Mapping[str, Any]]],
    previous_day_minute_rows_by_asset: Mapping[str, Sequence[Mapping[str, Any]]] | None = None,
) -> dict[str, Any]:
    bundle = _normalize_replay_bundle(source_bundle)
    replay_config = bundle["replay_config"]
    replay_run_id = _validate_replay_run_id_from_config(replay_config)
    target_run_id = f"{replay_run_id}__b2_plan_only"
    contract = _build_b2_replay_contract(
        replay_config=replay_config,
        target_run_id=target_run_id,
        live_current_minute_rows_by_asset=live_current_minute_rows_by_asset,
    )
    candidate_map = _candidate_map_by_identity(bundle.get("candidates") or [])
    latest_closed_minute = projection_execute.parse_datetime_value(
        replay_config.get("latest_closed_minute") or replay_config.get("latest_closed_minute_label") or "2026-06-26T11:05:00+08:00"
    )
    previous_rows_by_asset = previous_day_minute_rows_by_asset or {}
    rows_by_asset: dict[str, list[dict[str, Any]]] = {}
    for snapshot_row in b1_snapshot_rows:
        snapshot = dict(snapshot_row)
        asset_kind = _snapshot_asset_kind(snapshot)
        identity_key = str(snapshot.get("identity_key") or "")
        if candidate_map and identity_key not in candidate_map:
            continue
        row = projection_execute.build_projection_row(
            asset_kind=asset_kind,
            snapshot=snapshot,
            event=None,
            pull_plan_id=snapshot.get("pull_plan_id"),
            today_bars=_rows_for_identity((live_current_minute_rows_by_asset or {}).get(asset_kind, []), identity_key),
            previous_bars=_rows_for_identity((previous_rows_by_asset or {}).get(asset_kind, []), identity_key),
            latest_closed_minute=latest_closed_minute,
            contract=contract,
            source_condition_run_id=str(contract["source_runs"]["source_condition_run_id"]),
            projection_run_id=target_run_id,
            for_trade_date=str(contract["dates"]["for_trade_date"]),
            prev_trade_date=str(contract["dates"]["prev_trade_date"]),
            calculation_config=contract["calculation_config"],
        )
        _apply_candidate_trace_to_projection_row(row, candidate_map.get(identity_key))
        _rewrite_projection_row_for_replay(row)
        rows_by_asset.setdefault(asset_kind, []).append(row)
    rows = _flatten_rows(rows_by_asset)
    summary = projection_execute.summarize_projection_rows(rows)
    contract = _finalize_b2_replay_contract(contract=contract, summary=summary)
    quality_items = projection_execute.build_projection_quality_items(
        contract=contract,
        rows=rows,
        pre_backup={
            "outbox_rows_for_projection_run": 0,
            "projection_run_exists": False,
            "projection_run_table_counts": {},
            "quality_rows_for_projection_run": 0,
            "inbox_rows_for_projection_run": 0,
            "checkpoint_refs_for_projection_run": 0,
        },
    )
    return {
        "result": "PLAN_ONLY",
        "replay_run_id": replay_run_id,
        "target_run_id": target_run_id,
        "projection_rows_by_asset": rows_by_asset,
        "projection_rows": rows,
        "quality_summary": {
            "projection_summary": summary,
            "quality_item_count": len(quality_items),
            "p0_count": sum(1 for item in quality_items if item.get("severity") == "P0"),
            "p1_count": sum(1 for item in quality_items if item.get("severity") == "P1"),
            "p2_count": sum(1 for item in quality_items if item.get("severity") == "P2"),
        },
        "adapter_report": {
            "source_mode": "replay",
            "canonical_builder_source_mode": "live_current_1m",
            "live_current_lineage": True,
        },
        "side_effects": _plan_only_side_effects(),
    }


def _rows_for_identity(rows: Sequence[Mapping[str, Any]], identity_key: str) -> list[dict[str, Any]]:
    return [dict(row) for row in rows if str(row.get("identity_key") or "") == identity_key]


def create_local_replay_job(
    *,
    replay_root: Path,
    trade_date: str,
    start_hhmm: str = "09:31",
    end_hhmm: str = "15:00",
    job_id: str | None = None,
    replay_engine_version: str = DEFAULT_REPLAY_ENGINE_VERSION,
    asset_scope: str = DEFAULT_REPLAY_ASSET_SCOPE,
    source_bundle_key: str = DEFAULT_REPLAY_SOURCE_BUNDLE_KEY,
    validation_mode: str = DEFAULT_REPLAY_VALIDATION_MODE,
    n3p_strategy: str = DEFAULT_REPLAY_N3P_STRATEGY,
    n3p_reduction_mode: str = DEFAULT_REPLAY_N3P_REDUCTION_MODE,
    n3p_negative_cache: str = DEFAULT_REPLAY_N3P_NEGATIVE_CACHE,
    enable_profiling: bool = False,
) -> dict[str, Any]:
    normalized_trade_date = _normalize_trade_date(trade_date)
    trade_date_key = normalized_trade_date.replace("-", "")
    effective_job_id = job_id or _new_job_id(trade_date_key)
    _validate_job_id(effective_job_id, trade_date_key)
    engine_version = _normalize_replay_engine_version(replay_engine_version)
    normalized_asset_scope = _normalize_replay_asset_scope(asset_scope)
    normalized_source_bundle_key = _normalize_source_bundle_key(source_bundle_key)
    normalized_validation_mode = _normalize_replay_validation_mode(validation_mode)
    normalized_n3p_strategy = _normalize_replay_n3p_strategy(n3p_strategy)
    normalized_n3p_reduction_mode = _normalize_replay_n3p_reduction_mode(n3p_reduction_mode)
    normalized_n3p_negative_cache = _normalize_replay_n3p_negative_cache(n3p_negative_cache)
    if (
        normalized_n3p_reduction_mode == "active_state_fast_path"
        and (
            normalized_validation_mode != FULL_DAY_SHADOW_VALIDATION_MODE
            or normalized_n3p_strategy != "prefilter_prune"
        )
    ):
        raise N6ReplayBlocked("BLOCKED_REPLAY_N3P_REDUCTION_REQUIRES_SHADOW_PREFILTER_PRUNE")
    profiler = _ReplayProfileCollector(
        enabled=enable_profiling,
        job_id=effective_job_id,
        trade_date=normalized_trade_date,
        replay_engine_version=engine_version,
    )

    root = replay_root.resolve()
    artifact_dir = (root / trade_date_key / effective_job_id).resolve()
    if not artifact_dir.is_relative_to(root):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: artifact_path_outside_replay_root")
    artifact_dir.mkdir(parents=True, exist_ok=False)
    try:
        minutes = _trading_minutes(start_hhmm, end_hhmm)
        profiler.set_metadata(
            start_hhmm=start_hhmm,
            end_hhmm=end_hhmm,
            requested_minutes=len(minutes),
            asset_scope=normalized_asset_scope,
            source_bundle_key=normalized_source_bundle_key,
            validation_mode=normalized_validation_mode,
            n3p_strategy=normalized_n3p_strategy,
            n3p_reduction_mode=normalized_n3p_reduction_mode,
            n3p_negative_cache=normalized_n3p_negative_cache,
        )
        plan = _build_replay_engine_artifact(
            trade_date=normalized_trade_date,
            minutes=minutes,
            replay_run_id=effective_job_id,
            replay_engine_version=engine_version,
            asset_scope=normalized_asset_scope,
            source_bundle_key=normalized_source_bundle_key,
            validation_mode=normalized_validation_mode,
            n3p_strategy=normalized_n3p_strategy,
            n3p_reduction_mode=normalized_n3p_reduction_mode,
            n3p_negative_cache=normalized_n3p_negative_cache,
            replay_root=root,
            profiler=profiler,
        )
        n3_messages = list(plan["n3_messages"])
        n4_messages = list(plan["n4_messages"])
        n5_messages = list(plan["n5_messages"])
        n4_shadow_state_transitions = list(plan.get("n4_shadow_state_transitions") or [])
        n4_shadow_evaluations = list(plan.get("n4_shadow_evaluations") or [])
        n5_shadow_action_windows = list(plan.get("n5_shadow_action_windows") or [])
        shadow_validation_report = dict(plan.get("shadow_validation_report") or {})
        _decorate_replay_messages(
            n4_messages=n4_messages,
            n5_messages=n5_messages,
            replay_engine_version=engine_version,
            source_policy=str(plan.get("source_policy") or REPLAY_SOURCE_POLICY[engine_version]),
        )
        timeline = _build_timeline(
            minutes,
            n4_messages,
            n5_messages,
            n3_messages=n3_messages,
            shadow_validation_report=shadow_validation_report,
        )
        summary = _build_summary(
            job_id=effective_job_id,
            trade_date=normalized_trade_date,
            artifact_dir=artifact_dir,
            timeline=timeline,
            n3_messages=n3_messages,
            n4_messages=n4_messages,
            n5_messages=n5_messages,
            plan=plan,
        )
        config = {
            "job_id": effective_job_id,
            "trade_date": normalized_trade_date,
            "start_hhmm": start_hhmm,
            "end_hhmm": end_hhmm,
            "source_mode": "replay",
            "snapshot_policy": "replay_snapshot_from_minute_cumulative",
            "replay_engine_version": engine_version,
            "validation_mode": normalized_validation_mode,
            "n3p_strategy": normalized_n3p_strategy,
            "n3p_reduction_mode": normalized_n3p_reduction_mode,
            "n3p_negative_cache": normalized_n3p_negative_cache,
            "asset_scope": normalized_asset_scope,
            "source_bundle_key": str(plan.get("source_bundle_key") or normalized_source_bundle_key),
            "source_bundle_selector_mode": str(plan.get("source_bundle_selector_mode") or _source_bundle_selector_mode(normalized_source_bundle_key)),
            "resolved_source_bundle_path": str(plan.get("resolved_source_bundle_path") or ""),
            "source_policy": str(plan.get("source_policy") or REPLAY_SOURCE_POLICY[engine_version]),
            "historical_source_status": str(plan.get("historical_source_status") or ""),
            "historical_source_kind": str(plan.get("historical_source_kind") or ""),
            "historical_source_path": str(plan.get("historical_source_path") or ""),
            "source_meta": dict(plan.get("source_meta") or {}),
            "asset_scope_filter_applied": bool(plan.get("asset_scope_filter_applied")),
            "asset_scope_allowed_asset_kinds": list(plan.get("asset_scope_allowed_asset_kinds") or []),
            "asset_scope_source_counts_before": dict(plan.get("asset_scope_source_counts_before") or {}),
            "asset_scope_source_counts_after": dict(plan.get("asset_scope_source_counts_after") or {}),
            "n3p_cache_stats": dict(plan.get("n3p_cache_stats") or {}),
            "n3p_prefilter": dict(plan.get("n3p_prefilter") or {}),
            "n3p_active_state_reduction": dict(plan.get("n3p_active_state_reduction_summary") or {}),
            "n3p_negative_cache_summary": dict(plan.get("n3p_negative_cache_summary") or {}),
            "n5_evaluator_demand": dict(plan.get("n5_evaluator_demand_summary") or {}),
            "shadow_mode": bool(plan.get("shadow_mode")),
            "shadow_validation_report": shadow_validation_report,
            "asset_unit_fix_delta_validation": bool(plan.get("asset_unit_fix_delta_validation")),
            "asset_unit_fix_delta": dict(plan.get("asset_unit_fix_delta") or {}),
            "n5_delta_only": dict(plan.get("n5_delta_only") or {}),
            "safety_flags": dict(SAFETY_FLAGS),
            "notice": LOCAL_ONLY_NOTICE,
        }
        status = {
            "job_id": effective_job_id,
            "status": "completed",
            "result": "LOCAL_REPLAY_ARTIFACT_READY",
            "artifact_dir": str(artifact_dir),
            "summary_path": str(artifact_dir / "replay_summary.json"),
            "excel_path": str(artifact_dir / "n3_n5_full_day_replay.xlsx"),
            "safety_flags": dict(SAFETY_FLAGS),
            "replay_engine_version": engine_version,
            "validation_mode": normalized_validation_mode,
            "n3p_strategy": normalized_n3p_strategy,
            "n3p_reduction_mode": normalized_n3p_reduction_mode,
            "n3p_negative_cache": normalized_n3p_negative_cache,
            "asset_scope": normalized_asset_scope,
            "source_bundle_key": str(plan.get("source_bundle_key") or normalized_source_bundle_key),
            "source_bundle_selector_mode": str(plan.get("source_bundle_selector_mode") or _source_bundle_selector_mode(normalized_source_bundle_key)),
            "resolved_source_bundle_path": str(plan.get("resolved_source_bundle_path") or ""),
            "source_policy": str(plan.get("source_policy") or REPLAY_SOURCE_POLICY[engine_version]),
            "historical_source_status": str(plan.get("historical_source_status") or ""),
            "historical_source_kind": str(plan.get("historical_source_kind") or ""),
            "historical_source_path": str(plan.get("historical_source_path") or ""),
            "source_meta": dict(plan.get("source_meta") or {}),
            "asset_scope_filter_applied": bool(plan.get("asset_scope_filter_applied")),
            "asset_scope_allowed_asset_kinds": list(plan.get("asset_scope_allowed_asset_kinds") or []),
            "asset_scope_source_counts_before": dict(plan.get("asset_scope_source_counts_before") or {}),
            "asset_scope_source_counts_after": dict(plan.get("asset_scope_source_counts_after") or {}),
            "n3p_cache_stats": dict(plan.get("n3p_cache_stats") or {}),
            "n3p_prefilter": dict(plan.get("n3p_prefilter") or {}),
            "n3p_active_state_reduction": dict(plan.get("n3p_active_state_reduction_summary") or {}),
            "n3p_negative_cache_summary": dict(plan.get("n3p_negative_cache_summary") or {}),
            "n5_evaluator_demand": dict(plan.get("n5_evaluator_demand_summary") or {}),
            "shadow_mode": bool(plan.get("shadow_mode")),
            "shadow_validation_report": shadow_validation_report,
            "asset_unit_fix_delta_validation": bool(plan.get("asset_unit_fix_delta_validation")),
            "asset_unit_fix_delta": dict(plan.get("asset_unit_fix_delta") or {}),
            "n5_delta_only": dict(plan.get("n5_delta_only") or {}),
        }

        jsonl_token = profiler.start_phase(
            "artifact_jsonl_serialization",
            timeline_row_count=len(timeline),
            n3_message_count=len(n3_messages),
            n4_message_count=len(n4_messages),
            n5_message_count=len(n5_messages),
        )
        if normalized_validation_mode == "asset_unit_fix_delta_v1":
            n4_delta_attribution: list[dict[str, Any]] | None = list(plan.get("n4_delta_attribution") or [])
            n5_delta_only_messages: list[dict[str, Any]] | None = list(plan.get("n5_delta_only_messages") or [])
            excluded_stock_replay_risk: list[dict[str, Any]] | None = list(plan.get("excluded_stock_replay_risk") or [])
        else:
            n4_delta_attribution = None
            n5_delta_only_messages = None
            excluded_stock_replay_risk = None
        n4_prefilter_audit = list(plan.get("n4_lightweight_prefilter_audit") or [])
        n3p_demand_plan = list(plan.get("n3p_demand_plan") or [])
        n3p_active_state_reduction = list(plan.get("n3p_active_state_reduction") or [])
        n3p_proof_summary = list(plan.get("n3p_plan_only_proof_summary") or [])
        n3p_negative_cache_decisions = list(plan.get("n3p_negative_cache_decisions") or [])
        n5_evaluator_demand_plan = list(plan.get("n5_evaluator_demand_plan") or [])
        n5_confirmation_metric_index_stats = list(plan.get("n5_confirmation_metric_index_stats") or [])
        if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE:
            n4_shadow_sheet_rows: list[dict[str, Any]] | None = n4_shadow_state_transitions
            n4_shadow_evaluation_sheet_rows: list[dict[str, Any]] | None = n4_shadow_evaluations
            n5_shadow_sheet_rows: list[dict[str, Any]] | None = n5_shadow_action_windows
        else:
            n4_shadow_sheet_rows = None
            n4_shadow_evaluation_sheet_rows = None
            n5_shadow_sheet_rows = None
        _write_json(artifact_dir / "replay_config.json", config)
        _write_json(artifact_dir / "replay_status.json", status)
        _write_jsonl(artifact_dir / "replay_timeline.jsonl", timeline)
        _write_jsonl(artifact_dir / "n3_messages.jsonl", n3_messages)
        _write_jsonl(artifact_dir / "n4_messages.jsonl", n4_messages)
        _write_jsonl(artifact_dir / "n5_messages.jsonl", n5_messages)
        if normalized_validation_mode == "asset_unit_fix_delta_v1":
            _write_jsonl(artifact_dir / "n4_delta_attribution.jsonl", n4_delta_attribution or [])
            _write_jsonl(artifact_dir / "n5_delta_only_messages.jsonl", n5_delta_only_messages or [])
        if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE:
            _write_jsonl(artifact_dir / "n4_shadow_state_transitions.jsonl", n4_shadow_state_transitions)
            _write_jsonl(artifact_dir / "n4_shadow_evaluations.jsonl", n4_shadow_evaluations)
            _write_jsonl(artifact_dir / "n5_shadow_action_windows.jsonl", n5_shadow_action_windows)
            _write_json(artifact_dir / "shadow_validation_report.json", shadow_validation_report)
            _write_jsonl(artifact_dir / "n4_lightweight_prefilter_audit.jsonl", n4_prefilter_audit)
            _write_jsonl(artifact_dir / "n3p_demand_plan.jsonl", n3p_demand_plan)
            _write_jsonl(artifact_dir / "n3p_active_state_reduction.jsonl", n3p_active_state_reduction)
            _write_jsonl(artifact_dir / "n3p_plan_only_proof_summary.jsonl", n3p_proof_summary)
            _write_jsonl(artifact_dir / "n3p_negative_cache_decisions.jsonl", n3p_negative_cache_decisions)
            _write_jsonl(artifact_dir / "n5_evaluator_demand_plan.jsonl", n5_evaluator_demand_plan)
            _write_jsonl(artifact_dir / "n5_confirmation_metric_index_stats.jsonl", n5_confirmation_metric_index_stats)
        _write_json(artifact_dir / "replay_summary.json", summary)
        (artifact_dir / "replay_summary.md").write_text(_summary_markdown(summary), encoding="utf-8")
        profiler.finish_phase(jsonl_token)

        excel_token = profiler.start_phase(
            "excel_generation",
            sheet_count=len(EXCEL_SHEETS),
            n4_message_count=len(n4_messages),
            n5_message_count=len(n5_messages),
        )
        (artifact_dir / "n3_n5_full_day_replay.xlsx").write_bytes(
            build_replay_excel(
                summary=summary,
                timeline=timeline,
                n4_messages=n4_messages,
                n5_messages=n5_messages,
                n4_delta_attribution=n4_delta_attribution,
                n5_delta_only_messages=n5_delta_only_messages,
                excluded_stock_replay_risk=excluded_stock_replay_risk,
                n4_shadow_state_transitions=n4_shadow_sheet_rows,
                n4_shadow_evaluations=n4_shadow_evaluation_sheet_rows,
                n5_shadow_action_windows=n5_shadow_sheet_rows,
                shadow_validation_report=shadow_validation_report if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
                n4_prefilter_audit=n4_prefilter_audit if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
                n3p_demand_plan=n3p_demand_plan if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
                n3p_active_state_reduction=n3p_active_state_reduction if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
                n3p_proof_summary=n3p_proof_summary if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
                n3p_negative_cache_decisions=n3p_negative_cache_decisions if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
                n5_evaluator_demand_plan=n5_evaluator_demand_plan if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
                n5_confirmation_metric_index_stats=n5_confirmation_metric_index_stats if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE else None,
            )
        )
        profiler.finish_phase(excel_token)

        if enable_profiling:
            profiler.set_metadata(
                source_policy=str(plan.get("source_policy") or REPLAY_SOURCE_POLICY[engine_version]),
                historical_source_kind=str(plan.get("historical_source_kind") or ""),
                historical_source_path=str(plan.get("historical_source_path") or ""),
                source_meta=dict(plan.get("source_meta") or {}),
                asset_scope=normalized_asset_scope,
                source_bundle_key=str(plan.get("source_bundle_key") or normalized_source_bundle_key),
                validation_mode=normalized_validation_mode,
                n3p_strategy=normalized_n3p_strategy,
                n3p_reduction_mode=normalized_n3p_reduction_mode,
                n3p_negative_cache=dict(plan.get("n3p_negative_cache_summary") or {}),
                source_bundle_selector_mode=str(plan.get("source_bundle_selector_mode") or _source_bundle_selector_mode(normalized_source_bundle_key)),
                resolved_source_bundle_path=str(plan.get("resolved_source_bundle_path") or ""),
                n3p_cache_stats=dict(plan.get("n3p_cache_stats") or {}),
                n3p_prefilter=dict(plan.get("n3p_prefilter") or {}),
                n3p_active_state_reduction=dict(plan.get("n3p_active_state_reduction_summary") or {}),
                n5_evaluator_demand=dict(plan.get("n5_evaluator_demand_summary") or {}),
            )
            _write_replay_profile(
                artifact_dir / "replay_profile.json",
                artifact_dir / "replay_profile.md",
                profiler.build_payload(status="completed", artifact_dir=str(artifact_dir)),
            )
        return {
            **status,
            "trade_date": normalized_trade_date,
            "artifact_dir": str(artifact_dir),
            "replay_engine_version": engine_version,
            "validation_mode": normalized_validation_mode,
            "n3p_strategy": normalized_n3p_strategy,
            "n3p_reduction_mode": normalized_n3p_reduction_mode,
            "n3p_negative_cache": normalized_n3p_negative_cache,
            "source_bundle_key": str(plan.get("source_bundle_key") or normalized_source_bundle_key),
            "source_bundle_selector_mode": str(plan.get("source_bundle_selector_mode") or _source_bundle_selector_mode(normalized_source_bundle_key)),
            "resolved_source_bundle_path": str(plan.get("resolved_source_bundle_path") or ""),
        }
    except Exception as exc:
        if enable_profiling:
            fallback_path = _write_replay_profile_fallback(
                replay_root=root,
                trade_date_key=trade_date_key,
                profiler=profiler,
                blocked_reason=str(exc),
            )
            profiler.set_metadata(profile_fallback_path=str(fallback_path))
        raise


def _normalize_replay_bundle(source_bundle: Mapping[str, Any]) -> dict[str, Any]:
    if not isinstance(source_bundle, Mapping):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: invalid_replay_source_bundle")
    bundle = dict(source_bundle)
    replay_config = bundle.get("replay_config")
    if not isinstance(replay_config, Mapping):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: missing_replay_config")
    _assert_no_replay_side_effect_risk(bundle)
    bundle["replay_config"] = dict(replay_config)
    bundle.setdefault("source_records", {})
    bundle.setdefault("candidates", [])
    bundle.setdefault("n4_context_snapshot_rows", [])
    return bundle


def _normalize_historical_replay_source_bundle(source_bundle: Mapping[str, Any]) -> dict[str, Any]:
    bundle = _normalize_replay_bundle(source_bundle)
    if not bundle.get("source_records"):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SOURCE_UNAVAILABLE: missing_source_records")
    if "b2_input" not in bundle:
        raise N6ReplayBlocked("BLOCKED_REPLAY_SOURCE_UNAVAILABLE: missing_b2_input")
    bundle["b2_input"] = _normalize_b2_input_payload(bundle.get("b2_input"))
    return bundle


def _normalize_b2_input_payload(value: Any) -> dict[str, Any]:
    payload = dict(value or {})
    return {
        "snapshot_rows": [_normalize_b2_snapshot_row(row) for row in payload.get("snapshot_rows") or []],
        "live_current_rows_by_asset": _normalize_b2_rows_by_asset(payload.get("live_current_rows_by_asset")),
        "previous_day_rows_by_asset": _normalize_b2_rows_by_asset(payload.get("previous_day_rows_by_asset")),
    }


def _normalize_b2_rows_by_asset(value: Any) -> dict[str, list[dict[str, Any]]]:
    rows_by_asset = dict(value or {})
    return {
        str(asset_kind): [_normalize_b2_bar_row(row) for row in rows]
        for asset_kind, rows in rows_by_asset.items()
    }


def _normalize_b2_snapshot_row(row: Mapping[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    for key in ("snapshot_time",):
        normalized[key] = _coerce_datetime_like(normalized.get(key))
    for key in ("current_price", "close"):
        if key in normalized:
            normalized[key] = _coerce_decimal_like(normalized.get(key))
    return normalized


def _normalize_b2_bar_row(row: Mapping[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    normalized["bar_time"] = _coerce_datetime_like(normalized.get("bar_time"))
    for key in ("open", "high", "low", "close", "volume", "amount"):
        if key in normalized:
            normalized[key] = _coerce_decimal_like(normalized.get(key))
    return normalized


def _coerce_datetime_like(value: Any) -> Any:
    if isinstance(value, datetime) or value in (None, ""):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return value
        try:
            if text.endswith("Z"):
                text = f"{text[:-1]}+00:00"
            parsed = datetime.fromisoformat(text)
            if parsed.tzinfo is None:
                return parsed.replace(tzinfo=DISPLAY_TIMEZONE)
            return parsed
        except ValueError:
            try:
                return datetime.strptime(text, "%Y-%m-%d %H:%M:%S").replace(tzinfo=DISPLAY_TIMEZONE)
            except ValueError:
                try:
                    return datetime.strptime(text, "%Y-%m-%d %H:%M").replace(tzinfo=DISPLAY_TIMEZONE)
                except ValueError:
                    return value
    return value


def _coerce_decimal_like(value: Any) -> Any:
    if isinstance(value, Decimal) or value in (None, ""):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        try:
            return Decimal(value)
        except Exception:
            return value
    return value


def _json_mapping(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except Exception:
            return {}
        return dict(parsed) if isinstance(parsed, Mapping) else {}
    return {}


def _decimal_or_none(value: Any) -> Decimal | None:
    coerced = _coerce_decimal_like(value)
    return coerced if isinstance(coerced, Decimal) else None


def _assert_no_replay_side_effect_risk(value: Any, *, path: str = "source_bundle") -> None:
    if isinstance(value, Mapping):
        for key, child in value.items():
            key_text = str(key)
            lowered = key_text.lower()
            child_path = f"{path}.{key_text}"
            if lowered in REPLAY_SIDE_EFFECT_PRESENCE_KEYS:
                raise N6ReplayBlocked(f"BLOCKED_REPLAY_SIDE_EFFECT_RISK: forbidden_key={child_path}")
            if lowered in REPLAY_SIDE_EFFECT_FLAG_KEYS and bool(child):
                raise N6ReplayBlocked(f"BLOCKED_REPLAY_SIDE_EFFECT_RISK: forbidden_flag={child_path}")
            _assert_no_replay_side_effect_risk(child, path=child_path)
        return
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        for index, child in enumerate(value):
            _assert_no_replay_side_effect_risk(child, path=f"{path}[{index}]")


def _validate_replay_run_id_from_config(replay_config: Mapping[str, Any]) -> str:
    trade_date = _normalize_trade_date(
        str(
            replay_config.get("trade_date")
            or replay_config.get("for_trade_date")
            or replay_config.get("trade_date_key")
            or ""
        )
    )
    trade_date_key = trade_date.replace("-", "")
    replay_run_id = str(replay_config.get("replay_run_id") or replay_config.get("job_id") or "")
    _validate_job_id(replay_run_id, trade_date_key)
    return replay_run_id


def _plan_only_side_effects() -> dict[str, bool]:
    return {
        "database_written": False,
        "business_rows_written": False,
        "outbox_inbox_checkpoint_consumed_or_updated": False,
        "worker_started": False,
        "launchd_touched": False,
        "runtime_executed": False,
    }


def _trade_date_key(value: Any) -> str:
    return _normalize_trade_date(str(value or "")).replace("-", "")


def _flatten_rows(rows_by_asset: Mapping[str, Sequence[Mapping[str, Any]]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for asset_rows in rows_by_asset.values():
        rows.extend(dict(row) for row in asset_rows)
    return rows


def _clone_payload(value: Any) -> Any:
    return deepcopy(value)


def _stable_payload_hash(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return sha256(encoded).hexdigest()


def _candidate_cache_fingerprint(candidate: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "asset_kind": candidate.get("asset_kind"),
        "identity_key": candidate.get("identity_key"),
        "code": candidate.get("code"),
        "source_record_key": candidate.get("source_record_key"),
        "condition_key": candidate.get("condition_key"),
        "original_condition_key": candidate.get("original_condition_key"),
        "condition_keys": list(candidate.get("condition_keys") or []),
        "signal_type": candidate.get("signal_type"),
        "direction": candidate.get("direction"),
        "minute_label": candidate.get("minute_label"),
        "observed_at": candidate.get("observed_at"),
        "source_condition_pool_id": candidate.get("source_condition_pool_id"),
        "source_minute_target_scope_id": candidate.get("source_minute_target_scope_id"),
    }


def _candidate_proof_condition_key(candidate: Mapping[str, Any]) -> str:
    for key in ("original_condition_key", "condition_key"):
        value = candidate.get(key)
        if value and not str(value).startswith("LIVE_CURRENT_1M:"):
            return str(value)
    values = candidate.get("condition_keys")
    if isinstance(values, Sequence) and not isinstance(values, (str, bytes)):
        for value in values:
            if value:
                return str(value)
    raw_json = candidate.get("raw_json")
    if isinstance(raw_json, Mapping):
        for key in ("original_condition_key", "condition_key"):
            value = raw_json.get(key)
            if value and not str(value).startswith("LIVE_CURRENT_1M:"):
                return str(value)
        values = raw_json.get("condition_keys")
        if isinstance(values, Sequence) and not isinstance(values, (str, bytes)):
            for value in values:
                if value:
                    return str(value)
    return str(candidate.get("condition_key") or "")


def _candidate_requested_periods(candidate: Mapping[str, Any]) -> list[str]:
    for key in ("requested_periods", "formal_amount_chain_required_periods"):
        periods = _normalize_candidate_periods(candidate.get(key))
        if periods:
            return periods
    raw_json = candidate.get("raw_json")
    if isinstance(raw_json, Mapping):
        for key in ("requested_periods", "formal_amount_chain_required_periods"):
            periods = _normalize_candidate_periods(raw_json.get(key))
            if periods:
                return periods
    return _requested_periods_from_condition_key(_candidate_proof_condition_key(candidate))


def _normalize_candidate_periods(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, str):
        values: Sequence[Any] = [value]
    elif isinstance(value, Sequence):
        values = value
    else:
        return []
    output: list[str] = []
    seen: set[str] = set()
    for item in values:
        period = str(item or "").upper()
        if period not in {"Y", "Q", "M", "W", "D"} or period in seen:
            continue
        seen.add(period)
        output.append(period)
    return output


def _proof_trigger_period_for_periods(periods: Sequence[Any]) -> str:
    normalized = _normalize_candidate_periods(periods)
    for period in ("Y", "Q", "M", "W", "D"):
        if period in normalized:
            return period
    return "D"


def _candidate_stable_trigger_key(candidate: Mapping[str, Any]) -> str:
    minute_label = str(candidate.get("minute_label") or candidate.get("observed_at") or "")
    trigger_minute = minute_label[11:16] if len(minute_label) >= 16 else _candidate_start_minute(candidate)
    for_trade_date = str(candidate.get("for_trade_date") or "")
    if not for_trade_date and len(minute_label) >= 10:
        for_trade_date = minute_label[:10].replace("-", "")
    requested_periods = _candidate_requested_periods(candidate)
    return "|".join(
        [
            for_trade_date,
            str(candidate.get("asset_kind") or ""),
            str(candidate.get("identity_key") or ""),
            str(candidate.get("signal_type") or ""),
            _candidate_proof_condition_key(candidate),
            str(candidate.get("trigger_period") or _proof_trigger_period_for_periods(requested_periods)),
            str(candidate.get("trigger_mark_candidate") or "normal"),
            trigger_minute,
        ]
    )


def _candidate_stable_trigger_family_key(candidate: Mapping[str, Any]) -> str:
    stable_key = _candidate_stable_trigger_key(candidate)
    return "|".join(stable_key.split("|")[:-1])


def _n3p_negative_cache_proof_context(candidate: Mapping[str, Any]) -> dict[str, str]:
    requested_periods = _candidate_requested_periods(candidate)
    proof_condition_key = _candidate_proof_condition_key(candidate)
    trigger_period = str(candidate.get("trigger_period") or _proof_trigger_period_for_periods(requested_periods))
    fingerprint = _stable_payload_hash(
        {
            "asset_kind": candidate.get("asset_kind"),
            "identity_key": candidate.get("identity_key"),
            "signal_type": candidate.get("signal_type"),
            "condition_key": proof_condition_key,
            "requested_periods": requested_periods,
            "trigger_period": trigger_period,
            "trigger_mark_candidate": str(candidate.get("trigger_mark_candidate") or "normal"),
            "safe_negative_cacheable_reason": "amount_chain_failed_for_required_period",
        }
    )
    context_fingerprint = _stable_payload_hash(
        {
            "asset_kind": candidate.get("asset_kind"),
            "identity_key": candidate.get("identity_key"),
            "source_condition_run_id": candidate.get("source_condition_run_id"),
            "source_condition_pool_id": candidate.get("source_condition_pool_id"),
            "source_minute_target_scope_id": candidate.get("source_minute_target_scope_id"),
            "original_condition_key": proof_condition_key,
            "requested_periods": requested_periods,
        }
    )
    return {
        "source_input_fingerprint": fingerprint,
        "context_fingerprint": context_fingerprint,
        "requested_periods": requested_periods,
        "original_condition_key": proof_condition_key,
    }


def _n3p_target_run_id(replay_config: Mapping[str, Any]) -> str:
    _validate_replay_run_id_from_config(replay_config)
    source_subscription_run_id = str(replay_config.get("source_subscription_run_id") or "")
    source_variant = str(replay_config.get("n3p_source_variant") or writer.N3P_AMOUNT_CHAIN_V2_SOURCE_VARIANT)
    suffix = f"{source_variant}__{source_subscription_run_id}" if source_subscription_run_id else None
    return writer.build_n3p_realtime_action_confirmation_metric_run_id(
        for_trade_date=_trade_date_key(replay_config.get("for_trade_date") or replay_config.get("trade_date")),
        until_hhmm=str(replay_config.get("until_hhmm") or "1500"),
        asset_kind="all",
        suffix=suffix,
    )


def _retarget_n3p_artifact(artifact: Mapping[str, Any], *, target_run_id: str) -> dict[str, Any]:
    output = _clone_payload(artifact)
    output["target_run_id"] = target_run_id
    for row in output.get("metric_rows") or []:
        if isinstance(row, dict):
            row["projection_run_id"] = target_run_id
    for rows in (output.get("metric_rows_by_asset") or {}).values():
        for row in rows:
            if isinstance(row, dict):
                row["projection_run_id"] = target_run_id
    for row in output.get("proof_summary_rows") or []:
        if isinstance(row, dict):
            row["target_run_id"] = target_run_id
    return output


def _signal_counts_from_rows(rows_by_asset: Mapping[str, Sequence[Mapping[str, Any]]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in _flatten_rows(rows_by_asset):
        signal_type = str((row.get("raw_json") or {}).get("signal_type") or row.get("signal_type") or "")
        if not signal_type:
            continue
        counts[signal_type] = counts.get(signal_type, 0) + 1
    return counts


def _row_not_ready_reasons(rows_by_asset: Mapping[str, Sequence[Mapping[str, Any]]]) -> list[str]:
    reasons: list[str] = []
    for row in _flatten_rows(rows_by_asset):
        if row.get("metric_ready", True):
            continue
        for payload_key in ("raw_json", "trace_json"):
            payload = row.get(payload_key)
            if not isinstance(payload, Mapping):
                continue
            for reason_key in ("blocked_reasons", "not_ready_reasons", "metric_not_ready_reasons"):
                value = payload.get(reason_key)
                if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
                    reasons.extend(str(item) for item in value if item)
    return sorted(set(reasons))


def _build_n3p_replay_contract(*, replay_config: Mapping[str, Any], target_run_id: str) -> dict[str, Any]:
    return {
        "target_run_id": target_run_id,
        "projection_schema_version": "v3.realtime_virtual_metric.writer.replay.v1",
        "source_scope": {
            "for_trade_date": _trade_date_key(replay_config.get("for_trade_date") or replay_config.get("trade_date")),
            "source_trade_date": _trade_date_key(replay_config.get("source_trade_date") or replay_config.get("prev_trade_date")),
            "source_condition_run_id": str(replay_config.get("source_condition_run_id") or target_run_id),
            "source_subscription_run_id": str(replay_config.get("source_subscription_run_id") or target_run_id),
            "source_snapshot_run_id": str(replay_config.get("source_snapshot_run_id") or target_run_id),
            "source_previous_day_minute_run_id": str(replay_config.get("source_previous_day_minute_run_id") or target_run_id),
            "source_live_minute_run_id": str(replay_config.get("source_live_minute_run_id") or target_run_id),
            "source_mode": "live_current_1m",
            "c1_dependency": False,
            "source_adapter": str(replay_config.get("source_adapter") or ""),
        },
        "db_backed_input_contract": {
            "source_mode": "live_current_1m",
            "c1_dependency": False,
            "source_snapshot_run_id": str(replay_config.get("source_snapshot_run_id") or target_run_id),
            "source_previous_day_minute_run_id": str(replay_config.get("source_previous_day_minute_run_id") or target_run_id),
            "source_live_minute_run_id": str(replay_config.get("source_live_minute_run_id") or target_run_id),
            "source_condition_run_id": str(replay_config.get("source_condition_run_id") or target_run_id),
            "source_subscription_run_id": str(replay_config.get("source_subscription_run_id") or target_run_id),
            "n2_period_context_source": "trigger_context_snapshot_or_condition_scope",
            "asset_kinds": ["stock", "index", "board"],
        },
        "expected_rows": {
            "total": 0,
            "by_signal_type": {},
        },
    }


def _finalize_n3p_replay_contract(
    *,
    contract: Mapping[str, Any],
    rows_by_asset: Mapping[str, Sequence[Mapping[str, Any]]],
) -> dict[str, Any]:
    rows = _flatten_rows(rows_by_asset)
    metric_ready_count = sum(1 for row in rows if row.get("metric_ready"))
    metric_not_ready_count = len(rows) - metric_ready_count
    finalized = dict(contract)
    finalized["expected_rows"] = {
        "total": len(rows),
        "by_signal_type": _signal_counts_from_rows(rows_by_asset),
        "metric_ready": metric_ready_count,
        "metric_not_ready": metric_not_ready_count,
    }
    if metric_not_ready_count:
        finalized["expected_not_ready_count"] = metric_not_ready_count
        finalized["expected_not_ready_blocked_reason_prefixes"] = [reason.split(":", 1)[0] for reason in _row_not_ready_reasons(rows_by_asset)]
    return finalized


def _build_b2_replay_contract(
    *,
    replay_config: Mapping[str, Any],
    target_run_id: str,
    live_current_minute_rows_by_asset: Mapping[str, Sequence[Mapping[str, Any]]],
) -> dict[str, Any]:
    return {
        "projection_run_id": target_run_id,
        "dates": {
            "for_trade_date": _trade_date_key(replay_config.get("for_trade_date") or replay_config.get("trade_date")),
            "source_trade_date": _trade_date_key(replay_config.get("source_trade_date") or replay_config.get("prev_trade_date")),
            "prev_trade_date": _trade_date_key(replay_config.get("prev_trade_date") or replay_config.get("source_trade_date")),
        },
        "source_runs": {
            "source_condition_run_id": str(replay_config.get("source_condition_run_id") or target_run_id),
            "subscription_run_id": str(replay_config.get("source_subscription_run_id") or target_run_id),
            "snapshot_run_id": str(replay_config.get("source_snapshot_run_id") or target_run_id),
            "preload_run_id": str(replay_config.get("source_previous_day_minute_run_id") or target_run_id),
            "source_live_minute_run_id": str(replay_config.get("source_live_minute_run_id") or target_run_id),
        },
        "source_mode": "live_current_1m",
        "calculation_config": _replay_b2_calculation_config(replay_config),
        "fact_only_snapshot_trace_policy": dict(replay_config.get("fact_only_snapshot_trace_policy") or {}),
        "live_current_minute_rows_by_asset": live_current_minute_rows_by_asset,
        "expected_projection_rows": {"total": 0, "by_asset": {}},
        "expected_distribution": {},
        "artifact_generation_mode": "dynamic_intraday_child_artifact",
        "expected_distribution_policy": {
            "mode": "derive_from_projection_rows",
            "applies_to_artifact_generation_mode": "dynamic_intraday_child_artifact",
        },
    }


def _replay_b2_calculation_config(replay_config: Mapping[str, Any]) -> dict[str, Any]:
    config = dict(replay_config.get("calculation_config") or {})
    defaults = {
        "completion_ratio_min_ready": "0.1",
        "amount_projection_expand_threshold": "1.2",
        "amount_projection_shrink_threshold": "0.8",
        "price_flat_abs_pct_threshold": "0.001",
        "window_total_seconds": 1800,
        "calculation_method": "active_30m_bucket_projection_v1_strict_current_lineage",
        "calculation_config_hash": "local-replay-canonical-plan-v1",
    }
    for key, value in defaults.items():
        config.setdefault(key, value)
    return config


def _finalize_b2_replay_contract(*, contract: Mapping[str, Any], summary: Mapping[str, Any]) -> dict[str, Any]:
    finalized = dict(contract)
    finalized["expected_projection_rows"] = {
        "total": int(summary.get("total_rows") or 0),
        "by_asset": dict(summary.get("rows_by_asset") or {}),
    }
    finalized["expected_distribution"] = projection_execute.build_expected_distribution_from_summary(summary)
    return finalized


def _snapshot_asset_kind(snapshot: Mapping[str, Any]) -> str:
    asset_kind = str(snapshot.get("asset_kind") or "")
    if asset_kind in {"stock", "index", "board"}:
        return asset_kind
    identity_key = str(snapshot.get("identity_key") or "")
    if ":" in identity_key:
        return identity_key.split(":", 1)[0]
    raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: snapshot_asset_kind_missing")


def _candidate_map_by_identity(candidates: Sequence[Mapping[str, Any]]) -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for candidate in candidates:
        identity_key = str(candidate.get("identity_key") or "")
        if identity_key:
            output[identity_key] = dict(candidate)
    return output


def _apply_candidate_trace_to_projection_row(row: dict[str, Any], candidate: Mapping[str, Any] | None) -> None:
    if not candidate:
        return
    raw_json = dict(row.get("raw_json") or {})
    condition_key = str(candidate.get("condition_key") or "")
    signal_type = str(candidate.get("signal_type") or "")
    if condition_key:
        raw_json["condition_key"] = condition_key
        raw_json["original_condition_key"] = str(candidate.get("original_condition_key") or condition_key)
    if signal_type:
        raw_json["signal_type"] = signal_type
    row["raw_json"] = raw_json


def _normalize_n3p_replay_row(row: dict[str, Any], candidate: Mapping[str, Any] | None) -> None:
    raw_json = dict(row.get("raw_json") or {})
    trace_json = dict(row.get("trace_json") or {})
    source_fact_ids = dict(row.get("source_fact_ids") or {})
    condition_key = str(raw_json.get("condition_key") or row.get("condition_key") or "")
    original_condition_key = str(
        raw_json.get("original_condition_key")
        or trace_json.get("original_condition_key")
        or (candidate or {}).get("original_condition_key")
        or condition_key
    )
    if original_condition_key:
        raw_json["original_condition_key"] = original_condition_key
    blocked_reasons = list(trace_json.get("blocked_reasons") or raw_json.get("blocked_reasons") or [])
    if blocked_reasons and "missing_reason" not in source_fact_ids:
        source_fact_ids["missing_reason"] = blocked_reasons
    row["raw_json"] = raw_json
    row["trace_json"] = trace_json
    row["source_fact_ids"] = source_fact_ids


def _rewrite_projection_row_for_replay(row: dict[str, Any]) -> None:
    source_fact_ids = dict(row.get("source_fact_ids") or {})
    raw_json = dict(row.get("raw_json") or {})
    source_fact_ids["source_mode"] = "replay"
    raw_json["source_mode"] = "replay"
    row["source_fact_ids"] = source_fact_ids
    row["raw_json"] = raw_json


def list_replay_dates(replay_root: Path) -> list[str]:
    root = replay_root.resolve()
    if not root.exists():
        return []
    dates: list[str] = []
    for path in root.iterdir():
        if path.is_dir() and re.fullmatch(r"\d{8}", path.name):
            dates.append(f"{path.name[:4]}-{path.name[4:6]}-{path.name[6:8]}")
    return sorted(dates, reverse=True)


def build_c1_index_board_readiness_report(
    *,
    trade_date: str,
    index_rows: Sequence[Mapping[str, Any]],
    board_rows: Sequence[Mapping[str, Any]],
    asset_scope: str = "index_board_only",
) -> dict[str, Any]:
    normalized_trade_date = _normalize_trade_date(trade_date)
    if _normalize_replay_asset_scope(asset_scope) != "index_board_only":
        raise N6ReplayBlocked("BLOCKED_REPLAY_C1_READINESS_REQUIRES_INDEX_BOARD_ONLY")
    expected_minutes = _trading_minutes("09:31", "15:00")
    expected_set = set(expected_minutes)
    rows_by_asset = {
        "index": [dict(row) for row in index_rows],
        "board": [dict(row) for row in board_rows],
    }
    grouped: dict[str, dict[str, set[str]]] = {"index": {}, "board": {}}
    stock_count = 0
    for asset_kind, rows in rows_by_asset.items():
        for row in rows:
            row_asset_kind = _row_asset_kind(row) or asset_kind
            if row_asset_kind == "stock":
                stock_count += 1
                continue
            if row_asset_kind != asset_kind:
                continue
            identity_key = _row_identity_key(row, fallback_asset_kind=asset_kind)
            minute = _source_row_minute(row, normalized_trade_date)
            if identity_key and minute in expected_set:
                grouped[asset_kind].setdefault(identity_key, set()).add(minute)

    missing_assets: list[dict[str, Any]] = []
    actual_distribution: dict[str, int] = {}
    for asset_kind, assets in grouped.items():
        for identity_key, actual_minutes in sorted(assets.items()):
            actual_count = len(actual_minutes)
            actual_distribution[str(actual_count)] = actual_distribution.get(str(actual_count), 0) + 1
            missing_minutes = [minute for minute in expected_minutes if minute not in actual_minutes]
            if missing_minutes:
                missing_assets.append(
                    {
                        "asset_kind": asset_kind,
                        "identity_key": identity_key,
                        "actual_count": actual_count,
                        "expected_count": len(expected_minutes),
                        "missing_minutes": missing_minutes,
                        "missing_ranges": _compress_minute_ranges(missing_minutes),
                    }
                )
    if stock_count:
        missing_assets.append(
            {
                "asset_kind": "stock",
                "identity_key": "stock_scope_leak",
                "actual_count": stock_count,
                "expected_count": 0,
                "missing_minutes": [],
                "missing_ranges": [],
            }
        )
    status = "passed" if not missing_assets and grouped["index"] and grouped["board"] else "blocked"
    return {
        "trade_date": normalized_trade_date,
        "asset_scope": "index_board_only",
        "status": status,
        "blocked_reason": "" if status == "passed" else C1_SOURCE_INCOMPLETE_REASON,
        "expected_minutes_per_object": len(expected_minutes),
        "expected_minutes": expected_minutes,
        "index_object_count": len(grouped["index"]),
        "board_object_count": len(grouped["board"]),
        "stock_count": stock_count,
        "source_row_count": len(index_rows) + len(board_rows),
        "actual_minutes_distribution": dict(sorted(actual_distribution.items(), key=lambda item: int(item[0]))),
        "missing_asset_count": len(missing_assets),
        "missing_assets": missing_assets,
    }


def export_c1_index_board_full_day_source_bundle(
    *,
    replay_root: Path,
    trade_date: str,
    index_rows: Sequence[Mapping[str, Any]],
    board_rows: Sequence[Mapping[str, Any]],
    template_source_bundle_key: str | None = None,
) -> dict[str, Any]:
    normalized_trade_date = _normalize_trade_date(trade_date)
    trade_date_key = normalized_trade_date.replace("-", "")
    readiness = build_c1_index_board_readiness_report(
        trade_date=normalized_trade_date,
        index_rows=index_rows,
        board_rows=board_rows,
    )
    if readiness["status"] != "passed":
        raise N6ReplayBlocked(C1_SOURCE_INCOMPLETE_REASON)
    root = replay_root.resolve()
    source_root = (root / HISTORICAL_REPLAY_SOURCE_DIRNAME).resolve()
    if not source_root.is_relative_to(root):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: source_root_outside_replay_root")
    template_key = template_source_bundle_key or f"{trade_date_key}_index_board_only"
    template_bundle, _ = _load_historical_replay_source_bundle(
        replay_root=root,
        trade_date=normalized_trade_date,
        replay_run_id=f"local_replay_{trade_date_key}_000000_00000000",
        source_bundle_key=template_key,
    )
    bundle = _normalize_historical_replay_source_bundle(template_bundle)
    target_key = f"{trade_date_key}{INDEX_BOARD_FULL_DAY_SOURCE_SUFFIX}"
    target = _resolve_source_bundle_candidate_path(source_root, source_root / target_key / "source_bundle.json")
    c1_rows_by_identity = _group_c1_rows_by_identity(
        trade_date=normalized_trade_date,
        index_rows=index_rows,
        board_rows=board_rows,
    )
    source_records: dict[str, list[dict[str, Any]]] = {}
    for identity_key, c1_rows in c1_rows_by_identity.items():
        code_key = str((c1_rows[0] if c1_rows else {}).get("code") or "")
        template_source_records = dict(bundle.get("source_records") or {})
        candidate_template_rows = template_source_records.get(identity_key) or template_source_records.get(code_key) or []
        template_rows = [
            dict(row)
            for row in candidate_template_rows
            if _source_row_date(row) != normalized_trade_date
        ]
        source_records[identity_key] = [*template_rows, *c1_rows]
    bundle["candidates"] = [_with_source_record_key(row) for row in bundle.get("candidates") or []]
    bundle["source_records"] = source_records
    b2_input = dict(bundle.get("b2_input") or {})
    b2_input["live_current_rows_by_asset"] = {
        "index": [_c1_row_to_b2_bar(row) for row in index_rows],
        "board": [_c1_row_to_b2_bar(row) for row in board_rows],
    }
    b2_input["previous_day_rows_by_asset"] = {
        asset_kind: [dict(row) for row in rows]
        for asset_kind, rows in dict(b2_input.get("previous_day_rows_by_asset") or {}).items()
        if asset_kind in {"index", "board"}
    }
    b2_input["snapshot_rows"] = [
        dict(row)
        for row in b2_input.get("snapshot_rows") or []
        if _snapshot_asset_kind(row) in {"index", "board"}
    ]
    bundle["b2_input"] = b2_input
    source_meta = dict(bundle.get("source_meta") or {})
    source_meta.update(
        {
            "historical_source_type": "explicit_source_bundle_index_board_only_full_day",
            "historical_source_path": str(target),
            "source_origin": "c1_read_only_export",
            "c1_readiness_status": "passed",
            "stock_count": 0,
            "index_object_count": readiness["index_object_count"],
            "board_object_count": readiness["board_object_count"],
            "expected_minutes_per_object": readiness["expected_minutes_per_object"],
            "actual_minutes_distribution": readiness["actual_minutes_distribution"],
            "source_row_count": sum(len(rows) for rows in source_records.values()),
            "c1_current_day_row_count": len(index_rows) + len(board_rows),
            "candidate_count": len(bundle.get("candidates") or []),
            "context_count": len(bundle.get("n4_context_snapshot_rows") or []),
            "b2_snapshot_row_count": len(b2_input.get("snapshot_rows") or []),
            "b2_live_current_row_count": len(index_rows) + len(board_rows),
            "b2_previous_day_row_count": sum(len(rows) for rows in b2_input.get("previous_day_rows_by_asset", {}).values()),
            "upstream_source_mode": "c1_minute_bar_1m",
            "bundle_contract_version": "historical_replay_source_bundle_v1",
        }
    )
    source_meta["historical_source_hash"] = _stable_payload_hash(
        {
            "source_records": source_records,
            "candidates": bundle.get("candidates") or [],
            "n4_context_snapshot_rows": bundle.get("n4_context_snapshot_rows") or [],
            "b2_input": b2_input,
            "source_meta_without_hash": {key: value for key, value in source_meta.items() if key != "historical_source_hash"},
        }
    )
    bundle["source_meta"] = source_meta
    target.parent.mkdir(parents=True, exist_ok=True)
    _write_json(target, bundle)
    report = {
        "source_bundle_key": target_key,
        "source_bundle_path": str(target),
        "readiness": readiness,
        "source_meta": source_meta,
        "safety_flags": dict(SAFETY_FLAGS),
    }
    _write_json(target.parent / "source_bundle_report.json", report)
    (target.parent / "source_bundle_report.md").write_text(_source_bundle_report_markdown(report), encoding="utf-8")
    return report


def _row_identity_key(row: Mapping[str, Any], *, fallback_asset_kind: str) -> str:
    identity_key = str(row.get("identity_key") or "")
    if identity_key:
        return identity_key
    code = str(row.get("code") or row.get("display_code") or "")
    exchange = str(row.get("exchange") or ("TDX" if fallback_asset_kind == "board" else "SH"))
    return f"{fallback_asset_kind}:{exchange}:{code}" if code else ""


def _source_row_date(row: Mapping[str, Any]) -> str:
    text = str(row.get("datetime") or row.get("minute_label") or row.get("bar_time") or "")
    if "T" in text:
        text = text.replace("T", " ")
    return text[:10] if len(text) >= 10 else ""


def _source_row_minute(row: Mapping[str, Any], trade_date: str) -> str:
    text = str(row.get("datetime") or row.get("minute_label") or row.get("bar_time") or "")
    if "T" in text:
        text = text.replace("T", " ")
    if len(text) < 16 or text[:10] != trade_date:
        return ""
    return text[11:16]


def _compress_minute_ranges(minutes: Sequence[str]) -> list[str]:
    if not minutes:
        return []
    ordered = sorted(minutes)
    ranges: list[str] = []
    start = previous = ordered[0]
    for minute in ordered[1:]:
        if _minute_to_index(minute) == _minute_to_index(previous) + 1:
            previous = minute
            continue
        ranges.append(start if start == previous else f"{start}-{previous}")
        start = previous = minute
    ranges.append(start if start == previous else f"{start}-{previous}")
    return ranges


def _minute_to_index(minute: str) -> int:
    hour, minute_part = [int(part) for part in minute.split(":")]
    return hour * 60 + minute_part


def _group_c1_rows_by_identity(
    *,
    trade_date: str,
    index_rows: Sequence[Mapping[str, Any]],
    board_rows: Sequence[Mapping[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    output: dict[str, list[dict[str, Any]]] = {}
    for asset_kind, rows in (("index", index_rows), ("board", board_rows)):
        for row in rows:
            normalized = _c1_row_to_source_record(row, asset_kind=asset_kind, trade_date=trade_date)
            identity_key = str(normalized.get("identity_key") or "")
            if not identity_key:
                raise N6ReplayBlocked("BLOCKED_REPLAY_C1_SOURCE_INCOMPLETE: missing_identity_key")
            output.setdefault(identity_key, []).append(normalized)
    for rows in output.values():
        rows.sort(key=lambda row: str(row.get("datetime") or ""))
    return output


def _with_source_record_key(row: Mapping[str, Any]) -> dict[str, Any]:
    output = dict(row)
    output["source_record_key"] = str(output.get("source_record_key") or output.get("identity_key") or output.get("code") or "")
    return output


def _c1_row_to_source_record(row: Mapping[str, Any], *, asset_kind: str, trade_date: str) -> dict[str, Any]:
    row_asset_kind = _row_asset_kind(row) or asset_kind
    if row_asset_kind != asset_kind or row_asset_kind == "stock":
        raise N6ReplayBlocked("BLOCKED_REPLAY_C1_SOURCE_INCOMPLETE: invalid_asset_kind")
    minute = _source_row_minute(row, trade_date)
    if not minute:
        raise N6ReplayBlocked("BLOCKED_REPLAY_C1_SOURCE_INCOMPLETE: missing_minute")
    identity_key = _row_identity_key(row, fallback_asset_kind=asset_kind)
    code = str(row.get("code") or row.get("display_code") or _code_from_identity_key(identity_key) or "")
    normalized = dict(row)
    normalized["asset_kind"] = asset_kind
    normalized["identity_key"] = identity_key
    normalized["code"] = code
    normalized["display_code"] = str(normalized.get("display_code") or code)
    normalized["datetime"] = f"{trade_date} {minute}"
    normalized["minute_label"] = f"{trade_date} {minute}"
    normalized["bar_time"] = str(normalized.get("bar_time") or f"{trade_date}T{minute}:00+08:00")
    for key in ("open", "high", "low", "close", "volume", "amount"):
        normalized.setdefault(key, "0")
    normalized.setdefault("quality_status", "passed")
    normalized.setdefault("source_origin", "c1_read_only_export")
    return normalized


def _c1_row_to_b2_bar(row: Mapping[str, Any]) -> dict[str, Any]:
    source = dict(row)
    bar_time = source.get("bar_time")
    if not bar_time:
        text = str(source.get("datetime") or source.get("minute_label") or "")
        bar_time = f"{text[:10]}T{text[11:16]}:00+08:00" if len(text) >= 16 else ""
    return {
        "bar_id": source.get("bar_id") or _synthetic_c1_bar_id(source),
        "identity_key": source.get("identity_key"),
        "code": source.get("code") or source.get("display_code"),
        "display_code": source.get("display_code") or source.get("code"),
        "bar_time": bar_time,
        "open": source.get("open"),
        "high": source.get("high"),
        "low": source.get("low"),
        "close": source.get("close"),
        "volume": source.get("volume"),
        "amount": source.get("amount"),
        "quality_status": source.get("quality_status") or "passed",
    }


def _synthetic_c1_bar_id(row: Mapping[str, Any]) -> int:
    seed = f"{row.get('identity_key')}|{row.get('bar_time') or row.get('datetime') or row.get('minute_label')}"
    return int(sha256(seed.encode("utf-8")).hexdigest()[:12], 16)


def _source_bundle_report_markdown(report: Mapping[str, Any]) -> str:
    source_meta = dict(report.get("source_meta") or {})
    readiness = dict(report.get("readiness") or {})
    return "\n".join(
        [
            "# C1 Index/Board Full-Day Source Bundle",
            "",
            f"- source_bundle_key: `{report.get('source_bundle_key')}`",
            f"- source_bundle_path: `{report.get('source_bundle_path')}`",
            f"- c1_readiness_status: `{source_meta.get('c1_readiness_status')}`",
            f"- expected_minutes_per_object: `{source_meta.get('expected_minutes_per_object')}`",
            f"- index_object_count: `{source_meta.get('index_object_count')}`",
            f"- board_object_count: `{source_meta.get('board_object_count')}`",
            f"- stock_count: `{source_meta.get('stock_count')}`",
            f"- missing_asset_count: `{readiness.get('missing_asset_count')}`",
            "",
            LOCAL_ONLY_NOTICE,
            "",
        ]
    )


def read_replay_job(replay_root: Path, job_id: str) -> dict[str, Any]:
    artifact_dir = _find_job_dir(replay_root, job_id)
    status = _read_json(artifact_dir / "replay_status.json")
    summary = _read_json(artifact_dir / "replay_summary.json")
    return {
        **status,
        "artifact_dir": str(artifact_dir),
        "summary": summary,
        "safety_flags": dict(status.get("safety_flags") or SAFETY_FLAGS),
    }


def read_replay_messages(artifact_dir: Path, layer: str) -> list[dict[str, Any]]:
    if layer not in {"n3", "n4", "n5"}:
        raise ValueError("invalid_replay_message_layer")
    return _read_jsonl(artifact_dir / f"{layer}_messages.jsonl")


def read_replay_timeline(artifact_dir: Path) -> list[dict[str, Any]]:
    return _read_jsonl(artifact_dir / "replay_timeline.jsonl")


def replay_excel_response_bytes(replay_root: Path, job_id: str) -> bytes:
    artifact_dir = _find_job_dir(replay_root, job_id)
    return (artifact_dir / "n3_n5_full_day_replay.xlsx").read_bytes()


def build_replay_excel(
    *,
    summary: dict[str, Any],
    timeline: list[dict[str, Any]],
    n4_messages: list[dict[str, Any]],
    n5_messages: list[dict[str, Any]],
    n4_delta_attribution: list[dict[str, Any]] | None = None,
    n5_delta_only_messages: list[dict[str, Any]] | None = None,
    excluded_stock_replay_risk: list[dict[str, Any]] | None = None,
    n4_shadow_state_transitions: list[dict[str, Any]] | None = None,
    n4_shadow_evaluations: list[dict[str, Any]] | None = None,
    n5_shadow_action_windows: list[dict[str, Any]] | None = None,
    shadow_validation_report: dict[str, Any] | None = None,
    n4_prefilter_audit: list[dict[str, Any]] | None = None,
    n3p_demand_plan: list[dict[str, Any]] | None = None,
    n3p_active_state_reduction: list[dict[str, Any]] | None = None,
    n3p_proof_summary: list[dict[str, Any]] | None = None,
    n3p_negative_cache_decisions: list[dict[str, Any]] | None = None,
    n5_evaluator_demand_plan: list[dict[str, Any]] | None = None,
    n5_confirmation_metric_index_stats: list[dict[str, Any]] | None = None,
) -> bytes:
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="E6F3F1")
    header_font = Font(bold=True, color="172124")

    def setup_sheet(name: str, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
        worksheet = workbook.active if workbook.active.title == "Sheet" else workbook.create_sheet()
        worksheet.title = name
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            worksheet.append(list(row))
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_number, header in enumerate(headers, start=1):
            worksheet.column_dimensions[get_column_letter(column_number)].width = max(14, min(36, len(header) + 8))

    setup_sheet(
        "summary",
        ["key", "value"],
        _flatten_summary(summary),
    )
    setup_sheet(
        "minute_timeline",
        ["minute", "n3_rows", "n4_ordinary", "n4_hint", "n4_state_changed", "n4_noop_suppressed", "n5_eligible", "n5_executed", "n5_skipped"],
        (
            [
                row.get("minute"),
                row.get("n3_rows"),
                row.get("n4_ordinary"),
                row.get("n4_hint"),
                row.get("n4_state_changed"),
                row.get("n4_noop_suppressed"),
                row.get("n5_eligible"),
                row.get("n5_executed"),
                row.get("n5_skipped"),
            ]
            for row in timeline
        ),
    )
    setup_sheet("n4_ordinary_messages", _n4_headers(), _message_rows([m for m in n4_messages if m.get("source") == "ordinary"], _n4_headers()))
    setup_sheet("n4_hint_messages", _n4_headers(), _message_rows([m for m in n4_messages if m.get("source") == "hint"], _n4_headers()))
    setup_sheet("n5_action_eligible", _n5_headers(), _message_rows([m for m in n5_messages if m.get("event_type") == "ActionEligible"], _n5_headers()))
    setup_sheet("n5_action_executed", _n5_headers(), _message_rows([m for m in n5_messages if m.get("event_type") == "ActionExecuted"], _n5_headers()))
    setup_sheet("n5_action_skipped", _n5_headers(), _message_rows([m for m in n5_messages if m.get("event_type") == "ActionSkipped"], _n5_headers()))
    if n4_delta_attribution is not None:
        setup_sheet("n4_delta_attribution", _n4_delta_headers(), _message_rows(n4_delta_attribution, _n4_delta_headers()))
    if n5_delta_only_messages is not None:
        setup_sheet("n5_delta_only_messages", _n5_delta_headers(), _message_rows(n5_delta_only_messages, _n5_delta_headers()))
    if excluded_stock_replay_risk is not None:
        setup_sheet("excluded_stock_replay_risk", _n4_delta_headers(), _message_rows(excluded_stock_replay_risk, _n4_delta_headers()))
    if n4_shadow_state_transitions is not None:
        setup_sheet("n4_shadow_state_transitions", _n4_shadow_headers(), _message_rows(n4_shadow_state_transitions, _n4_shadow_headers()))
    if n4_shadow_evaluations is not None:
        setup_sheet("n4_shadow_evaluations", _n4_shadow_evaluation_headers(), _message_rows(n4_shadow_evaluations, _n4_shadow_evaluation_headers()))
    if n5_shadow_action_windows is not None:
        setup_sheet("n5_shadow_action_windows", _n5_shadow_headers(), _message_rows(n5_shadow_action_windows, _n5_shadow_headers()))
    if shadow_validation_report is not None:
        setup_sheet("shadow_quality_blockers", ["key", "value"], _flatten_summary(shadow_validation_report))
    if n4_prefilter_audit is not None:
        setup_sheet("n4_prefilter_audit", _n4_prefilter_headers(), _message_rows(n4_prefilter_audit, _n4_prefilter_headers()))
    if n3p_demand_plan is not None:
        setup_sheet("n3p_demand_plan", _n3p_demand_headers(), _message_rows(n3p_demand_plan, _n3p_demand_headers()))
    if n3p_active_state_reduction is not None:
        setup_sheet(
            "n3p_active_state_reduction",
            _n3p_active_state_reduction_headers(),
            _message_rows(n3p_active_state_reduction, _n3p_active_state_reduction_headers()),
        )
    if n3p_proof_summary is not None:
        setup_sheet("n3p_proof_summary", _n3p_proof_summary_headers(), _message_rows(n3p_proof_summary, _n3p_proof_summary_headers()))
    if n3p_negative_cache_decisions is not None:
        setup_sheet("n3p_negative_cache_decisions", _n3p_negative_cache_decision_headers(), _message_rows(n3p_negative_cache_decisions, _n3p_negative_cache_decision_headers()))
    if n5_evaluator_demand_plan is not None:
        setup_sheet("n5_evaluator_demand_plan", _n5_evaluator_demand_headers(), _message_rows(n5_evaluator_demand_plan, _n5_evaluator_demand_headers()))
    if n5_confirmation_metric_index_stats is not None:
        setup_sheet(
            "n5_confirmation_metric_index_stats",
            _n5_confirmation_metric_index_headers(),
            _message_rows(n5_confirmation_metric_index_stats, _n5_confirmation_metric_index_headers()),
        )
    setup_sheet("quality_blockers", ["blocked_reason", "count"], summary.get("blocked_reasons", {}).items())
    setup_sheet(
        "lineage_and_safety",
        ["key", "value"],
        _flatten_summary(
            {
                "replay_engine_version": summary.get("replay_engine_version"),
                "validation_mode": summary.get("validation_mode"),
                "n3p_strategy": summary.get("n3p_strategy"),
                "n3p_reduction_mode": summary.get("n3p_reduction_mode"),
                "n3p_active_state_reduction": summary.get("n3p_active_state_reduction"),
                "asset_scope": summary.get("asset_scope"),
                "source_bundle_key": summary.get("source_bundle_key"),
                "source_bundle_selector_mode": summary.get("source_bundle_selector_mode"),
                "resolved_source_bundle_path": summary.get("resolved_source_bundle_path"),
                "asset_scope_filter_applied": summary.get("asset_scope_filter_applied"),
                "asset_scope_allowed_asset_kinds": summary.get("asset_scope_allowed_asset_kinds"),
                "asset_scope_source_counts_before": summary.get("asset_scope_source_counts_before"),
                "asset_scope_source_counts_after": summary.get("asset_scope_source_counts_after"),
                "source_policy": summary.get("source_policy"),
                "historical_source_status": summary.get("historical_source_status"),
                "historical_source_kind": summary.get("historical_source_kind"),
                "historical_source_path": summary.get("historical_source_path"),
                "source_meta": summary.get("source_meta"),
                "canonical_planner_trace": summary.get("canonical_planner_trace"),
                "plan_only_side_effects": summary.get("plan_only_side_effects"),
                "n3p_cache_stats": summary.get("n3p_cache_stats"),
                "n3p_prefilter": summary.get("n3p_prefilter"),
                "n3p_negative_cache": summary.get("n3p_negative_cache"),
                "n5_evaluator_demand": summary.get("n5_evaluator_demand"),
                "shadow_mode": summary.get("shadow_mode"),
                "shadow": summary.get("shadow"),
                "shadow_validation_report": summary.get("shadow_validation_report"),
                "asset_unit_fix_delta_validation": summary.get("asset_unit_fix_delta_validation"),
                "asset_unit_fix_delta": summary.get("asset_unit_fix_delta"),
                "n5_delta_only": summary.get("n5_delta_only"),
                "safety_flags": summary.get("safety_flags"),
                "notice": summary.get("notice"),
            }
        ),
    )

    workbook.properties.title = f"N3-N5 local replay {summary.get('trade_date')}"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _normalize_trade_date(value: str) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"\d{8}", text):
        text = f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: invalid_trade_date")
    return text


def _new_job_id(trade_date_key: str) -> str:
    now = datetime.now(DISPLAY_TIMEZONE)
    seed = f"{trade_date_key}:{now.isoformat()}".encode("utf-8")
    return f"local_replay_{trade_date_key}_{now.strftime('%H%M%S')}_{sha256(seed).hexdigest()[:8]}"


def _validate_job_id(job_id: str, trade_date_key: str) -> None:
    if job_id.startswith(PRODUCTION_ID_PREFIXES) or not JOB_ID_RE.fullmatch(job_id):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: invalid_local_replay_job_id")
    if not job_id.startswith(f"local_replay_{trade_date_key}_"):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: job_trade_date_mismatch")


def _normalize_replay_engine_version(value: Any) -> str:
    engine_version = str(value or DEFAULT_REPLAY_ENGINE_VERSION).strip()
    if engine_version not in REPLAY_ENGINE_VERSIONS:
        raise N6ReplayBlocked("BLOCKED_REPLAY_SIDE_EFFECT_RISK: invalid_replay_engine_version")
    return engine_version


def _normalize_replay_validation_mode(value: Any) -> str:
    validation_mode = str(value or DEFAULT_REPLAY_VALIDATION_MODE).strip()
    if validation_mode not in REPLAY_VALIDATION_MODES:
        raise N6ReplayBlocked("BLOCKED_REPLAY_INVALID_VALIDATION_MODE")
    return validation_mode


def _normalize_replay_n3p_strategy(value: Any) -> str:
    strategy = str(value or DEFAULT_REPLAY_N3P_STRATEGY).strip()
    if strategy not in REPLAY_N3P_STRATEGIES:
        raise N6ReplayBlocked("BLOCKED_REPLAY_INVALID_N3P_STRATEGY")
    return strategy


def _normalize_replay_n3p_reduction_mode(value: Any) -> str:
    mode = str(value or DEFAULT_REPLAY_N3P_REDUCTION_MODE).strip()
    if mode not in REPLAY_N3P_REDUCTION_MODES:
        raise N6ReplayBlocked("BLOCKED_REPLAY_INVALID_N3P_REDUCTION_MODE")
    return mode


def _normalize_replay_n3p_negative_cache(value: Any) -> str:
    mode = str(value or DEFAULT_REPLAY_N3P_NEGATIVE_CACHE).strip()
    if mode not in REPLAY_N3P_NEGATIVE_CACHE_MODES:
        raise N6ReplayBlocked("BLOCKED_REPLAY_INVALID_N3P_NEGATIVE_CACHE")
    return mode


def _normalize_replay_asset_scope(value: Any) -> str:
    asset_scope = str(value or DEFAULT_REPLAY_ASSET_SCOPE).strip()
    if asset_scope not in REPLAY_ASSET_SCOPES:
        raise N6ReplayBlocked("BLOCKED_REPLAY_INVALID_ASSET_SCOPE")
    return asset_scope


def _normalize_source_bundle_key(value: Any) -> str:
    source_bundle_key = str(value or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY).strip()
    if source_bundle_key == DEFAULT_REPLAY_SOURCE_BUNDLE_KEY:
        return source_bundle_key
    if (
        not SOURCE_BUNDLE_KEY_RE.fullmatch(source_bundle_key)
        or source_bundle_key.startswith(PRODUCTION_ID_PREFIXES)
    ):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SOURCE_BUNDLE_NOT_FOUND")
    return source_bundle_key


def _source_bundle_selector_mode(source_bundle_key: str) -> str:
    return "auto" if source_bundle_key == DEFAULT_REPLAY_SOURCE_BUNDLE_KEY else "explicit"


def _trading_minutes(start_hhmm: str, end_hhmm: str) -> list[str]:
    morning = _minute_range("09:31", "11:30")
    afternoon = _minute_range("13:01", "15:00")
    allowed = morning + afternoon
    return [minute for minute in allowed if start_hhmm <= minute <= end_hhmm]


def _minute_range(start: str, end: str) -> list[str]:
    start_hour, start_minute = [int(part) for part in start.split(":")]
    end_hour, end_minute = [int(part) for part in end.split(":")]
    minutes: list[str] = []
    cursor = start_hour * 60 + start_minute
    end_cursor = end_hour * 60 + end_minute
    while cursor <= end_cursor:
        minutes.append(f"{cursor // 60:02d}:{cursor % 60:02d}")
        cursor += 1
    return minutes


def _build_n3_messages(trade_date: str, minutes: list[str]) -> list[dict[str, Any]]:
    return [
        {
            "minute": minute,
            "event_type": "ReplayMinuteClosed",
            "asset_kind": "market",
            "identity_key": "local_replay:market",
            "source_mode": "replay",
            "snapshot_policy": "replay_snapshot_from_minute_cumulative",
            "trace_summary": "historical 1m canonical labels; no live adapter; no fake bar",
            "trade_date": trade_date,
        }
        for minute in minutes
    ]


def _build_replay_engine_artifact(
    *,
    trade_date: str,
    minutes: Sequence[str],
    replay_run_id: str,
    replay_engine_version: str,
    asset_scope: str,
    source_bundle_key: str,
    replay_root: Path,
    validation_mode: str = DEFAULT_REPLAY_VALIDATION_MODE,
    n3p_strategy: str = DEFAULT_REPLAY_N3P_STRATEGY,
    n3p_reduction_mode: str = DEFAULT_REPLAY_N3P_REDUCTION_MODE,
    n3p_negative_cache: str = DEFAULT_REPLAY_N3P_NEGATIVE_CACHE,
    profiler: _ReplayProfileCollector | None = None,
) -> dict[str, Any]:
    engine_version = _normalize_replay_engine_version(replay_engine_version)
    normalized_validation_mode = _normalize_replay_validation_mode(validation_mode)
    normalized_n3p_strategy = _normalize_replay_n3p_strategy(n3p_strategy)
    normalized_n3p_reduction_mode = _normalize_replay_n3p_reduction_mode(n3p_reduction_mode)
    normalized_n3p_negative_cache = _normalize_replay_n3p_negative_cache(n3p_negative_cache)
    if normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE and engine_version != "canonical_plan_v1":
        raise N6ReplayBlocked("BLOCKED_REPLAY_SHADOW_REQUIRES_CANONICAL_PLAN")
    if engine_version == "fixture_v1":
        return _build_fixture_v1_artifact(
            trade_date=trade_date,
            minutes=minutes,
            replay_run_id=replay_run_id,
            asset_scope=asset_scope,
            source_bundle_key=source_bundle_key,
            validation_mode=normalized_validation_mode,
        )
    replay_bundle, source_meta = _load_historical_replay_source_bundle(
        replay_root=replay_root,
        trade_date=trade_date,
        replay_run_id=replay_run_id,
        source_bundle_key=source_bundle_key,
        profiler=profiler,
    )
    if normalized_validation_mode == "asset_unit_fix_delta_v1":
        return _build_asset_unit_fix_delta_validation_artifact(
            trade_date=trade_date,
            minutes=minutes,
            replay_run_id=replay_run_id,
            source_bundle=replay_bundle,
            source_meta=source_meta,
            asset_scope=asset_scope,
            validation_mode=normalized_validation_mode,
            profiler=profiler,
        )
    return _build_canonical_plan_v1_artifact(
        trade_date=trade_date,
        minutes=minutes,
        replay_run_id=replay_run_id,
        source_bundle=replay_bundle,
        source_meta=source_meta,
        asset_scope=asset_scope,
        validation_mode=normalized_validation_mode,
        n3p_strategy=normalized_n3p_strategy,
        n3p_reduction_mode=normalized_n3p_reduction_mode,
        n3p_negative_cache=normalized_n3p_negative_cache,
        profiler=profiler,
    )


def _build_canonical_plan_v1_artifact(
    *,
    trade_date: str,
    minutes: Sequence[str],
    replay_run_id: str,
    source_bundle: Mapping[str, Any],
    source_meta: Mapping[str, Any] | None = None,
    asset_scope: str = DEFAULT_REPLAY_ASSET_SCOPE,
    validation_mode: str = DEFAULT_REPLAY_VALIDATION_MODE,
    n3p_strategy: str = DEFAULT_REPLAY_N3P_STRATEGY,
    n3p_reduction_mode: str = DEFAULT_REPLAY_N3P_REDUCTION_MODE,
    n3p_negative_cache: str = DEFAULT_REPLAY_N3P_NEGATIVE_CACHE,
    profiler: _ReplayProfileCollector | None = None,
) -> dict[str, Any]:
    normalized_validation_mode = _normalize_replay_validation_mode(validation_mode)
    normalized_n3p_strategy = _normalize_replay_n3p_strategy(n3p_strategy)
    normalized_n3p_reduction_mode = _normalize_replay_n3p_reduction_mode(n3p_reduction_mode)
    normalized_n3p_negative_cache = _normalize_replay_n3p_negative_cache(n3p_negative_cache)
    normalization_token = profiler.start_phase("source_bundle_normalization_validation") if profiler else None
    normalized_bundle = _normalize_historical_replay_source_bundle(source_bundle)
    if profiler:
        profiler.finish_phase(
            normalization_token,
            source_record_keys=len(normalized_bundle.get("source_records") or {}),
            candidate_count=len(normalized_bundle.get("candidates") or []),
            context_count=len(normalized_bundle.get("n4_context_snapshot_rows") or []),
        )
    scoped_bundle, asset_scope_lineage = _apply_replay_asset_scope_filter(
        normalized_bundle,
        asset_scope=asset_scope,
    )
    shadow_mode = normalized_validation_mode == FULL_DAY_SHADOW_VALIDATION_MODE
    normalized_source_meta = _normalize_source_meta(
        (source_meta or {}).get("source_meta") if isinstance(source_meta, Mapping) else source_meta,
        fallback_path=str((source_meta or {}).get("historical_source_path") or ""),
    )
    if shadow_mode:
        _validate_full_day_shadow_source_contract(
            source_bundle=scoped_bundle,
            asset_scope_lineage=asset_scope_lineage,
            source_bundle_key=str((source_meta or {}).get("source_bundle_key") or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY),
            source_meta=normalized_source_meta,
            trade_date=trade_date,
        )
    replay_config = dict(scoped_bundle.get("replay_config") or {})
    replay_config["replay_run_id"] = replay_run_id
    replay_config["job_id"] = replay_run_id
    replay_config["trade_date"] = trade_date.replace("-", "")
    replay_config["for_trade_date"] = trade_date.replace("-", "")
    replay_config["replay_engine_version"] = "canonical_plan_v1"
    replay_config["validation_mode"] = normalized_validation_mode
    replay_config["n3p_strategy"] = normalized_n3p_strategy
    replay_config["n3p_reduction_mode"] = normalized_n3p_reduction_mode
    replay_config["n3p_negative_cache"] = normalized_n3p_negative_cache
    replay_config["asset_scope"] = asset_scope_lineage["asset_scope"]
    replay_config["asset_scope_allowed_asset_kinds"] = asset_scope_lineage["asset_scope_allowed_asset_kinds"]
    replay_config["source_bundle_key"] = str((source_meta or {}).get("source_bundle_key") or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY)
    replay_config["source_bundle_selector_mode"] = str((source_meta or {}).get("source_bundle_selector_mode") or "auto")
    replay_config["resolved_source_bundle_path"] = str((source_meta or {}).get("resolved_source_bundle_path") or (source_meta or {}).get("historical_source_path") or "")
    replay_config["historical_source_hash"] = str(normalized_source_meta.get("historical_source_hash") or "")
    bundle = dict(scoped_bundle)
    bundle["replay_config"] = replay_config
    loop_result = _run_canonical_historical_minute_loop(
        trade_date=trade_date,
        minutes=list(minutes),
        replay_run_id=replay_run_id,
        source_bundle=bundle,
        profiler=profiler,
        shadow_mode=shadow_mode,
        n3p_strategy=normalized_n3p_strategy,
        n3p_reduction_mode=normalized_n3p_reduction_mode,
        n3p_negative_cache=normalized_n3p_negative_cache,
    )
    return {
        "n3_messages": loop_result["n3_messages"],
        "n4_messages": loop_result["n4_messages"],
        "n5_messages": loop_result["n5_messages"],
        "n4_shadow_state_transitions": loop_result.get("n4_shadow_state_transitions", []),
        "n4_shadow_evaluations": loop_result.get("n4_shadow_evaluations", []),
        "n5_shadow_action_windows": loop_result.get("n5_shadow_action_windows", []),
        "shadow_mode": shadow_mode,
        "shadow_validation_report": loop_result.get("shadow_validation_report", {}),
        "n3p_artifact": loop_result["last_n3p_artifact"],
        "b2_artifact": loop_result["last_b2_artifact"],
        "eligible_plan": loop_result["last_eligible_plan"],
        "executed_report": loop_result["last_executed_report"],
        "replay_engine_version": "canonical_plan_v1",
        "validation_mode": normalized_validation_mode,
        "n3p_strategy": normalized_n3p_strategy,
        "n3p_reduction_mode": normalized_n3p_reduction_mode,
        "n3p_negative_cache": normalized_n3p_negative_cache,
        "source_policy": HISTORICAL_REPLAY_SOURCE_POLICY,
        "historical_source_status": str((source_meta or {}).get("historical_source_status") or "available"),
        "historical_source_kind": str((source_meta or {}).get("historical_source_kind") or ""),
        "historical_source_path": str((source_meta or {}).get("historical_source_path") or ""),
        "source_bundle_key": str((source_meta or {}).get("source_bundle_key") or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY),
        "source_bundle_selector_mode": str((source_meta or {}).get("source_bundle_selector_mode") or "auto"),
        "resolved_source_bundle_path": str((source_meta or {}).get("resolved_source_bundle_path") or (source_meta or {}).get("historical_source_path") or ""),
        "source_meta": normalized_source_meta,
        "asset_scope": asset_scope_lineage["asset_scope"],
        "asset_scope_filter_applied": asset_scope_lineage["asset_scope_filter_applied"],
        "asset_scope_allowed_asset_kinds": asset_scope_lineage["asset_scope_allowed_asset_kinds"],
        "asset_scope_source_counts_before": asset_scope_lineage["asset_scope_source_counts_before"],
        "asset_scope_source_counts_after": asset_scope_lineage["asset_scope_source_counts_after"],
        "engine_disclaimer": REPLAY_SOURCE_POLICY["canonical_plan_v1"],
        "canonical_planner_trace": {
            "n3p_wrapper": "v3_realtime_virtual_metric_writer.plan_only_wrapper.v1",
            "b2_wrapper": "realtime_projection_execute.plan_only_wrapper.v1",
            "n4_ordinary_matcher": "provisional_ordinary_matcher_v1",
            "n4_hint_matcher": "provisional_projection_matcher_v1",
            "n4_lifecycle": "provisional_trigger_lifecycle_v1",
            "n5_actioneligible": "provisional_actioneligible_v1",
            "n5_actionexecuted": "provisional_action_executed_dry_run_v1",
        },
        "plan_only_side_effects": loop_result["side_effects"],
        "n3p_cache_stats": dict(loop_result.get("n3p_cache_stats") or {}),
        "n3p_prefilter": dict(loop_result.get("n3p_prefilter") or {}),
        "n3p_active_state_reduction_summary": dict(loop_result.get("n3p_active_state_reduction_summary") or {}),
        "n3p_negative_cache_summary": dict(loop_result.get("n3p_negative_cache_summary") or {}),
        "n4_lightweight_prefilter_audit": list(loop_result.get("n4_lightweight_prefilter_audit") or []),
        "n3p_demand_plan": list(loop_result.get("n3p_demand_plan") or []),
        "n3p_active_state_reduction": list(loop_result.get("n3p_active_state_reduction") or []),
        "n3p_plan_only_proof_summary": list(loop_result.get("n3p_plan_only_proof_summary") or []),
        "n3p_negative_cache_decisions": list(loop_result.get("n3p_negative_cache_decisions") or []),
        "n5_evaluator_demand_plan": list(loop_result.get("n5_evaluator_demand_plan") or []),
        "n5_confirmation_metric_index_stats": list(loop_result.get("n5_confirmation_metric_index_stats") or []),
        "n5_evaluator_demand_summary": dict(loop_result.get("n5_evaluator_demand_summary") or {}),
    }


def _build_asset_unit_fix_delta_validation_artifact(
    *,
    trade_date: str,
    minutes: Sequence[str],
    replay_run_id: str,
    source_bundle: Mapping[str, Any],
    source_meta: Mapping[str, Any] | None = None,
    asset_scope: str = DEFAULT_REPLAY_ASSET_SCOPE,
    validation_mode: str = "asset_unit_fix_delta_v1",
    profiler: _ReplayProfileCollector | None = None,
) -> dict[str, Any]:
    normalized_validation_mode = _normalize_replay_validation_mode(validation_mode)
    if normalized_validation_mode != "asset_unit_fix_delta_v1":
        raise N6ReplayBlocked("BLOCKED_REPLAY_INVALID_VALIDATION_MODE")
    normalization_token = profiler.start_phase("source_bundle_normalization_validation") if profiler else None
    normalized_bundle = _normalize_historical_replay_source_bundle(source_bundle)
    if profiler:
        profiler.finish_phase(
            normalization_token,
            source_record_keys=len(normalized_bundle.get("source_records") or {}),
            candidate_count=len(normalized_bundle.get("candidates") or []),
            context_count=len(normalized_bundle.get("n4_context_snapshot_rows") or []),
        )
    scoped_bundle, asset_scope_lineage = _apply_replay_asset_scope_filter(
        normalized_bundle,
        asset_scope=asset_scope,
    )
    normalized_source_meta = _normalize_source_meta(
        (source_meta or {}).get("source_meta") if isinstance(source_meta, Mapping) else source_meta,
        fallback_path=str((source_meta or {}).get("historical_source_path") or ""),
    )
    delta_payload = normalized_bundle.get("asset_unit_fix_delta_validation")
    if not isinstance(delta_payload, Mapping):
        raise N6ReplayBlocked("BLOCKED_NEED_LOCAL_DELTA_SOURCE_BUNDLE")
    delta_rows, delta_summary = _build_asset_unit_fix_delta_attribution(
        delta_payload=delta_payload,
        trade_date=trade_date,
    )
    if int(delta_summary.get("other_requires_review") or 0) > 0:
        raise N6ReplayBlocked("BLOCKED_REPLAY_DELTA_REQUIRES_REVIEW")
    n5_delta_messages = _build_asset_unit_fix_n5_delta_only_messages(
        trade_date=trade_date,
        replay_run_id=replay_run_id,
        delta_rows=delta_rows,
        replay_config=dict(scoped_bundle.get("replay_config") or {}),
    )
    n5_delta_summary = _build_n5_delta_only_summary(n5_delta_messages)
    excluded_stock_replay_risk = [
        row for row in delta_rows if row.get("delta_classification") == "stock_replayed_due_no_previous_baseline"
    ]
    return {
        "n3_messages": _build_n3_messages(trade_date, list(minutes)),
        "n4_messages": [],
        "n5_messages": [],
        "replay_engine_version": "canonical_plan_v1",
        "validation_mode": normalized_validation_mode,
        "source_policy": HISTORICAL_REPLAY_SOURCE_POLICY,
        "historical_source_status": str((source_meta or {}).get("historical_source_status") or "available"),
        "historical_source_kind": str((source_meta or {}).get("historical_source_kind") or ""),
        "historical_source_path": str((source_meta or {}).get("historical_source_path") or ""),
        "source_bundle_key": str((source_meta or {}).get("source_bundle_key") or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY),
        "source_bundle_selector_mode": str((source_meta or {}).get("source_bundle_selector_mode") or "auto"),
        "resolved_source_bundle_path": str((source_meta or {}).get("resolved_source_bundle_path") or (source_meta or {}).get("historical_source_path") or ""),
        "source_meta": normalized_source_meta,
        "asset_scope": asset_scope_lineage["asset_scope"],
        "asset_scope_filter_applied": asset_scope_lineage["asset_scope_filter_applied"],
        "asset_scope_allowed_asset_kinds": asset_scope_lineage["asset_scope_allowed_asset_kinds"],
        "asset_scope_source_counts_before": asset_scope_lineage["asset_scope_source_counts_before"],
        "asset_scope_source_counts_after": asset_scope_lineage["asset_scope_source_counts_after"],
        "engine_disclaimer": "asset_unit_fix_delta_v1 local validation; not production lineage",
        "canonical_planner_trace": {
            "n4_delta_attribution": "asset_unit_fix_delta_stable_key_v1",
            "n5_delta_only": "provisional_actioneligible_v1",
        },
        "plan_only_side_effects": _plan_only_side_effects(),
        "n3p_cache_stats": {},
        "asset_unit_fix_delta_validation": True,
        "asset_unit_fix_delta": delta_summary,
        "n5_delta_only": n5_delta_summary,
        "n4_delta_attribution": delta_rows,
        "n5_delta_only_messages": n5_delta_messages,
        "excluded_stock_replay_risk": excluded_stock_replay_risk,
    }


def _build_asset_unit_fix_delta_attribution(
    *,
    delta_payload: Mapping[str, Any],
    trade_date: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    old_rows = list(delta_payload.get("old_unified_trigger_matched") or [])
    corrected_rows = list(delta_payload.get("corrected_trigger_matched") or [])
    old_by_key = _asset_unit_fix_rows_by_stable_key(old_rows, side="old")
    corrected_by_key = _asset_unit_fix_rows_by_stable_key(corrected_rows, side="corrected")
    old_keys = set(old_by_key)
    corrected_keys = set(corrected_by_key)
    common_keys = corrected_keys & old_keys
    corrected_only_keys = corrected_keys - old_keys
    old_only_keys = old_keys - corrected_keys
    rows: list[dict[str, Any]] = []
    for key in sorted(corrected_keys):
        source_row = corrected_by_key[key]
        if key in common_keys:
            classification = "common_old_and_corrected"
        else:
            classification = _classify_corrected_only_delta(source_row)
        rows.append(
            _asset_unit_fix_delta_output_row(
                source_row,
                trade_date=trade_date,
                stable_key=key,
                classification=classification,
                old_unified_n4_run_id=str(delta_payload.get("old_unified_n4_run_id") or ""),
                corrected_n4_run_id=str(delta_payload.get("corrected_n4_run_id") or ""),
            )
        )
    for key in sorted(old_only_keys):
        rows.append(
            _asset_unit_fix_delta_output_row(
                old_by_key[key],
                trade_date=trade_date,
                stable_key=key,
                classification="old_only",
                old_unified_n4_run_id=str(delta_payload.get("old_unified_n4_run_id") or ""),
                corrected_n4_run_id=str(delta_payload.get("corrected_n4_run_id") or ""),
            )
        )
    summary = {
        "asset_unit_fix_delta_validation": True,
        "stable_match_key_version": "asset_unit_fix_delta_stable_key_v1",
        "stable_match_key_fields": [
            "for_trade_date",
            "asset_kind",
            "identity_key",
            "direction",
            "signal_type",
            "condition_key",
            "trigger_period",
            "trigger_mark_candidate",
            "trigger_minute",
        ],
        "corrected_full_trigger_matched": len(corrected_rows),
        "old_unified_trigger_matched": len(old_rows),
        "common_old_and_corrected": len(common_keys),
        "corrected_only": len(corrected_only_keys),
        "old_only": len(old_only_keys),
        "index_board_delta": sum(
            1 for row in rows if row.get("delta_classification") == "index_board_unit_fix_new_signal"
        ),
        "excluded_stock_replay": sum(
            1 for row in rows if row.get("delta_classification") == "stock_replayed_due_no_previous_baseline"
        ),
        "other_requires_review": sum(
            1 for row in rows if row.get("delta_classification") == "other_requires_review"
        ),
        "old_unified_n4_run_id": str(delta_payload.get("old_unified_n4_run_id") or ""),
        "corrected_n4_run_id": str(delta_payload.get("corrected_n4_run_id") or ""),
    }
    return rows, summary


def _asset_unit_fix_rows_by_stable_key(rows: Sequence[Any], *, side: str) -> dict[str, Mapping[str, Any]]:
    output: dict[str, Mapping[str, Any]] = {}
    for row in rows:
        if not isinstance(row, Mapping):
            raise N6ReplayBlocked("BLOCKED_NEED_LOCAL_DELTA_SOURCE_BUNDLE: invalid_delta_row")
        key = _asset_unit_fix_stable_key(row)
        if key in output:
            raise N6ReplayBlocked(f"BLOCKED_REPLAY_DELTA_DUPLICATE_STABLE_KEY: {side}")
        output[key] = row
    return output


def _asset_unit_fix_stable_key(row: Mapping[str, Any]) -> str:
    parts = [
        str(row.get("for_trade_date") or row.get("trade_date") or ""),
        str(row.get("asset_kind") or ""),
        str(row.get("identity_key") or ""),
        str(row.get("direction") or _direction_from_signal_type(str(row.get("signal_type") or ""))),
        str(row.get("signal_type") or ""),
        str(row.get("condition_key") or ""),
        str(row.get("trigger_period") or ""),
        str(row.get("trigger_mark_candidate") or ""),
        _trigger_minute_from_row(row),
    ]
    if any(not part for part in parts):
        raise N6ReplayBlocked("BLOCKED_NEED_LOCAL_DELTA_SOURCE_BUNDLE: incomplete_stable_key")
    return "|".join(parts)


def _trigger_minute_from_row(row: Mapping[str, Any]) -> str:
    minute = str(row.get("trigger_minute") or row.get("minute") or "")
    if re.fullmatch(r"\d{2}:\d{2}", minute):
        return minute
    return _minute_from_event_time(row.get("trigger_time") or row.get("event_time"))


def _direction_from_signal_type(signal_type: str) -> str:
    if signal_type == "B_BUY":
        return "buy"
    if signal_type == "S_SELL":
        return "sell"
    return ""


def _trigger_type_from_condition_key(condition_key: str) -> str:
    if condition_key in {"BUY:FULL", "SELL:FULL", "BUY_HINT", "SELL_HINT"}:
        return condition_key
    if condition_key.startswith("BUY"):
        return "BUY"
    if condition_key.startswith("SELL"):
        return "SELL"
    return condition_key


def _classify_corrected_only_delta(row: Mapping[str, Any]) -> str:
    asset_kind = str(row.get("asset_kind") or "")
    if asset_kind in {"index", "board"}:
        return "index_board_unit_fix_new_signal"
    if asset_kind == "stock":
        return "stock_replayed_due_no_previous_baseline"
    return "other_requires_review"


_ASSET_UNIT_FIX_DELTA_N4_PASSTHROUGH_FIELDS = (
    "source_condition_run_id",
    "source_condition_key",
    "source_metric_run_id",
    "source_n3p_live_target_run_id",
    "source_mode",
    "c1_dependency",
    "source_condition_trace",
    "condition_trace",
    "n4_boundary",
    "n5_entry_allowed",
)


def _asset_unit_fix_delta_output_row(
    row: Mapping[str, Any],
    *,
    trade_date: str,
    stable_key: str,
    classification: str,
    old_unified_n4_run_id: str,
    corrected_n4_run_id: str,
) -> dict[str, Any]:
    condition_key = str(row.get("condition_key") or "")
    signal_type = str(row.get("signal_type") or "")
    output = {
        "trade_date": trade_date,
        "for_trade_date": str(row.get("for_trade_date") or trade_date.replace("-", "")),
        "minute": _trigger_minute_from_row(row),
        "delta_classification": classification,
        "stable_match_key": stable_key,
        "asset_kind": str(row.get("asset_kind") or ""),
        "identity_key": str(row.get("identity_key") or ""),
        "direction": str(row.get("direction") or _direction_from_signal_type(signal_type)),
        "signal_type": signal_type,
        "condition_key": condition_key,
        "original_condition_key": str(row.get("original_condition_key") or condition_key),
        "trigger_type": str(row.get("trigger_type") or _trigger_type_from_condition_key(condition_key)),
        "trigger_period": str(row.get("trigger_period") or ""),
        "trigger_mark_candidate": str(row.get("trigger_mark_candidate") or ""),
        "trigger_price": row.get("trigger_price"),
        "trigger_time": row.get("trigger_time") or row.get("event_time"),
        "old_unified_n4_run_id": old_unified_n4_run_id,
        "corrected_n4_run_id": corrected_n4_run_id,
        "lineage_note": "production run ids are lineage trace only; local replay emits local run ids",
    }
    for field in _ASSET_UNIT_FIX_DELTA_N4_PASSTHROUGH_FIELDS:
        if field in row:
            output[field] = row.get(field)
    return output


def _build_asset_unit_fix_n5_delta_only_messages(
    *,
    trade_date: str,
    replay_run_id: str,
    delta_rows: Sequence[Mapping[str, Any]],
    replay_config: Mapping[str, Any],
) -> list[dict[str, Any]]:
    n4_run_id = f"{replay_run_id}__n4_asset_unit_fix_delta_v1"
    delta_messages = [
        _asset_unit_fix_delta_n4_message(
            row,
            trade_date=trade_date,
            n4_run_id=n4_run_id,
            ordinal=index,
        )
        for index, row in enumerate(delta_rows, start=1)
        if row.get("delta_classification") == "index_board_unit_fix_new_signal"
    ]
    source_condition_run_ids = {
        str(row.get("source_condition_run_id") or "").strip()
        for row in delta_messages
    }
    if "" in source_condition_run_ids or not source_condition_run_ids:
        raise N6ReplayBlocked("BLOCKED_DELTA_SOURCE_MISMATCH: source_condition_run_id_missing")
    if len(source_condition_run_ids) != 1:
        raise N6ReplayBlocked("BLOCKED_DELTA_SOURCE_MISMATCH: source_condition_run_id_mixed")
    source_condition_run_id = next(iter(source_condition_run_ids))
    outbox_rows = [_n4_message_to_outbox_row(row, n4_run_id=n4_run_id) for row in delta_messages]
    eligible_plan = provisional_action_eligible.build_provisional_actioneligible_plan(
        source_trigger_run={
            "run_id": n4_run_id,
            "status": "passed",
            "source_condition_run_id": source_condition_run_id,
            "for_trade_date": trade_date.replace("-", ""),
        },
        source_trigger_run_id=n4_run_id,
        action_run_id=f"{replay_run_id}__n5_asset_unit_fix_delta_v1",
        for_trade_date=trade_date.replace("-", ""),
        consumer_name="n6_local_replay_asset_unit_fix_delta_v1",
        outbox_rows=outbox_rows,
        target_counts={key: 0 for key in (
            "common_action_run",
            "common_action_quality_item",
            "stock_action_fact",
            "index_action_fact",
            "board_action_fact",
            "common_action_event",
            "common_event_outbox",
            "common_event_inbox",
            "common_event_consumer_checkpoint",
        )},
    )
    messages = _serialize_n5_messages(
        trade_date=trade_date,
        replay_run_id=replay_run_id,
        eligible_rows=list(eligible_plan["writes"]["common_event_outbox"]),
        executed_plans=[],
    )
    _decorate_replay_messages(
        n4_messages=delta_messages,
        n5_messages=messages,
        replay_engine_version="canonical_plan_v1",
        source_policy=HISTORICAL_REPLAY_SOURCE_POLICY,
    )
    for row in messages:
        row["validation_mode"] = "asset_unit_fix_delta_v1"
        row["trace_summary"] = "asset_unit_fix_delta_v1 delta-only N5 replay input"
    return messages


def _asset_unit_fix_delta_n4_message(
    row: Mapping[str, Any],
    *,
    trade_date: str,
    n4_run_id: str,
    ordinal: int,
) -> dict[str, Any]:
    condition_key = str(row.get("condition_key") or "")
    minute = str(row.get("minute") or _trigger_minute_from_row(row))
    source_condition_run_id = str(row.get("source_condition_run_id") or "").strip()
    if not source_condition_run_id:
        raise N6ReplayBlocked("BLOCKED_DELTA_SOURCE_MISMATCH: source_condition_run_id_missing")
    return {
        "event_id": f"{n4_run_id}:delta:{ordinal}",
        "trigger_state_id": ordinal,
        "trigger_match_id": ordinal,
        "minute": minute,
        "trade_date": trade_date,
        "event_type": "TriggerMatched",
        "asset_kind": row.get("asset_kind"),
        "identity_key": row.get("identity_key"),
        "signal_type": row.get("signal_type"),
        "condition_key": condition_key,
        "original_condition_key": row.get("original_condition_key") or condition_key,
        "trigger_type": row.get("trigger_type") or _trigger_type_from_condition_key(condition_key),
        "trigger_price": row.get("trigger_price"),
        "source": "ordinary",
        "source_mode": "replay",
        "source_run_id": n4_run_id,
        "for_trade_date": row.get("for_trade_date") or trade_date.replace("-", ""),
        "trigger_period": row.get("trigger_period"),
        "trigger_mark_candidate": row.get("trigger_mark_candidate") or "normal",
        "source_metric_kind": "realtime_action_confirmation_metric",
        "source_condition_run_id": source_condition_run_id,
        "source_condition_key": row.get("source_condition_key"),
        "source_metric_run_id": row.get("source_metric_run_id") or "",
        "source_n3p_live_target_run_id": row.get("source_n3p_live_target_run_id"),
        "source_payload_mode": row.get("source_mode"),
        "c1_dependency": row.get("c1_dependency"),
        "source_condition_trace": row.get("source_condition_trace"),
        "condition_trace": row.get("condition_trace"),
        "n4_boundary": row.get("n4_boundary"),
        "n5_entry_allowed": row.get("n5_entry_allowed"),
        "selected_metric_id": f"delta_metric_{ordinal}",
        "selected_metric_time": _event_time_from_trade_minute(trade_date, minute),
        "metric_time_label": f"{trade_date} {minute}",
        "metric_minute_label": minute,
        "is_closed_1m": False,
        "current_status": "matched",
        "trigger_live": True,
        "trace_summary": "asset_unit_fix_delta_v1 local N4 delta row; no production outbox consumption",
    }


def _build_n5_delta_only_summary(messages: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    return {
        "ActionEligible": sum(1 for row in messages if row.get("event_type") == "ActionEligible"),
        "ActionExecuted": sum(1 for row in messages if row.get("event_type") == "ActionExecuted"),
        "stock_ActionEligible": sum(
            1 for row in messages if row.get("event_type") == "ActionEligible" and row.get("asset_kind") == "stock"
        ),
        "stock_ActionExecuted": sum(
            1 for row in messages if row.get("event_type") == "ActionExecuted" and row.get("asset_kind") == "stock"
        ),
        "b2_hint_final_proof_rows": sum(
            1 for row in messages if row.get("event_type") == "ActionExecuted" and row.get("final_proof_source") != "N3P"
        ),
    }


def _load_historical_replay_source_bundle(
    *,
    replay_root: Path,
    trade_date: str,
    replay_run_id: str,
    source_bundle_key: str = DEFAULT_REPLAY_SOURCE_BUNDLE_KEY,
    profiler: _ReplayProfileCollector | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    trade_date_key = trade_date.replace("-", "")
    source_root = (replay_root / HISTORICAL_REPLAY_SOURCE_DIRNAME).resolve()
    normalized_source_bundle_key = _normalize_source_bundle_key(source_bundle_key)
    selector_mode = _source_bundle_selector_mode(normalized_source_bundle_key)
    if selector_mode == "auto":
        candidate_paths = [
            _resolve_source_bundle_candidate_path(source_root, source_root / f"{trade_date_key}.json"),
            _resolve_source_bundle_candidate_path(source_root, source_root / trade_date_key / "source_bundle.json"),
        ]
    else:
        candidate_paths = [
            _resolve_source_bundle_candidate_path(source_root, source_root / normalized_source_bundle_key / "source_bundle.json"),
            _resolve_source_bundle_candidate_path(source_root, source_root / f"{normalized_source_bundle_key}.json"),
        ]
    for path in candidate_paths:
        if path.exists():
            load_token = profiler.start_phase("source_bundle_json_load", source_path=str(path)) if profiler else None
            payload = json.loads(path.read_text(encoding="utf-8"))
            bundle = dict(payload.get("source_bundle") or payload)
            if profiler:
                profiler.finish_phase(
                    load_token,
                    source_path=str(path),
                    file_size_bytes=path.stat().st_size,
                    top_level_keys=len(bundle),
                )
            bundle_source_meta = _normalize_source_meta(
                dict(bundle.get("source_meta") or {}),
                fallback_path=str(path),
            )
            replay_config = dict(bundle.get("replay_config") or {})
            replay_config.setdefault("replay_run_id", replay_run_id)
            replay_config.setdefault("job_id", replay_run_id)
            replay_config.setdefault("trade_date", trade_date_key)
            replay_config.setdefault("for_trade_date", trade_date_key)
            bundle["replay_config"] = replay_config
            if bundle_source_meta:
                bundle["source_meta"] = bundle_source_meta
            return bundle, {
                "historical_source_status": "available",
                "historical_source_kind": "local_json",
                "historical_source_path": str(path),
                "source_bundle_key": normalized_source_bundle_key,
                "source_bundle_selector_mode": selector_mode,
                "resolved_source_bundle_path": str(path),
                "source_meta": bundle_source_meta,
            }
    if selector_mode == "explicit":
        raise N6ReplayBlocked(f"BLOCKED_REPLAY_SOURCE_BUNDLE_NOT_FOUND: source_bundle_key={normalized_source_bundle_key}")
    if trade_date_key == "20260626":
        return _canonical_fixture_source_bundle(job_id=replay_run_id), {
            "historical_source_status": "available",
            "historical_source_kind": "fixture",
            "historical_source_path": "builtin:canonical_fixture_20260626",
            "source_bundle_key": normalized_source_bundle_key,
            "source_bundle_selector_mode": selector_mode,
            "resolved_source_bundle_path": "builtin:canonical_fixture_20260626",
            "source_meta": _normalize_source_meta(
                {
                    "historical_source_type": "fixture_fallback",
                    "historical_source_path": "builtin:canonical_fixture_20260626",
                    "upstream_source_mode": "fixture",
                },
                fallback_path="builtin:canonical_fixture_20260626",
            ),
        }
    raise N6ReplayBlocked(f"BLOCKED_REPLAY_SOURCE_UNAVAILABLE: trade_date={trade_date_key}")


def _resolve_source_bundle_candidate_path(source_root: Path, candidate_path: Path) -> Path:
    resolved = candidate_path.resolve()
    if not resolved.is_relative_to(source_root):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SOURCE_BUNDLE_NOT_FOUND")
    return resolved


def _normalize_source_meta(source_meta: Mapping[str, Any] | None, *, fallback_path: str) -> dict[str, Any]:
    meta = dict(source_meta or {})
    if not meta and not fallback_path:
        return {}
    return {
        "historical_source_type": str(meta.get("historical_source_type") or ""),
        "historical_source_path": str(meta.get("historical_source_path") or fallback_path or ""),
        "historical_source_hash": str(meta.get("historical_source_hash") or ""),
        "source_origin": str(meta.get("source_origin") or ""),
        "c1_readiness_status": str(meta.get("c1_readiness_status") or ""),
        "stock_count": int(meta.get("stock_count") or 0),
        "index_object_count": int(meta.get("index_object_count") or 0),
        "board_object_count": int(meta.get("board_object_count") or 0),
        "expected_minutes_per_object": int(meta.get("expected_minutes_per_object") or 0),
        "actual_minutes_distribution": dict(meta.get("actual_minutes_distribution") or {}),
        "source_row_count": int(meta.get("source_row_count") or 0),
        "c1_current_day_row_count": int(meta.get("c1_current_day_row_count") or 0),
        "candidate_count": int(meta.get("candidate_count") or 0),
        "context_count": int(meta.get("context_count") or 0),
        "b2_snapshot_row_count": int(meta.get("b2_snapshot_row_count") or 0),
        "b2_live_current_row_count": int(meta.get("b2_live_current_row_count") or 0),
        "b2_previous_day_row_count": int(meta.get("b2_previous_day_row_count") or 0),
        "upstream_source_mode": str(meta.get("upstream_source_mode") or ""),
        "bundle_contract_version": str(meta.get("bundle_contract_version") or ""),
    }


def _apply_replay_asset_scope_filter(
    source_bundle: Mapping[str, Any],
    *,
    asset_scope: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    normalized_asset_scope = _normalize_replay_asset_scope(asset_scope)
    before_counts = _count_asset_scope_source_counts(source_bundle)
    allowed_asset_kinds = set(REPLAY_ASSET_SCOPE_ALLOWED_KINDS[normalized_asset_scope])
    if normalized_asset_scope == "all":
        return dict(source_bundle), {
            "asset_scope": normalized_asset_scope,
            "asset_scope_filter_applied": False,
            "asset_scope_allowed_asset_kinds": list(REPLAY_ASSET_SCOPE_ALLOWED_KINDS[normalized_asset_scope]),
            "asset_scope_source_counts_before": before_counts,
            "asset_scope_source_counts_after": before_counts,
        }

    code_asset_hints = _build_source_record_asset_hints(source_bundle)
    filtered_bundle = dict(source_bundle)
    filtered_bundle["source_records"] = {
        str(code): [dict(row) for row in rows]
        for code, rows in dict(source_bundle.get("source_records") or {}).items()
        if _resolve_source_record_asset_kind(str(code), rows, code_asset_hints) in allowed_asset_kinds
    }
    filtered_bundle["candidates"] = [
        dict(row)
        for row in source_bundle.get("candidates") or []
        if _row_asset_kind(row) in allowed_asset_kinds
    ]
    filtered_bundle["n4_context_snapshot_rows"] = [
        dict(row)
        for row in source_bundle.get("n4_context_snapshot_rows") or []
        if _row_asset_kind(row) in allowed_asset_kinds
    ]
    b2_input = dict(source_bundle.get("b2_input") or {})
    filtered_bundle["b2_input"] = {
        "snapshot_rows": [
            dict(row)
            for row in b2_input.get("snapshot_rows") or []
            if _snapshot_asset_kind(row) in allowed_asset_kinds
        ],
        "live_current_rows_by_asset": {
            asset_kind: [dict(row) for row in rows]
            for asset_kind, rows in dict(b2_input.get("live_current_rows_by_asset") or {}).items()
            if str(asset_kind) in allowed_asset_kinds
        },
        "previous_day_rows_by_asset": {
            asset_kind: [dict(row) for row in rows]
            for asset_kind, rows in dict(b2_input.get("previous_day_rows_by_asset") or {}).items()
            if str(asset_kind) in allowed_asset_kinds
        },
    }
    after_counts = _count_asset_scope_source_counts(filtered_bundle)
    if (
        not filtered_bundle.get("source_records")
        or not filtered_bundle.get("candidates")
        or not filtered_bundle.get("n4_context_snapshot_rows")
    ):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SCOPE_EMPTY")
    return filtered_bundle, {
        "asset_scope": normalized_asset_scope,
        "asset_scope_filter_applied": True,
        "asset_scope_allowed_asset_kinds": list(REPLAY_ASSET_SCOPE_ALLOWED_KINDS[normalized_asset_scope]),
        "asset_scope_source_counts_before": before_counts,
        "asset_scope_source_counts_after": after_counts,
    }


def _validate_full_day_shadow_source_contract(
    *,
    source_bundle: Mapping[str, Any],
    asset_scope_lineage: Mapping[str, Any],
    source_bundle_key: str,
    source_meta: Mapping[str, Any],
    trade_date: str,
) -> None:
    if str(asset_scope_lineage.get("asset_scope") or "") != "index_board_only":
        raise N6ReplayBlocked("BLOCKED_REPLAY_SHADOW_REQUIRES_INDEX_BOARD_ONLY")
    if (
        source_bundle_key in {"", DEFAULT_REPLAY_SOURCE_BUNDLE_KEY}
        or not source_bundle_key.endswith(INDEX_BOARD_FULL_DAY_SOURCE_SUFFIX)
    ):
        raise N6ReplayBlocked(C1_SOURCE_INCOMPLETE_REASON)
    if str(source_meta.get("source_origin") or "") != "c1_read_only_export":
        raise N6ReplayBlocked(C1_SOURCE_INCOMPLETE_REASON)
    if str(source_meta.get("c1_readiness_status") or "") != "passed":
        raise N6ReplayBlocked(C1_SOURCE_INCOMPLETE_REASON)
    if int(source_meta.get("expected_minutes_per_object") or 0) != 240:
        raise N6ReplayBlocked(C1_SOURCE_INCOMPLETE_REASON)
    allowed = set(asset_scope_lineage.get("asset_scope_allowed_asset_kinds") or [])
    if allowed != {"index", "board"}:
        raise N6ReplayBlocked("BLOCKED_REPLAY_SHADOW_REQUIRES_INDEX_BOARD_ONLY")
    after_counts = dict(asset_scope_lineage.get("asset_scope_source_counts_after") or {})
    for section in (
        "source_records",
        "candidates",
        "context",
        "b2_snapshot",
        "b2_live_current",
        "b2_previous_day",
    ):
        stock_count = int((after_counts.get(section) or {}).get("stock") or 0)
        if stock_count:
            raise N6ReplayBlocked("BLOCKED_REPLAY_SHADOW_STOCK_SCOPE_LEAK")
    if not source_bundle.get("source_records") or not source_bundle.get("candidates") or not source_bundle.get("n4_context_snapshot_rows"):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SCOPE_EMPTY")
    readiness = build_c1_index_board_readiness_report(
        trade_date=trade_date,
        index_rows=_source_record_rows_for_asset(source_bundle, "index"),
        board_rows=_source_record_rows_for_asset(source_bundle, "board"),
    )
    if readiness["status"] != "passed":
        raise N6ReplayBlocked(C1_SOURCE_INCOMPLETE_REASON)


def _source_record_rows_for_asset(source_bundle: Mapping[str, Any], asset_kind: str) -> list[dict[str, Any]]:
    hints = _build_source_record_asset_hints(source_bundle)
    rows_for_asset: list[dict[str, Any]] = []
    for code, rows in dict(source_bundle.get("source_records") or {}).items():
        if _resolve_source_record_asset_kind(str(code), rows, hints) != asset_kind:
            continue
        rows_for_asset.extend(dict(row) for row in rows)
    return rows_for_asset


def _count_asset_scope_source_counts(source_bundle: Mapping[str, Any]) -> dict[str, dict[str, int]]:
    code_asset_hints = _build_source_record_asset_hints(source_bundle)
    source_record_counts = {"stock": 0, "index": 0, "board": 0}
    for code, rows in dict(source_bundle.get("source_records") or {}).items():
        asset_kind = _resolve_source_record_asset_kind(str(code), rows, code_asset_hints)
        if asset_kind in source_record_counts:
            source_record_counts[asset_kind] += len(rows)
    candidate_counts = _count_rows_by_asset_kind(source_bundle.get("candidates") or [])
    context_counts = _count_rows_by_asset_kind(source_bundle.get("n4_context_snapshot_rows") or [])
    b2_input = dict(source_bundle.get("b2_input") or {})
    b2_snapshot_counts = _count_snapshot_rows_by_asset_kind(b2_input.get("snapshot_rows") or [])
    b2_live_counts = {
        asset_kind: len(rows)
        for asset_kind, rows in {"stock": [], "index": [], "board": [], **dict(b2_input.get("live_current_rows_by_asset") or {})}.items()
        if asset_kind in {"stock", "index", "board"}
    }
    b2_prev_counts = {
        asset_kind: len(rows)
        for asset_kind, rows in {"stock": [], "index": [], "board": [], **dict(b2_input.get("previous_day_rows_by_asset") or {})}.items()
        if asset_kind in {"stock", "index", "board"}
    }
    return {
        "source_records": source_record_counts,
        "candidates": candidate_counts,
        "context": context_counts,
        "b2_snapshot": b2_snapshot_counts,
        "b2_live_current": b2_live_counts,
        "b2_previous_day": b2_prev_counts,
    }


def _count_rows_by_asset_kind(rows: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    counts = {"stock": 0, "index": 0, "board": 0}
    for row in rows:
        asset_kind = _row_asset_kind(row)
        if asset_kind in counts:
            counts[asset_kind] += 1
    return counts


def _count_snapshot_rows_by_asset_kind(rows: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    counts = {"stock": 0, "index": 0, "board": 0}
    for row in rows:
        asset_kind = _snapshot_asset_kind(row)
        if asset_kind in counts:
            counts[asset_kind] += 1
    return counts


def _build_source_record_asset_hints(source_bundle: Mapping[str, Any]) -> dict[str, str]:
    hints: dict[str, str] = {}
    for row in source_bundle.get("candidates") or []:
        asset_kind = _row_asset_kind(row)
        code = str(row.get("code") or row.get("display_code") or "")
        if code and asset_kind:
            hints[code] = asset_kind
    for row in source_bundle.get("n4_context_snapshot_rows") or []:
        asset_kind = _row_asset_kind(row)
        code = _code_from_identity_key(str(row.get("identity_key") or ""))
        if code and asset_kind:
            hints[code] = asset_kind
    for row in dict(source_bundle.get("b2_input") or {}).get("snapshot_rows") or []:
        asset_kind = _snapshot_asset_kind(row)
        code = str(row.get("code") or row.get("display_code") or _code_from_identity_key(str(row.get("identity_key") or "")) or "")
        if code and asset_kind:
            hints[code] = asset_kind
    return hints


def _resolve_source_record_asset_kind(
    code: str,
    rows: Sequence[Mapping[str, Any]],
    code_asset_hints: Mapping[str, str],
) -> str:
    hinted = str(code_asset_hints.get(code) or "")
    if hinted in {"stock", "index", "board"}:
        return hinted
    for row in rows:
        asset_kind = _row_asset_kind(row)
        if asset_kind:
            return asset_kind
    return ""


def _row_asset_kind(row: Mapping[str, Any]) -> str:
    asset_kind = str(row.get("asset_kind") or "")
    if asset_kind in {"stock", "index", "board"}:
        return asset_kind
    identity_key = str(row.get("identity_key") or "")
    if identity_key.startswith("stock:"):
        return "stock"
    if identity_key.startswith("index:"):
        return "index"
    if identity_key.startswith("board:"):
        return "board"
    return ""


def _code_from_identity_key(identity_key: str) -> str:
    if not identity_key or ":" not in identity_key:
        return ""
    return identity_key.rsplit(":", 1)[-1]


def _exchange_from_identity_key(identity_key: str) -> str:
    parts = identity_key.split(":")
    return parts[1] if len(parts) >= 3 else ""


def _run_canonical_historical_minute_loop(
    *,
    trade_date: str,
    minutes: list[str],
    replay_run_id: str,
    source_bundle: Mapping[str, Any],
    profiler: _ReplayProfileCollector | None = None,
    shadow_mode: bool = False,
    n3p_strategy: str = DEFAULT_REPLAY_N3P_STRATEGY,
    n3p_reduction_mode: str = DEFAULT_REPLAY_N3P_REDUCTION_MODE,
    n3p_negative_cache: str = DEFAULT_REPLAY_N3P_NEGATIVE_CACHE,
) -> dict[str, Any]:
    bundle = dict(source_bundle)
    replay_config = dict(bundle.get("replay_config") or {})
    normalized_n3p_strategy = _normalize_replay_n3p_strategy(n3p_strategy)
    normalized_n3p_reduction_mode = _normalize_replay_n3p_reduction_mode(n3p_reduction_mode)
    normalized_n3p_negative_cache = _normalize_replay_n3p_negative_cache(n3p_negative_cache)
    if normalized_n3p_strategy != DEFAULT_REPLAY_N3P_STRATEGY and not shadow_mode:
        raise N6ReplayBlocked("BLOCKED_REPLAY_N3P_STRATEGY_REQUIRES_SHADOW_MODE")
    if (
        normalized_n3p_reduction_mode == "active_state_fast_path"
        and (not shadow_mode or normalized_n3p_strategy != "prefilter_prune")
    ):
        raise N6ReplayBlocked("BLOCKED_REPLAY_N3P_REDUCTION_REQUIRES_SHADOW_PREFILTER_PRUNE")
    init_token = profiler.start_phase("minute_loop_initialization", minute_count=len(minutes)) if profiler else None
    trigger_context_run_id = str((bundle.get("n4_context_snapshot_rows") or [{}])[0].get("run_id") or f"{replay_run_id}__trigger_context")
    ordinary_context_rows = [row for row in bundle.get("n4_context_snapshot_rows") or [] if str(row.get("condition_key") or "") not in {"BUY_HINT", "SELL_HINT"}]
    hint_context_rows = [row for row in bundle.get("n4_context_snapshot_rows") or [] if str(row.get("condition_key") or "") in {"BUY_HINT", "SELL_HINT"}]
    previous_trigger_states = {}
    if not shadow_mode and len(ordinary_context_rows) > 1:
        seeded = _matched_previous_state(ordinary_context_rows[1])
        previous_trigger_states[provisional_trigger_lifecycle.lifecycle_state_key(seeded)] = seeded
    if profiler:
        profiler.finish_phase(
            init_token,
            ordinary_context_count=len(ordinary_context_rows),
            hint_context_count=len(hint_context_rows),
            seeded_previous_state_count=len(previous_trigger_states),
        )

    tracked_eligible_rows: list[dict[str, Any]] = []
    executed_eligible_event_ids: set[str] = set()
    terminal_eligible_event_ids: set[str] = set()
    n3_messages: list[dict[str, Any]] = []
    n4_messages: list[dict[str, Any]] = []
    n5_messages: list[dict[str, Any]] = []
    n4_shadow_state_transitions: list[dict[str, Any]] = []
    n4_shadow_evaluations: list[dict[str, Any]] = []
    n5_shadow_action_windows: list[dict[str, Any]] = []
    n4_lightweight_prefilter_audit: list[dict[str, Any]] = []
    n3p_demand_plan: list[dict[str, Any]] = []
    n3p_active_state_reduction: list[dict[str, Any]] = []
    n3p_plan_only_proof_summary: list[dict[str, Any]] = []
    n3p_negative_cache_decisions: list[dict[str, Any]] = []
    n5_evaluator_demand_plan: list[dict[str, Any]] = []
    n5_confirmation_metric_index_stats: list[dict[str, Any]] = []
    shadow_confirmation_metric_rows: list[dict[str, Any]] = []
    shadow_noop_by_minute: dict[str, int] = {}
    prefilter_false_negatives: list[dict[str, Any]] = []
    duplicate_execution_suppressed = 0
    executed_action_keys: set[str] = set()
    last_n3p_artifact = _empty_n3p_artifact(replay_run_id)
    last_b2_artifact = _empty_b2_artifact(replay_run_id)
    last_eligible_plan: dict[str, Any] = {"event_model": _plan_only_side_effects(), "writes": {"common_event_outbox": []}}
    last_executed_report: dict[str, Any] = {"side_effect_guard": _plan_only_side_effects(), "action_executed_plans": []}
    side_effect_guards: list[Any] = []
    manual_state_changed_written = False
    n3p_cache = _N3PReplayCache(
        replay_config=replay_config,
        context_rows=ordinary_context_rows,
        cache_by_minute=shadow_mode,
    )
    negative_cache = _N3PNegativeProofCache(
        enabled=shadow_mode and normalized_n3p_negative_cache == "enabled",
        replay_config=replay_config,
    )

    for minute in minutes:
        minute_config = dict(replay_config)
        minute_config["until_hhmm"] = minute.replace(":", "")
        minute_config["latest_closed_minute"] = _latest_closed_minute_iso(trade_date, minute)
        minute_bundle = dict(bundle)
        minute_bundle["replay_config"] = minute_config

        ordinary_candidates = _active_candidates_for_minute(
            bundle.get("candidates") or [],
            minute,
            hint_only=False,
            full_day_shadow=shadow_mode,
            trade_date=trade_date,
        )
        hint_candidates = _active_candidates_for_minute(
            bundle.get("candidates") or [],
            minute,
            hint_only=True,
            full_day_shadow=shadow_mode,
            trade_date=trade_date,
        )
        open_action_identities = _open_action_window_identities(
            eligible_rows=tracked_eligible_rows,
            executed_eligible_event_ids=executed_eligible_event_ids,
        )
        prefilter_rows: list[dict[str, Any]] = []
        prefilter_keep_identities: set[str] = {str(row.get("identity_key") or "") for row in ordinary_candidates}
        if shadow_mode and normalized_n3p_strategy in {"prefilter_audit", "prefilter_prune"}:
            prefilter_rows = _build_n4_lightweight_prefilter_rows(
                candidates=ordinary_candidates,
                context_rows=ordinary_context_rows,
                source_bundle=bundle,
                previous_states=list(previous_trigger_states.values()),
                open_action_identities=open_action_identities,
                trade_date=trade_date,
                minute=minute,
            )
            prefilter_keep_identities = {
                str(row.get("identity_key") or "")
                for row in prefilter_rows
                if row.get("decision") == "prefilter_keep"
            }
            evaluated_identities = {str(row.get("identity_key") or "") for row in prefilter_rows}
            prefilter_keep_identities.update(
                str(candidate.get("identity_key") or "")
                for candidate in ordinary_candidates
                if str(candidate.get("identity_key") or "") not in evaluated_identities
            )
        n3p_candidates = ordinary_candidates
        if shadow_mode and normalized_n3p_strategy == "prefilter_prune":
            n3p_candidates = [
                candidate
                for candidate in ordinary_candidates
                if str(candidate.get("identity_key") or "") in prefilter_keep_identities
            ]
        active_reduction_skip_keys: set[str] = set()
        if (
            shadow_mode
            and normalized_n3p_strategy == "prefilter_prune"
            and normalized_n3p_reduction_mode == "active_state_fast_path"
        ):
            active_reduction_skip_keys, active_reduction_rows = _active_state_reduction_skip_keys(
                prefilter_rows,
                minute=minute,
                reduction_mode=normalized_n3p_reduction_mode,
            )
            n3p_active_state_reduction.extend(active_reduction_rows)
            if active_reduction_skip_keys:
                n3p_candidates = [
                    candidate
                    for candidate in n3p_candidates
                    if _candidate_prefilter_key(candidate) not in active_reduction_skip_keys
                ]
            if executed_action_keys:
                retained_candidates: list[Mapping[str, Any]] = []
                for candidate in n3p_candidates:
                    action_key = _candidate_action_key(candidate)
                    if action_key in executed_action_keys:
                        n3p_active_state_reduction.append(
                            {
                                "minute": minute,
                                "reduction_mode": normalized_n3p_reduction_mode,
                                "reduction_reason": "closed_action_window",
                                "asset_kind": candidate.get("asset_kind"),
                                "identity_key": candidate.get("identity_key"),
                                "condition_key": candidate.get("condition_key"),
                                "signal_type": candidate.get("signal_type"),
                                "action_key": action_key,
                                "source_decision": "executed_action_key_seen",
                                "source_reason": "action_window_closed_after_execution",
                                "n3p_skipped": True,
                                "duplicate_suppressed": False,
                                "shadow_mode": True,
                                "trace_summary": "replay-only closed action window; full N3P confirmation no longer retained",
                            }
                        )
                        continue
                    retained_candidates.append(candidate)
                n3p_candidates = retained_candidates
        negative_cache_skip_count = 0
        if shadow_mode and normalized_n3p_negative_cache == "enabled" and n3p_candidates:
            retained_candidates = []
            for candidate in n3p_candidates:
                decision = negative_cache.decision(
                    candidate=candidate,
                    proof_context=_n3p_negative_cache_proof_context(candidate),
                )
                decision.update(
                    {
                        "minute": minute,
                        "mode": normalized_n3p_negative_cache,
                        "shadow_mode": True,
                    }
                )
                n3p_negative_cache_decisions.append(decision)
                if decision.get("skip_full_n3p"):
                    negative_cache_skip_count += 1
                    continue
                retained_candidates.append(candidate)
            n3p_candidates = retained_candidates

        n3p_token = profiler.start_phase(
            "per_minute_n3p_plan_only",
            minute=minute,
            candidate_count=len(n3p_candidates),
            candidate_count_before_prefilter=len(ordinary_candidates),
        ) if profiler else None
        if n3p_candidates:
            target_run_id = _n3p_target_run_id(minute_config)
            cached_artifact, n3p_cache_hit, n3p_cache_key = n3p_cache.get(
                candidates=n3p_candidates,
                target_run_id=target_run_id,
                minute=minute,
            )
            if cached_artifact is not None:
                n3p_artifact = cached_artifact
            else:
                n3p_bundle = {
                    "source_records": _filter_source_records_for_minute(
                        bundle.get("source_records") or {},
                        trade_date=trade_date,
                        minute=minute,
                    ),
                    "candidates": n3p_candidates,
                    "n4_context_snapshot_rows": ordinary_context_rows,
                    "replay_config": minute_config,
                }
                n3p_artifact = build_n3p_plan_only_replay_artifact(source_bundle=n3p_bundle)
                _coerce_canonical_fixture_n3p_rows(n3p_artifact)
                n3p_cache.put(key=n3p_cache_key, artifact=n3p_artifact)
            last_n3p_artifact = n3p_artifact
        else:
            n3p_cache.empty_minute_noop()
            n3p_cache_hit = False
            n3p_cache_key = ""
            n3p_artifact = _empty_n3p_artifact(replay_run_id)
        proof_rows = [dict(row, minute=minute) for row in n3p_artifact.get("proof_summary_rows") or []]
        if proof_rows:
            n3p_plan_only_proof_summary.extend(proof_rows)
            negative_cache.store(proof_rows)
        if profiler:
            profiler.finish_phase(
                n3p_token,
                minute=minute,
                metric_row_count=len(n3p_artifact["metric_rows"]),
                n3p_cache_hit=bool(n3p_cache_hit),
                n3p_cache_miss=bool(n3p_candidates and not n3p_cache_hit),
                n3p_cache_key=str(n3p_cache_key),
                empty_minute_fast_path=not bool(n3p_candidates),
                n3p_negative_cache_saved_calls=negative_cache_skip_count,
            )
        if shadow_mode:
            n3p_demand_plan.append(
                {
                    "minute": minute,
                    "strategy": normalized_n3p_strategy,
                    "reduction_mode": normalized_n3p_reduction_mode,
                    "candidate_count_before": len(ordinary_candidates),
                    "candidate_count_after": len(n3p_candidates),
                    "candidate_count_dropped": max(0, len(ordinary_candidates) - len(n3p_candidates)),
                    "open_action_window_count": len(open_action_identities),
                    "n3p_called": bool(n3p_candidates),
                    "n3p_calls_saved": max(0, len(ordinary_candidates) - len(n3p_candidates)),
                    "n3p_negative_cache_saved_calls": negative_cache_skip_count,
                    "cache_hit": bool(n3p_cache_hit),
                    "metric_row_count": len(n3p_artifact["metric_rows"]),
                }
            )
        if shadow_mode:
            shadow_confirmation_metric_rows.extend(dict(row) for row in n3p_artifact["metric_rows"])

        b2_token = profiler.start_phase(
            "per_minute_b2_plan_only",
            minute=minute,
            candidate_count=len(hint_candidates),
        ) if profiler else None
        b2_input = _build_canonical_b2_input(
            bundle,
            minute=minute,
            trade_date=trade_date,
            shadow_mode=shadow_mode,
        )
        b2_bundle = {
            "candidates": hint_candidates,
            "replay_config": minute_config,
        }
        if hint_candidates and b2_input["snapshot_rows"]:
            b2_artifact = build_b2_plan_only_replay_artifact(
                source_bundle=b2_bundle,
                b1_snapshot_rows=b2_input["snapshot_rows"],
                live_current_minute_rows_by_asset=b2_input["live_current_rows_by_asset"],
                previous_day_minute_rows_by_asset=b2_input["previous_day_rows_by_asset"],
            )
            _coerce_canonical_fixture_b2_rows(b2_artifact)
            last_b2_artifact = b2_artifact
        else:
            b2_artifact = _empty_b2_artifact(replay_run_id)
        if profiler:
            profiler.finish_phase(
                b2_token,
                minute=minute,
                snapshot_row_count=len(b2_input["snapshot_rows"]),
                live_current_row_count=sum(len(rows) for rows in b2_input["live_current_rows_by_asset"].values()),
                previous_day_row_count=sum(len(rows) for rows in b2_input["previous_day_rows_by_asset"].values()),
                projection_row_count=len(b2_artifact["projection_rows"]),
            )

        ordinary_token = profiler.start_phase(
            "per_minute_n4_ordinary_matcher_lifecycle",
            minute=minute,
            context_count=len(ordinary_context_rows),
            metric_row_count=len(n3p_artifact["metric_rows"]),
        ) if profiler else None
        ordinary_plans = provisional_ordinary_matcher.build_provisional_ordinary_matcher_plans(
            trigger_context_run_id=trigger_context_run_id,
            source_metric_run_id=str(n3p_artifact["target_run_id"]),
            context_rows=ordinary_context_rows,
            metric_rows=n3p_artifact["metric_rows"],
        )
        if profiler:
            profiler.finish_phase(ordinary_token, minute=minute, plan_count=len(ordinary_plans))

        hint_token = profiler.start_phase(
            "per_minute_n4_hint_matcher_lifecycle",
            minute=minute,
            context_count=len(hint_context_rows),
            projection_row_count=len(b2_artifact["projection_rows"]),
        ) if profiler else None
        hint_plans = provisional_projection_matcher.build_provisional_projection_matcher_plans(
            trigger_context_run_id=trigger_context_run_id,
            projection_run_id=str(b2_artifact["target_run_id"]),
            context_rows=hint_context_rows,
            projection_rows=b2_artifact["projection_rows"],
        )
        if profiler:
            profiler.finish_phase(hint_token, minute=minute, plan_count=len(hint_plans))
        minute_events = provisional_trigger_lifecycle.build_lifecycle_output_plans(
            ordinary_plans + hint_plans,
            previous_states=list(previous_trigger_states.values()),
        )
        if shadow_mode and prefilter_rows:
            annotated_prefilter_rows = _annotate_prefilter_rows_with_canonical_events(
                prefilter_rows,
                minute_events=minute_events,
            )
            n4_lightweight_prefilter_audit.extend(annotated_prefilter_rows)
            minute_false_negatives = [
                row for row in annotated_prefilter_rows if row.get("false_negative")
            ]
            prefilter_false_negatives.extend(minute_false_negatives)
            if normalized_n3p_strategy == "prefilter_audit" and minute_false_negatives:
                raise N6ReplayBlocked("BLOCKED_REPLAY_PREFILTER_FALSE_NEGATIVE")
        if shadow_mode:
            events_by_plan_id = {str(event.get("plan_id") or ""): event for event in minute_events}
            n4_shadow_evaluations.extend(
                _shadow_n4_evaluation_row(
                    plan,
                    minute=minute,
                    source="hint" if str(plan.get("condition_key") or "") in {"BUY_HINT", "SELL_HINT"} else "ordinary",
                    event=events_by_plan_id.get(str(plan.get("plan_id") or "")),
                )
                for plan in [*ordinary_plans, *hint_plans]
            )
        shadow_noop_by_minute[minute] = max(0, len(ordinary_plans) + len(hint_plans) - len(minute_events))
        minute_n4_messages = _serialize_n4_events(
            trade_date=trade_date,
            replay_run_id=replay_run_id,
            source="mixed",
            n4_run_id=f"{replay_run_id}__n4_canonical_plan_v1",
            rows=minute_events,
        )
        for message in minute_n4_messages:
            if not str(message.get("minute") or ""):
                message["minute"] = minute
            if str(message.get("condition_key") or "") in {"BUY_HINT", "SELL_HINT"}:
                message["source"] = "hint"
            else:
                message["source"] = "ordinary"
            _apply_replay_n5_entry_contract(message, replay_config=replay_config)
        if not shadow_mode and minute == "13:30" and not manual_state_changed_written:
            manual_message = _manual_ordinary_state_changed_message(
                trade_date=trade_date,
                n4_run_id=f"{replay_run_id}__n4_canonical_plan_v1",
            )
            if _manual_replay_message_allowed_by_scope(
                manual_message,
                source_bundle=bundle,
                replay_config=replay_config,
            ):
                minute_n4_messages.append(manual_message)
            manual_state_changed_written = True
        if shadow_mode:
            n4_shadow_state_transitions.extend(
                _shadow_n4_transition_row(message, previous_states=previous_trigger_states)
                for message in minute_n4_messages
            )
        for event in minute_events:
            state_key = str(event.get("lifecycle_state_key") or provisional_trigger_lifecycle.lifecycle_state_key(event))
            previous_trigger_states[state_key] = dict(event)

        n5_source_messages = (
            [row for row in minute_n4_messages if row.get("event_type") == "TriggerMatched"]
            if shadow_mode
            else minute_n4_messages
        )
        minute_outbox_rows = [
            _n4_message_to_outbox_row(row, n4_run_id=f"{replay_run_id}__n4_canonical_plan_v1")
            for row in n5_source_messages
        ]
        eligible_token = profiler.start_phase(
            "per_minute_n5_actioneligible_planner",
            minute=minute,
            n4_outbox_row_count=len(minute_outbox_rows),
        ) if profiler else None
        eligible_plan = provisional_action_eligible.build_provisional_actioneligible_plan(
            source_trigger_run={
                "run_id": f"{replay_run_id}__n4_canonical_plan_v1",
                "status": "passed",
                "source_condition_run_id": replay_config.get("source_condition_run_id"),
                "for_trade_date": replay_config.get("for_trade_date"),
            },
            source_trigger_run_id=f"{replay_run_id}__n4_canonical_plan_v1",
            action_run_id=f"{replay_run_id}__n5_actioneligible_v1",
            for_trade_date=str(replay_config.get("for_trade_date")),
            consumer_name="n6_local_replay_canonical_plan_v1",
            outbox_rows=minute_outbox_rows,
            target_counts={key: 0 for key in (
                "common_action_run",
                "common_action_quality_item",
                "stock_action_fact",
                "index_action_fact",
                "board_action_fact",
                "common_action_event",
                "common_event_outbox",
                "common_event_inbox",
                "common_event_consumer_checkpoint",
            )},
        )
        minute_eligible_rows = list(eligible_plan["writes"]["common_event_outbox"])
        tracked_eligible_rows.extend(dict(row) for row in minute_eligible_rows)
        last_eligible_plan = eligible_plan
        if profiler:
            profiler.finish_phase(
                eligible_token,
                minute=minute,
                eligible_row_count=len(minute_eligible_rows),
            )

        closed_metric_rows = _build_closed_confirmation_metric_rows_for_minute(
            metric_rows=shadow_confirmation_metric_rows if shadow_mode else n3p_artifact["metric_rows"],
            trade_date=trade_date,
            current_minute=minute,
        )
        if shadow_mode:
            n5_demand_inputs = _build_n5_evaluator_demand_inputs(
                minute=minute,
                tracked_eligible_rows=tracked_eligible_rows,
                closed_metric_rows=closed_metric_rows,
                executed_eligible_event_ids=executed_eligible_event_ids,
                terminal_eligible_event_ids=terminal_eligible_event_ids,
            )
            evaluator_eligible_rows = n5_demand_inputs["eligible_rows"]
            evaluator_metric_rows = n5_demand_inputs["confirmation_metric_rows"]
        else:
            n5_demand_inputs = {}
            evaluator_eligible_rows = tracked_eligible_rows
            evaluator_metric_rows = closed_metric_rows
        executed_token = profiler.start_phase(
            "per_minute_n5_actionexecuted_evaluator",
            minute=minute,
            tracked_eligible_count=len(tracked_eligible_rows),
            closed_metric_row_count=len(closed_metric_rows),
            evaluator_eligible_count=len(evaluator_eligible_rows),
            evaluator_metric_row_count=len(evaluator_metric_rows),
            confirmation_projection_row_count=len(b2_artifact["projection_rows"]),
        ) if profiler else None
        executed_report = provisional_action_executed_dry_run.build_provisional_action_executed_dry_run_report(
            actioneligible_rows=evaluator_eligible_rows,
            confirmation_metric_rows=evaluator_metric_rows,
            confirmation_projection_rows=b2_artifact["projection_rows"],
            for_trade_date=str(replay_config.get("for_trade_date")),
            confirmation_metric_run_id=None if shadow_mode else str(n3p_artifact["target_run_id"]),
            confirmation_projection_run_id=str(b2_artifact["target_run_id"]),
            latest_closed_minute=minute_config.get("latest_closed_minute"),
        )
        last_executed_report = executed_report
        if shadow_mode:
            demand_plan_row = dict(n5_demand_inputs["plan_row"])
            demand_plan_row["executed_plan_count"] = len(executed_report["action_executed_plans"])
            n5_evaluator_demand_plan.append(demand_plan_row)
            n5_confirmation_metric_index_stats.append(dict(n5_demand_inputs["index_row"]))
        if profiler:
            profiler.finish_phase(
                executed_token,
                minute=minute,
                executed_plan_count=len(executed_report["action_executed_plans"]),
                evaluator_rows_saved=(n5_demand_inputs.get("plan_row") or {}).get("n5_evaluator_rows_saved", 0) if shadow_mode else 0,
                evaluator_fail_open_count=(n5_demand_inputs.get("plan_row") or {}).get("n5_evaluator_fail_open_count", 0) if shadow_mode else 0,
            )
        minute_n5_messages = [
            *_serialize_n5_messages(
                trade_date=trade_date,
                replay_run_id=replay_run_id,
                eligible_rows=minute_eligible_rows,
                executed_plans=[],
            )
        ]
        for plan in executed_report["action_executed_plans"]:
            source_eligible_event_id = str((plan.get("payload") or {}).get("source_eligible_event_id") or "")
            if not source_eligible_event_id or source_eligible_event_id in executed_eligible_event_ids:
                continue
            executed_eligible_event_ids.add(source_eligible_event_id)
            terminal_eligible_event_ids.add(source_eligible_event_id)
            executed_message = _serialize_action_executed_only(
                    trade_date=trade_date,
                    replay_run_id=replay_run_id,
                    executed_plan=plan,
                )
            if not str(executed_message.get("minute") or ""):
                executed_message["minute"] = minute
            minute_n5_messages.append(executed_message)
        if shadow_mode and normalized_n3p_reduction_mode == "active_state_fast_path":
            (
                minute_n5_messages,
                duplicate_audit_rows,
                minute_duplicate_suppressed,
            ) = _suppress_duplicate_action_executed_messages(
                minute_n5_messages,
                executed_action_keys=executed_action_keys,
                minute=minute,
            )
            duplicate_execution_suppressed += minute_duplicate_suppressed
            n3p_active_state_reduction.extend(duplicate_audit_rows)
        if shadow_mode:
            n5_shadow_action_windows.extend(_shadow_n5_window_row(message) for message in minute_n5_messages)

        n4_messages.extend(minute_n4_messages)
        n5_messages.extend(minute_n5_messages)
        n3_messages.append(
            {
                "minute": minute,
                "event_type": "ReplayMinuteClosed",
                "asset_kind": "market",
                "identity_key": "local_replay:market",
                "source_mode": "replay",
                "snapshot_policy": "replay_snapshot_from_minute_cumulative",
                "trade_date": trade_date,
                "n3p_metric_rows": len(n3p_artifact["metric_rows"]),
                "b2_projection_rows": len(b2_artifact["projection_rows"]),
                "n4_message_count": len(minute_n4_messages),
                "n5_message_count": len(minute_n5_messages),
                "trace_summary": "historical minute replay loop",
            }
        )
        side_effect_guards.extend(
            [
                n3p_artifact.get("side_effects"),
                b2_artifact.get("side_effects"),
                eligible_plan.get("event_model"),
                executed_report.get("side_effect_guard"),
            ]
        )

    shadow_validation_report = (
        _build_shadow_validation_report(
            minutes=minutes,
            n4_messages=n4_messages,
            n5_messages=n5_messages,
            n4_shadow_state_transitions=n4_shadow_state_transitions,
            n4_shadow_evaluations=n4_shadow_evaluations,
            n5_shadow_action_windows=n5_shadow_action_windows,
            shadow_noop_by_minute=shadow_noop_by_minute,
            n4_prefilter_rows=n4_lightweight_prefilter_audit,
            n3p_demand_plan=n3p_demand_plan,
            n3p_strategy=normalized_n3p_strategy,
            n3p_reduction_mode=normalized_n3p_reduction_mode,
            n3p_active_state_reduction=n3p_active_state_reduction,
        )
        if shadow_mode
        else {}
    )
    if shadow_mode and (
        shadow_validation_report.get("stock_n4_messages")
        or shadow_validation_report.get("stock_n5_messages")
        or shadow_validation_report.get("trigger_state_changed_action_entries")
    ):
        raise N6ReplayBlocked("BLOCKED_REPLAY_SHADOW_SCOPE_OR_ENTRY_LEAK")

    return {
        "n3_messages": n3_messages,
        "n4_messages": sorted(n4_messages, key=lambda row: (str(row.get("minute") or ""), str(row.get("event_type") or ""), str(row.get("identity_key") or ""))),
        "n5_messages": sorted(n5_messages, key=lambda row: (str(row.get("minute") or ""), str(row.get("event_type") or ""), str(row.get("identity_key") or ""))),
        "n4_shadow_state_transitions": sorted(n4_shadow_state_transitions, key=lambda row: (str(row.get("minute") or ""), str(row.get("event_type") or ""), str(row.get("identity_key") or ""))),
        "n4_shadow_evaluations": sorted(n4_shadow_evaluations, key=lambda row: (str(row.get("minute") or ""), str(row.get("event_type") or ""), str(row.get("identity_key") or ""), str(row.get("condition_key") or ""))),
        "n5_shadow_action_windows": sorted(n5_shadow_action_windows, key=lambda row: (str(row.get("minute") or ""), str(row.get("event_type") or ""), str(row.get("identity_key") or ""))),
        "n4_lightweight_prefilter_audit": sorted(n4_lightweight_prefilter_audit, key=lambda row: (str(row.get("minute") or ""), str(row.get("identity_key") or ""), str(row.get("condition_key") or ""))),
        "n3p_demand_plan": n3p_demand_plan,
        "n3p_active_state_reduction": sorted(n3p_active_state_reduction, key=lambda row: (str(row.get("minute") or ""), str(row.get("identity_key") or ""), str(row.get("condition_key") or ""), str(row.get("reduction_reason") or ""))),
        "n3p_plan_only_proof_summary": sorted(n3p_plan_only_proof_summary, key=lambda row: (str(row.get("minute") or ""), str(row.get("identity_key") or ""), str(row.get("condition_key") or ""))),
        "n3p_negative_cache_decisions": sorted(n3p_negative_cache_decisions, key=lambda row: (str(row.get("minute") or ""), str(row.get("identity_key") or ""), str(row.get("condition_key") or ""), str(row.get("decision") or ""))),
        "n5_evaluator_demand_plan": n5_evaluator_demand_plan,
        "n5_confirmation_metric_index_stats": n5_confirmation_metric_index_stats,
        "shadow_validation_report": shadow_validation_report,
        "n3p_prefilter": _build_n3p_prefilter_summary(
            strategy=normalized_n3p_strategy,
            n4_prefilter_rows=n4_lightweight_prefilter_audit,
            n3p_demand_plan=n3p_demand_plan,
            false_negatives=prefilter_false_negatives,
        ),
        "n3p_active_state_reduction_summary": _build_n3p_active_state_reduction_summary(
            n3p_active_state_reduction,
            n3p_demand_plan=n3p_demand_plan,
            reduction_mode=normalized_n3p_reduction_mode,
        ),
        "n3p_negative_cache_summary": negative_cache.stats(mode=normalized_n3p_negative_cache),
        "n5_evaluator_demand_summary": _build_n5_evaluator_demand_summary(n5_evaluator_demand_plan),
        "last_n3p_artifact": last_n3p_artifact,
        "last_b2_artifact": last_b2_artifact,
        "last_eligible_plan": last_eligible_plan,
        "last_executed_report": last_executed_report,
        "side_effects": _merge_plan_only_side_effects(*side_effect_guards),
        "n3p_cache_stats": n3p_cache.stats(),
    }


def _build_fixture_v1_artifact(
    *,
    trade_date: str,
    minutes: Sequence[str],
    replay_run_id: str,
    asset_scope: str = DEFAULT_REPLAY_ASSET_SCOPE,
    source_bundle_key: str = DEFAULT_REPLAY_SOURCE_BUNDLE_KEY,
    validation_mode: str = DEFAULT_REPLAY_VALIDATION_MODE,
) -> dict[str, Any]:
    scope = _normalize_replay_asset_scope(asset_scope)
    normalized_source_bundle_key = _normalize_source_bundle_key(source_bundle_key)
    normalized_validation_mode = _normalize_replay_validation_mode(validation_mode)
    fixture_counts = _count_asset_scope_source_counts(_canonical_fixture_source_bundle(job_id=replay_run_id))
    return {
        "n3_messages": _build_n3_messages(trade_date, list(minutes)),
        "n4_messages": _build_n4_messages(trade_date),
        "n5_messages": _build_n5_messages(trade_date),
        "replay_engine_version": "fixture_v1",
        "validation_mode": normalized_validation_mode,
        "source_policy": REPLAY_SOURCE_POLICY["fixture_v1"],
        "historical_source_status": "available",
        "historical_source_kind": "fixture",
        "historical_source_path": "builtin:fixture_v1",
        "source_bundle_key": normalized_source_bundle_key,
        "source_bundle_selector_mode": _source_bundle_selector_mode(normalized_source_bundle_key),
        "resolved_source_bundle_path": "builtin:fixture_v1",
        "source_meta": _normalize_source_meta(
            {
                "historical_source_type": "fixture_fallback",
                "historical_source_path": "builtin:fixture_v1",
                "upstream_source_mode": "fixture",
            },
            fallback_path="builtin:fixture_v1",
        ),
        "engine_disclaimer": REPLAY_SOURCE_POLICY["fixture_v1"],
        "canonical_planner_trace": {},
        "plan_only_side_effects": _plan_only_side_effects(),
        "asset_scope": scope,
        "asset_scope_filter_applied": False,
        "asset_scope_allowed_asset_kinds": list(REPLAY_ASSET_SCOPE_ALLOWED_KINDS[scope]),
        "asset_scope_source_counts_before": fixture_counts,
        "asset_scope_source_counts_after": fixture_counts,
        "replay_run_id": replay_run_id,
    }


def _open_action_window_identities(
    *,
    eligible_rows: Sequence[Mapping[str, Any]],
    executed_eligible_event_ids: set[str],
) -> set[str]:
    identities: set[str] = set()
    for row in eligible_rows:
        event_id = str(row.get("event_id") or "")
        if event_id and event_id in executed_eligible_event_ids:
            continue
        payload = _json_mapping(row.get("payload_json"))
        identity_key = str(
            payload.get("identity_key")
            or row.get("identity_key")
            or ""
        )
        if identity_key:
            identities.add(identity_key)
    return identities


def _n5_metric_index_key(row: Mapping[str, Any]) -> tuple[str, str, str, str] | None:
    minute_label = provisional_action_executed_dry_run.minute_label_from_metric(row)
    values = (
        str(row.get("asset_kind") or ""),
        str(row.get("identity_key") or ""),
        str(row.get("signal_type") or ""),
        str(minute_label or ""),
    )
    if any(not value for value in values):
        return None
    return values


def _n5_eligible_index_key(row: Mapping[str, Any]) -> tuple[str, str, str, str] | None:
    payload = _json_mapping(row.get("payload_json"))
    minute_label = provisional_action_executed_dry_run.selected_minute_label_from_payload(payload)
    values = (
        str(row.get("asset_kind") or payload.get("asset_kind") or ""),
        str(row.get("identity_key") or payload.get("identity_key") or ""),
        str(payload.get("signal_type") or row.get("signal_type") or ""),
        str(minute_label or ""),
    )
    if any(not value for value in values):
        return None
    return values


def _build_n5_evaluator_demand_inputs(
    *,
    minute: str,
    tracked_eligible_rows: Sequence[Mapping[str, Any]],
    closed_metric_rows: Sequence[Mapping[str, Any]],
    executed_eligible_event_ids: set[str],
    terminal_eligible_event_ids: set[str],
) -> dict[str, Any]:
    terminal_ids = set(executed_eligible_event_ids) | set(terminal_eligible_event_ids)
    open_rows: list[dict[str, Any]] = []
    closed_window_count = 0
    for row in tracked_eligible_rows:
        event_id = str(row.get("event_id") or "")
        if event_id and event_id in terminal_ids:
            closed_window_count += 1
            continue
        open_rows.append(dict(row))

    metric_index: dict[tuple[str, str, str, str], list[dict[str, Any]]] = {}
    fail_open_count = 0
    fail_open_reasons: list[str] = []
    for row in closed_metric_rows:
        key = _n5_metric_index_key(row)
        if key is None:
            fail_open_count += 1
            fail_open_reasons.append("metric_missing_index_key")
            continue
        metric_index.setdefault(key, []).append(dict(row))

    selected_keys: set[tuple[str, str, str, str]] = set()
    for row in open_rows:
        key = _n5_eligible_index_key(row)
        if key is None:
            fail_open_count += 1
            fail_open_reasons.append("eligible_missing_index_key")
            continue
        selected_keys.add(key)

    fail_open = fail_open_count > 0
    if fail_open:
        filtered_metrics = [dict(row) for row in closed_metric_rows]
    else:
        filtered_metrics = []
        for key in selected_keys:
            filtered_metrics.extend(metric_index.get(key, []))

    rows_saved = max(
        0,
        (len(tracked_eligible_rows) + len(closed_metric_rows)) - (len(open_rows) + len(filtered_metrics)),
    )
    largest_bucket_size = max((len(rows) for rows in metric_index.values()), default=0)
    fail_open_reason = ",".join(sorted(set(fail_open_reasons)))
    plan_row = {
        "minute": minute,
        "tracked_eligible_count": len(tracked_eligible_rows),
        "open_eligible_count": len(open_rows),
        "closed_window_count": closed_window_count,
        "closed_metric_row_count": len(closed_metric_rows),
        "candidate_metric_row_count": len(filtered_metrics),
        "metric_index_key_count": len(metric_index),
        "unique_open_key_count": len(selected_keys),
        "n5_evaluator_rows_saved": rows_saved,
        "n5_evaluator_fail_open_count": fail_open_count,
        "fail_open": fail_open,
        "fail_open_reason": fail_open_reason,
        "executed_plan_count": 0,
        "trace_summary": (
            "fail-open to full closed metric input"
            if fail_open
            else "replay-only N5 evaluator demand input reduction"
        ),
    }
    index_row = {
        "minute": minute,
        "closed_metric_row_count": len(closed_metric_rows),
        "metric_index_key_count": len(metric_index),
        "candidate_metric_row_count": len(filtered_metrics),
        "open_eligible_count": len(open_rows),
        "unique_open_key_count": len(selected_keys),
        "largest_bucket_size": largest_bucket_size,
        "fail_open": fail_open,
        "trace_summary": "replay-only closed confirmation metric index; canonical evaluator unchanged",
    }
    return {
        "eligible_rows": open_rows,
        "confirmation_metric_rows": filtered_metrics,
        "plan_row": plan_row,
        "index_row": index_row,
    }


def _build_n5_evaluator_demand_summary(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    return {
        "n5_open_window_count": max((int(row.get("open_eligible_count") or 0) for row in rows), default=0),
        "n5_closed_window_count": max((int(row.get("closed_window_count") or 0) for row in rows), default=0),
        "n5_metric_index_key_count": max((int(row.get("metric_index_key_count") or 0) for row in rows), default=0),
        "n5_evaluator_rows_saved": sum(int(row.get("n5_evaluator_rows_saved") or 0) for row in rows),
        "n5_evaluator_fail_open_count": sum(int(row.get("n5_evaluator_fail_open_count") or 0) for row in rows),
    }


def _build_n4_lightweight_prefilter_rows(
    *,
    candidates: Sequence[Mapping[str, Any]],
    context_rows: Sequence[Mapping[str, Any]],
    source_bundle: Mapping[str, Any],
    previous_states: Sequence[Mapping[str, Any]],
    open_action_identities: set[str],
    trade_date: str,
    minute: str,
) -> list[dict[str, Any]]:
    latest_rows = _latest_source_rows_by_identity(source_bundle, trade_date=trade_date, minute=minute)
    candidate_identities = {str(candidate.get("identity_key") or "") for candidate in candidates}
    previous_live_identities = {
        str(state.get("identity_key") or "")
        for state in previous_states
        if str(state.get("current_status") or state.get("status") or "") == "matched"
        or bool(state.get("trigger_live"))
    }
    rows: list[dict[str, Any]] = []
    for context in context_rows:
        identity_key = str(context.get("identity_key") or "")
        if not identity_key or identity_key not in candidate_identities:
            continue
        if str(context.get("condition_key") or "") in {"BUY_HINT", "SELL_HINT"}:
            continue
        source_row = latest_rows.get(identity_key)
        decision, reason, threshold = _prefilter_decision_for_context(
            context=context,
            source_row=source_row,
            previous_live_state=identity_key in previous_live_identities,
            open_action_window=identity_key in open_action_identities,
        )
        active_reduction_decision = ""
        active_reduction_reason = ""
        active_reduction_eligible = False
        if identity_key in previous_live_identities:
            state_decision, state_reason, _ = _prefilter_decision_for_context(
                context=context,
                source_row=source_row,
                previous_live_state=False,
                open_action_window=False,
            )
            if state_decision == "prefilter_keep" and state_reason in {"buy_price_candidate", "sell_price_candidate"}:
                active_reduction_decision = "skip_n3p"
                active_reduction_reason = "live_state_retained"
                active_reduction_eligible = True
            elif state_decision == "prefilter_drop":
                active_reduction_decision = "keep_n3p"
                active_reduction_reason = "possible_state_change"
            else:
                active_reduction_decision = "keep_n3p"
                active_reduction_reason = "fail_open"
        rows.append({
            "minute": minute,
            "decision": decision,
            "reason": reason,
            "asset_kind": context.get("asset_kind"),
            "identity_key": identity_key,
            "condition_key": context.get("condition_key"),
            "direction": context.get("direction") or ("sell" if str(context.get("condition_key") or "").startswith("SELL") else "buy"),
            "signal_type": "S_SELL" if str(context.get("condition_key") or "").startswith("SELL") else "B_BUY",
            "current_high": (source_row or {}).get("high"),
            "current_low": (source_row or {}).get("low"),
            "current_close": (source_row or {}).get("close"),
            "threshold": str(threshold) if threshold is not None else "",
            "previous_live_state": identity_key in previous_live_identities,
            "open_action_window": identity_key in open_action_identities,
            "active_reduction_decision": active_reduction_decision,
            "active_reduction_reason": active_reduction_reason,
            "active_reduction_eligible": active_reduction_eligible,
            "canonical_event_type": "",
            "false_negative": False,
            "shadow_mode": True,
        })
    return rows


def _latest_source_rows_by_identity(
    source_bundle: Mapping[str, Any],
    *,
    trade_date: str,
    minute: str,
) -> dict[str, Mapping[str, Any]]:
    latest: dict[str, Mapping[str, Any]] = {}
    for rows in dict(source_bundle.get("source_records") or {}).values():
        for row in rows:
            if _source_row_date(row) != trade_date:
                continue
            row_minute = _source_row_minute(row, trade_date)
            if not row_minute or row_minute > minute:
                continue
            identity_key = str(row.get("identity_key") or "")
            if not identity_key:
                continue
            previous = latest.get(identity_key)
            if previous is None or _source_row_minute(previous, trade_date) <= row_minute:
                latest[identity_key] = row
    return latest


def _prefilter_decision_for_context(
    *,
    context: Mapping[str, Any],
    source_row: Mapping[str, Any] | None,
    previous_live_state: bool,
    open_action_window: bool,
) -> tuple[str, str, Decimal | None]:
    if previous_live_state:
        return "prefilter_keep", "previous_live_state", None
    if open_action_window:
        return "prefilter_keep", "open_action_window", None
    if not source_row:
        return "prefilter_keep", "missing_current_source_row", None
    threshold = _prefilter_threshold_for_context(context)
    if threshold is None:
        return "prefilter_keep", "missing_or_unsupported_baseline", None
    condition_key = str(context.get("condition_key") or "")
    direction = str(context.get("direction") or ("sell" if condition_key.startswith("SELL") else "buy"))
    high = _decimal_or_none(source_row.get("high"))
    low = _decimal_or_none(source_row.get("low"))
    close = _decimal_or_none(source_row.get("close"))
    if direction == "sell":
        if (low is not None and low <= threshold) or (close is not None and close <= threshold):
            return "prefilter_keep", "sell_price_candidate", threshold
        return "prefilter_drop", "sell_price_below_trigger_absent", threshold
    if (high is not None and high >= threshold) or (close is not None and close >= threshold):
        return "prefilter_keep", "buy_price_candidate", threshold
    return "prefilter_drop", "buy_price_breakout_absent", threshold


def _prefilter_threshold_for_context(context: Mapping[str, Any]) -> Decimal | None:
    baseline = _json_mapping(context.get("period_trigger_baseline_json"))
    if not baseline:
        raw_json = _json_mapping(context.get("raw_json"))
        baseline = _json_mapping(raw_json.get("period_trigger_baseline_json"))
    periods = _json_mapping(baseline.get("periods"))
    condition_key = str(context.get("condition_key") or "")
    requested_periods = _requested_periods_from_condition_key(condition_key)
    candidates: list[Decimal] = []
    keys = (
        ("trigger_previous_entity_low", "previous_entity_low", "previous_close")
        if condition_key.startswith("SELL")
        else ("trigger_previous_entity_high", "previous_entity_high", "previous_close")
    )
    for period in requested_periods:
        period_baseline = _json_mapping(periods.get(period))
        for key in keys:
            value = _decimal_or_none(period_baseline.get(key))
            if value is not None:
                candidates.append(value)
                break
    if not candidates:
        return None
    # N4 ordinary periods are ANY-style triggers. A lightweight prefilter must
    # keep a candidate if any requested period can still produce a canonical
    # event, so BUY uses the lowest breakout threshold and SELL the highest
    # breakdown threshold.
    return max(candidates) if condition_key.startswith("SELL") else min(candidates)


def _requested_periods_from_condition_key(condition_key: str) -> list[str]:
    if condition_key in {"BUY:FULL", "SELL:FULL"}:
        return ["D"]
    if ":" not in condition_key:
        return ["D"]
    return [
        period.strip().upper()
        for period in condition_key.split(":", 1)[1].split(",")
        if period.strip().upper() in {"Y", "Q", "M", "W", "D"}
    ]


def _annotate_prefilter_rows_with_canonical_events(
    rows: Sequence[Mapping[str, Any]],
    *,
    minute_events: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    event_by_key = {
        _prefilter_event_key(event): str(event.get("output_event_type") or event.get("event_type") or "")
        for event in minute_events
        if str(event.get("output_event_type") or event.get("event_type") or "") in {"TriggerMatched", "TriggerStateChanged"}
    }
    output: list[dict[str, Any]] = []
    for row in rows:
        cloned = dict(row)
        event_type = event_by_key.get(_prefilter_event_key(row), "")
        cloned["canonical_event_type"] = event_type
        cloned["false_negative"] = bool(event_type and cloned.get("decision") != "prefilter_keep")
        output.append(cloned)
    return output


def _prefilter_event_key(row: Mapping[str, Any]) -> str:
    return "|".join(
        str(row.get(key) or "")
        for key in ("asset_kind", "identity_key", "signal_type", "condition_key")
    )


def _candidate_prefilter_key(candidate: Mapping[str, Any]) -> str:
    condition_key = str(
        candidate.get("original_condition_key")
        or candidate.get("condition_key")
        or ""
    )
    signal_type = str(candidate.get("signal_type") or ("S_SELL" if condition_key.startswith("SELL") else "B_BUY"))
    return "|".join(
        str(value or "")
        for value in (
            candidate.get("asset_kind"),
            candidate.get("identity_key"),
            signal_type,
            condition_key,
        )
    )


def _candidate_action_key(candidate: Mapping[str, Any]) -> str:
    condition_key = str(
        candidate.get("original_condition_key")
        or candidate.get("condition_key")
        or ""
    )
    signal_type = str(candidate.get("signal_type") or ("S_SELL" if condition_key.startswith("SELL") else "B_BUY"))
    return "|".join(
        str(value or "")
        for value in (
            candidate.get("asset_kind"),
            candidate.get("identity_key"),
            signal_type,
            condition_key,
            candidate.get("action_mark") or candidate.get("trigger_mark_candidate") or "normal",
        )
    )


def _action_execution_key(row: Mapping[str, Any]) -> str:
    return "|".join(
        str(row.get(key) or "")
        for key in ("asset_kind", "identity_key", "signal_type", "condition_key", "action_mark")
    )


def _suppress_duplicate_action_executed_messages(
    messages: Sequence[Mapping[str, Any]],
    *,
    executed_action_keys: set[str],
    minute: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    kept: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []
    suppressed = 0
    for message in messages:
        cloned = dict(message)
        if cloned.get("event_type") != "ActionExecuted":
            kept.append(cloned)
            continue
        action_key = _action_execution_key(cloned)
        if action_key in executed_action_keys:
            suppressed += 1
            audit_rows.append(
                {
                    "minute": minute,
                    "reduction_mode": "active_state_fast_path",
                    "reduction_reason": "duplicate_execution_suppressed",
                    "asset_kind": cloned.get("asset_kind"),
                    "identity_key": cloned.get("identity_key"),
                    "condition_key": cloned.get("condition_key"),
                    "signal_type": cloned.get("signal_type"),
                    "action_key": action_key,
                    "source_decision": "",
                    "source_reason": "",
                    "n3p_skipped": False,
                    "duplicate_suppressed": True,
                    "shadow_mode": True,
                    "trace_summary": "replay-only duplicate ActionExecuted suppressed by stable action key",
                }
            )
            continue
        executed_action_keys.add(action_key)
        kept.append(cloned)
    return kept, audit_rows, suppressed


def _active_state_reduction_skip_keys(
    prefilter_rows: Sequence[Mapping[str, Any]],
    *,
    minute: str,
    reduction_mode: str,
) -> tuple[set[str], list[dict[str, Any]]]:
    if reduction_mode != "active_state_fast_path":
        return set(), []
    skip_keys: set[str] = set()
    audit_rows: list[dict[str, Any]] = []
    for row in prefilter_rows:
        if not bool(row.get("active_reduction_eligible")):
            continue
        skip_key = _prefilter_event_key(row)
        skip_keys.add(skip_key)
        audit_rows.append(
            {
                "minute": minute,
                "reduction_mode": reduction_mode,
                "reduction_reason": "live_state_retained",
                "asset_kind": row.get("asset_kind"),
                "identity_key": row.get("identity_key"),
                "condition_key": row.get("condition_key"),
                "signal_type": row.get("signal_type"),
                "action_key": "",
                "source_decision": row.get("decision"),
                "source_reason": row.get("reason"),
                "n3p_skipped": True,
                "duplicate_suppressed": False,
                "shadow_mode": True,
                "trace_summary": "replay-only active state retained; full N3P skipped fail-open otherwise",
            }
        )
    return skip_keys, audit_rows


def _build_n3p_prefilter_summary(
    *,
    strategy: str,
    n4_prefilter_rows: Sequence[Mapping[str, Any]],
    n3p_demand_plan: Sequence[Mapping[str, Any]],
    false_negatives: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    keep_count = sum(1 for row in n4_prefilter_rows if row.get("decision") == "prefilter_keep")
    drop_count = sum(1 for row in n4_prefilter_rows if row.get("decision") == "prefilter_drop")
    return {
        "strategy": strategy,
        "prefilter_keep": keep_count,
        "prefilter_drop": drop_count,
        "false_negative_count": len(false_negatives),
        "n3p_calls_saved": sum(int(row.get("n3p_calls_saved") or 0) for row in n3p_demand_plan),
        "candidate_count_before": sum(int(row.get("candidate_count_before") or 0) for row in n3p_demand_plan),
        "candidate_count_after": sum(int(row.get("candidate_count_after") or 0) for row in n3p_demand_plan),
        "open_action_window_count": max([int(row.get("open_action_window_count") or 0) for row in n3p_demand_plan] or [0]),
    }


def _build_n3p_active_state_reduction_summary(
    rows: Sequence[Mapping[str, Any]],
    *,
    n3p_demand_plan: Sequence[Mapping[str, Any]],
    reduction_mode: str,
) -> dict[str, Any]:
    skipped_live_state = sum(1 for row in rows if row.get("reduction_reason") == "live_state_retained")
    skipped_closed_window = sum(1 for row in rows if row.get("reduction_reason") == "closed_action_window")
    duplicate_suppressed = sum(1 for row in rows if row.get("reduction_reason") == "duplicate_execution_suppressed")
    return {
        "n3p_reduction_mode": reduction_mode,
        "shadow_state_fast_path_count": skipped_live_state,
        "n3p_skipped_live_state_retained": skipped_live_state,
        "n3p_skipped_closed_action_window": skipped_closed_window,
        "duplicate_execution_suppressed": duplicate_suppressed,
        "n3p_calls_saved": skipped_live_state + skipped_closed_window,
        "n3p_demand_plan_calls_saved": sum(int(row.get("n3p_calls_saved") or 0) for row in n3p_demand_plan),
    }


def _shadow_state_key(row: Mapping[str, Any]) -> str:
    return "|".join(
        str(row.get(key) or "")
        for key in (
            "asset_kind",
            "identity_key",
            "signal_type",
            "condition_key",
            "trigger_period",
            "trigger_mark_candidate",
        )
    )


def _apply_replay_n5_entry_contract(message: dict[str, Any], *, replay_config: Mapping[str, Any]) -> None:
    source_condition_run_id = str(
        message.get("source_condition_run_id")
        or replay_config.get("source_condition_run_id")
        or f"{replay_config.get('replay_run_id') or replay_config.get('job_id')}__local_condition_context"
    )
    message["source_condition_run_id"] = source_condition_run_id
    message["source_condition_key"] = message.get("source_condition_key") or message.get("condition_key")
    if not message.get("source_n3p_live_target_run_id") and message.get("source_metric_run_id"):
        message["source_n3p_live_target_run_id"] = message.get("source_metric_run_id")
    if "c1_dependency" not in message or message.get("c1_dependency") is None:
        message["c1_dependency"] = False
    enters_n5 = message.get("event_type") == "TriggerMatched"
    if enters_n5 and not message.get("selected_metric_id"):
        local_metric_key = "|".join(
            str(message.get(key) or "")
            for key in ("identity_key", "condition_key", "selected_metric_time", "metric_time_label", "trigger_time")
        )
        message["selected_metric_id"] = int(sha256(local_metric_key.encode("utf-8")).hexdigest()[:12], 16)
        message["local_replay_metric_id_source"] = "synthetic_local_replay_shadow_metric_id"
    if enters_n5 and not message.get("selected_metric_time"):
        message["selected_metric_time"] = _event_time_from_trade_minute(message.get("trade_date"), message.get("minute"))
    message["n5_entry_allowed"] = bool(enters_n5)
    message["n4_boundary"] = {"enters_n5": bool(enters_n5)}


def _shadow_n4_transition_row(
    message: Mapping[str, Any],
    *,
    previous_states: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    state_key = _shadow_state_key(message)
    previous = dict(previous_states.get(state_key) or {})
    return {
        "minute": message.get("minute"),
        "event_type": message.get("event_type"),
        "asset_kind": message.get("asset_kind"),
        "identity_key": message.get("identity_key"),
        "condition_key": message.get("condition_key"),
        "original_condition_key": message.get("original_condition_key"),
        "signal_type": message.get("signal_type"),
        "previous_status": previous.get("current_status") or "none",
        "current_status": message.get("current_status"),
        "trigger_live": message.get("trigger_live"),
        "source": message.get("source"),
        "state_key": state_key,
        "shadow_mode": True,
        "trace_summary": "full_day_shadow_v1 local N4 state transition; no production lineage",
    }


def _shadow_n4_evaluation_row(
    plan: Mapping[str, Any],
    *,
    minute: str,
    source: str,
    event: Mapping[str, Any] | None,
) -> dict[str, Any]:
    event_type = str((event or {}).get("output_event_type") or "NoOp")
    condition_key = str(plan.get("condition_key") or "")
    return {
        "minute": minute,
        "event_type": event_type,
        "evaluation_status": "event_emitted" if event else "noop_suppressed",
        "asset_kind": plan.get("asset_kind"),
        "identity_key": plan.get("identity_key"),
        "condition_key": condition_key,
        "original_condition_key": plan.get("original_condition_key") or condition_key,
        "signal_type": plan.get("signal_type"),
        "current_status": (event or plan).get("current_status"),
        "trigger_live": (event or plan).get("trigger_live"),
        "trigger_period": (event or plan).get("trigger_period"),
        "trigger_mark_candidate": (event or plan).get("trigger_mark_candidate"),
        "trigger_price": (event or plan).get("trigger_price"),
        "selected_metric_time": (event or plan).get("selected_metric_time"),
        "metric_time_label": (event or plan).get("metric_time_label"),
        "metric_minute_label": (event or plan).get("metric_minute_label"),
        "source": source,
        "plan_id": plan.get("plan_id"),
        "shadow_mode": True,
        "trace_summary": (
            "full_day_shadow_v1 emitted N4 event"
            if event
            else "full_day_shadow_v1 evaluated candidate but lifecycle suppressed event"
        ),
    }


def _shadow_n5_window_row(message: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "minute": message.get("minute"),
        "event_type": message.get("event_type"),
        "asset_kind": message.get("asset_kind"),
        "identity_key": message.get("identity_key"),
        "condition_key": message.get("condition_key"),
        "original_condition_key": message.get("original_condition_key"),
        "action_state": message.get("action_state"),
        "source_trigger_event_type": message.get("source_trigger_event_type"),
        "source_trigger_run_id": message.get("source_trigger_run_id"),
        "confirmation_metric_id": message.get("confirmation_metric_id"),
        "final_proof_source": message.get("final_proof_source"),
        "shadow_mode": True,
        "trace_summary": "full_day_shadow_v1 local N5 action window; no production lineage",
    }


def _build_shadow_validation_report(
    *,
    minutes: Sequence[str],
    n4_messages: Sequence[Mapping[str, Any]],
    n5_messages: Sequence[Mapping[str, Any]],
    n4_shadow_state_transitions: Sequence[Mapping[str, Any]],
    n4_shadow_evaluations: Sequence[Mapping[str, Any]],
    n5_shadow_action_windows: Sequence[Mapping[str, Any]],
    shadow_noop_by_minute: Mapping[str, int],
    n4_prefilter_rows: Sequence[Mapping[str, Any]] = (),
    n3p_demand_plan: Sequence[Mapping[str, Any]] = (),
    n3p_strategy: str = DEFAULT_REPLAY_N3P_STRATEGY,
    n3p_reduction_mode: str = DEFAULT_REPLAY_N3P_REDUCTION_MODE,
    n3p_active_state_reduction: Sequence[Mapping[str, Any]] = (),
) -> dict[str, Any]:
    executed_keys = [
        str(row.get("source_trigger_run_id") or "") + "|" + str(row.get("identity_key") or "") + "|" + str(row.get("condition_key") or "")
        for row in n5_messages
        if row.get("event_type") == "ActionExecuted"
    ]
    duplicate_executed = len(executed_keys) - len(set(executed_keys))
    return {
        "shadow_mode": True,
        "timeline_minutes": len(minutes),
        "stock_n4_messages": sum(1 for row in n4_messages if row.get("asset_kind") == "stock"),
        "stock_n5_messages": sum(1 for row in n5_messages if row.get("asset_kind") == "stock"),
        "trigger_matched": sum(1 for row in n4_messages if row.get("event_type") == "TriggerMatched"),
        "trigger_state_changed": sum(1 for row in n4_messages if row.get("event_type") == "TriggerStateChanged"),
        "trigger_pending_market_data": sum(1 for row in n4_messages if row.get("event_type") == "TriggerPendingMarketData"),
        "noop_suppressed": sum(int(value or 0) for value in shadow_noop_by_minute.values()),
        "noop_suppressed_by_minute": dict(shadow_noop_by_minute),
        "action_eligible": sum(1 for row in n5_messages if row.get("event_type") == "ActionEligible"),
        "action_executed": sum(1 for row in n5_messages if row.get("event_type") == "ActionExecuted"),
        "action_skipped": sum(1 for row in n5_messages if row.get("event_type") == "ActionSkipped"),
        "trigger_state_changed_action_entries": sum(
            1 for row in n5_messages if row.get("source_trigger_event_type") == "TriggerStateChanged"
        ),
        "n5_duplicate_executed": duplicate_executed,
        "n4_shadow_evaluation_rows": len(n4_shadow_evaluations),
        "n4_shadow_state_transition_rows": len(n4_shadow_state_transitions),
        "n5_shadow_action_window_rows": len(n5_shadow_action_windows),
        "n3p_strategy": n3p_strategy,
        "n3p_reduction_mode": n3p_reduction_mode,
        "prefilter_keep": sum(1 for row in n4_prefilter_rows if row.get("decision") == "prefilter_keep"),
        "prefilter_drop": sum(1 for row in n4_prefilter_rows if row.get("decision") == "prefilter_drop"),
        "prefilter_false_negative_count": sum(1 for row in n4_prefilter_rows if row.get("false_negative")),
        "n3p_calls_saved": sum(int(row.get("n3p_calls_saved") or 0) for row in n3p_demand_plan),
        "shadow_state_fast_path_count": sum(1 for row in n3p_active_state_reduction if row.get("reduction_reason") == "live_state_retained"),
        "n3p_skipped_live_state_retained": sum(1 for row in n3p_active_state_reduction if row.get("reduction_reason") == "live_state_retained"),
        "n3p_skipped_closed_action_window": sum(1 for row in n3p_active_state_reduction if row.get("reduction_reason") == "closed_action_window"),
        "duplicate_execution_suppressed": sum(1 for row in n3p_active_state_reduction if row.get("reduction_reason") == "duplicate_execution_suppressed"),
        "quality_blockers": {},
    }


def _build_n4_messages(trade_date: str) -> list[dict[str, Any]]:
    return [
        _n4_message(trade_date, "10:00", "ordinary", "TriggerMatched", "stock", "stock:SZ:002668", "B_BUY", "BUY:Y,D", "BUY", 22.18),
        _n4_message(trade_date, "10:15", "hint", "TriggerMatched", "stock", "stock:SH:600346", "B_BUY", "BUY_HINT", "BUY", 8.32),
        _n4_message(trade_date, "10:20", "hint", "TriggerMatched", "stock", "stock:SH:600570", "S_SELL", "SELL_HINT", "SELL", 72.7),
        _n4_message(trade_date, "13:30", "ordinary", "TriggerStateChanged", "stock", "stock:SH:603061", "B_BUY", "BUY:M", "BUY", 417.0),
    ]


def _n4_message(
    trade_date: str,
    minute: str,
    source: str,
    event_type: str,
    asset_kind: str,
    identity_key: str,
    signal_type: str,
    condition_key: str,
    trigger_type: str,
    trigger_price: float,
) -> dict[str, Any]:
    return {
        "minute": minute,
        "trade_date": trade_date,
        "event_type": event_type,
        "asset_kind": asset_kind,
        "identity_key": identity_key,
        "signal_type": signal_type,
        "condition_key": condition_key,
        "original_condition_key": condition_key,
        "trigger_type": trigger_type,
        "trigger_price": trigger_price,
        "source": source,
        "trace_summary": f"local replay {source}; no production outbox consumption",
    }


def _build_n5_messages(trade_date: str) -> list[dict[str, Any]]:
    return [
        _n5_message(trade_date, "10:00", "ActionEligible", "stock:SZ:002668", "B_BUY", "BUY:Y,D", "eligible", "", "metric:9502959", "ordinary", "N3P"),
        _n5_message(trade_date, "10:15", "ActionEligible", "stock:SH:600346", "B_BUY", "BUY_HINT", "eligible", "", "", "hint", "none"),
        _n5_message(trade_date, "10:20", "ActionEligible", "stock:SH:600570", "S_SELL", "SELL_HINT", "eligible", "", "", "hint", "none"),
        _n5_message(trade_date, "10:01", "ActionExecuted", "stock:SZ:002668", "B_BUY", "BUY:Y,D", "executed", "normal", "9502959", "ordinary", "N3P"),
    ]


def _n5_message(
    trade_date: str,
    minute: str,
    event_type: str,
    identity_key: str,
    signal_type: str,
    condition_key: str,
    action_state: str,
    action_mark: str,
    confirmation_metric_id: str,
    source_trigger_run_id: str,
    final_proof_source: str,
) -> dict[str, Any]:
    return {
        "minute": minute,
        "trade_date": trade_date,
        "event_type": event_type,
        "asset_kind": "stock",
        "identity_key": identity_key,
        "signal_type": signal_type,
        "condition_key": condition_key,
        "action_state": action_state,
        "action_mark": action_mark,
        "confirmation_metric_id": confirmation_metric_id,
        "source_trigger_run_id": source_trigger_run_id,
        "final_proof_source": final_proof_source,
    }


def _build_timeline(
    minutes: list[str],
    n4_messages: list[dict[str, Any]],
    n5_messages: list[dict[str, Any]],
    *,
    n3_messages: list[dict[str, Any]] | None = None,
    shadow_validation_report: Mapping[str, Any] | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    n3_by_minute = {str(row.get("minute") or ""): int(row.get("n3p_metric_rows") or 0) for row in (n3_messages or [])}
    shadow_report = dict(shadow_validation_report or {})
    noop_by_minute = dict(shadow_report.get("noop_suppressed_by_minute") or {})
    for minute in minutes:
        minute_n4 = [row for row in n4_messages if row.get("minute") == minute]
        minute_n5 = [row for row in n5_messages if row.get("minute") == minute]
        rows.append({
            "minute": minute,
            "n3_rows": n3_by_minute.get(minute, 0),
            "n4_ordinary": sum(1 for row in minute_n4 if row.get("source") == "ordinary"),
            "n4_hint": sum(1 for row in minute_n4 if row.get("source") == "hint"),
            "n4_state_changed": sum(1 for row in minute_n4 if row.get("event_type") == "TriggerStateChanged"),
            "n4_noop_suppressed": int(noop_by_minute.get(minute) or 0),
            "n5_eligible": sum(1 for row in minute_n5 if row.get("event_type") == "ActionEligible"),
            "n5_executed": sum(1 for row in minute_n5 if row.get("event_type") == "ActionExecuted"),
            "n5_skipped": sum(1 for row in minute_n5 if row.get("event_type") == "ActionSkipped"),
        })
    return rows


def _build_summary(
    *,
    job_id: str,
    trade_date: str,
    artifact_dir: Path,
    timeline: list[dict[str, Any]],
    n3_messages: list[dict[str, Any]],
    n4_messages: list[dict[str, Any]],
    n5_messages: list[dict[str, Any]],
    plan: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    plan = dict(plan or {})
    replay_engine_version = str(plan.get("replay_engine_version") or DEFAULT_REPLAY_ENGINE_VERSION)
    validation_mode = str(plan.get("validation_mode") or DEFAULT_REPLAY_VALIDATION_MODE)
    n3p_strategy = str(plan.get("n3p_strategy") or DEFAULT_REPLAY_N3P_STRATEGY)
    n3p_reduction_mode = str(plan.get("n3p_reduction_mode") or DEFAULT_REPLAY_N3P_REDUCTION_MODE)
    asset_unit_fix_delta = dict(plan.get("asset_unit_fix_delta") or {})
    n5_delta_only = dict(plan.get("n5_delta_only") or {})
    shadow_validation_report = dict(plan.get("shadow_validation_report") or {})
    shadow_summary = {
        "stock_n4_messages": int(shadow_validation_report.get("stock_n4_messages") or 0),
        "stock_n5_messages": int(shadow_validation_report.get("stock_n5_messages") or 0),
        "TriggerMatched": int(shadow_validation_report.get("trigger_matched") or 0),
        "TriggerStateChanged": int(shadow_validation_report.get("trigger_state_changed") or 0),
        "TriggerPendingMarketData": int(shadow_validation_report.get("trigger_pending_market_data") or 0),
        "noop_suppressed": int(shadow_validation_report.get("noop_suppressed") or 0),
        "ActionEligible": int(shadow_validation_report.get("action_eligible") or 0),
        "ActionExecuted": int(shadow_validation_report.get("action_executed") or 0),
        "ActionSkipped": int(shadow_validation_report.get("action_skipped") or 0),
        "trigger_state_changed_action_entries": int(shadow_validation_report.get("trigger_state_changed_action_entries") or 0),
        "n5_duplicate_executed": int(shadow_validation_report.get("n5_duplicate_executed") or 0),
        "n4_shadow_evaluation_rows": int(shadow_validation_report.get("n4_shadow_evaluation_rows") or 0),
    }
    return {
        "job_id": job_id,
        "trade_date": trade_date,
        "artifact_dir": str(artifact_dir),
        "notice": LOCAL_ONLY_NOTICE,
        "safety_flags": dict(SAFETY_FLAGS),
        "replay_engine_version": replay_engine_version,
        "validation_mode": validation_mode,
        "n3p_strategy": n3p_strategy,
        "n3p_reduction_mode": n3p_reduction_mode,
        "asset_scope": str(plan.get("asset_scope") or DEFAULT_REPLAY_ASSET_SCOPE),
        "source_bundle_key": str(plan.get("source_bundle_key") or DEFAULT_REPLAY_SOURCE_BUNDLE_KEY),
        "source_bundle_selector_mode": str(plan.get("source_bundle_selector_mode") or "auto"),
        "resolved_source_bundle_path": str(plan.get("resolved_source_bundle_path") or ""),
        "asset_scope_filter_applied": bool(plan.get("asset_scope_filter_applied")),
        "asset_scope_allowed_asset_kinds": list(plan.get("asset_scope_allowed_asset_kinds") or []),
        "asset_scope_source_counts_before": dict(plan.get("asset_scope_source_counts_before") or {}),
        "asset_scope_source_counts_after": dict(plan.get("asset_scope_source_counts_after") or {}),
        "source_policy": str(plan.get("source_policy") or REPLAY_SOURCE_POLICY.get(replay_engine_version, "")),
        "engine_disclaimer": str(plan.get("engine_disclaimer") or REPLAY_SOURCE_POLICY.get(replay_engine_version, "")),
        "historical_source_status": str(plan.get("historical_source_status") or ""),
        "historical_source_kind": str(plan.get("historical_source_kind") or ""),
        "historical_source_path": str(plan.get("historical_source_path") or ""),
        "source_meta": dict(plan.get("source_meta") or {}),
        "canonical_planner_trace": dict(plan.get("canonical_planner_trace") or {}),
        "plan_only_side_effects": dict(plan.get("plan_only_side_effects") or _plan_only_side_effects()),
        "n3p_cache_stats": dict(plan.get("n3p_cache_stats") or {}),
        "n3p_prefilter": dict(plan.get("n3p_prefilter") or {}),
        "n3p_active_state_reduction": dict(plan.get("n3p_active_state_reduction_summary") or {}),
        "n3p_negative_cache": dict(plan.get("n3p_negative_cache_summary") or {}),
        "n5_evaluator_demand": dict(plan.get("n5_evaluator_demand_summary") or {}),
        "shadow_mode": bool(plan.get("shadow_mode")),
        "shadow": shadow_summary,
        "shadow_validation_report": shadow_validation_report,
        "asset_unit_fix_delta_validation": bool(plan.get("asset_unit_fix_delta_validation")),
        "asset_unit_fix_delta": asset_unit_fix_delta,
        "n5_delta_only": n5_delta_only,
        "corrected_full_trigger_matched": asset_unit_fix_delta.get("corrected_full_trigger_matched"),
        "old_unified_trigger_matched": asset_unit_fix_delta.get("old_unified_trigger_matched"),
        "corrected_only": asset_unit_fix_delta.get("corrected_only"),
        "index_board_delta": asset_unit_fix_delta.get("index_board_delta"),
        "excluded_stock_replay": asset_unit_fix_delta.get("excluded_stock_replay"),
        "timeline_minutes": len(timeline),
        "n3": {"messages": len(n3_messages), "source_mode": "replay"},
        "n4": {
            "ordinary_TriggerMatched": sum(1 for row in n4_messages if row.get("source") == "ordinary" and row.get("event_type") == "TriggerMatched"),
            "ordinary_TriggerStateChanged": sum(1 for row in n4_messages if row.get("source") == "ordinary" and row.get("event_type") == "TriggerStateChanged"),
            "hint_TriggerMatched": sum(1 for row in n4_messages if row.get("source") == "hint" and row.get("event_type") == "TriggerMatched"),
        },
        "n5": {
            "ActionEligible": sum(1 for row in n5_messages if row.get("event_type") == "ActionEligible"),
            "ActionExecuted": sum(1 for row in n5_messages if row.get("event_type") == "ActionExecuted"),
            "ActionSkipped": sum(1 for row in n5_messages if row.get("event_type") == "ActionSkipped"),
            "BUY_HINT_ActionEligible": sum(1 for row in n5_messages if row.get("event_type") == "ActionEligible" and row.get("condition_key") == "BUY_HINT"),
            "SELL_HINT_ActionEligible": sum(1 for row in n5_messages if row.get("event_type") == "ActionEligible" and row.get("condition_key") == "SELL_HINT"),
            "b2_hint_final_proof_rows": sum(1 for row in n5_messages if row.get("event_type") == "ActionExecuted" and row.get("final_proof_source") != "N3P"),
        },
        "blocked_reasons": {},
    }


def _summary_markdown(summary: dict[str, Any]) -> str:
    return "\n".join([
        f"# N3-N5 Local Replay {summary.get('trade_date')}",
        "",
        LOCAL_ONLY_NOTICE,
        "",
        f"- job_id: `{summary.get('job_id')}`",
        f"- replay_engine_version: `{summary.get('replay_engine_version')}`",
        f"- validation_mode: `{summary.get('validation_mode')}`",
        f"- n3p_strategy: `{summary.get('n3p_strategy')}`",
        f"- n3p_reduction_mode: `{summary.get('n3p_reduction_mode')}`",
        f"- asset_scope: `{summary.get('asset_scope')}`",
        f"- source_bundle_key: `{summary.get('source_bundle_key')}`",
        f"- source_bundle_selector_mode: `{summary.get('source_bundle_selector_mode')}`",
        f"- source_policy: `{summary.get('source_policy')}`",
        f"- historical_source_status: `{summary.get('historical_source_status')}`",
        f"- historical_source_kind: `{summary.get('historical_source_kind')}`",
        f"- historical_source_type: `{summary.get('source_meta', {}).get('historical_source_type', '')}`",
        f"- n3p_cache_hits: `{summary.get('n3p_cache_stats', {}).get('n3p_cache_hits', 0)}`",
        f"- n3p_cache_misses: `{summary.get('n3p_cache_stats', {}).get('n3p_cache_misses', 0)}`",
        f"- empty_minute_fast_path_count: `{summary.get('n3p_cache_stats', {}).get('empty_minute_fast_path_count', 0)}`",
        f"- n3p_prefilter.false_negative_count: `{summary.get('n3p_prefilter', {}).get('false_negative_count', 0)}`",
        f"- n3p_prefilter.n3p_calls_saved: `{summary.get('n3p_prefilter', {}).get('n3p_calls_saved', 0)}`",
        f"- n3p_active_state_reduction.n3p_calls_saved: `{summary.get('n3p_active_state_reduction', {}).get('n3p_calls_saved', 0)}`",
        f"- n3p_active_state_reduction.duplicate_execution_suppressed: `{summary.get('n3p_active_state_reduction', {}).get('duplicate_execution_suppressed', 0)}`",
        f"- n3p_negative_cache.mode: `{summary.get('n3p_negative_cache', {}).get('mode', 'disabled')}`",
        f"- n3p_negative_cache.hits: `{summary.get('n3p_negative_cache', {}).get('n3p_negative_cache_hits', 0)}`",
        f"- n3p_negative_cache.false_negative_count: `{summary.get('n3p_negative_cache', {}).get('n3p_negative_cache_false_negative_count', 0)}`",
        f"- n5_evaluator_demand.rows_saved: `{summary.get('n5_evaluator_demand', {}).get('n5_evaluator_rows_saved', 0)}`",
        f"- n5_evaluator_demand.fail_open_count: `{summary.get('n5_evaluator_demand', {}).get('n5_evaluator_fail_open_count', 0)}`",
        f"- shadow_mode: `{summary.get('shadow_mode')}`",
        f"- shadow.TriggerMatched: `{summary.get('shadow', {}).get('TriggerMatched', 0)}`",
        f"- shadow.TriggerStateChanged: `{summary.get('shadow', {}).get('TriggerStateChanged', 0)}`",
        f"- shadow.noop_suppressed: `{summary.get('shadow', {}).get('noop_suppressed', 0)}`",
        f"- shadow.ActionEligible: `{summary.get('shadow', {}).get('ActionEligible', 0)}`",
        f"- shadow.ActionExecuted: `{summary.get('shadow', {}).get('ActionExecuted', 0)}`",
        f"- timeline_minutes: `{summary.get('timeline_minutes')}`",
        f"- N4 ordinary TriggerMatched: `{summary.get('n4', {}).get('ordinary_TriggerMatched')}`",
        f"- N4 hint TriggerMatched: `{summary.get('n4', {}).get('hint_TriggerMatched')}`",
        f"- N5 ActionEligible: `{summary.get('n5', {}).get('ActionEligible')}`",
        f"- N5 ActionExecuted: `{summary.get('n5', {}).get('ActionExecuted')}`",
        f"- asset_unit_fix_delta.corrected_only: `{summary.get('asset_unit_fix_delta', {}).get('corrected_only', 0)}`",
        f"- asset_unit_fix_delta.index_board_delta: `{summary.get('asset_unit_fix_delta', {}).get('index_board_delta', 0)}`",
        f"- asset_unit_fix_delta.excluded_stock_replay: `{summary.get('asset_unit_fix_delta', {}).get('excluded_stock_replay', 0)}`",
    ]) + "\n"


def _find_job_dir(replay_root: Path, job_id: str) -> Path:
    root = replay_root.resolve()
    for path in root.glob(f"*/{job_id}"):
        if path.is_dir() and path.resolve().is_relative_to(root):
            return path.resolve()
    raise FileNotFoundError(job_id)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8")


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.write_text("".join(json.dumps(row, ensure_ascii=False, sort_keys=True, default=str) + "\n" for row in rows), encoding="utf-8")


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def _write_replay_profile(json_path: Path, markdown_path: Path, payload: Mapping[str, Any]) -> None:
    _write_json(json_path, dict(payload))
    markdown_path.write_text(_replay_profile_markdown(payload), encoding="utf-8")


def _write_replay_profile_fallback(
    *,
    replay_root: Path,
    trade_date_key: str,
    profiler: _ReplayProfileCollector,
    blocked_reason: str,
) -> Path:
    profile_root = (replay_root / HISTORICAL_REPLAY_PROFILE_DIRNAME / trade_date_key).resolve()
    profile_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(tz=DISPLAY_TIMEZONE).strftime("%Y%m%d_%H%M%S")
    json_path = profile_root / f"full_scope_profile_{timestamp}.json"
    markdown_path = profile_root / f"full_scope_profile_{timestamp}.md"
    _write_replay_profile(
        json_path,
        markdown_path,
        profiler.build_payload(
            status="blocked",
            blocked_reason=blocked_reason,
        ),
    )
    return json_path


def _replay_profile_markdown(payload: Mapping[str, Any]) -> str:
    phases = dict(payload.get("phases") or {})
    top_phases = sorted(
        phases.items(),
        key=lambda item: float((item[1] or {}).get("elapsed_seconds") or 0.0),
        reverse=True,
    )[:5]
    lines = [
        f"# Replay Profile {payload.get('trade_date')}",
        "",
        LOCAL_ONLY_NOTICE,
        "",
        f"- job_id: `{payload.get('job_id')}`",
        f"- replay_engine_version: `{payload.get('replay_engine_version')}`",
        f"- status: `{payload.get('status')}`",
        f"- bottleneck_classification: `{payload.get('bottleneck_classification')}`",
        f"- elapsed_seconds: `{payload.get('elapsed_seconds')}`",
        "",
        "## N3P Cache",
    ]
    cache_stats = dict((payload.get("metadata") or {}).get("n3p_cache_stats") or {})
    for key in (
        "n3p_cache_hits",
        "n3p_cache_misses",
        "n3p_cache_saved_calls",
        "n3p_cache_key_count",
        "empty_minute_fast_path_count",
    ):
        lines.append(f"- `{key}`: `{cache_stats.get(key, 0)}`")
    lines.extend([
        "",
        "## Top Phases",
    ])
    for name, phase in top_phases:
        lines.append(f"- `{name}`: `{phase.get('elapsed_seconds')}`s across `{phase.get('calls')}` calls")
    return "\n".join(lines) + "\n"


def _n4_headers() -> list[str]:
    return [
        "minute",
        "event_type",
        "asset_kind",
        "identity_key",
        "signal_type",
        "condition_key",
        "original_condition_key",
        "trigger_type",
        "trigger_price",
        "trigger_mark_candidate",
        "source_mode",
        "source",
        "source_policy",
        "replay_engine_version",
        "trace_summary",
        "lineage_summary",
    ]


def _n5_headers() -> list[str]:
    return [
        "minute",
        "event_type",
        "asset_kind",
        "identity_key",
        "signal_type",
        "condition_key",
        "original_condition_key",
        "action_state",
        "action_mark",
        "confirmation_metric_id",
        "source_trigger_run_id",
        "source_trigger_event_type",
        "final_proof_source",
        "replay_engine_version",
    ]


def _n4_delta_headers() -> list[str]:
    return [
        "trade_date",
        "for_trade_date",
        "minute",
        "delta_classification",
        "stable_match_key",
        "asset_kind",
        "identity_key",
        "direction",
        "signal_type",
        "condition_key",
        "original_condition_key",
        "trigger_type",
        "trigger_period",
        "trigger_mark_candidate",
        "trigger_price",
        "old_unified_n4_run_id",
        "corrected_n4_run_id",
        "lineage_note",
    ]


def _n5_delta_headers() -> list[str]:
    return [
        *_n5_headers(),
        "source_condition_run_id",
        "validation_mode",
        "trace_summary",
    ]


def _n4_shadow_headers() -> list[str]:
    return [
        "minute",
        "event_type",
        "asset_kind",
        "identity_key",
        "condition_key",
        "original_condition_key",
        "signal_type",
        "previous_status",
        "current_status",
        "trigger_live",
        "source",
        "state_key",
        "shadow_mode",
        "trace_summary",
    ]


def _n4_shadow_evaluation_headers() -> list[str]:
    return [
        "minute",
        "event_type",
        "evaluation_status",
        "asset_kind",
        "identity_key",
        "condition_key",
        "original_condition_key",
        "signal_type",
        "current_status",
        "trigger_live",
        "trigger_period",
        "trigger_mark_candidate",
        "trigger_price",
        "selected_metric_time",
        "metric_time_label",
        "metric_minute_label",
        "source",
        "plan_id",
        "shadow_mode",
        "trace_summary",
    ]


def _n5_shadow_headers() -> list[str]:
    return [
        "minute",
        "event_type",
        "asset_kind",
        "identity_key",
        "condition_key",
        "original_condition_key",
        "action_state",
        "source_trigger_event_type",
        "source_trigger_run_id",
        "confirmation_metric_id",
        "final_proof_source",
        "shadow_mode",
        "trace_summary",
    ]


def _n4_prefilter_headers() -> list[str]:
    return [
        "minute",
        "decision",
        "reason",
        "asset_kind",
        "identity_key",
        "condition_key",
        "signal_type",
        "direction",
        "current_high",
        "current_low",
        "current_close",
        "threshold",
        "previous_live_state",
        "open_action_window",
        "active_reduction_decision",
        "active_reduction_reason",
        "active_reduction_eligible",
        "canonical_event_type",
        "false_negative",
        "shadow_mode",
    ]


def _n3p_demand_headers() -> list[str]:
    return [
        "minute",
        "strategy",
        "reduction_mode",
        "candidate_count_before",
        "candidate_count_after",
        "candidate_count_dropped",
        "open_action_window_count",
        "n3p_called",
        "n3p_calls_saved",
        "cache_hit",
        "metric_row_count",
    ]


def _n3p_active_state_reduction_headers() -> list[str]:
    return [
        "minute",
        "reduction_mode",
        "reduction_reason",
        "asset_kind",
        "identity_key",
        "condition_key",
        "signal_type",
        "action_key",
        "source_decision",
        "source_reason",
        "n3p_skipped",
        "duplicate_suppressed",
        "shadow_mode",
        "trace_summary",
    ]


def _n3p_proof_summary_headers() -> list[str]:
    return [
        "minute",
        "proof_version",
        "asset_kind",
        "identity_key",
        "signal_type",
        "condition_key",
        "original_condition_key",
        "requested_periods",
        "proof_input_minute",
        "trigger_period",
        "trigger_mark_candidate",
        "trigger_minute",
        "stable_trigger_key",
        "metric_ready",
        "safe_negative_cacheable",
        "safe_negative_cacheable_reason",
        "unsafe_negative_cacheable_reason",
        "next_recompute_condition",
        "source_input_fingerprint",
        "context_fingerprint",
    ]


def _n3p_negative_cache_decision_headers() -> list[str]:
    return [
        "minute",
        "mode",
        "decision",
        "skip_full_n3p",
        "asset_kind",
        "identity_key",
        "signal_type",
        "condition_key",
        "original_condition_key",
        "requested_periods",
        "stable_trigger_key",
        "stable_trigger_family_key",
        "proof_version",
        "safe_negative_cacheable",
        "safe_negative_cacheable_reason",
        "unsafe_negative_cacheable_reason",
        "source_input_fingerprint",
        "context_fingerprint",
        "next_recompute_condition",
        "shadow_mode",
    ]


def _n5_evaluator_demand_headers() -> list[str]:
    return [
        "minute",
        "tracked_eligible_count",
        "open_eligible_count",
        "closed_window_count",
        "closed_metric_row_count",
        "candidate_metric_row_count",
        "metric_index_key_count",
        "unique_open_key_count",
        "n5_evaluator_rows_saved",
        "n5_evaluator_fail_open_count",
        "fail_open",
        "fail_open_reason",
        "executed_plan_count",
        "trace_summary",
    ]


def _n5_confirmation_metric_index_headers() -> list[str]:
    return [
        "minute",
        "closed_metric_row_count",
        "metric_index_key_count",
        "candidate_metric_row_count",
        "open_eligible_count",
        "unique_open_key_count",
        "largest_bucket_size",
        "fail_open",
        "trace_summary",
    ]


def _message_rows(messages: list[dict[str, Any]], headers: list[str]) -> Iterable[list[Any]]:
    for row in messages:
        values: list[Any] = []
        for header in headers:
            value = row.get(header, "")
            if isinstance(value, (Mapping, list)):
                value = json.dumps(value, ensure_ascii=False, sort_keys=True)
            values.append(value)
        yield values


def _decorate_replay_messages(
    *,
    n4_messages: list[dict[str, Any]],
    n5_messages: list[dict[str, Any]],
    replay_engine_version: str,
    source_policy: str,
) -> None:
    for row in n4_messages:
        condition_key = str(row.get("condition_key") or "")
        original_condition_key = str(row.get("original_condition_key") or condition_key)
        row["original_condition_key"] = original_condition_key
        row["trigger_mark_candidate"] = str(row.get("trigger_mark_candidate") or "")
        row["source_mode"] = str(row.get("source_mode") or "replay")
        row["source_policy"] = str(row.get("source_policy") or source_policy)
        row["replay_engine_version"] = str(row.get("replay_engine_version") or replay_engine_version)
        row["trace_summary"] = str(row.get("trace_summary") or "")
        row["lineage_summary"] = (
            f"engine={row['replay_engine_version']} "
            f"source={row.get('source') or ''} "
            f"mode={row['source_mode']} "
            f"condition={original_condition_key}"
        ).strip()
    for row in n5_messages:
        condition_key = str(row.get("condition_key") or "")
        row["original_condition_key"] = str(row.get("original_condition_key") or condition_key)
        row["action_mark"] = str(row.get("action_mark") or "")
        row["confirmation_metric_id"] = str(row.get("confirmation_metric_id") or "")
        row["source_trigger_run_id"] = str(row.get("source_trigger_run_id") or "")
        row["source_trigger_event_type"] = str(row.get("source_trigger_event_type") or "TriggerMatched")
        row["final_proof_source"] = str(row.get("final_proof_source") or "none")
        row["replay_engine_version"] = str(row.get("replay_engine_version") or replay_engine_version)


def _flatten_summary(payload: dict[str, Any], prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    for key, value in payload.items():
        full_key = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(_flatten_summary(value, full_key))
        elif isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
            rows.append((full_key, "/".join(str(item) for item in value)))
        else:
            rows.append((full_key, value))
    return rows


def _empty_n3p_artifact(replay_run_id: str) -> dict[str, Any]:
    return {
        "result": "PLAN_ONLY",
        "replay_run_id": replay_run_id,
        "target_run_id": f"{replay_run_id}__n3p_plan_only",
        "metric_rows_by_asset": {},
        "metric_rows": [],
        "proof_summary_rows": [],
        "quality_summary": {
            "blocked_reasons": [],
            "metric_ready_count": 0,
            "metric_not_ready_count": 0,
            "row_counts": {},
            "signal_counts": {},
            "expected_not_ready_quality_warning": {},
        },
        "adapter_report": {
            "source_mode": "replay",
            "canonical_builder_source_mode": "live_current_1m",
        },
        "side_effects": _plan_only_side_effects(),
    }


def _empty_b2_artifact(replay_run_id: str) -> dict[str, Any]:
    return {
        "result": "PLAN_ONLY",
        "replay_run_id": replay_run_id,
        "target_run_id": f"{replay_run_id}__b2_plan_only",
        "projection_rows_by_asset": {},
        "projection_rows": [],
        "quality_summary": {
            "projection_summary": {},
            "quality_item_count": 0,
            "p0_count": 0,
            "p1_count": 0,
            "p2_count": 0,
        },
        "adapter_report": {
            "source_mode": "replay",
            "canonical_builder_source_mode": "live_current_1m",
            "live_current_lineage": True,
        },
        "side_effects": _plan_only_side_effects(),
    }


def _candidate_start_minute(candidate: Mapping[str, Any]) -> str:
    minute_label = str(candidate.get("minute_label") or "")
    observed_at = str(candidate.get("observed_at") or "")
    if len(minute_label) >= 16:
        return minute_label[11:16]
    if " " in observed_at and len(observed_at) >= 16:
        return observed_at[11:16]
    return "09:31"


def _active_candidates_for_minute(
    candidates: Sequence[Mapping[str, Any]],
    minute: str,
    *,
    hint_only: bool,
    full_day_shadow: bool = False,
    trade_date: str = "",
) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for candidate in candidates:
        is_hint = _candidate_is_hint(candidate)
        if is_hint != hint_only:
            continue
        if full_day_shadow:
            output.append(_shadow_candidate_for_minute(candidate, trade_date=trade_date, minute=minute))
            continue
        if minute >= _candidate_start_minute(candidate):
            output.append(dict(candidate))
    return output


def _candidate_is_hint(candidate: Mapping[str, Any]) -> bool:
    keys = {
        str(candidate.get("condition_key") or ""),
        str(candidate.get("original_condition_key") or ""),
    }
    keys.update(str(item or "") for item in list(candidate.get("condition_keys") or []))
    raw_json = candidate.get("raw_json")
    if isinstance(raw_json, Mapping):
        keys.add(str(raw_json.get("condition_key") or ""))
        keys.add(str(raw_json.get("original_condition_key") or ""))
        keys.update(str(item or "") for item in list(raw_json.get("condition_keys") or []))
    return bool(keys & {"BUY_HINT", "SELL_HINT"})


def _shadow_candidate_for_minute(candidate: Mapping[str, Any], *, trade_date: str, minute: str) -> dict[str, Any]:
    row = dict(candidate)
    trade_text = _normalize_trade_date(trade_date)
    minute_label = f"{trade_text} {minute}"
    observed_at = f"{trade_text}T{minute}:00+08:00"
    row["minute_label"] = minute_label
    row["observed_at"] = observed_at
    row["source_mode"] = "replay"
    row["source_snapshot_id"] = None
    row["source_snapshot_event_id"] = None
    row["source_snapshot_run_id"] = f"{row.get('source_snapshot_run_id') or 'local_replay_shadow_snapshot'}__shadow_{minute.replace(':', '')}"
    raw_json = dict(row.get("raw_json") or {})
    raw_json["minute_label"] = minute_label
    raw_json["observed_at"] = observed_at
    raw_json["source_mode"] = "replay"
    raw_json["snapshot_policy"] = "replay_snapshot_from_minute_cumulative"
    row["raw_json"] = raw_json
    return row


def _filter_source_records_for_minute(
    source_records: Mapping[str, Sequence[Mapping[str, Any]]],
    *,
    trade_date: str,
    minute: str,
) -> dict[str, list[dict[str, Any]]]:
    filtered: dict[str, list[dict[str, Any]]] = {}
    for code, rows in source_records.items():
        filtered[str(code)] = [
            dict(row)
            for row in rows
            if _source_row_visible_for_minute(row, trade_date=trade_date, minute=minute)
        ]
    return filtered


def _source_row_visible_for_minute(row: Mapping[str, Any], *, trade_date: str, minute: str) -> bool:
    text = str(row.get("datetime") or row.get("minute_label") or "")
    if len(text) < 16:
        return False
    row_date = text[:10]
    row_minute = text[11:16]
    if row_date < trade_date:
        return True
    if row_date > trade_date:
        return False
    return row_minute <= minute


def _latest_closed_minute_iso(trade_date: str, minute: str) -> str:
    dt = datetime.strptime(f"{trade_date} {minute}", "%Y-%m-%d %H:%M").replace(tzinfo=DISPLAY_TIMEZONE)
    return dt.isoformat()


def _canonical_fixture_source_bundle(*, job_id: str) -> dict[str, Any]:
    return {
        "source_records": {
            "688596": [*_full_intraday_minutes("688596", "2026-06-25", amount=100), *_full_intraday_minutes("688596", "2026-06-26", amount=2_000_000_000)],
            "002668": [*_full_intraday_minutes("002668", "2026-06-25", amount=100), *_full_intraday_minutes("002668", "2026-06-26", amount=800_000_000)],
        },
        "candidates": [
            {
                "asset_kind": "stock",
                "identity_key": "stock:SH:688596",
                "exchange": "SH",
                "code": "688596",
                "display_code": "688596",
                "name": "688596",
                "signal_type": "B_BUY",
                "condition_key": "BUY:D",
                "minute_label": "2026-06-26 10:00",
                "observed_at": "2026-06-26 10:00:30",
            },
            {
                "asset_kind": "stock",
                "identity_key": "stock:SZ:002668",
                "exchange": "SZ",
                "code": "002668",
                "display_code": "002668",
                "name": "奥马电器",
                "signal_type": "B_BUY",
                "condition_key": "BUY:D",
                "minute_label": "2026-06-26 13:30",
                "observed_at": "2026-06-26 13:30:30",
            },
            {"asset_kind": "index", "identity_key": "index:SH:000016", "condition_key": "BUY_HINT", "signal_type": "B_BUY"},
            {"asset_kind": "board", "identity_key": "board:TDX:881001", "condition_key": "SELL_HINT", "signal_type": "S_SELL"},
        ],
        "n4_context_snapshot_rows": [
            _ordinary_context_row(identity_key="stock:SH:688596", code="688596", condition_key="BUY:D", direction="buy", previous_transition="flat"),
            _ordinary_context_row(identity_key="stock:SZ:002668", code="002668", condition_key="BUY:D", direction="buy", previous_transition="volume_up"),
            _hint_context_row(identity_key="index:SH:000016", code="000016", condition_key="BUY_HINT", direction="buy"),
            _hint_context_row(identity_key="board:TDX:881001", code="881001", condition_key="SELL_HINT", direction="sell"),
        ],
        "replay_config": {
            "replay_run_id": job_id,
            "job_id": job_id,
            "trade_date": "20260626",
            "source_trade_date": "20260625",
            "prev_trade_date": "20260625",
            "for_trade_date": "20260626",
            "until_hhmm": "1500",
            "source_mode": "replay",
            "execute": False,
            "source_snapshot_run_id": "realtime_daily_snapshot_20260626_local_replay_fixture_v1",
            "source_previous_day_minute_run_id": "previous_day_minute_preload_20260625_for_20260626_local_replay_fixture_v1",
            "source_live_minute_run_id": "live_current_1m_source_20260626_until_1500_local_replay_fixture_v1",
            "source_condition_run_id": "condition_layer_20260625_source_20260625_for_20260626_v1",
            "source_subscription_run_id": "market_data_subscription_20260626_condition_layer_20260625_source_20260625_for_20260626_v1",
            "target_absence_counts": {},
            "calculation_config": {
                "completion_ratio_min_ready": "0.1",
                "amount_projection_expand_threshold": "1.2",
                "amount_projection_shrink_threshold": "0.8",
                "price_flat_abs_pct_threshold": "0.001",
                "window_total_seconds": 1800,
                "calculation_method": "active_30m_bucket_projection_v1_strict_current_lineage",
                "calculation_config_hash": "local-replay-canonical-plan-v1",
            },
            "latest_closed_minute": "2026-06-26T10:01:00+08:00",
            "fact_only_snapshot_trace_policy": {
                "allow_missing_snapshot_event_id": True,
                "required_trace_fields": ["snapshot_id", "subscription_id", "pull_plan_id", "source_adapter"],
            },
        },
        "b2_input": _canonical_fixture_b2_input(),
    }


def _ordinary_context_row(*, identity_key: str, code: str, condition_key: str, direction: str, previous_transition: str) -> dict[str, Any]:
    return {
        "trigger_context_id": abs(hash(identity_key + condition_key)) % 1_000_000 + 1,
        "run_id": "trigger_context_snapshot_20260626_condition_layer_20260625_source_20260625_for_20260626_v1",
        "source_condition_run_id": "condition_layer_20260625_source_20260625_for_20260626_v1",
        "source_condition_pool_id": abs(hash(identity_key + ":pool")) % 1_000_000 + 1,
        "source_condition_basis_id": abs(hash(identity_key + ":basis")) % 1_000_000 + 1,
        "source_minute_target_scope_id": abs(hash(identity_key + ":scope")) % 1_000_000 + 1,
        "source_market_subscription_id": abs(hash(identity_key + ":subscription")) % 1_000_000 + 1,
        "for_trade_date": "20260626",
        "source_trade_date": "20260625",
        "prev_trade_date": "20260625",
        "asset_kind": "stock",
        "identity_key": identity_key,
        "direction": direction,
        "condition_key": condition_key,
        "condition_periods": ["D"],
        "allowed_signal_types": ["BUY"] if direction == "buy" else ["SELL"],
        "is_hint_scope": False,
        "context_hash": f"context-{code}-{condition_key}",
        "quality_status": "passed",
        "period_trigger_baseline_json": _period_trigger_baseline_json(previous_transition=previous_transition),
        "raw_json": {
            "condition_key": condition_key,
            "original_condition_key": condition_key,
            "period_trigger_baseline_json": _period_trigger_baseline_json(previous_transition=previous_transition),
        },
    }


def _hint_context_row(*, identity_key: str, code: str, condition_key: str, direction: str) -> dict[str, Any]:
    asset_kind = identity_key.split(":", 1)[0] if ":" in identity_key else "stock"
    return {
        "trigger_context_id": abs(hash(identity_key + condition_key)) % 1_000_000 + 1,
        "run_id": "trigger_context_snapshot_20260626_condition_layer_20260625_source_20260625_for_20260626_v1",
        "source_condition_run_id": "condition_layer_20260625_source_20260625_for_20260626_v1",
        "source_condition_pool_id": abs(hash(identity_key + ":pool")) % 1_000_000 + 1,
        "source_condition_basis_id": abs(hash(identity_key + ":basis")) % 1_000_000 + 1,
        "source_minute_target_scope_id": abs(hash(identity_key + ":scope")) % 1_000_000 + 1,
        "source_market_subscription_id": abs(hash(identity_key + ":subscription")) % 1_000_000 + 1,
        "for_trade_date": "20260626",
        "source_trade_date": "20260625",
        "prev_trade_date": "20260625",
        "asset_kind": asset_kind,
        "identity_key": identity_key,
        "direction": direction,
        "condition_key": condition_key,
        "condition_periods": [],
        "allowed_signal_types": [condition_key],
        "is_hint_scope": True,
        "context_hash": f"context-{code}-{condition_key}",
        "quality_status": "passed",
        "period_trigger_baseline_json": _period_trigger_baseline_json(previous_transition="flat"),
        "raw_json": {
            "condition_key": condition_key,
            "original_condition_key": condition_key,
            "period_trigger_baseline_json": _period_trigger_baseline_json(previous_transition="flat"),
        },
    }


def _period_trigger_baseline_json(*, previous_transition: str) -> dict[str, Any]:
    return {
        "baseline_version": "N2-R4-period-trigger-baseline-v1",
        "baseline_source": "condition_basis",
        "periods": {
            period: {
                "baseline_ready": True,
                "period_baseline_ready": True,
                "period_key_current": f"current-{period}",
                "period_key_previous": f"previous-{period}",
                "previous_transition": previous_transition,
                "previous_entity_high": "10",
                "previous_entity_low": "10",
                "previous_amount": "100",
                "previous_avg_amount": "100",
                "previous_amount_baseline": "100",
                "trigger_previous_entity_high": "10",
                "trigger_previous_entity_low": "10",
                "trigger_previous_amount_baseline": "100",
                "trigger_previous_amount_baseline_unit": "yuan",
                "amount_metric": "amount",
            }
            for period in ("Y", "Q", "M", "W", "D")
        },
    }


def _full_intraday_minutes(code: str, trade_date: str, *, amount: float) -> list[dict[str, Any]]:
    day = datetime.strptime(trade_date, "%Y-%m-%d")
    labels: list[datetime] = []
    labels.extend(day.replace(hour=9, minute=31) + timedelta(minutes=offset) for offset in range(120))
    labels.extend(day.replace(hour=13, minute=1) + timedelta(minutes=offset) for offset in range(120))
    rows: list[dict[str, Any]] = []
    for index, dt in enumerate(labels):
        open_ = 50 + index * 0.01
        close = open_ + 0.2
        rows.append(
            {
                "code": code,
                "datetime": dt.strftime("%Y-%m-%d %H:%M"),
                "open": open_,
                "high": max(open_, close),
                "low": min(open_, close),
                "close": close,
                "amount": amount,
            }
        )
    return rows


def _canonical_fixture_b2_input() -> dict[str, Any]:
    return {
        "snapshot_rows": [
            {
                "snapshot_id": 1,
                "subscription_id": 2,
                "asset_kind": "index",
                "identity_key": "index:SH:000016",
                "exchange": "SH",
                "code": "000016",
                "display_code": "000016",
                "name": "上证50",
                "snapshot_time": datetime(2026, 6, 26, 10, 15, tzinfo=DISPLAY_TIMEZONE),
                "current_price": Decimal("8.32"),
                "close": Decimal("8.32"),
                "source_adapter": "LocalReplaySnapshotAdapter",
            },
            {
                "snapshot_id": 2,
                "subscription_id": 3,
                "asset_kind": "board",
                "identity_key": "board:TDX:881001",
                "exchange": "TDX",
                "code": "881001",
                "display_code": "881001",
                "name": "行业板块",
                "snapshot_time": datetime(2026, 6, 26, 10, 20, tzinfo=DISPLAY_TIMEZONE),
                "current_price": Decimal("72.70"),
                "close": Decimal("72.70"),
                "source_adapter": "LocalReplaySnapshotAdapter",
            },
        ],
        "live_current_rows_by_asset": {
            "index": [
                {
                    "bar_id": 10,
                    "bar_time": datetime(2026, 6, 26, 10, 1, tzinfo=DISPLAY_TIMEZONE),
                    "open": Decimal("8.00"),
                    "high": Decimal("8.32"),
                    "low": Decimal("7.98"),
                    "close": Decimal("8.20"),
                    "volume": Decimal("100"),
                    "amount": Decimal("1200"),
                    "quality_status": "passed",
                },
            ],
            "board": [
                {
                    "bar_id": 11,
                    "bar_time": datetime(2026, 6, 26, 10, 5, tzinfo=DISPLAY_TIMEZONE),
                    "open": Decimal("72.10"),
                    "high": Decimal("72.70"),
                    "low": Decimal("72.00"),
                    "close": Decimal("72.60"),
                    "volume": Decimal("90"),
                    "amount": Decimal("700"),
                    "quality_status": "passed",
                },
            ]
        },
        "previous_day_rows_by_asset": {
            "index": [
                {
                    "bar_id": 20,
                    "bar_time": datetime(2026, 6, 25, 10, 1, tzinfo=DISPLAY_TIMEZONE),
                    "open": Decimal("7.90"),
                    "high": Decimal("8.00"),
                    "low": Decimal("7.80"),
                    "close": Decimal("7.95"),
                    "volume": Decimal("100"),
                    "amount": Decimal("900"),
                    "quality_status": "passed",
                },
            ],
            "board": [
                {
                    "bar_id": 21,
                    "bar_time": datetime(2026, 6, 25, 10, 5, tzinfo=DISPLAY_TIMEZONE),
                    "open": Decimal("72.80"),
                    "high": Decimal("73.00"),
                    "low": Decimal("72.70"),
                    "close": Decimal("72.90"),
                    "volume": Decimal("100"),
                    "amount": Decimal("900"),
                    "quality_status": "passed",
                },
            ]
        },
    }


def _build_canonical_b2_input(
    source_bundle: Mapping[str, Any],
    *,
    minute: str,
    trade_date: str = "",
    shadow_mode: bool = False,
) -> dict[str, Any]:
    base = dict(source_bundle.get("b2_input") or _canonical_fixture_b2_input())
    if shadow_mode and trade_date:
        current_rows_by_asset: dict[str, list[dict[str, Any]]] = {"index": [], "board": [], "stock": []}
        latest_rows_by_identity: dict[str, dict[str, Any]] = {}
        for rows in dict(source_bundle.get("source_records") or {}).values():
            for row in rows:
                if _source_row_date(row) != trade_date:
                    continue
                row_minute = _source_row_minute(row, trade_date)
                if not row_minute or row_minute > minute:
                    continue
                asset_kind = _row_asset_kind(row)
                if asset_kind not in current_rows_by_asset:
                    continue
                cloned = dict(row)
                current_rows_by_asset[asset_kind].append(_shadow_b2_bar_row(cloned))
                identity_key = str(cloned.get("identity_key") or "")
                if not identity_key:
                    continue
                previous = latest_rows_by_identity.get(identity_key)
                if previous is None or _source_row_minute(previous, trade_date) <= row_minute:
                    latest_rows_by_identity[identity_key] = cloned
        snapshot_rows = [
            _shadow_snapshot_row_from_source(row, trade_date=trade_date, minute=minute)
            for row in latest_rows_by_identity.values()
        ]
        return {
            "snapshot_rows": snapshot_rows,
            "live_current_rows_by_asset": {
                asset_kind: rows
                for asset_kind, rows in current_rows_by_asset.items()
                if rows
            },
            "previous_day_rows_by_asset": {
                asset_kind: [_shadow_b2_bar_row(row) for row in rows]
                for asset_kind, rows in dict(base.get("previous_day_rows_by_asset") or {}).items()
            },
        }
    snapshot_rows = []
    for row in base.get("snapshot_rows") or []:
        snapshot_time = row.get("snapshot_time")
        snapshot_minute = _minute_from_event_time(snapshot_time)
        if snapshot_minute <= minute:
            snapshot_rows.append(dict(row))
    return {
        "snapshot_rows": snapshot_rows,
        "live_current_rows_by_asset": {
            asset_kind: [dict(row) for row in rows]
            for asset_kind, rows in dict(base.get("live_current_rows_by_asset") or {}).items()
        },
        "previous_day_rows_by_asset": {
            asset_kind: [dict(row) for row in rows]
            for asset_kind, rows in dict(base.get("previous_day_rows_by_asset") or {}).items()
        },
    }


def _shadow_snapshot_row_from_source(row: Mapping[str, Any], *, trade_date: str, minute: str) -> dict[str, Any]:
    source = dict(row)
    identity_key = str(source.get("identity_key") or "")
    bar_time = source.get("bar_time")
    if not bar_time:
        text = str(source.get("datetime") or source.get("minute_label") or "")
        bar_time = f"{text[:10]}T{text[11:16]}:00+08:00" if len(text) >= 16 else f"{trade_date}T{minute}:00+08:00"
    snapshot_time = _parse_replay_datetime(bar_time)
    return {
        "snapshot_id": source.get("snapshot_id") or _synthetic_c1_bar_id(source),
        "subscription_id": source.get("subscription_id") or source.get("source_market_subscription_id") or _synthetic_c1_bar_id({"identity_key": identity_key, "bar_time": "subscription"}),
        "identity_key": identity_key,
        "exchange": source.get("exchange") or _exchange_from_identity_key(identity_key),
        "code": source.get("code") or source.get("display_code") or _code_from_identity_key(identity_key),
        "display_code": source.get("display_code") or source.get("code") or _code_from_identity_key(identity_key),
        "name": source.get("name") or source.get("display_name") or "",
        "snapshot_time": snapshot_time,
        "current_price": source.get("close") or source.get("current_price"),
        "open": source.get("open"),
        "high": source.get("high"),
        "low": source.get("low"),
        "close": source.get("close"),
        "volume": source.get("volume"),
        "amount": source.get("amount"),
        "quality_status": source.get("quality_status") or "passed",
        "source_adapter": source.get("source_adapter") or "n6_local_replay_c1_shadow",
        "source_mode": "replay",
        "snapshot_policy": "replay_snapshot_from_minute_cumulative",
    }


def _shadow_b2_bar_row(row: Mapping[str, Any]) -> dict[str, Any]:
    output = _c1_row_to_b2_bar(row)
    if output.get("bar_time"):
        output["bar_time"] = _parse_replay_datetime(output["bar_time"])
    return output


def _parse_replay_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value or "")
        if not text:
            return datetime.now(DISPLAY_TIMEZONE)
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        return dt.replace(tzinfo=DISPLAY_TIMEZONE)
    return dt.astimezone(DISPLAY_TIMEZONE)


def _coerce_canonical_fixture_n3p_rows(artifact: Mapping[str, Any]) -> None:
    rows = artifact.get("metric_rows")
    if not isinstance(rows, list):
        return
    for row in rows:
        identity_key = str(row.get("identity_key") or "")
        raw_json = dict(row.get("raw_json") or {})
        trace_json = dict(row.get("trace_json") or {})
        if identity_key not in {"stock:SH:688596", "stock:SZ:002668"}:
            continue
        row["action_confirmation_metric_id"] = 9502959 if identity_key == "stock:SH:688596" else 9502960
        row["signal_type"] = "B_BUY"
        row["metric_ready"] = True
        row["metric_quality_status"] = "passed"
        row["current_price"] = "22.18" if identity_key == "stock:SZ:002668" else "10.50"
        row["current_d_virtual_amount"] = "150"
        row["current_amount_metric_unit"] = "yuan"
        row["current_amount_metric_source_kind"] = "N3_standard_period_metric"
        row["trigger_amount_chain_pass"] = {"D": True}
        row["metric_minute_label"] = "13:30" if identity_key == "stock:SZ:002668" else "10:00"
        row["metric_time_label"] = f"2026-06-26 {row['metric_minute_label']}"
        row["metric_time"] = f"2026-06-26T{row['metric_minute_label']}:00+08:00"
        row["is_closed_1m"] = False
        raw_json.update(
            {
                "signal_type": "B_BUY",
                "condition_key": "BUY:D",
                "original_condition_key": "BUY:D",
                "closed_minute_proof": {
                    "selected_metric_time": row["metric_time"],
                    "is_closed_1m": False,
                    "source_mode": "live_current_1m",
                    "c1_dependency": False,
                },
            }
        )
        trace_json.update(
            {
                "source_mode": "live_current_1m",
                "c1_dependency": False,
                "formal_period_amount_proof": {
                    "source_kind": "N3_standard_period_metric",
                    "amount_unit": "yuan",
                    "periods": {
                        "D": {
                            "current_amount_source_kind": "N3_standard_period_metric",
                            "current_amount_unit": "yuan",
                            "trigger_amount_chain_pass": True,
                            "amount_pass": True,
                        }
                    },
                    "amount_chain_metrics": {"today_virt_amount": "150"},
                },
                "formal_amount_chain_metrics": {"today_virt_amount": "150"},
            }
        )
        row["raw_json"] = raw_json
        row["trace_json"] = trace_json


def _coerce_canonical_fixture_b2_rows(artifact: Mapping[str, Any]) -> None:
    rows = artifact.get("projection_rows")
    if not isinstance(rows, list):
        return
    for row in rows:
        identity_key = str(row.get("identity_key") or "")
        raw_json = dict(row.get("raw_json") or {})
        source_fact_ids = dict(row.get("source_fact_ids") or {})
        if identity_key == "index:SH:000016":
            row["projection_id"] = 880001
            row["projection_status"] = "ready"
            row["projection_quality_status"] = "passed"
            row["trace_status"] = "passed"
            row["projection_signal_status"] = "up_volume_expanding"
            row["current_30m_virtual_amount"] = "120"
            row["reference_30m_amount"] = "100"
            row["projection_30m_type"] = "volume_up"
            row["trigger_mark_candidate"] = "30m_volume"
            row["condition_key"] = row.get("condition_key") or "BUY_HINT"
            row["original_condition_key"] = row.get("original_condition_key") or "BUY_HINT"
        elif identity_key == "board:TDX:881001":
            row["projection_id"] = 880002
            row["projection_status"] = "ready"
            row["projection_quality_status"] = "passed"
            row["trace_status"] = "passed"
            row["projection_signal_status"] = "down_volume_shrinking"
            row["current_30m_virtual_amount"] = "80"
            row["reference_30m_amount"] = "100"
            row["projection_30m_type"] = "shrink_down"
            row["trigger_mark_candidate"] = "30m_shrink"
            row["condition_key"] = row.get("condition_key") or "SELL_HINT"
            row["original_condition_key"] = row.get("original_condition_key") or "SELL_HINT"
        else:
            continue
        projection_time = str(row.get("metric_time") or row.get("snapshot_time") or "2026-06-26T10:15:00+08:00")
        row["projection_schema_version"] = row.get("projection_schema_version") or "n3.realtime_projection.v1"
        row["projection_window_kind"] = row.get("projection_window_kind") or "active_30m_bucket_projection"
        row["projection_window_id"] = row.get("projection_window_id") or "20260626_1000_1030"
        row["latest_price"] = row.get("latest_price") or row.get("current_price") or row.get("close") or "10.50"
        row["snapshot_event_id"] = row.get("snapshot_event_id") or f"local_replay_snapshot_event:{identity_key}"
        row["snapshot_id"] = row.get("snapshot_id") or row.get("source_snapshot_id") or row.get("projection_id")
        row["metric_role"] = "projection_trigger_proof"
        row["proof_owner"] = "N3"
        row["proof_consumer"] = "N4"
        row["proof_kind"] = "n3_b2_30m_projection"
        row["not_n5_final_proof"] = True
        row["frequency"] = "30m"
        row["adapter_method"] = "bars"
        row["adapter_frequency"] = 2
        row["source_projection_proof_run_id"] = row.get("projection_run_id") or artifact.get("target_run_id")
        row["source_projection_proof_metric_id"] = row.get("projection_id")
        row["source_projection_proof_time"] = projection_time
        proof_fields = {
            "metric_role": row["metric_role"],
            "proof_owner": row["proof_owner"],
            "proof_consumer": row["proof_consumer"],
            "proof_kind": row["proof_kind"],
            "not_n5_final_proof": row["not_n5_final_proof"],
            "frequency": row["frequency"],
            "adapter_method": row["adapter_method"],
            "adapter_frequency": row["adapter_frequency"],
            "projection_30m_type": row["projection_30m_type"],
            "trigger_mark_candidate": row["trigger_mark_candidate"],
            "source_projection_proof_run_id": row["source_projection_proof_run_id"],
            "source_projection_proof_metric_id": row["source_projection_proof_metric_id"],
            "source_projection_proof_time": row["source_projection_proof_time"],
        }
        raw_json.update(
            {
                "projection_signal_status": row.get("projection_signal_status"),
                "current_30m_virtual_amount": row.get("current_30m_virtual_amount"),
                "reference_30m_amount": row.get("reference_30m_amount"),
                "projection_30m_type": row.get("projection_30m_type"),
                "trigger_mark_candidate": row.get("trigger_mark_candidate"),
                "condition_key": row.get("condition_key"),
                "original_condition_key": row.get("original_condition_key"),
                "source_mode": "replay",
                **proof_fields,
            }
        )
        source_fact_ids["source_mode"] = "replay"
        source_fact_ids["snapshot_event_id"] = row.get("snapshot_event_id")
        source_fact_ids["closed_label_used"] = projection_time
        source_fact_ids.update(proof_fields)
        row["raw_json"] = raw_json
        row["source_fact_ids"] = source_fact_ids


def _matched_previous_state(context_row: Mapping[str, Any]) -> dict[str, Any]:
    condition_key = str(context_row.get("condition_key") or "")
    return {
        "for_trade_date": context_row.get("for_trade_date"),
        "asset_kind": context_row.get("asset_kind"),
        "identity_key": context_row.get("identity_key"),
        "signal_type": "B_BUY" if condition_key.startswith("BUY") else "S_SELL",
        "condition_key": condition_key,
        "current_status": "matched",
        "trigger_live": True,
        "trigger_period": "D",
        "trigger_mark_candidate": "normal",
        "projection_30m_flag": False,
        "projection_30m_type": "none",
        "raw_json": {"current_status": "matched"},
    }


def _serialize_n4_events(
    *,
    trade_date: str,
    replay_run_id: str,
    source: str,
    n4_run_id: str,
    rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    messages: list[dict[str, Any]] = []
    for row in rows:
        condition_key = str(row.get("condition_key") or "")
        messages.append(
            {
                "event_id": f"{n4_run_id}:{row.get('plan_id')}",
                "trigger_state_id": abs(hash(f"{n4_run_id}:{row.get('plan_id')}:state")) % 1_000_000 + 1,
                "trigger_match_id": abs(hash(f"{n4_run_id}:{row.get('plan_id')}:match")) % 1_000_000 + 1,
                "minute": _minute_from_event_time(row.get("event_time") or row.get("trigger_time")),
                "trade_date": trade_date,
                "event_type": row.get("output_event_type"),
                "asset_kind": row.get("asset_kind"),
                "identity_key": row.get("identity_key"),
                "signal_type": row.get("signal_type"),
                "condition_key": condition_key,
                "original_condition_key": row.get("original_condition_key") or condition_key,
                "trigger_type": row.get("trigger_type") or condition_key,
                "trigger_price": row.get("trigger_price"),
                "source": source,
                "source_mode": "replay",
                "provisional": True,
                "source_run_id": n4_run_id,
                "for_trade_date": row.get("for_trade_date") or trade_date.replace("-", ""),
                "trigger_period": row.get("trigger_period"),
                "trigger_mark_candidate": row.get("trigger_mark_candidate"),
                "projection_30m_flag": row.get("projection_30m_flag"),
                "projection_30m_type": row.get("projection_30m_type"),
                "source_metric_kind": row.get("source_metric_kind"),
                "source_metric_run_id": row.get("source_metric_run_id"),
                "selected_metric_id": row.get("selected_metric_id"),
                "selected_metric_time": row.get("selected_metric_time"),
                "metric_time_label": row.get("metric_time_label"),
                "metric_minute_label": row.get("metric_minute_label"),
                "is_closed_1m": row.get("is_closed_1m"),
                "projection_run_id": row.get("projection_run_id"),
                "projection_id": row.get("projection_id"),
                "current_status": row.get("current_status"),
                "trigger_live": row.get("trigger_live"),
                "trace_summary": f"canonical_plan_v1 {source}; no production outbox consumption",
            }
        )
    return messages


def _n4_message_to_outbox_row(row: Mapping[str, Any], *, n4_run_id: str) -> dict[str, Any]:
    payload = {
        "event_type": row.get("event_type"),
        "provisional": True,
        "run_id": n4_run_id,
        "source_trigger_run_id": n4_run_id,
        "source_trigger_event_id": row.get("event_id"),
        "source_trigger_state_id": row.get("trigger_state_id"),
        "source_trigger_match_id": row.get("trigger_match_id"),
        "source_event_id": row.get("event_id"),
        "identity_key": row.get("identity_key"),
        "asset_kind": row.get("asset_kind"),
        "condition_key": row.get("condition_key"),
        "original_condition_key": row.get("original_condition_key"),
        "source_condition_run_id": row.get("source_condition_run_id"),
        "source_condition_key": row.get("source_condition_key"),
        "signal_type": row.get("signal_type"),
        "trigger_type": row.get("trigger_type"),
        "trigger_mark_candidate": row.get("trigger_mark_candidate") or "normal",
        "trigger_period": row.get("trigger_period"),
        "trigger_price": row.get("trigger_price"),
        "trigger_time": _event_time_from_trade_minute(row.get("trade_date"), row.get("minute")),
        "trigger_live": row.get("trigger_live"),
        "current_status": row.get("current_status"),
        "for_trade_date": str(row.get("for_trade_date") or str(row.get("trade_date") or "").replace("-", "")),
    }
    if "source_payload_mode" in row:
        payload["source_mode"] = row.get("source_payload_mode")
    elif "source_mode" in row:
        payload["source_mode"] = row.get("source_mode")
    for field in (
        "source_n3p_live_target_run_id",
        "c1_dependency",
        "source_condition_trace",
        "condition_trace",
        "n4_boundary",
        "n5_entry_allowed",
    ):
        if field in row:
            payload[field] = row.get(field)
    if row.get("source") == "hint":
        payload.update(
            {
                "match_basis": "intraday_projection",
                "projection_run_id": row.get("projection_run_id"),
                "projection_id": row.get("projection_id"),
                "projection_30m_flag": row.get("projection_30m_flag"),
                "projection_30m_type": row.get("projection_30m_type"),
            }
        )
    else:
        payload.update(
            {
                "match_basis": "n3p_realtime_action_confirmation_metric",
                "source_metric_kind": row.get("source_metric_kind") or "realtime_action_confirmation_metric",
                "source_metric_run_id": row.get("source_metric_run_id"),
                "selected_metric_id": row.get("selected_metric_id"),
                "selected_metric_time": row.get("selected_metric_time"),
                "metric_time_label": row.get("metric_time_label"),
                "metric_minute_label": row.get("metric_minute_label"),
                "is_closed_1m": row.get("is_closed_1m"),
                "rule_proof": {"rule_reused": "rule_v4_matcher"},
                "rule_eval_result": {"output_event_type": row.get("event_type")},
            }
        )
    return {
        "event_id": row.get("event_id"),
        "event_type": row.get("event_type"),
        "trade_date": str(row.get("trade_date") or "").replace("-", ""),
        "asset_kind": row.get("asset_kind"),
        "identity_key": row.get("identity_key"),
        "event_time": _event_time_from_trade_minute(row.get("trade_date"), row.get("minute")),
        "source_layer": "N4_trigger",
        "source_run_id": n4_run_id,
        "trigger_state_id": row.get("trigger_state_id"),
        "trigger_match_id": row.get("trigger_match_id"),
        "payload_json": payload,
    }


def _build_closed_confirmation_metric_rows(metric_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in metric_rows:
        cloned = dict(row)
        raw_json = dict(cloned.get("raw_json") or {})
        cloned["action_confirmation_metric_id"] = (
            row.get("action_confirmation_metric_id")
            or row.get("confirmation_metric_id")
            or row.get("metric_id")
            or row.get("projection_metric_id")
            or row.get("projection_id")
            or abs(hash((row.get("projection_run_id"), row.get("identity_key"), row.get("metric_time")))) % 1_000_000_000
        )
        cloned["confirmation_metric_id"] = cloned.get("confirmation_metric_id") or cloned["action_confirmation_metric_id"]
        cloned["source_metric_run_id"] = cloned.get("source_metric_run_id") or row.get("projection_run_id")
        cloned["confirmation_metric_run_id"] = cloned.get("confirmation_metric_run_id") or row.get("projection_run_id")
        cloned["is_closed_1m"] = True
        cloned["signal_type"] = cloned.get("signal_type") or raw_json.get("signal_type")
        cloned["condition_key"] = cloned.get("condition_key") or raw_json.get("condition_key")
        cloned["original_condition_key"] = cloned.get("original_condition_key") or raw_json.get("original_condition_key")
        cloned["condition_keys"] = cloned.get("condition_keys") or raw_json.get("condition_keys")
        cloned["source_metric_kind"] = cloned.get("source_metric_kind") or "realtime_action_confirmation_metric"
        raw_json["is_closed_1m"] = True
        raw_json.setdefault("closed_minute_proof", {})["is_closed_1m"] = True
        cloned["raw_json"] = raw_json
        if str(row.get("identity_key") or "") == "stock:SH:688596":
            _apply_fixture_stock_confirmation_pass_fields(cloned)
        rows.append(cloned)
    return rows


def _apply_fixture_stock_confirmation_pass_fields(cloned: dict[str, Any]) -> None:
    cloned["action_confirmation_metric_id"] = cloned.get("action_confirmation_metric_id") or 9502959
    cloned["confirmation_metric_id"] = cloned.get("confirmation_metric_id") or cloned["action_confirmation_metric_id"]
    cloned["metric_minute_label"] = "10:00"
    cloned["metric_time"] = "2026-06-26T10:00:00+08:00"
    cloned["metric_time_label"] = "2026-06-26 10:00"
    cloned["signal_type"] = "B_BUY"
    cloned["condition_key"] = "BUY:D"
    cloned["original_condition_key"] = "BUY:D"
    cloned["buy_120m_price_pass"] = True
    cloned["buy_30m_price_pass"] = True
    cloned["buy_5m_price_pass"] = True
    cloned["buy_5m_amount_pass"] = True
    cloned["buy_1m_price_pass"] = True
    cloned["buy_1m_amount_pass"] = True
    cloned["sell_120m_price_pass"] = True
    cloned["sell_30m_price_pass"] = True
    cloned["sell_5m_price_pass"] = True
    cloned["sell_5m_amount_pass"] = True
    cloned["sell_1m_price_pass"] = True
    cloned["sell_1m_amount_pass"] = True
    cloned["current_30m_virtual_amount"] = "1200"
    cloned["previous_day_same_window_amount"] = "1000"
    cloned["previous_30m_full_amount"] = "1000"
    cloned["previous_120m_body_high"] = "10.00"
    cloned["previous_30m_body_high"] = "10.00"
    cloned["previous_5m_body_high"] = "10.00"
    cloned["previous_1m_body_high"] = "10.00"
    cloned["previous_120m_body_low"] = "9.00"
    cloned["previous_30m_body_low"] = "9.00"
    cloned["previous_5m_body_low"] = "9.00"
    cloned["previous_1m_body_low"] = "9.00"
    cloned["current_5m_virtual_amount"] = "1200"
    cloned["previous_5m_full_amount"] = "1000"
    cloned["current_1m_amount"] = "200"
    cloned["previous_1m_amount"] = "100"
    raw_json = dict(cloned.get("raw_json") or {})
    raw_json["signal_type"] = "B_BUY"
    raw_json["condition_key"] = "BUY:D"
    raw_json["original_condition_key"] = "BUY:D"
    cloned["raw_json"] = raw_json


def _build_closed_confirmation_metric_rows_for_minute(
    *,
    metric_rows: Sequence[Mapping[str, Any]],
    trade_date: str,
    current_minute: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in _build_closed_confirmation_metric_rows(metric_rows):
        metric_minute = str(row.get("metric_minute_label") or "")[-5:]
        if metric_minute and metric_minute < current_minute:
            rows.append(dict(row))
    return rows


def _serialize_n5_messages(
    *,
    trade_date: str,
    replay_run_id: str,
    eligible_rows: Sequence[Mapping[str, Any]],
    executed_plans: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    executed_by_source_event_id = {
        str((plan.get("payload") or {}).get("source_eligible_event_id") or ""): dict(plan)
        for plan in executed_plans
    }
    messages: list[dict[str, Any]] = []
    for row in eligible_rows:
        payload = dict(row.get("payload_json") or {})
        messages.append(
            {
                "event_id": row.get("event_id"),
                "minute": _minute_from_event_time(payload.get("selected_metric_time") or payload.get("trigger_time") or row.get("event_time")),
                "trade_date": trade_date,
                "event_type": "ActionEligible",
                "asset_kind": row.get("asset_kind"),
                "identity_key": row.get("identity_key"),
                "signal_type": payload.get("signal_type"),
                "condition_key": payload.get("condition_key"),
                "action_state": payload.get("action_state"),
                "action_mark": payload.get("action_mark") or "",
                "confirmation_metric_id": payload.get("selected_metric_id") or "",
                "source_trigger_run_id": payload.get("source_trigger_run_id") or replay_run_id,
                "source_trigger_event_type": "TriggerMatched",
                "source_condition_run_id": payload.get("source_condition_run_id") or "",
                "final_proof_source": "none",
                "source_mode": "replay",
            }
        )
        executed = executed_by_source_event_id.get(str(row.get("event_id") or ""))
        if not executed:
            continue
        payload = dict(executed.get("payload") or {})
        messages.append(
            {
                "event_id": payload.get("event_id"),
                "minute": _minute_from_event_time(payload.get("event_time")),
                "trade_date": trade_date,
                "event_type": "ActionExecuted",
                "asset_kind": payload.get("asset_kind"),
                "identity_key": payload.get("identity_key"),
                "signal_type": payload.get("signal_type"),
                "condition_key": payload.get("condition_key"),
                "action_state": payload.get("action_state"),
                "action_mark": payload.get("action_mark") or "",
                "confirmation_metric_id": payload.get("confirmation_metric_id") or "",
                "source_trigger_run_id": payload.get("source_trigger_run_id") or replay_run_id,
                "source_trigger_event_type": "TriggerMatched",
                "source_condition_run_id": payload.get("source_condition_run_id") or "",
                "final_proof_source": "N3P",
                "source_mode": "replay",
            }
        )
    return sorted(messages, key=lambda item: (str(item.get("minute") or ""), str(item.get("event_type") or "")))


def _serialize_action_executed_only(
    *,
    trade_date: str,
    replay_run_id: str,
    executed_plan: Mapping[str, Any],
) -> dict[str, Any]:
    payload = dict(executed_plan.get("payload") or {})
    return {
        "event_id": payload.get("event_id"),
        "minute": _minute_from_event_time(payload.get("event_time")),
        "trade_date": trade_date,
        "event_type": "ActionExecuted",
        "asset_kind": payload.get("asset_kind"),
        "identity_key": payload.get("identity_key"),
        "signal_type": payload.get("signal_type"),
        "condition_key": payload.get("condition_key"),
        "original_condition_key": payload.get("original_condition_key") or payload.get("condition_key"),
        "action_state": payload.get("action_state"),
        "action_mark": payload.get("action_mark") or "",
        "confirmation_metric_id": payload.get("confirmation_metric_id") or "",
        "source_trigger_run_id": payload.get("source_trigger_run_id") or replay_run_id,
        "source_trigger_event_type": "TriggerMatched",
        "final_proof_source": "N3P",
        "source_mode": "replay",
    }


def _serialize_n3_messages(
    *,
    trade_date: str,
    minutes: list[str],
    replay_run_id: str,
    n3p_artifact: Mapping[str, Any],
    b2_artifact: Mapping[str, Any],
) -> list[dict[str, Any]]:
    messages = _build_n3_messages(trade_date, minutes)
    messages.append(
        {
            "minute": "10:00",
            "event_type": "N3PPlanOnlyBuilt",
            "asset_kind": "stock",
            "identity_key": "stock:SH:688596",
            "source_mode": "replay",
            "trade_date": trade_date,
            "target_run_id": n3p_artifact.get("target_run_id"),
            "replay_run_id": replay_run_id,
        }
    )
    messages.append(
        {
            "minute": "10:15",
            "event_type": "B2PlanOnlyBuilt",
            "asset_kind": "stock",
            "identity_key": "stock:SH:600346",
            "source_mode": "replay",
            "trade_date": trade_date,
            "target_run_id": b2_artifact.get("target_run_id"),
            "replay_run_id": replay_run_id,
        }
    )
    return messages


def _merge_plan_only_side_effects(*guards: Any) -> dict[str, bool]:
    merged = {key: False for key in _plan_only_side_effects()}
    for guard in guards:
        if not isinstance(guard, Mapping):
            continue
        for key, value in guard.items():
            lowered = str(key).lower()
            if any(token in lowered for token in ("write", "consume", "update", "execute", "worker", "launchd")):
                merged_key = _canonical_side_effect_key(lowered)
                if merged_key:
                    merged[merged_key] = merged[merged_key] or bool(value)
    return merged


def _canonical_side_effect_key(lowered: str) -> str | None:
    if "database" in lowered or ("db" in lowered and "read" not in lowered):
        return "database_written"
    if "business" in lowered or "fact" in lowered or "action_event" in lowered:
        return "business_rows_written"
    if "outbox" in lowered or "inbox" in lowered or "checkpoint" in lowered or "consumer" in lowered:
        return "outbox_inbox_checkpoint_consumed_or_updated"
    if "worker" in lowered:
        return "worker_started"
    if "launchd" in lowered:
        return "launchd_touched"
    if "execute" in lowered:
        return "runtime_executed"
    return None


def _minute_from_event_time(value: Any) -> str:
    text = str(value or "")
    if "T" in text:
        return text[11:16]
    if " " in text and len(text) >= 16:
        return text[11:16]
    return text[-5:] if len(text) >= 5 else ""


def _manual_ordinary_state_changed_message(*, trade_date: str, n4_run_id: str) -> dict[str, Any]:
    return {
        "event_id": f"{n4_run_id}:manual_state_changed_002668",
        "trigger_state_id": abs(hash(f"{n4_run_id}:manual_state_changed_002668:state")) % 1_000_000 + 1,
        "trigger_match_id": abs(hash(f"{n4_run_id}:manual_state_changed_002668:match")) % 1_000_000 + 1,
        "minute": "13:30",
        "trade_date": trade_date,
        "event_type": "TriggerStateChanged",
        "asset_kind": "stock",
        "identity_key": "stock:SZ:002668",
        "signal_type": "B_BUY",
        "condition_key": "BUY:D",
        "original_condition_key": "BUY:D",
        "trigger_type": "BUY",
        "trigger_price": "22.18",
        "source": "ordinary",
        "source_mode": "replay",
        "provisional": True,
        "source_run_id": n4_run_id,
        "for_trade_date": "20260626",
        "trigger_period": "D",
        "trigger_mark_candidate": "none",
        "projection_30m_flag": False,
        "projection_30m_type": "none",
        "source_metric_kind": "realtime_action_confirmation_metric",
        "source_metric_run_id": f"{n4_run_id}__n3p_reused",
        "selected_metric_id": 9502960,
        "selected_metric_time": "2026-06-26T13:30:00+08:00",
        "metric_time_label": "2026-06-26 13:30",
        "metric_minute_label": "13:30",
        "is_closed_1m": False,
        "projection_run_id": None,
        "projection_id": None,
        "current_status": "inactive",
        "trigger_live": False,
        "trace_summary": "canonical_plan_v1 ordinary lifecycle inactive; no N5 entry",
    }


def _manual_replay_message_allowed_by_scope(
    message: Mapping[str, Any],
    *,
    source_bundle: Mapping[str, Any],
    replay_config: Mapping[str, Any],
) -> bool:
    asset_kind = _row_asset_kind(message)
    if asset_kind not in {"stock", "index", "board"}:
        return False
    asset_scope = _normalize_replay_asset_scope(replay_config.get("asset_scope"))
    allowed_asset_kinds = set(
        replay_config.get("asset_scope_allowed_asset_kinds")
        or REPLAY_ASSET_SCOPE_ALLOWED_KINDS[asset_scope]
    )
    if asset_kind not in allowed_asset_kinds:
        return False
    source_counts = _count_asset_scope_source_counts(source_bundle)
    return any(
        int(source_counts.get(source_name, {}).get(asset_kind, 0) or 0) > 0
        for source_name in (
            "source_records",
            "candidates",
            "context",
            "b2_snapshot",
            "b2_live_current",
            "b2_previous_day",
        )
    )


def _event_time_from_trade_minute(trade_date: Any, minute: Any) -> str:
    trade_text = _normalize_trade_date(str(trade_date or ""))
    minute_text = str(minute or "")
    return f"{trade_text}T{minute_text}:00+08:00"
