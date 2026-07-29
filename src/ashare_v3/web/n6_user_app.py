"""FastAPI app for the N6 user login and read-only projection MVP."""

from __future__ import annotations

import asyncio
import base64
import binascii
import copy
from contextlib import nullcontext
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
import hmac
from io import BytesIO
import json
import os
from pathlib import Path
import re
import secrets
import stat
import threading
import time
from typing import Any, Callable, Mapping, Protocol
from urllib.parse import parse_qs, quote, urlsplit, urlunsplit
from zoneinfo import ZoneInfo

import psycopg
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

from fastapi import FastAPI, HTTPException, Request
from fastapi.encoders import jsonable_encoder
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ashare_v3.user.admin_bootstrap import HashResult, hash_password, validate_password
from ashare_v3.user.membership_drilldown import (
    get_board_membership_stocks,
    get_index_membership_stocks,
)
from ashare_v3.user.stale_active_lineage import stale_source_action_run_ids
from ashare_v3.web.n6_app_v1 import (
    POSTGRES_SIGNED_BIGINT_MAX,
    DEFAULT_REALTIME_SCOPE_INDEXES,
    app_account_model,
    app_ai_agent_public_decision_detail_model,
    app_ai_agent_public_model,
    app_ai_agent_public_section_model,
    app_ai_users_model,
    app_dashboard_model,
    app_empty_planned_model,
    app_leaderboard_model,
    app_locked_future_module_model,
    app_me_model,
    app_message_asset_kind,
    app_message_asset_kinds,
    app_page_model,
    app_pnl_model,
    app_portfolio_model,
    app_realtime_scope_model,
    app_v2_filter_center_model,
    app_v2_expected_return_threshold,
    app_v2_expected_return_value_text,
    app_v2_filter_linked_stocks_model,
    app_v2_filter_members_model,
    app_v2_filter_api_model,
    app_v2_filter_model,
    app_v2_level_up_recommendation_value,
    app_v2_buy_messages_model,
    app_v2_membership_drilldown_model,
    app_v2_message_dashboard_model,
    app_v2_message_groups_model,
    app_v2_message_projection_status_model,
    app_v2_monitor_model,
    app_signal_detail_model,
    app_signal_item,
    app_signal_sse_data_model,
    app_signals_model,
    app_status_monitor_model,
    app_trade_proposals_model,
    app_user_guide_model,
    app_virtual_trades_model,
    app_watchlist_model,
    canonical_bigint_id,
    V2_FILTER_VISIBLE_FIELDS_BY_ASSET,
)
from ashare_v3.web.n6_ui_v1 import (
    artifacts_model,
    cash_ledger_model,
    cash_snapshot_model,
    dashboard_metrics_model,
    lineage_stats_model,
    message_dashboard_model,
    n2_condition_basis_model,
    n2_condition_basis_item,
    n3_messages_model,
    n4_messages_model,
    n5_messages_model,
    post_close_fastlane_status_model,
    rag_search_model,
    READ_ONLY_SIDE_EFFECTS,
    runtime_archive_status_model as _runtime_archive_status_model,
    input_messages_model,
    rollback_summary_model,
    signal_detail_model,
    signal_list_model,
    status_monitor_model,
    virtual_account_summary_model,
)
from ashare_v3.web.n6_btrack_authority import (
    N6BTrackAuthorityRepository,
    PostgresN6BTrackAuthorityRepository,
)
from ashare_v3.web.post_close_fastlane_status import read_post_close_fastlane_status
from ashare_v3.web.rag_status import read_rag_status_answer
from ashare_v3.web.runtime_archive_status import read_runtime_archive_status


COOKIE_NAME = "ashare_v3_n6_session"
SESSION_HASH_ALGO = "sha256"
TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"
PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_DSN = "postgresql://ashare_v3_user@127.0.0.1:5432/ashare_v3"
N6_BTRACK_WEB_DB_SERVICE = "n6_btrack_web"
N6_CSRF_SECRET_MAX_BYTES = 4096
N6_SCOPE_BULK_SELECTION_TTL_SECONDS = 5 * 60
N6_SCOPE_BULK_MAX_IDENTITIES = 10000
DEFAULT_B_TRACK_ENTRY = "/n6/app"
DEFAULT_A_TRACK_ADMIN_ENTRY = "/n6/post-close-fastlane-status"
DEFAULT_POST_CLOSE_FASTLANE_DOCS_ROOT = str(PROJECT_ROOT / "docs/post_close_fastlane")
DEFAULT_RUNTIME_ARCHIVE_DOCS_ROOT = str(PROJECT_ROOT / "docs/runtime_archive")
DEFAULT_RUNTIME_ARCHIVE_ROOT = "/Volumes/MacRaid/stock_db_archive/v3_runtime"
DEFAULT_RAG_DOCS_ROOT = str(PROJECT_ROOT / "docs")
DEFAULT_RAG_SQL_ROOT = str(PROJECT_ROOT / "sql")
DEFAULT_REPLAY_DOCS_ROOT = str(PROJECT_ROOT / "docs/replay")
DEFAULT_PROFILE_NAME = "MVP default"
DEFAULT_SIM_ACCOUNT_NAME = "MVP T+1 shadow account"
DEFAULT_INITIAL_CASH = 1000000000
DISPLAY_TIMEZONE = ZoneInfo("Asia/Shanghai")
APP_V3_PUBLIC_BIGINT_FIELDS = frozenset(
    {
        "monitor_id",
        "principal_id",
        "user_id",
        "realtime_scope_id",
        "proposal_id",
        "source_signal_projection_id",
        "source_virtual_position_id",
        "virtual_account_id",
        "virtual_position_id",
        "virtual_position_lot_id",
        "virtual_order_id",
        "virtual_trade_id",
        "source_virtual_trade_id",
        "virtual_cash_snapshot_id",
        "virtual_cash_ledger_id",
        "virtual_quote_snapshot_id",
        "fill_quote_snapshot_id",
        "stop_loss_source_quote_snapshot_id",
    }
)


def app_v3_public_payload(value: Any) -> Any:
    if isinstance(value, list):
        return [app_v3_public_payload(item) for item in value]
    if isinstance(value, tuple):
        return [app_v3_public_payload(item) for item in value]
    if not isinstance(value, dict):
        return value
    output: dict[str, Any] = {}
    for key, item in value.items():
        if key in APP_V3_PUBLIC_BIGINT_FIELDS:
            output[key] = canonical_bigint_id(item, field_name=key)
        else:
            output[key] = app_v3_public_payload(item)
    return output


def runtime_archive_status_model(data: dict[str, Any]) -> dict[str, Any]:
    """Extend the read-only archive model with the combined local file cleanup."""

    model = _runtime_archive_status_model(data)
    hot_cleanup = dict(data.get("hot_cleanup") or {})
    local_cleanup = dict(hot_cleanup.get("local_file_cleanup") or {})
    per_layer = dict(local_cleanup.get("per_layer") or {})
    model["local_file_cleanup"] = {
        "result": str(local_cleanup.get("result") or "NO_LOCAL_FILE_CLEANUP_STATUS"),
        "mode": str(local_cleanup.get("mode") or "dry_run"),
        "started_at": str(local_cleanup.get("started_at") or ""),
        "finished_at": str(local_cleanup.get("finished_at") or ""),
        "duration_ms": float(local_cleanup.get("duration_ms") or 0),
        "retention_trade_days": int(local_cleanup.get("retention_trade_days") or 5),
        "retained_trade_dates": [str(item) for item in list(local_cleanup.get("retained_trade_dates") or [])],
        "cleanup_trade_dates": [str(item) for item in list(local_cleanup.get("cleanup_trade_dates") or [])],
        "deleted_file_count": int(local_cleanup.get("deleted_file_count") or 0),
        "deleted_directory_count": int(local_cleanup.get("deleted_directory_count") or 0),
        "released_bytes": int(local_cleanup.get("released_bytes") or 0),
        "per_layer": {
            layer: {
                "deleted_file_count": int(dict(per_layer.get(layer) or {}).get("deleted_file_count") or 0),
                "deleted_directory_count": int(dict(per_layer.get(layer) or {}).get("deleted_directory_count") or 0),
                "released_bytes": int(dict(per_layer.get(layer) or {}).get("released_bytes") or 0),
            }
            for layer in ("n3", "n4", "n5")
        },
        "errors": [str(item) for item in list(local_cleanup.get("errors") or [])],
        "blockers": [str(item) for item in list(local_cleanup.get("blockers") or [])],
        "status_path": str(data.get("hot_cleanup_source_path") or ""),
    }
    return model
N6_UI_V1_LINEAGE_N4_RUN_ID = "trigger_execute_20260605_condition_layer_20260604_source_20260604_v1"
N6_UI_V1_LINEAGE_N5_RUN_ID = (
    "action_consumer_action_pipeline_20260605_trigger_execute_20260605_condition_layer_20260604_source_20260604_v1"
)
N6_UI_V1_LINEAGE_N6_RUN_ID = (
    "user_projection_shadow_20260605__action_consumer_action_pipeline_20260605_trigger_execute_20260605_condition_layer_20260604_source_20260604_v1"
)
N4_STANDARD_EVENT_TYPES = ("TriggerMatched", "TriggerPendingMarketData", "TriggerStateChanged")
N4_LEGACY_EVENT_TYPES = ("TriggerCleared", "TriggerLiveChanged")
N4_ALL_EVENT_TYPES = N4_STANDARD_EVENT_TYPES + N4_LEGACY_EVENT_TYPES
N4_MESSAGE_DEFAULT_LIMIT = 200
N3_MESSAGE_EVENT_TYPES = (
    "MarketSnapshotUpdated",
    "MinuteBarClosed",
    "MinuteBarCorrected",
    "MarketDataDelayed",
    "MarketDataMissing",
    "MarketDisplaySnapshotUpdated",
)
N3_MESSAGE_DEFAULT_LIMIT = 200
N5_STANDARD_EVENT_TYPES = ("ActionEligible", "ActionBlocked", "ActionExecuted", "ActionSkipped")
N5_LEGACY_EVENT_TYPES = ("ActionEvent", "HintEvent", "RiskEvent", "PositionEvent")
N5_ALL_EVENT_TYPES = N5_STANDARD_EVENT_TYPES + N5_LEGACY_EVENT_TYPES
N5_ACTION_DISPLAY_EVENT_TYPES = ("ActionExecuted", "ActionEligible")
N5_MESSAGE_DEFAULT_LIMIT = 200
N6_SIGNAL_PAGE_DEFAULT_LIMIT = 50
N6_SIGNAL_PAGE_MAX_LIMIT = 100
N6_SIGNAL_SSE_BATCH_LIMIT = 100
N6_SIGNAL_SSE_DB_READ_INTERVAL_SECONDS = 2.0
N6_SIGNAL_SSE_HEARTBEAT_SECONDS = 15.0
N6_SIGNAL_SSE_RETRY_MILLISECONDS = 5000
N6_SIGNAL_SSE_MAX_CURSOR = POSTGRES_SIGNED_BIGINT_MAX
N6_SIGNAL_SSE_MAX_CURSOR_TEXT = str(POSTGRES_SIGNED_BIGINT_MAX)
N6_SIGNAL_COMPACT_FIELDS = (
    "user_signal_projection_id",
    "user_signal_card_id",
    "user_projection_run_id",
    "event_type",
    "event_time",
    "event_label",
    "trade_date",
    "identity_key",
    "asset_kind",
    "asset_kind_label",
    "display_code",
    "display_name",
    "industry_code",
    "industry_name",
    "direction",
    "direction_label",
    "signal_type",
    "condition_key",
    "triggered_periods",
    "primary_trigger_period",
    "target_price",
    "buy_target",
    "buy_target_price",
    "sell_target_price",
    "display_values",
    "current_price",
    "trigger_price",
    "action_price",
    "expected_return_pct",
    "buy_return",
    "secondary_return",
    "sell_return",
    "up_ref",
    "down_ref",
    "score",
    "pe_core",
    "trigger_pct",
    "action_pct",
    "projection_message_status",
    "action_state",
    "action_state_label",
    "action_mark",
    "blocked_reason",
    "blocked_reason_label",
    "quality_status",
    "source_run_id",
    "projection_run_id",
)
def encode_n6_signal_page_cursor(row: Mapping[str, Any]) -> str:
    projection_id = canonical_bigint_id(
        row.get("user_signal_projection_id"),
        field_name="user_signal_projection_id",
        required=True,
    )
    created_at = row.get("created_at")
    if not isinstance(created_at, datetime):
        raise ValueError("invalid_signal_page_cursor")
    payload = json.dumps(
        [created_at.isoformat(), projection_id],
        ensure_ascii=True,
        separators=(",", ":"),
    ).encode("ascii")
    return base64.urlsafe_b64encode(payload).decode("ascii").rstrip("=")


def parse_n6_signal_page_cursor(value: Any) -> tuple[datetime, int] | None:
    text = str(value or "").strip()
    if not text:
        return None
    if len(text) > 160 or not re.fullmatch(r"[A-Za-z0-9_-]+", text):
        raise ValueError("invalid_signal_page_cursor")
    try:
        padding = "=" * (-len(text) % 4)
        decoded = base64.urlsafe_b64decode((text + padding).encode("ascii"))
        created_at_text, projection_id_raw = json.loads(decoded.decode("ascii"))
        created_at = datetime.fromisoformat(str(created_at_text))
        projection_id = int(projection_id_raw)
    except (ValueError, TypeError, binascii.Error, json.JSONDecodeError) as exc:
        raise ValueError("invalid_signal_page_cursor") from exc
    if (
        created_at.tzinfo is None
        or projection_id <= 0
        or projection_id > N6_SIGNAL_SSE_MAX_CURSOR
    ):
        raise ValueError("invalid_signal_page_cursor")
    return created_at, projection_id


def n6_signal_keyset_page(
    rows: list[dict[str, Any]],
    *,
    limit: int,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    page_limit = max(1, min(int(limit), N6_SIGNAL_PAGE_MAX_LIMIT))
    items = rows[:page_limit]
    has_more = len(rows) > page_limit
    next_cursor = encode_n6_signal_page_cursor(items[-1]) if has_more and items else ""
    watermark = max(
        (
            int(row["user_signal_projection_id"])
            for row in items
            if row.get("user_signal_projection_id") is not None
        ),
        default=0,
    )
    return items, {
        "limit": page_limit,
        "has_more": has_more,
        "next_cursor": next_cursor,
        "watermark": str(watermark),
    }


def n6_compact_signal_item(item: Mapping[str, Any]) -> dict[str, Any]:
    compact = {
        field: item.get(field)
        for field in N6_SIGNAL_COMPACT_FIELDS
        if field in item
    }
    condition_trace = item.get("condition_trace")
    if isinstance(condition_trace, Mapping):
        compact["condition_rendering_policy"] = condition_trace.get("rendering_policy")
    elif "condition_rendering_policy" in item:
        compact["condition_rendering_policy"] = item.get("condition_rendering_policy")
    return compact


def n6_compact_signal_payload(payload: Mapping[str, Any]) -> dict[str, Any]:
    compact = dict(payload)
    compact["items"] = [
        n6_compact_signal_item(item)
        for item in list(payload.get("items") or [])
    ]
    return compact


def n6_compact_message_dashboard_payload(payload: Mapping[str, Any]) -> dict[str, Any]:
    compact = dict(payload)
    compact.pop("items_preview", None)
    compact["groups"] = [
        {
            key: value
            for key, value in dict(group).items()
            if key != "items_preview"
        }
        for group in list(payload.get("groups") or [])
    ]
    return compact


def n6_signal_response_etag(
    *,
    principal_id: int,
    filters: Mapping[str, Any],
    pagination: Mapping[str, Any],
) -> str:
    fingerprint = json.dumps(
        {
            "principal_id": int(principal_id),
            "filters": {
                key: value
                for key, value in sorted(filters.items())
                if key not in {"before_created_at", "before_id"}
            },
            "pagination": dict(pagination),
        },
        ensure_ascii=True,
        separators=(",", ":"),
        sort_keys=True,
        default=str,
    )
    return f'"{hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()}"'


def n6_signal_json_response(
    request: Request,
    payload: dict[str, Any],
    *,
    etag: str,
    watermark: str,
) -> Response:
    headers = {
        "ETag": etag,
        "X-N6-Watermark": watermark,
        "Cache-Control": "private, no-cache",
    }
    if request.headers.get("if-none-match") == etag:
        return Response(status_code=304, headers=headers)
    return JSONResponse(payload, headers=headers)


def parse_n6_signal_sse_cursor(*, last_event_id: Any, after_id: Any) -> int:
    candidate = last_event_id if last_event_id is not None and str(last_event_id).strip() else after_id
    if candidate is None or not str(candidate).strip():
        return 0
    text = str(candidate).strip()
    if not re.fullmatch(r"[0-9]+", text):
        raise ValueError("invalid_signal_sse_cursor")
    value = int(text)
    if value > N6_SIGNAL_SSE_MAX_CURSOR:
        raise ValueError("invalid_signal_sse_cursor")
    return value


def encode_n6_signal_sse_event(payload: dict[str, Any]) -> str:
    signal = dict(payload.get("signal") or {})
    raw_projection_id = signal.get("user_signal_projection_id")
    if (
        not isinstance(raw_projection_id, str)
        or not re.fullmatch(r"[1-9][0-9]*", raw_projection_id)
        or len(raw_projection_id) > len(N6_SIGNAL_SSE_MAX_CURSOR_TEXT)
        or (
            len(raw_projection_id) == len(N6_SIGNAL_SSE_MAX_CURSOR_TEXT)
            and raw_projection_id > N6_SIGNAL_SSE_MAX_CURSOR_TEXT
        )
    ):
        raise ValueError("invalid_signal_sse_event_id")
    data = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return f"event: signal\nid: {raw_projection_id}\ndata: {data}\n\n"


def encode_n6_signal_sse_heartbeat(current_time: datetime) -> str:
    return f": heartbeat {current_time.astimezone(timezone.utc).isoformat()}\n\n"


async def iter_n6_signal_sse(
    *,
    after_id: int,
    read_batch: Callable[[int, int], Any],
    is_disconnected: Callable[[], Any],
    sleep: Callable[[float], Any] = asyncio.sleep,
    monotonic: Callable[[], float] = time.monotonic,
    now: Callable[[], datetime] | None = None,
) -> Any:
    current_time = now or (lambda: datetime.now(timezone.utc))
    cursor = int(after_id)
    heartbeat_at = monotonic() + N6_SIGNAL_SSE_HEARTBEAT_SECONDS
    yield f"retry: {N6_SIGNAL_SSE_RETRY_MILLISECONDS}\n\n"
    while True:
        if await is_disconnected():
            return
        try:
            rows = await read_batch(cursor, N6_SIGNAL_SSE_BATCH_LIMIT)
        except Exception:
            return
        for row in rows:
            projection_id = row.get("user_signal_projection_id")
            if isinstance(projection_id, bool) or not isinstance(projection_id, int):
                return
            if projection_id <= cursor or projection_id > N6_SIGNAL_SSE_MAX_CURSOR:
                return
            payload = app_signal_sse_data_model(row)
            yield encode_n6_signal_sse_event(payload)
            cursor = projection_id
        if len(rows) >= N6_SIGNAL_SSE_BATCH_LIMIT:
            continue
        current_tick = monotonic()
        if current_tick >= heartbeat_at:
            yield encode_n6_signal_sse_heartbeat(current_time())
            heartbeat_at = current_tick + N6_SIGNAL_SSE_HEARTBEAT_SECONDS
        await sleep(N6_SIGNAL_SSE_DB_READ_INTERVAL_SECONDS)


def _sql_text_literal(value: object) -> str:
    return "'" + str(value).replace("'", "''") + "'"


V3_20260612_ACTIVE_N4_SOURCE_RUN_ID = "v3_n4_trigger_replay_20260612_after_trigger_period_baseline_fix_v1"
V3_20260612_ACTIVE_N5_SOURCE_RUN_ID = "v3_n5_action_replay_20260612_after_n4_trigger_period_baseline_fix_v1"
V3_20260622_ACTIVE_N4_SOURCE_RUN_ID = (
    "trigger_replay_phase2d_20260622_formal_unitfix_dseed_periodguard_until_1500__"
    "condition_layer_20260618_source_20260618_for_20260622_v1"
)
ACTIVE_RAW_MESSAGE_SOURCE_RUN_BY_LAYER_DATE = {
    ("N4_trigger", "2026-06-12"): V3_20260612_ACTIVE_N4_SOURCE_RUN_ID,
    ("N5_action", "2026-06-12"): V3_20260612_ACTIVE_N5_SOURCE_RUN_ID,
    ("N4_trigger", "2026-06-22"): V3_20260622_ACTIVE_N4_SOURCE_RUN_ID,
}
N3_DISPLAY_INPUT_EVENT_TYPES = ("MarketDisplaySnapshotUpdated",)
N6_INPUT_MESSAGE_DEFAULT_LIMIT = 200
N2_CONDITION_BASIS_DEFAULT_LIMIT = 200
N2_CONDITION_BASIS_EXPORT_FILENAME_PREFIX = "N2条件基础表_最近日期"
N2_CONDITION_BASIS_EXPORT_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
N2_CONDITION_BASIS_ASSET_ORDER = ("index", "board", "stock")
N2_CONDITION_BASIS_ASSET_META = {
    "index": {
        "label": "指数",
        "source_table": "index_condition_basis",
        "id_column": "index_condition_basis_id",
        "identity_column": "index_identity_key",
        "code_column": "code",
        "name_column": "name",
        "exchange_expr": "t.exchange",
        "board_type_expr": "NULL::text",
    },
    "board": {
        "label": "板块",
        "source_table": "board_condition_basis",
        "id_column": "board_condition_basis_id",
        "identity_column": "board_identity_key",
        "code_column": "board_code",
        "name_column": "board_name",
        "exchange_expr": "NULL::text",
        "board_type_expr": "t.board_type",
    },
    "stock": {
        "label": "个股",
        "source_table": "stock_condition_basis",
        "id_column": "stock_condition_basis_id",
        "identity_column": "stock_identity_key",
        "code_column": "code",
        "name_column": "name",
        "exchange_expr": "t.exchange",
        "board_type_expr": "NULL::text",
    },
}
N2_CONDITION_BASIS_EXPORT_COLUMNS = (
    ("asset_label", "资产类型"),
    ("source_table", "来源表"),
    ("condition_basis_id", "条件基础ID"),
    ("run_id", "运行批次"),
    ("source_trade_date", "来源日期"),
    ("for_trade_date", "适用交易日"),
    ("identity_key", "标识"),
    ("code", "代码"),
    ("exchange", "交易所"),
    ("name", "名称"),
    ("board_type", "板块类型"),
    ("direction_scope_text", "方向范围"),
    ("condition_keys_text", "条件键"),
    ("buy_necessary_key", "买入必要条件"),
    ("sell_necessary_key", "卖出必要条件"),
    ("buy_full_necessary_key", "买入FULL条件"),
    ("sell_full_necessary_key", "卖出FULL条件"),
    ("oversold_hint_key", "超跌提示条件"),
    ("overbought_hint_key", "超涨提示条件"),
    ("period_keys.Y", "年周期"),
    ("period_keys.Q", "季周期"),
    ("period_keys.M", "月周期"),
    ("period_keys.W", "周周期"),
    ("period_keys.D", "日周期"),
    ("period_grades.Y", "年分级"),
    ("period_grades.Q", "季分级"),
    ("period_grades.M", "月分级"),
    ("period_grades.W", "周分级"),
    ("period_grades.D", "日分级"),
    ("amount_quality_status", "金额质量"),
    ("buy_target_price", "买入目标价"),
    ("buy_expected_return_pct", "买入预期收益率"),
    ("sell_target_price", "卖出目标价"),
    ("sell_expected_return_pct", "卖出预期收益率"),
    ("up_sell_reference_period", "上涨卖出参考周期"),
    ("down_buy_reference_period", "下跌买入参考周期"),
    ("quality_status", "质量状态"),
    ("quality_reason", "质量原因"),
    ("source_version", "来源版本"),
    ("source_batch_id", "来源批次"),
    ("created_at", "创建时间"),
    ("updated_at", "更新时间"),
    ("row_json_text", "原始行JSON"),
    ("raw_json_text", "原始JSON"),
    ("period_trigger_baseline_json_text", "周期触发基准JSON"),
    ("target_price_trace_json_text", "目标价追溯JSON"),
)


def raw_message_event_date_bounds(event_date: str) -> tuple[datetime, datetime]:
    start = datetime.strptime(event_date, "%Y-%m-%d").replace(tzinfo=DISPLAY_TIMEZONE)
    return start, start + timedelta(days=1)


def raw_message_identity_candidates(keyword: Any) -> list[str]:
    text = str(keyword or "").strip()
    if not text:
        return []
    candidates: list[str] = []

    def add(value: str) -> None:
        if value and value not in candidates:
            candidates.append(value)

    normalized = text.upper()
    if re.fullmatch(r"(STOCK|INDEX|BOARD):[A-Z]+:[0-9A-Z]+", normalized):
        asset, exchange, code = normalized.split(":")
        add(f"{asset.lower()}:{exchange}:{code}")
        return candidates

    if re.fullmatch(r"[0-9]{6}", text):
        add(f"stock:SH:{text}")
        add(f"stock:SZ:{text}")
        add(f"stock:BJ:{text}")
        add(f"index:SH:{text}")
        add(f"index:SZ:{text}")
        if text.startswith("88"):
            add(f"board:TDX:{text}")
    return candidates


def raw_message_keyword_predicate(keyword: Any) -> tuple[str, list[Any]]:
    text = str(keyword or "").strip()
    candidates = raw_message_identity_candidates(text)
    if candidates:
        return (
            """
            partition_key = ANY(%s)
            """,
            [candidates],
        )
    pattern = f"%{text}%"
    return (
        """
        (
            identity_key ILIKE %s
            OR event_id ILIKE %s
            OR source_run_id ILIKE %s
        )
        """,
        [pattern, pattern, pattern],
    )


def n5_actions_read_only_side_effects() -> dict[str, Any]:
    return {
        "writes_database": False,
        "outbox_status_updates": 0,
        "inbox_writes": 0,
        "checkpoint_writes": 0,
        "user_projection_writes": 0,
        "sim_written": False,
        "real_trade_submitted": False,
        "voice_triggered": False,
        "mobile_triggered": False,
    }


def n5_actions_empty_model(*, filters: dict[str, Any], limit: int) -> dict[str, Any]:
    return {
        "ok": True,
        "component": "N5 Actions",
        "title": "N5动作",
        "source_layer": "N5_action",
        "action_run_id": "",
        "source_run_id": "",
        "total_count": 0,
        "filtered_count": 0,
        "returned_count": 0,
        "default_limit": limit,
        "filters": {key: value for key, value in filters.items() if value},
        "filter_inputs": {
            "action_run_id": str(filters.get("action_run_id") or ""),
            "source_run_id": str(filters.get("source_run_id") or ""),
            "event_type": str(filters.get("event_type") or ""),
            "status": str(filters.get("status") or ""),
            "asset_kind": str(filters.get("asset_kind") or ""),
            "action_state": str(filters.get("action_state") or ""),
            "q": str(filters.get("q") or ""),
        },
        "event_types": list(N5_ACTION_DISPLAY_EVENT_TYPES),
        "action_states": ["executed", "eligible"],
        "summary": {"total": 0, "pending": 0, "ActionExecuted": 0, "ActionEligible": 0},
        "items": [],
        "side_effects": n5_actions_read_only_side_effects(),
    }


def n5_action_display_item(row: dict[str, Any]) -> dict[str, Any]:
    payload = row.get("payload_json") or {}
    if not isinstance(payload, dict):
        payload = {}
    trace_json = payload.get("trace_json") or {}
    if not isinstance(trace_json, dict):
        trace_json = {}
    live_window = trace_json.get("live_window_confirmation") or {}
    if not isinstance(live_window, dict):
        live_window = {}
    selected_metric_id = live_window.get("selected_metric_id") or payload.get("selected_metric_id") or ""
    selected_metric_time = (
        live_window.get("executed_metric_time")
        or live_window.get("selected_metric_time")
        or payload.get("executed_metric_time")
        or payload.get("selected_metric_time")
        or ""
    )
    trigger_metric_time = live_window.get("trigger_metric_time") or payload.get("trigger_metric_time") or ""
    live_window_confirmation = live_window.get("live_window_confirmation")
    if live_window_confirmation is None:
        live_window_confirmation = payload.get("live_window_confirmation")
    multi_action_window = live_window.get("multi_action_window")
    if multi_action_window is None:
        multi_action_window = payload.get("multi_action_window")
    return {
        "outbox_id": row.get("outbox_id"),
        "event_id": str(row.get("event_id") or ""),
        "event_time": format_datetime(row.get("event_time")),
        "trade_date": str(row.get("trade_date") or ""),
        "asset_kind": str(row.get("asset_kind") or ""),
        "identity_key": str(row.get("identity_key") or ""),
        "event_type": str(row.get("event_type") or ""),
        "source_layer": str(row.get("source_layer") or ""),
        "source_run_id": str(row.get("source_run_id") or ""),
        "status": str(row.get("status") or ""),
        "action_state": str(payload.get("action_state") or ""),
        "action_type": str(payload.get("action_type") or ""),
        "condition_key": str(payload.get("condition_key") or ""),
        "selected_metric_id": str(selected_metric_id) if selected_metric_id else "",
        "selected_metric_time": str(selected_metric_time) if selected_metric_time else "",
        "trigger_metric_time": str(trigger_metric_time) if trigger_metric_time else "",
        "live_window_confirmation": bool(live_window_confirmation),
        "multi_action_window": bool(multi_action_window),
        "action_key": str(payload.get("action_key") or ""),
        "dedup_key": str(payload.get("dedup_key") or row.get("dedup_key") or ""),
        "payload_json": payload,
        "payload_json_text": json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True, default=str),
    }


STRATEGY_FILTER_OPTIONS = (
    {"key": "chase", "label": "追涨", "profile_key": "enable_chase"},
    {"key": "ultra_short", "label": "超短", "profile_key": "enable_ultra_short"},
    {"key": "short", "label": "短线", "profile_key": "enable_short"},
    {"key": "mid", "label": "中线", "profile_key": "enable_mid"},
    {"key": "long", "label": "长线", "profile_key": "enable_long"},
)
VALID_STRATEGY_FILTER_KEYS = {str(option["key"]) for option in STRATEGY_FILTER_OPTIONS}
APP_FILTER_CENTER_ASSET_BY_PAGE_KEY = {
    "filter-center": "index",
    "filter-center:indexes": "index",
    "filter-center:boards": "board",
    "filter-center:stocks": "stock",
}
APP_FILTER_CENTER_PAGE_KEY_BY_SLUG = {
    "indexes": "filter-center:indexes",
    "boards": "filter-center:boards",
    "stocks": "filter-center:stocks",
}
APP_FILTER_CENTER_PAGE_BY_ASSET = {
    "index": "indexes",
    "board": "boards",
    "stock": "stocks",
}
APP_MONITOR_ASSET_BY_PAGE_KEY = {
    "my-monitor": "stock",
    "my-monitor:stocks": "stock",
    "my-monitor:boards": "board",
    "my-monitor:indexes": "index",
}
APP_MONITOR_PAGE_KEY_BY_SLUG = {
    "stocks": "my-monitor:stocks",
    "boards": "my-monitor:boards",
    "indexes": "my-monitor:indexes",
}
APP_MONITOR_PAGE_BY_ASSET = {
    "stock": "stocks",
    "board": "boards",
    "index": "indexes",
}
APP_V2_MONITOR_TABLE_BY_ASSET = {
    "stock": "user_monitor_stock",
    "board": "user_monitor_board",
    "index": "user_monitor_index",
}
APP_REALTIME_SCOPE_TABLE = "user_realtime_monitor_scope"
APP_V2_FILTER_DISPLAY_TABLE_BY_ASSET = {
    "stock": "v_n6_stock_condition_display_basis",
    "board": "v_n6_board_condition_display_basis",
    "index": "v_n6_index_condition_display_basis",
}
APP_V2_VALID_DIRECTIONS = {"buy", "sell"}
APP_V2_MONITOR_STATUS_FILTERS = {"active", "expired", "all"}
APP_V2_PERIOD_GRADE_FILTER_VALUES = {
    "volume_up",
    "volume_down",
    "low_volume_up",
    "low_volume_down",
    "flat",
}
APP_V2_FILTER_PAGE_MAX_ROWS = 10000
APP_V2_FILTER_DEFAULT_ROWS_BY_ASSET = {
    "index": 200,
    "board": 200,
    "stock": 200,
}
N6_TRADING_SESSION_HISTORY_BLOCKER = "historical_query_disabled_during_trading_session"
N6_TRADING_SESSION_HISTORY_MESSAGE = "实时消息页仅显示当前交易日；历史消息必须使用独立只读归档入口"
N6_TRADING_SESSION_START = (9, 25)
N6_TRADING_SESSION_END = (15, 0)


def app_v2_filter_default_limit(asset_kind: str) -> int:
    return APP_V2_FILTER_DEFAULT_ROWS_BY_ASSET.get(asset_kind, 200)


@dataclass(frozen=True)
class N6UserWebConfig:
    dsn: str = DEFAULT_DSN
    cookie_secure: bool = False
    session_ttl_seconds: int = 8 * 60 * 60
    card_limit: int = 500
    notification_limit: int = 500
    action_event_limit: int = 500
    ui_signal_limit: int = 500
    signal_source_user_id: int = 1
    post_close_fastlane_docs_root: str = DEFAULT_POST_CLOSE_FASTLANE_DOCS_ROOT
    runtime_archive_docs_root: str = DEFAULT_RUNTIME_ARCHIVE_DOCS_ROOT
    runtime_archive_root: str = DEFAULT_RUNTIME_ARCHIVE_ROOT
    rag_docs_root: str = DEFAULT_RAG_DOCS_ROOT
    rag_sql_root: str = DEFAULT_RAG_SQL_ROOT
    replay_docs_root: str = DEFAULT_REPLAY_DOCS_ROOT
    scope_write_enabled: bool = False
    scope_bulk_write_enabled: bool = False
    proposal_write_enabled: bool = False
    csrf_secret_file: str = ""


def build_app_v2_message_dashboard(
    principal: dict[str, Any],
    *,
    user: dict[str, Any],
    rows: list[dict[str, Any]],
    filters: dict[str, Any] | None = None,
    scope_metadata: dict[str, Any],
    limit: int = 100,
) -> dict[str, Any]:
    return app_v2_message_dashboard_model(
        principal,
        user=user,
        rows=rows,
        filters=filters,
        scope_metadata=scope_metadata,
        limit=limit,
    )


def n6_trading_session_now() -> datetime:
    return datetime.now(DISPLAY_TIMEZONE)


def n6_is_trading_session_for_trade_date(current_trade_date: str | None, *, now: datetime | None = None) -> bool:
    trade_date = normalize_filter_value(current_trade_date)
    if not trade_date:
        return False
    current_time = now or n6_trading_session_now()
    if current_time.tzinfo is None:
        current_time = current_time.replace(tzinfo=DISPLAY_TIMEZONE)
    current_time = current_time.astimezone(DISPLAY_TIMEZONE)
    if current_time.strftime("%Y%m%d") != trade_date:
        return False
    start_hour, start_minute = N6_TRADING_SESSION_START
    end_hour, end_minute = N6_TRADING_SESSION_END
    session_start = current_time.replace(hour=start_hour, minute=start_minute, second=0, microsecond=0)
    session_end = current_time.replace(hour=end_hour, minute=end_minute, second=0, microsecond=0)
    return session_start <= current_time <= session_end


def n6_trade_date_access_policy(
    *,
    current_trade_date: str | None,
    requested_trade_date: str | None,
    now: datetime | None = None,
    live_only: bool = False,
) -> dict[str, Any]:
    current = normalize_filter_value(current_trade_date)
    requested = normalize_filter_value(requested_trade_date)
    effective = requested or current or ""
    blocked = bool(
        current
        and requested
        and requested != current
        and (live_only or n6_is_trading_session_for_trade_date(current, now=now))
    )
    if blocked:
        effective = current
    return {
        "blocked": blocked,
        "blocker": N6_TRADING_SESSION_HISTORY_BLOCKER if blocked else "",
        "message": N6_TRADING_SESSION_HISTORY_MESSAGE if blocked else "",
        "current_trade_date": current or "",
        "requested_trade_date": requested or "",
        "effective_trade_date": effective,
        "historical_query_allowed": not blocked,
    }


def n6_trading_session_blocker_response(policy: dict[str, Any]) -> JSONResponse:
    return JSONResponse(
        {
            "ok": False,
            "error": N6_TRADING_SESSION_HISTORY_BLOCKER,
            "blockers": [N6_TRADING_SESSION_HISTORY_BLOCKER],
            "message": N6_TRADING_SESSION_HISTORY_MESSAGE,
            "current_trade_date": policy.get("current_trade_date") or "",
            "requested_trade_date": policy.get("requested_trade_date") or "",
            "effective_trade_date": policy.get("effective_trade_date") or "",
        },
        status_code=409,
    )


def n6_json_response(content: Any, *, status_code: int = 200) -> JSONResponse:
    return JSONResponse(jsonable_encoder(content), status_code=status_code)


def build_app_v2_message_groups(
    principal: dict[str, Any],
    *,
    user: dict[str, Any],
    rows: list[dict[str, Any]],
    scope_metadata: dict[str, Any],
) -> dict[str, Any]:
    return app_v2_message_groups_model(
        principal,
        user=user,
        rows=rows,
        scope_metadata=scope_metadata,
    )


def build_app_v2_projection_status(
    principal: dict[str, Any],
    *,
    user: dict[str, Any],
    rows: list[dict[str, Any]],
    scope_metadata: dict[str, Any],
) -> dict[str, Any]:
    return app_v2_message_projection_status_model(
        principal,
        user=user,
        rows=rows,
        scope_metadata=scope_metadata,
    )


@dataclass(frozen=True)
class UserAccount:
    user_id: int
    login_name: str
    display_name: str | None
    role: str
    status: str
    password_hash: str
    password_hash_algo: str


@dataclass(frozen=True)
class AuthSession:
    user_session_id: int
    user_id: int
    login_name: str
    display_name: str | None
    role: str
    status: str
    session_token_hash: str
    expires_at: datetime
    revoked_at: datetime | None


class N6UserRepository(Protocol):
    def fetch_user_for_login(self, login_name: str) -> UserAccount | None:
        ...

    def create_session(
        self,
        *,
        user_id: int,
        session_token_hash: str,
        session_token_hash_algo: str,
        expires_at: datetime,
        client_info: dict[str, Any],
    ) -> AuthSession:
        ...

    def fetch_session(self, session_token_hash: str) -> AuthSession | None:
        ...

    def revoke_session(self, session_token_hash: str, revoked_at: datetime) -> bool:
        ...

    def fetch_cards(self, user_id: int, limit: int) -> list[dict[str, Any]]:
        ...

    def fetch_notifications(self, user_id: int, limit: int) -> list[dict[str, Any]]:
        ...

    def fetch_n5_action_events(self, limit: int) -> list[dict[str, Any]]:
        ...

    def fetch_message_dashboard(self, limit: int) -> dict[str, Any]:
        ...

    def fetch_top_index_strategy(self) -> dict[str, Any] | None:
        ...

    def fetch_strong_boards(self, limit: int) -> list[dict[str, Any]]:
        ...

    def fetch_filter_profile(self, user_id: int) -> dict[str, Any] | None:
        ...

    def fetch_users_for_admin(self) -> list[dict[str, Any]]:
        ...

    def create_user_with_defaults(
        self,
        *,
        login_name: str,
        display_name: str | None,
        role: str,
        password_hash: str,
        password_hash_algo: str,
        created_by_user_id: int,
    ) -> dict[str, Any]:
        ...

    def delete_user(self, *, target_user_id: int, deleted_by_user_id: int) -> dict[str, Any]:
        ...

    def fetch_sim_account(self, user_id: int) -> dict[str, Any] | None:
        ...

    def fetch_sim_positions(self, user_id: int) -> list[dict[str, Any]]:
        ...

    def fetch_ui_v1_signals(
        self,
        user_id: int,
        filters: dict[str, Any],
        limit: int,
        offset: int = 0,
    ) -> list[dict[str, Any]]:
        ...

    def count_ui_v1_signals(self, user_id: int, filters: dict[str, Any]) -> int:
        ...

    def fetch_ui_v1_signal_statistics(self, user_id: int) -> dict[str, Any]:
        ...

    def fetch_ui_v1_lineage_stats(self, user_id: int) -> dict[str, Any]:
        ...

    def fetch_ui_v1_n3_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        ...

    def fetch_ui_v1_n4_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        ...

    def fetch_ui_v1_n5_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        ...

    def fetch_ui_v1_n5_actions(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
    ) -> dict[str, Any]:
        ...

    def fetch_index_board_c1_minute_rows(self, trade_date: str) -> dict[str, list[dict[str, Any]]]:
        ...

    def fetch_ui_v1_input_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        ...

    def fetch_ui_v1_n2_condition_basis(
        self,
        *,
        asset_kind: str,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        ...

    def fetch_ui_v1_n2_condition_basis_latest_export(self) -> dict[str, Any]:
        ...

    def fetch_ui_v1_status_monitor(
        self,
        user_id: int,
        filters: dict[str, Any],
        limit: int,
        offset: int = 0,
    ) -> dict[str, Any]:
        ...

    def fetch_ui_v1_signal_detail(self, user_id: int, user_signal_projection_id: int) -> dict[str, Any] | None:
        ...

    def fetch_ui_v1_dashboard_metrics(self, user_id: int) -> dict[str, Any]:
        ...

    def fetch_ui_v1_virtual_account(self, user_id: int) -> dict[str, Any] | None:
        ...

    def fetch_ui_v1_cash_snapshot(self, user_id: int) -> dict[str, Any] | None:
        ...

    def fetch_ui_v1_cash_ledger(self, user_id: int, limit: int) -> list[dict[str, Any]]:
        ...

    def fetch_ui_v1_artifacts(self) -> dict[str, Any]:
        ...

    def fetch_ui_v1_rollback_summary(self) -> dict[str, Any]:
        ...

    def fetch_app_principals(self, user_id: int) -> list[dict[str, Any]]:
        ...

    def fetch_app_virtual_account(self, principal_id: int, principal_type: str) -> dict[str, Any] | None:
        ...

    def fetch_app_cash_snapshot(self, virtual_account_id: int) -> dict[str, Any] | None:
        ...

    def fetch_app_positions(self, principal_id: int, principal_type: str) -> list[dict[str, Any]]:
        ...

    def fetch_app_signals(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        filters: dict[str, Any],
        limit: int,
    ) -> list[dict[str, Any]]:
        ...

    def fetch_app_signal_events(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        filters: dict[str, Any],
        after_id: int,
        limit: int,
    ) -> list[dict[str, Any]]:
        ...

    def fetch_app_signal_detail(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        user_signal_projection_id: int,
        trade_date: str | None = None,
    ) -> dict[str, Any] | None:
        ...

    def fetch_app_current_signal_trade_date(self) -> str | None:
        ...

    def fetch_app_signal_scope_metadata(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        trade_date: str | None = None,
    ) -> dict[str, Any]:
        ...

    def fetch_app_realtime_scope(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
    ) -> dict[str, Any]:
        ...

    def fetch_app_filter_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        filters: dict[str, Any],
        limit: int,
    ) -> dict[str, Any]:
        ...

    def fetch_app_current_filter_identity(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
    ) -> dict[str, Any]:
        ...

    def fetch_app_filter_members(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        membership_kind: str,
        parent_identity_key: str,
        limit: int,
    ) -> dict[str, Any]:
        ...

    def fetch_app_filter_linked_stocks(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        membership_kind: str,
        parent_identity_key: str,
        limit: int,
        view: str = "matched",
    ) -> dict[str, Any]:
        ...

    def fetch_app_membership_stocks(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        entity_type: str,
        identity_key: str,
        limit: int,
    ) -> dict[str, Any]:
        ...

    def fetch_app_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str | None = None,
        limit: int = 500,
        monitor_status: str = "active",
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        ...

    def add_app_monitor_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
        direction: str,
        source: str = "single_row",
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        ...

    def add_app_monitor_directions(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
        directions: tuple[str, ...],
        source: str = "single_row",
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        ...

    def bulk_add_app_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        direction: str,
        filters: dict[str, Any],
    ) -> dict[str, Any]:
        ...

    def selected_add_app_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        direction: str,
        identity_keys: list[str],
    ) -> dict[str, Any]:
        ...

    def add_app_linked_stock_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        parent_asset_kind: str,
        parent_identity_key: str,
        mode: str,
        stock_identity_keys: list[str],
        direction: str,
    ) -> dict[str, Any]:
        ...

    def remove_app_monitor_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        monitor_id: int,
    ) -> dict[str, Any]:
        ...

    def add_app_realtime_scope_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
        for_trade_date: str = "",
        source: str = "single_row",
    ) -> dict[str, Any]:
        ...

    def selected_add_app_realtime_scope_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_keys: list[str],
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        ...

    def bulk_add_app_realtime_scope_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        filters: dict[str, Any],
    ) -> dict[str, Any]:
        ...

    def remove_app_realtime_scope_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        realtime_scope_id: int,
    ) -> dict[str, Any]:
        ...

    def fetch_app_trade_proposals(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        limit: int = 100,
    ) -> dict[str, Any]:
        ...

    def create_app_trade_proposal(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        source_type: str,
        source_id: str,
    ) -> dict[str, Any]:
        ...

    def confirm_app_trade_proposal(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        proposal_id: int,
        idempotency_key: str,
    ) -> dict[str, Any]:
        ...

    def fetch_app_virtual_trades(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        limit: int = 200,
    ) -> dict[str, Any]:
        ...


PasswordVerifier = Callable[[str, str, str], bool]
PasswordHasher = Callable[[str], HashResult]


class UserManagementError(ValueError):
    def __init__(self, code: str) -> None:
        super().__init__(code)
        self.code = code


class PostgresN6UserRepository:
    def __init__(self, dsn: str) -> None:
        self.dsn = dsn
        self._app_v2_monitor_column_cache: dict[str, set[str]] = {}
        self._app_v2_filter_column_cache: dict[str, set[str]] = {}
        self._app_v2_relation_existence_cache: dict[str, bool] = {}
        self._app_v1_signal_schema_capability_cache: dict[str, frozenset[str]] | None = None
        self._app_v1_signal_schema_capability_lock = threading.Lock()
        self._app_user_scope_cache: dict[tuple[int, str, int, str], tuple[float, dict[str, Any]]] = {}
        self._app_user_scope_cache_lock = threading.Lock()
        self._app_filter_result_cache: dict[tuple[Any, ...], dict[str, Any]] = {}
        self._app_membership_result_cache: dict[tuple[Any, ...], dict[str, Any]] = {}
        self._app_shared_cache_lock = threading.Lock()

    def _app_shared_cache_get(
        self,
        cache: dict[tuple[Any, ...], dict[str, Any]],
        key: tuple[Any, ...],
    ) -> dict[str, Any] | None:
        with self._app_shared_cache_lock:
            value = cache.get(key)
            return copy.deepcopy(value) if value is not None else None

    def _app_shared_cache_set(
        self,
        cache: dict[tuple[Any, ...], dict[str, Any]],
        key: tuple[Any, ...],
        value: dict[str, Any],
    ) -> None:
        with self._app_shared_cache_lock:
            cache[key] = copy.deepcopy(value)
            while len(cache) > 256:
                cache.pop(next(iter(cache)))

    def invalidate_app_user_scope_cache(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
    ) -> None:
        with self._app_user_scope_cache_lock:
            for key in list(self._app_user_scope_cache):
                if key[:3] == (int(principal_id), str(principal_type), int(user_id)):
                    self._app_user_scope_cache.pop(key, None)

    def fetch_user_for_login(self, login_name: str) -> UserAccount | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT user_id, login_name, display_name, role, status,
                       password_hash, password_hash_algo
                FROM user_account
                WHERE login_name = %s
                """,
                (login_name,),
            )
            row = cur.fetchone()
        return UserAccount(**dict(row)) if row else None

    def create_session(
        self,
        *,
        user_id: int,
        session_token_hash: str,
        session_token_hash_algo: str,
        expires_at: datetime,
        client_info: dict[str, Any],
    ) -> AuthSession:
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO user_session (
                      user_id,
                      session_token_hash,
                      session_token_hash_algo,
                      expires_at,
                      client_info_json
                    )
                    VALUES (%s, %s, %s, %s, %s::jsonb)
                    RETURNING user_session_id, session_token_hash, expires_at, revoked_at
                    """,
                    (
                        user_id,
                        session_token_hash,
                        session_token_hash_algo,
                        expires_at,
                        json.dumps(client_info, ensure_ascii=True),
                    ),
                )
                session_row = dict(cur.fetchone())
                cur.execute(
                    """
                    SELECT user_id, login_name, display_name, role, status
                    FROM user_account
                    WHERE user_id = %s
                    """,
                    (user_id,),
                )
                user_row = dict(cur.fetchone())
        return AuthSession(**user_row, **session_row)

    def fetch_session(self, session_token_hash: str) -> AuthSession | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT s.user_session_id,
                       s.user_id,
                       u.login_name,
                       u.display_name,
                       u.role,
                       u.status,
                       s.session_token_hash,
                       s.expires_at,
                       s.revoked_at
                FROM user_session s
                JOIN user_account u ON u.user_id = s.user_id
                WHERE s.session_token_hash = %s
                """,
                (session_token_hash,),
            )
            row = cur.fetchone()
        return AuthSession(**dict(row)) if row else None

    def revoke_session(self, session_token_hash: str, revoked_at: datetime) -> bool:
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE user_session
                    SET revoked_at = %s,
                        updated_at = now()
                    WHERE session_token_hash = %s
                      AND revoked_at IS NULL
                    RETURNING user_session_id
                    """,
                    (revoked_at, session_token_hash),
                )
                return cur.fetchone() is not None

    def fetch_cards(self, user_id: int, limit: int) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT c.user_signal_card_id,
                       c.card_type,
                       c.card_status,
                       c.display_priority,
                       c.title,
                       c.summary,
                       c.asset_kind,
                       c.identity_key,
                       c.code,
                       c.name,
                       c.direction,
                       c.signal_type,
                       c.target_price,
                       c.current_price,
                       c.expected_return_pct,
                       c.board_code,
                       c.board_name,
                       c.source_action_run_id,
                       c.source_event_id,
                       c.created_at,
                       s.period_transition_y,
                       s.period_transition_q,
                       s.period_transition_m,
                       s.period_transition_w,
                       s.period_transition_d
                FROM user_signal_card c
                LEFT JOIN user_signal_projection p
                  ON p.user_signal_projection_id = c.user_signal_projection_id
                 AND p.user_id = c.user_id
                LEFT JOIN stock_condition_display_basis s
                  ON c.asset_kind = 'stock'
                 AND p.source_display_table = 'stock_condition_display_basis'
                 AND p.source_condition_display_basis_id = s.stock_condition_display_basis_id
                WHERE c.user_id = %s
                ORDER BY c.display_priority ASC, c.created_at DESC, c.user_signal_card_id ASC
                LIMIT %s
                """,
                (user_id, limit),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_app_signal_events(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        filters: dict[str, Any],
        after_id: int,
        limit: int,
    ) -> list[dict[str, Any]]:
        if not self._app_v1_signal_scope_relations_ready():
            return []
        where_sql, params = self._app_v1_signal_where(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
            filters=filters,
            include_monitor_scope=False,
        )
        params["after_id"] = int(after_id)
        params["limit"] = min(max(int(limit), 1), N6_SIGNAL_SSE_BATCH_LIMIT)
        scope_cte_sql = self._app_v1_web_signal_scope_cte(
            include_expired=False,
            include_realtime_scope=True,
        )
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                WITH {scope_cte_sql}
                SELECT {self._app_v1_signal_sse_select_list()}
                FROM user_signal_projection p
                JOIN user_projection_run r
                  ON r.user_projection_run_id = p.user_projection_run_id
                 AND r.status IN ('passed', 'ready')
                LEFT JOIN user_signal_card c
                  ON c.user_signal_projection_id = p.user_signal_projection_id
                 AND c.user_projection_run_id = p.user_projection_run_id
                 AND c.user_id = p.user_id
                {self._app_v1_web_signal_scope_join()}
                WHERE p.user_id = %(user_id)s
                  AND p.user_signal_projection_id > %(after_id)s
                  AND {where_sql}
                ORDER BY p.user_signal_projection_id ASC
                LIMIT %(limit)s
                """,
                params,
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_notifications(self, user_id: int, limit: int) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT user_notification_queue_id,
                       user_projection_run_id,
                       notification_source,
                       queue_status,
                       channel,
                       title,
                       message,
                       priority,
                       source_event_id,
                       source_action_run_id,
                       asset_kind,
                       identity_key,
                       queued_at
                FROM user_notification_queue
                WHERE user_id = %s
                ORDER BY priority ASC, queued_at DESC, user_notification_queue_id ASC
                LIMIT %s
                """,
                (user_id, limit),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_n5_action_events(self, limit: int) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT outbox_id,
                       event_id,
                       event_type,
                       event_schema_version,
                       trade_date,
                       asset_kind,
                       identity_key,
                       event_time,
                       source_run_id,
                       dedup_key,
                       partition_key,
                       payload_json,
                       payload_json->>'direction' AS direction,
                       payload_json->>'signal_type' AS signal_type,
                       payload_json->>'action_state' AS action_state,
                       payload_json->>'action_mark' AS action_mark,
                       payload_json->>'condition_key' AS condition_key,
                       payload_json->>'original_condition_key' AS original_condition_key,
                       status,
                       created_at
                FROM common_event_outbox
                WHERE source_layer = 'N5_action'
                  AND event_type IN (
                    'ActionEligible',
                    'ActionBlocked',
                    'ActionExecuted',
                    'ActionSkipped',
                    'ActionEvent',
                    'HintEvent'
                  )
                ORDER BY created_at DESC, outbox_id DESC
                LIMIT %s
                """,
                (limit,),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_message_dashboard(self, limit: int) -> dict[str, Any]:
        event_types = (
            "TriggerMatched",
            "ActionBlocked",
            "ActionExecuted",
            "ActionEligible",
            "ActionSkipped",
        )
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                  count(*) FILTER (
                    WHERE source_layer = 'N4_trigger'
                      AND event_type = 'TriggerMatched'
                      AND status = 'pending'
                      AND (created_at AT TIME ZONE 'Asia/Shanghai')::date =
                          (now() AT TIME ZONE 'Asia/Shanghai')::date
                  )::int AS today_n4_trigger_matched_pending,
                  count(*) FILTER (
                    WHERE source_layer = 'N5_action'
                      AND event_type = 'ActionBlocked'
                      AND (created_at AT TIME ZONE 'Asia/Shanghai')::date =
                          (now() AT TIME ZONE 'Asia/Shanghai')::date
                  )::int AS today_n5_action_blocked,
                  count(*) FILTER (
                    WHERE source_layer = 'N5_action'
                      AND event_type = 'ActionExecuted'
                      AND (created_at AT TIME ZONE 'Asia/Shanghai')::date =
                          (now() AT TIME ZONE 'Asia/Shanghai')::date
                  )::int AS today_n5_action_executed,
                  count(*) FILTER (
                    WHERE source_layer = 'N5_action'
                      AND status = 'pending'
                      AND event_type IN ('ActionBlocked', 'ActionExecuted', 'ActionEligible', 'ActionSkipped')
                  )::int AS n5_outbox_pending
                FROM common_event_outbox
                WHERE event_type = ANY(%s)
                """,
                (list(event_types),),
            )
            admin_dashboard = dict(cur.fetchone() or {})

            cur.execute(
                """
                SELECT event_type, status, count(*)::int AS count
                FROM common_event_outbox
                WHERE event_type = ANY(%s)
                GROUP BY event_type, status
                ORDER BY event_type, status
                """,
                (list(event_types),),
            )
            event_distribution = [dict(row) for row in cur.fetchall()]

            cur.execute(
                """
                SELECT
                  count(DISTINCT p.user_projection_run_id)::int AS projection_run_count,
                  count(DISTINCT p.user_signal_projection_id)::int AS signal_projection_count,
                  count(DISTINCT c.user_signal_card_id)::int AS signal_card_count,
                  count(DISTINCT q.user_notification_queue_id)::int AS notification_queue_count,
                  count(DISTINCT q.user_notification_queue_id)
                    FILTER (WHERE q.queue_status = 'queued_only')::int AS queued_only,
                  count(DISTINCT q.user_notification_queue_id)
                    FILTER (WHERE q.queue_status = 'ready_for_future_push')::int AS ready_for_future_push,
                  (
                    SELECT r.user_projection_run_id
                    FROM user_projection_run r
                    ORDER BY r.created_at DESC NULLS LAST, r.finished_at DESC NULLS LAST
                    LIMIT 1
                  ) AS latest_projection_run_id
                FROM user_signal_projection p
                LEFT JOIN user_signal_card c
                  ON c.user_signal_projection_id = p.user_signal_projection_id
                 AND c.user_id = p.user_id
                LEFT JOIN user_notification_queue q
                  ON q.user_signal_projection_id = p.user_signal_projection_id
                 AND q.user_id = p.user_id
                """
            )
            n6_shadow = dict(cur.fetchone() or {})

            cur.execute(
                """
                SELECT COALESCE(
                         c.card_payload_json->>'blocked_reason',
                         p.display_payload_json->>'blocked_reason',
                         p.trace_json->>'blocked_reason',
                         p.source_payload_json->>'blocked_reason',
                         'unknown'
                       ) AS blocked_reason,
                       count(*)::int AS count
                FROM user_signal_projection p
                LEFT JOIN user_signal_card c
                  ON c.user_signal_projection_id = p.user_signal_projection_id
                 AND c.user_id = p.user_id
                WHERE COALESCE(
                        NULLIF(p.action_state, ''),
                        NULLIF(c.action_state, ''),
                        CASE
                          WHEN p.source_action_event_type = 'ActionBlocked' THEN 'blocked'
                          WHEN p.source_action_event_type = 'ActionExecuted' THEN 'executed'
                          WHEN p.source_action_event_type = 'ActionEligible' THEN 'eligible'
                          WHEN p.source_action_event_type = 'ActionSkipped' THEN 'skipped'
                          ELSE NULL
                        END,
                        p.projection_status,
                        c.card_status
                      ) = 'blocked'
                GROUP BY blocked_reason
                ORDER BY count DESC, blocked_reason ASC
                LIMIT 20
                """
            )
            blocked_reasons = [dict(row) for row in cur.fetchall()]

            cur.execute(
                """
                SELECT layer, run_id, status, created_at, finished_at
                FROM (
                  SELECT 'N4' AS layer, run_id, status, created_at, finished_at
                  FROM common_trigger_run
                  UNION ALL
                  SELECT 'N5' AS layer, run_id, status, created_at, finished_at
                  FROM common_action_run
                  UNION ALL
                  SELECT 'N6' AS layer, user_projection_run_id AS run_id, status, created_at, finished_at
                  FROM user_projection_run
                ) runs
                ORDER BY created_at DESC NULLS LAST, finished_at DESC NULLS LAST
                LIMIT 10
                """
            )
            recent_runs = [dict(row) for row in cur.fetchall()]

            cur.execute(
                """
                SELECT outbox_id,
                       event_id,
                       event_type,
                       event_schema_version,
                       trade_date,
                       asset_kind,
                       identity_key,
                       event_time,
                       source_layer,
                       source_run_id,
                       dedup_key,
                       partition_key,
                       payload_json,
                       payload_json->>'direction' AS direction,
                       payload_json->>'signal_type' AS signal_type,
                       payload_json->>'action_state' AS action_state,
                       payload_json->>'action_mark' AS action_mark,
                       payload_json->>'condition_key' AS condition_key,
                       payload_json->>'original_condition_key' AS original_condition_key,
                       status,
                       created_at
                FROM common_event_outbox
                WHERE event_type = ANY(%s)
                ORDER BY created_at DESC, outbox_id DESC
                LIMIT %s
                """,
                (list(event_types), max(1, min(int(limit), 1000))),
            )
            messages = [dict(row) for row in cur.fetchall()]

        return {
            "admin_dashboard": {
                **admin_dashboard,
                "n6_shadow_projection_count": int(n6_shadow.get("signal_projection_count") or 0),
                "n6_shadow_card_count": int(n6_shadow.get("signal_card_count") or 0),
                "n6_shadow_queue_count": int(n6_shadow.get("notification_queue_count") or 0),
                "n6_queued_only": int(n6_shadow.get("queued_only") or 0),
                "n6_ready_for_future_push": int(n6_shadow.get("ready_for_future_push") or 0),
                "latest_projection_run_id": n6_shadow.get("latest_projection_run_id") or "—",
            },
            "message_dashboard": {
                "event_distribution": event_distribution,
                "total_messages": sum(int(row.get("count") or 0) for row in event_distribution),
            },
            "blocked_reason_distribution": blocked_reasons,
            "recent_runs": recent_runs,
            "messages": messages,
        }

    def fetch_top_index_strategy(self) -> dict[str, Any] | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT code,
                       name,
                       display_title,
                       display_summary,
                       selected_signal_types,
                       selected_condition_keys,
                       period_transition_y,
                       period_transition_q,
                       period_transition_m,
                       period_transition_w,
                       period_transition_d,
                       buy_target_price,
                       sell_target_price
                FROM v_n6_index_condition_display_basis
                WHERE display_status = 'visible'
                  AND quality_status IN ('passed', 'warning', 'pending')
                ORDER BY
                  CASE
                    WHEN period_transition_y = 'volume_up'
                     AND period_transition_q = 'volume_up'
                     AND period_transition_m = 'volume_up'
                     AND period_transition_w = 'volume_up'
                     AND period_transition_d = 'volume_up' THEN 60
                    WHEN period_transition_y = 'volume_up'
                     AND period_transition_q = 'volume_up'
                     AND period_transition_m = 'volume_up'
                     AND period_transition_w = 'volume_up'
                     AND COALESCE(period_transition_d, '') <> 'volume_up' THEN 50
                    WHEN period_transition_y = 'volume_up'
                     AND period_transition_q = 'volume_up'
                     AND period_transition_m = 'volume_up'
                     AND COALESCE(period_transition_w, '') <> 'volume_up'
                     AND COALESCE(period_transition_d, '') <> 'volume_up' THEN 40
                    WHEN period_transition_y = 'volume_up'
                     AND period_transition_q = 'volume_up'
                     AND COALESCE(period_transition_m, '') <> 'volume_up'
                     AND COALESCE(period_transition_w, '') <> 'volume_up'
                     AND COALESCE(period_transition_d, '') <> 'volume_up' THEN 30
                    WHEN period_transition_y = 'volume_up'
                     AND COALESCE(period_transition_q, '') <> 'volume_up'
                     AND COALESCE(period_transition_m, '') <> 'volume_up'
                     AND COALESCE(period_transition_w, '') <> 'volume_up'
                     AND COALESCE(period_transition_d, '') <> 'volume_up' THEN 20
                    WHEN COALESCE(period_transition_y, '') <> 'volume_up'
                     AND COALESCE(period_transition_q, '') <> 'volume_up'
                     AND COALESCE(period_transition_m, '') <> 'volume_up'
                     AND COALESCE(period_transition_w, '') <> 'volume_up'
                     AND COALESCE(period_transition_d, '') <> 'volume_up' THEN 10
                    ELSE 0
                  END DESC,
                  updated_at DESC
                LIMIT 1
                """
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_strong_boards(self, limit: int) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT board_code,
                       board_name,
                       display_title,
                       display_summary,
                       selected_signal_types,
                       period_transition_y,
                       period_transition_q,
                       period_transition_m,
                       period_transition_w,
                       period_transition_d,
                       buy_target_price,
                       sell_target_price
                FROM v_n6_board_condition_display_basis
                WHERE display_status = 'visible'
                  AND quality_status IN ('passed', 'warning', 'pending')
                  AND board_code LIKE '881%%'
                  AND period_transition_y = 'volume_up'
                  AND period_transition_q = 'volume_up'
                  AND period_transition_m = 'volume_up'
                ORDER BY updated_at DESC, board_code ASC
                LIMIT %s
                """,
                (limit,),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_filter_profile(self, user_id: int) -> dict[str, Any] | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT enable_chase,
                       enable_ultra_short,
                       enable_short,
                       enable_mid,
                       enable_long
                FROM user_filter_profile
                WHERE user_id = %s
                  AND is_default = true
                  AND status = 'active'
                ORDER BY user_filter_profile_id ASC
                LIMIT 1
                """,
                (user_id,),
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_users_for_admin(self) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT u.user_id,
                       u.login_name,
                       u.display_name,
                       u.role,
                       u.status,
                       u.last_login_at,
                       u.created_at,
                       count(DISTINCT fp.user_filter_profile_id)::int AS filter_profile_count,
                       count(DISTINCT sa.user_sim_account_id)::int AS sim_account_count,
                       count(DISTINCT sp.user_sim_position_id)::int AS active_position_count
                FROM user_account u
                LEFT JOIN user_filter_profile fp
                  ON fp.user_id = u.user_id
                 AND fp.status = 'active'
                LEFT JOIN user_sim_account sa
                  ON sa.user_id = u.user_id
                 AND sa.account_status = 'active'
                LEFT JOIN user_sim_position sp
                  ON sp.user_id = u.user_id
                 AND sp.total_qty > 0
                GROUP BY u.user_id, u.login_name, u.display_name, u.role, u.status, u.last_login_at, u.created_at
                ORDER BY u.user_id ASC
                LIMIT 100
                """
            )
            return [dict(row) for row in cur.fetchall()]

    def create_user_with_defaults(
        self,
        *,
        login_name: str,
        display_name: str | None,
        role: str,
        password_hash: str,
        password_hash_algo: str,
        created_by_user_id: int,
    ) -> dict[str, Any]:
        try:
            with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
                with conn.transaction(), conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO user_account (
                          login_name,
                          display_name,
                          password_hash,
                          password_hash_algo,
                          role,
                          status,
                          created_by_user_id,
                          user_policy_json
                        )
                        VALUES (%s, %s, %s, %s, %s, 'active', %s, %s)
                        RETURNING user_id, login_name, display_name, role, status, created_at, last_login_at
                        """,
                        (
                            login_name,
                            display_name,
                            password_hash,
                            password_hash_algo,
                            role,
                            created_by_user_id,
                            Jsonb({"n6_web_created": True}),
                        ),
                    )
                    user_row = dict(cur.fetchone())
                    user_id = int(user_row["user_id"])
                    cur.execute(
                        """
                        INSERT INTO n6_principal (
                          principal_type,
                          owner_user_id,
                          principal_status,
                          principal_label,
                          principal_policy_json
                        )
                        VALUES (%s, %s, 'active', %s, %s)
                        RETURNING principal_id
                        """,
                        (
                            "admin" if role == "admin" else "human_user",
                            user_id,
                            display_name or login_name,
                            Jsonb(
                                {
                                    "source": "n6_web_user_create",
                                    "contract_version": "n6-web-user-principal-v1",
                                    "created_by_user_id": created_by_user_id,
                                }
                            ),
                        ),
                    )
                    principal_id = int(cur.fetchone()["principal_id"])
                    cur.execute(
                        """
                        INSERT INTO user_filter_profile (
                          user_id,
                          profile_name,
                          is_default,
                          enable_chase,
                          enable_ultra_short,
                          enable_short,
                          enable_mid,
                          enable_long,
                          permission_scope,
                          status
                        )
                        VALUES (%s, %s, true, true, true, true, true, true, 'self', 'active')
                        """,
                        (user_id, DEFAULT_PROFILE_NAME),
                    )
                    cur.execute(
                        """
                        INSERT INTO user_sim_account (
                          user_id,
                          account_name,
                          initial_cash,
                          cash_balance,
                          frozen_cash,
                          settlement_mode,
                          account_status,
                          sim_policy_json
                        )
                        VALUES (%s, %s, %s, %s, 0, 'T_PLUS_1', 'active', %s)
                        """,
                        (
                            user_id,
                            DEFAULT_SIM_ACCOUNT_NAME,
                            DEFAULT_INITIAL_CASH,
                            DEFAULT_INITIAL_CASH,
                            Jsonb({"shadow_only": True, "real_trade_submitted": False}),
                        ),
                    )
                    provisioning_status = "not_applicable"
                    if role == "user":
                        try:
                            cur.execute(
                                "SELECT public.n6_provision_human_virtual_account(%s) AS result",
                                (principal_id,),
                            )
                        except psycopg.Error as exc:
                            raise UserManagementError("virtual_account_provisioning_failed") from exc
                        provisioning_result = dict(cur.fetchone() or {}).get("result")
                        if (
                            not isinstance(provisioning_result, dict)
                            or provisioning_result.get("ok") is not True
                            or provisioning_result.get("status") != "created"
                        ):
                            raise UserManagementError("virtual_account_provisioning_failed")
                        provisioning_status = "created"
        except psycopg.errors.UniqueViolation as exc:
            raise UserManagementError("login_name_exists") from exc
        except UserManagementError:
            raise
        user_row["filter_profile_count"] = 1
        user_row["sim_account_count"] = 1
        user_row["active_position_count"] = 0
        user_row["virtual_account_provisioning_status"] = provisioning_status
        return user_row

    def delete_user(self, *, target_user_id: int, deleted_by_user_id: int) -> dict[str, Any]:
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE user_account
                    SET status = 'deleted',
                        deleted_by_user_id = %s,
                        deleted_at = now(),
                        updated_at = now()
                    WHERE user_id = %s
                      AND status <> 'deleted'
                    RETURNING user_id, login_name, display_name, role, status, created_at, last_login_at
                    """,
                    (deleted_by_user_id, target_user_id),
                )
                row = cur.fetchone()
                if row is None:
                    raise UserManagementError("user_not_found")
                cur.execute(
                    """
                    UPDATE user_session
                    SET revoked_at = COALESCE(revoked_at, now()),
                        updated_at = now()
                    WHERE user_id = %s
                      AND revoked_at IS NULL
                    """,
                    (target_user_id,),
                )
                cur.execute(
                    """
                    UPDATE user_filter_profile
                    SET status = 'deleted',
                        updated_at = now()
                    WHERE user_id = %s
                      AND status <> 'deleted'
                    """,
                    (target_user_id,),
                )
                cur.execute(
                    """
                    UPDATE user_sim_account
                    SET account_status = 'deleted',
                        updated_at = now()
                    WHERE user_id = %s
                      AND account_status <> 'deleted'
                    """,
                    (target_user_id,),
                )
        result = dict(row)
        result["filter_profile_count"] = 0
        result["sim_account_count"] = 0
        result["active_position_count"] = 0
        return result

    def fetch_sim_account(self, user_id: int) -> dict[str, Any] | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT user_sim_account_id,
                       account_name,
                       initial_cash,
                       cash_balance,
                       frozen_cash,
                       settlement_mode,
                       account_status
                FROM user_sim_account
                WHERE user_id = %s
                  AND account_status = 'active'
                ORDER BY user_sim_account_id ASC
                LIMIT 1
                """,
                (user_id,),
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_sim_positions(self, user_id: int) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT code,
                       name,
                       total_qty,
                       available_qty,
                       t_plus_one_locked_qty,
                       avg_cost,
                       last_price,
                       market_value,
                       unrealized_pnl
                FROM user_sim_position
                WHERE user_id = %s
                  AND total_qty > 0
                ORDER BY updated_at DESC, user_sim_position_id ASC
                LIMIT 200
                """,
                (user_id,),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_ui_v1_signals(
        self,
        user_id: int,
        filters: dict[str, Any],
        limit: int,
        offset: int = 0,
    ) -> list[dict[str, Any]]:
        where_sql, params = self._ui_v1_signal_where(user_id, filters)
        params["limit"] = max(1, min(int(limit), 500))
        params["offset"] = max(0, int(offset))
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT {self._ui_v1_signal_select_list()}
                {self._ui_v1_signal_from_sql()}
                WHERE {where_sql}
                ORDER BY p.created_at DESC, p.user_signal_projection_id DESC
                LIMIT %(limit)s
                OFFSET %(offset)s
                """,
                params,
            )
            return [dict(row) for row in cur.fetchall()]

    def count_ui_v1_signals(self, user_id: int, filters: dict[str, Any]) -> int:
        where_sql, params = self._ui_v1_signal_where(user_id, filters)
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT count(*)::int AS count
                {self._ui_v1_signal_from_sql()}
                WHERE {where_sql}
                """,
                params,
            )
            row = cur.fetchone() or {}
        return int(row.get("count") or 0)

    def fetch_ui_v1_signal_statistics(self, user_id: int) -> dict[str, Any]:
        where_sql, params = self._ui_v1_signal_where(user_id, {})
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT count(*)::int AS total_count,
                       count(*) FILTER (WHERE {self._ui_v1_event_type_expr()} = 'ActionExecuted')::int AS action_executed,
                       count(*) FILTER (WHERE {self._ui_v1_event_type_expr()} = 'ActionBlocked')::int AS action_blocked,
                       0::int AS trigger_matched
                {self._ui_v1_signal_from_sql()}
                WHERE {where_sql}
                """,
                params,
            )
            counts = dict(cur.fetchone() or {})

            cur.execute(
                f"""
                SELECT {self._ui_v1_blocked_reason_expr()} AS blocked_reason,
                       count(*)::int AS count
                {self._ui_v1_signal_from_sql()}
                WHERE {where_sql}
                  AND {self._ui_v1_blocked_reason_expr()} IS NOT NULL
                GROUP BY blocked_reason
                """,
                params,
            )
            reasons = {
                str(row["blocked_reason"]): int(row["count"])
                for row in cur.fetchall()
                if row.get("blocked_reason")
            }
        return {
            "total_count": int(counts.get("total_count") or 0),
            "ActionExecuted": int(counts.get("action_executed") or 0),
            "ActionBlocked": int(counts.get("action_blocked") or 0),
            "TriggerMatched": int(counts.get("trigger_matched") or 0),
            "blocked_reason_distribution": reasons,
        }

    def fetch_ui_v1_lineage_stats(self, user_id: int) -> dict[str, Any]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT event_type, status, count(*)::int AS count
                FROM common_event_outbox
                WHERE source_run_id = %s
                  AND event_type IN (
                      'TriggerMatched',
                      'TriggerPendingMarketData',
                      'TriggerStateChanged',
                      'TriggerCleared',
                      'TriggerLiveChanged'
                  )
                  AND status = 'pending'
                GROUP BY event_type, status
                """,
                (N6_UI_V1_LINEAGE_N4_RUN_ID,),
            )
            n4_counts = {
                (str(row["event_type"]), str(row["status"])): int(row["count"])
                for row in cur.fetchall()
            }
            cur.execute(
                """
                SELECT event_type, status, count(*)::int AS count
                FROM common_event_outbox
                WHERE source_run_id = %s
                  AND event_type IN ('ActionExecuted', 'ActionBlocked')
                  AND status = 'pending'
                GROUP BY event_type, status
                """,
                (N6_UI_V1_LINEAGE_N5_RUN_ID,),
            )
            n5_counts = {
                (str(row["event_type"]), str(row["status"])): int(row["count"])
                for row in cur.fetchall()
            }
            cur.execute(
                f"""
                SELECT {self._ui_v1_blocked_reason_expr()} AS blocked_reason,
                       count(*)::int AS count
                {self._ui_v1_signal_from_sql()}
                WHERE p.user_id = %s
                  AND p.user_projection_run_id = %s
                  AND p.source_action_run_id = %s
                  AND {self._ui_v1_event_type_expr()} = 'ActionBlocked'
                  AND {self._ui_v1_blocked_reason_expr()} IS NOT NULL
                GROUP BY blocked_reason
                """,
                (user_id, N6_UI_V1_LINEAGE_N6_RUN_ID, N6_UI_V1_LINEAGE_N5_RUN_ID),
            )
            blocked_reason = {
                str(row["blocked_reason"]): int(row["count"])
                for row in cur.fetchall()
                if row.get("blocked_reason")
            }
        return {
            "source_runs": {
                "N4": N6_UI_V1_LINEAGE_N4_RUN_ID,
                "N5": N6_UI_V1_LINEAGE_N5_RUN_ID,
                "N6": N6_UI_V1_LINEAGE_N6_RUN_ID,
            },
            "lineage_stats": {
                "N4": {
                    "TriggerMatched": {
                        "pending": n4_counts.get(("TriggerMatched", "pending"), 0)
                    },
                    "TriggerPendingMarketData": {
                        "pending": n4_counts.get(("TriggerPendingMarketData", "pending"), 0)
                    },
                    "TriggerStateChanged": {
                        "pending": n4_counts.get(("TriggerStateChanged", "pending"), 0)
                    },
                },
                "N5": {
                    "ActionExecuted": {
                        "pending": n5_counts.get(("ActionExecuted", "pending"), 0)
                    },
                    "ActionBlocked": {
                        "pending": n5_counts.get(("ActionBlocked", "pending"), 0)
                    },
                },
            },
            "legacy": {
                "N4": {
                    "TriggerCleared": {
                        "pending": n4_counts.get(("TriggerCleared", "pending"), 0),
                        "display": "hidden_by_default",
                    },
                    "TriggerLiveChanged": {
                        "pending": n4_counts.get(("TriggerLiveChanged", "pending"), 0),
                        "display": "hidden_by_default",
                    },
                }
            },
            "blocked_reason": blocked_reason,
        }

    def fetch_ui_v1_n3_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        effective_limit = max(1, min(int(limit or N3_MESSAGE_DEFAULT_LIMIT), 5000))
        filters = filters or {}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            latest_event_date = self._raw_message_latest_event_date(
                cur, "N3_market_data", N3_MESSAGE_EVENT_TYPES, filters
            )
            effective_filters, date_filter_defaulted = self._raw_message_effective_filters(
                filters,
                include_all=include_all,
                latest_event_date=latest_event_date,
            )
            where, params = self._raw_message_where("N3_market_data", N3_MESSAGE_EVENT_TYPES, effective_filters)
            cur.execute(
                """
                SELECT count(*)::int AS count
                FROM common_event_outbox
                WHERE source_layer = 'N3_market_data'
                  AND event_type = ANY(%s)
                """,
                (list(N3_MESSAGE_EVENT_TYPES),),
            )
            total_count = int((cur.fetchone() or {}).get("count") or 0)
            filtered_count = self._raw_message_count(cur, where, params)
            summary = self._n3_message_summary(cur, where, params)

            where_sql = " AND ".join(where)
            if include_all:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            else:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            rows = [dict(row) for row in cur.fetchall()]
        return {
            "total_count": total_count,
            "filtered_count": filtered_count,
            "returned_count": len(rows),
            "default_limit": effective_limit,
            "include_all": include_all,
            "filters": {key: value for key, value in effective_filters.items() if value},
            "latest_event_date": latest_event_date,
            "date_filter_defaulted": date_filter_defaulted,
            "event_types": list(N3_MESSAGE_EVENT_TYPES),
            "summary": summary,
            "items": rows,
        }

    def fetch_ui_v1_n4_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        effective_limit = max(1, min(int(limit or N4_MESSAGE_DEFAULT_LIMIT), 5000))
        filters = filters or {}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            latest_event_date = self._raw_message_latest_event_date(cur, "N4_trigger", N4_ALL_EVENT_TYPES, filters)
            effective_filters, date_filter_defaulted = self._raw_message_effective_filters(
                filters,
                include_all=include_all,
                latest_event_date=latest_event_date,
            )
            effective_filters = self._raw_message_active_lineage_filters(
                "N4_trigger",
                effective_filters,
                include_all=include_all,
            )
            where, params = self._raw_message_where("N4_trigger", N4_ALL_EVENT_TYPES, effective_filters)
            cur.execute(
                """
                SELECT count(*)::int AS count
                FROM common_event_outbox
                WHERE source_layer = 'N4_trigger'
                  AND event_type = ANY(%s)
                """,
                (list(N4_ALL_EVENT_TYPES),),
            )
            total_count = int((cur.fetchone() or {}).get("count") or 0)
            filtered_count = self._raw_message_count(cur, where, params)
            summary = self._raw_message_summary(cur, where, params, N5_LEGACY_EVENT_TYPES)

            where_sql = " AND ".join(where)
            if include_all:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            else:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            rows = [dict(row) for row in cur.fetchall()]
        return {
            "total_count": total_count,
            "filtered_count": filtered_count,
            "returned_count": len(rows),
            "default_limit": effective_limit,
            "include_all": include_all,
            "filters": {key: value for key, value in effective_filters.items() if value},
            "latest_event_date": latest_event_date,
            "date_filter_defaulted": date_filter_defaulted,
            "event_types": list(N4_ALL_EVENT_TYPES),
            "standard_event_types": list(N4_STANDARD_EVENT_TYPES),
            "legacy_event_types": list(N4_LEGACY_EVENT_TYPES),
            "items": rows,
        }

    def fetch_ui_v1_n5_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        effective_limit = max(1, min(int(limit or N5_MESSAGE_DEFAULT_LIMIT), 5000))
        filters = filters or {}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            latest_event_date = self._raw_message_latest_event_date(cur, "N5_action", N5_ALL_EVENT_TYPES, filters)
            effective_filters, date_filter_defaulted = self._raw_message_effective_filters(
                filters,
                include_all=include_all,
                latest_event_date=latest_event_date,
            )
            effective_filters = self._raw_message_active_lineage_filters(
                "N5_action",
                effective_filters,
                include_all=include_all,
            )
            where, params = self._raw_message_where("N5_action", N5_ALL_EVENT_TYPES, effective_filters)
            cur.execute(
                """
                SELECT count(*)::int AS count
                FROM common_event_outbox
                WHERE source_layer = 'N5_action'
                  AND event_type = ANY(%s)
                """,
                (list(N5_ALL_EVENT_TYPES),),
            )
            total_count = int((cur.fetchone() or {}).get("count") or 0)
            filtered_count = self._raw_message_count(cur, where, params)
            summary = self._raw_message_summary(cur, where, params, N5_LEGACY_EVENT_TYPES)

            where_sql = " AND ".join(where)
            if include_all:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            else:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            rows = [dict(row) for row in cur.fetchall()]
        return {
            "total_count": total_count,
            "filtered_count": filtered_count,
            "returned_count": len(rows),
            "default_limit": effective_limit,
            "include_all": include_all,
            "filters": {key: value for key, value in effective_filters.items() if value},
            "latest_event_date": latest_event_date,
            "date_filter_defaulted": date_filter_defaulted,
            "event_types": list(N5_ALL_EVENT_TYPES),
            "standard_event_types": list(N5_STANDARD_EVENT_TYPES),
            "legacy_event_types": list(N5_LEGACY_EVENT_TYPES),
            "summary": summary,
            "items": rows,
        }

    def _latest_passed_n5_action_run_id(self, cur: Any) -> str | None:
        cur.execute(
            """
            SELECT run_id
            FROM common_action_run
            WHERE status = 'passed'
            ORDER BY finished_at DESC NULLS LAST,
                     updated_at DESC NULLS LAST,
                     started_at DESC NULLS LAST,
                     created_at DESC NULLS LAST,
                     run_id DESC
            LIMIT 1
            """
        )
        row = cur.fetchone() or {}
        return row.get("run_id")

    def _n5_action_display_where(
        self,
        *,
        target_run_id: str,
        filters: dict[str, Any],
    ) -> tuple[list[str], list[Any]]:
        where = [
            "source_layer = 'N5_action'",
            "event_type = ANY(%s)",
            "source_run_id = %s",
        ]
        params: list[Any] = [list(N5_ACTION_DISPLAY_EVENT_TYPES), target_run_id]
        event_type = filters.get("event_type")
        if event_type:
            where.append("event_type = %s")
            params.append(event_type)
        status = filters.get("status")
        if status:
            where.append("status = %s")
            params.append(status)
        asset_kind = filters.get("asset_kind")
        if asset_kind:
            where.append("asset_kind = %s")
            params.append(asset_kind)
        action_state = filters.get("action_state")
        if action_state:
            where.append("payload_json ->> 'action_state' = %s")
            params.append(action_state)
        keyword = str(filters.get("q") or "").strip()
        if keyword:
            like = f"%{keyword}%"
            where.append(
                """
                (
                  identity_key ILIKE %s
                  OR event_id ILIKE %s
                  OR source_run_id ILIKE %s
                  OR COALESCE(payload_json ->> 'action_key', '') ILIKE %s
                  OR COALESCE(payload_json ->> 'condition_key', '') ILIKE %s
                  OR COALESCE(payload_json #>> '{trace_json,live_window_confirmation,selected_metric_id}', '') ILIKE %s
                )
                """
            )
            params.extend([like, like, like, like, like, like])
        return where, params

    def fetch_ui_v1_n5_actions(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
    ) -> dict[str, Any]:
        effective_limit = max(1, min(int(limit or N5_MESSAGE_DEFAULT_LIMIT), 5000))
        filters = filters or {}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            target_run_id = (
                filters.get("action_run_id")
                or filters.get("source_run_id")
                or self._latest_passed_n5_action_run_id(cur)
            )
            if not target_run_id:
                return n5_actions_empty_model(filters=filters, limit=effective_limit)
            base_where = [
                "source_layer = 'N5_action'",
                "event_type = ANY(%s)",
                "source_run_id = %s",
            ]
            base_params = [list(N5_ACTION_DISPLAY_EVENT_TYPES), target_run_id]
            cur.execute(
                f"""
                SELECT count(*)::int AS count
                FROM common_event_outbox
                WHERE {" AND ".join(base_where)}
                """,
                tuple(base_params),
            )
            total_count = int((cur.fetchone() or {}).get("count") or 0)
            where, params = self._n5_action_display_where(target_run_id=target_run_id, filters=filters)
            where_sql = " AND ".join(where)
            cur.execute(
                f"""
                SELECT count(*)::int AS total,
                       count(*) FILTER (WHERE status = 'pending')::int AS pending,
                       count(*) FILTER (WHERE event_type = 'ActionExecuted')::int AS action_executed,
                       count(*) FILTER (WHERE event_type = 'ActionEligible')::int AS action_eligible
                FROM common_event_outbox
                WHERE {where_sql}
                """,
                tuple(params),
            )
            summary_row = cur.fetchone() or {}
            filtered_count = int(summary_row.get("total") or 0)
            cur.execute(
                f"""
                SELECT outbox_id,
                       event_id,
                       event_type,
                       event_schema_version,
                       trade_date,
                       asset_kind,
                       identity_key,
                       event_time,
                       source_layer,
                       source_run_id,
                       dedup_key,
                       partition_key,
                       payload_json,
                       status,
                       attempt_count,
                       created_at,
                       updated_at
                FROM common_event_outbox
                WHERE {where_sql}
                ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                LIMIT %s
                """,
                tuple(params + [effective_limit]),
            )
            rows = [dict(row) for row in cur.fetchall()]
        effective_filters = {
            key: value
            for key, value in dict(filters, action_run_id=target_run_id, source_run_id=target_run_id).items()
            if value
        }
        items = [n5_action_display_item(row) for row in rows]
        return {
            "ok": True,
            "component": "N5 Actions",
            "title": "N5动作",
            "source_layer": "N5_action",
            "action_run_id": target_run_id,
            "source_run_id": target_run_id,
            "total_count": total_count,
            "filtered_count": filtered_count,
            "returned_count": len(items),
            "default_limit": effective_limit,
            "filters": effective_filters,
            "filter_inputs": {
                "action_run_id": str(filters.get("action_run_id") or ""),
                "source_run_id": str(filters.get("source_run_id") or ""),
                "event_type": str(filters.get("event_type") or ""),
                "status": str(filters.get("status") or ""),
                "asset_kind": str(filters.get("asset_kind") or ""),
                "action_state": str(filters.get("action_state") or ""),
                "q": str(filters.get("q") or ""),
            },
            "event_types": list(N5_ACTION_DISPLAY_EVENT_TYPES),
            "action_states": ["executed", "eligible"],
            "summary": {
                "total": filtered_count,
                "pending": int(summary_row.get("pending") or 0),
                "ActionExecuted": int(summary_row.get("action_executed") or 0),
                "ActionEligible": int(summary_row.get("action_eligible") or 0),
            },
            "items": items,
            "side_effects": n5_actions_read_only_side_effects(),
        }

    def fetch_ui_v1_input_messages(
        self,
        *,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        effective_limit = max(1, min(int(limit or N6_INPUT_MESSAGE_DEFAULT_LIMIT), 5000))
        filters = filters or {}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            latest_event_date = self._input_message_latest_event_date(cur, filters)
            effective_filters, date_filter_defaulted = self._raw_message_effective_filters(
                filters,
                include_all=include_all,
                latest_event_date=latest_event_date,
            )
            where, params = self._input_message_where(effective_filters)
            total_where, total_params = self._input_message_where({})
            total_count = self._raw_message_count(cur, total_where, total_params)
            filtered_count = self._raw_message_count(cur, where, params)
            summary = self._input_message_summary(cur, where, params)
            where_sql = " AND ".join(where)
            if include_all:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            else:
                cur.execute(
                    f"""
                    SELECT outbox_id,
                           event_id,
                           event_type,
                           event_schema_version,
                           trade_date,
                           asset_kind,
                           identity_key,
                           event_time,
                           source_layer,
                           source_run_id,
                           dedup_key,
                           partition_key,
                           payload_json,
                           status,
                           attempt_count,
                           created_at,
                           updated_at
                    FROM common_event_outbox
                    WHERE {where_sql}
                    ORDER BY event_time DESC, created_at DESC, outbox_id DESC
                    LIMIT %s
                    """,
                    tuple(params + [effective_limit]),
                )
            rows = [dict(row) for row in cur.fetchall()]
        return {
            "total_count": total_count,
            "filtered_count": filtered_count,
            "returned_count": len(rows),
            "default_limit": effective_limit,
            "include_all": include_all,
            "filters": {key: value for key, value in effective_filters.items() if value},
            "latest_event_date": latest_event_date,
            "date_filter_defaulted": date_filter_defaulted,
            "source_layers": ["N5_action", "N3_market_data"],
            "event_types": list(N5_ALL_EVENT_TYPES + N3_DISPLAY_INPUT_EVENT_TYPES),
            "n5_canonical_event_types": list(N5_STANDARD_EVENT_TYPES),
            "n5_legacy_event_types": list(N5_LEGACY_EVENT_TYPES),
            "n3_display_event_types": list(N3_DISPLAY_INPUT_EVENT_TYPES),
            "summary": summary,
            "items": rows,
        }

    def fetch_ui_v1_n2_condition_basis(
        self,
        *,
        asset_kind: str,
        filters: dict[str, Any] | None = None,
        limit: int,
        include_all: bool = False,
    ) -> dict[str, Any]:
        meta = n2_condition_basis_asset_meta(asset_kind)
        effective_limit = max(1, min(int(limit or N2_CONDITION_BASIS_DEFAULT_LIMIT), 5000))
        filters = filters or {}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            latest_source_trade_date = self._n2_condition_basis_latest_source_trade_date(cur, meta, filters)
            effective_filters, date_filter_defaulted = self._n2_condition_basis_effective_filters(
                filters,
                include_all=include_all,
                latest_source_trade_date=latest_source_trade_date,
            )
            where, params = self._n2_condition_basis_where(meta, effective_filters)
            source_table = str(meta["source_table"])
            id_column = str(meta["id_column"])
            identity_column = str(meta["identity_column"])
            code_column = str(meta["code_column"])
            name_column = str(meta["name_column"])
            exchange_expr = str(meta["exchange_expr"])
            board_type_expr = str(meta["board_type_expr"])
            cur.execute(f"SELECT count(*)::int AS count FROM {source_table}")
            total_count = int((cur.fetchone() or {}).get("count") or 0)
            cur.execute(
                f"""
                SELECT count(*)::int AS count
                FROM {source_table} t
                WHERE {" AND ".join(where)}
                """,
                tuple(params),
            )
            filtered_count = int((cur.fetchone() or {}).get("count") or 0)

            where_sql = " AND ".join(where)
            select_sql = f"""
                SELECT '{asset_kind}' AS asset_kind,
                       '{meta["label"]}' AS asset_label,
                       '{source_table}' AS source_table,
                       t.{id_column} AS condition_basis_id,
                       t.run_id,
                       t.for_trade_date,
                       t.source_trade_date,
                       t.prev_trade_date,
                       t.{identity_column} AS identity_key,
                       t.{code_column} AS code,
                       {exchange_expr} AS exchange,
                       t.{name_column} AS name,
                       {board_type_expr} AS board_type,
                       t.lane,
                       t.monitor_type,
                       t.monitor_status,
                       t.direction_scope,
                       t.period_key_y,
                       t.period_key_q,
                       t.period_key_m,
                       t.period_key_w,
                       t.period_key_d,
                       t.period_grade_y,
                       t.period_grade_q,
                       t.period_grade_m,
                       t.period_grade_w,
                       t.period_grade_d,
                       t.amount_quality_status,
                       t.buy_target_price,
                       t.buy_expected_return_pct,
                       t.sell_target_price,
                       t.sell_expected_return_pct,
                       t.up_sell_reference_period,
                       t.down_buy_reference_period,
                       t.buy_necessary_key,
                       t.sell_necessary_key,
                       t.buy_full_necessary_key,
                       t.sell_full_necessary_key,
                       t.oversold_hint_key,
                       t.overbought_hint_key,
                       t.quality_status,
                       t.quality_reason,
                       t.source_version,
                       t.source_batch_id,
                       t.raw_json,
                       t.missing_fields_json,
                       t.period_trigger_baseline_json,
                       t.target_price_trace_json,
                       t.created_at,
                       t.updated_at,
                       to_jsonb(t) AS row_json
                FROM {source_table} t
                WHERE {where_sql}
                ORDER BY t.source_trade_date DESC, t.created_at DESC, t.{id_column} DESC
            """
            if include_all:
                cur.execute(select_sql, tuple(params))
            else:
                cur.execute(f"{select_sql} LIMIT %s", tuple(params + [effective_limit]))
            rows = [dict(row) for row in cur.fetchall()]
        return {
            "asset_kind": asset_kind,
            "asset_label": str(meta["label"]),
            "source_table": source_table,
            "total_count": total_count,
            "filtered_count": filtered_count,
            "returned_count": len(rows),
            "default_limit": effective_limit,
            "include_all": include_all,
            "filters": {key: value for key, value in effective_filters.items() if value},
            "latest_source_trade_date": latest_source_trade_date,
            "date_filter_defaulted": date_filter_defaulted,
            "asset_tabs": n2_condition_basis_asset_tabs(),
            "items": rows,
        }

    def fetch_ui_v1_n2_condition_basis_latest_export(self) -> dict[str, Any]:
        latest_dates: list[str] = []
        with self._readonly_connection() as conn, conn.cursor() as cur:
            for asset_kind in N2_CONDITION_BASIS_ASSET_ORDER:
                meta = n2_condition_basis_asset_meta(asset_kind)
                latest_source_trade_date = self._n2_condition_basis_latest_source_trade_date(cur, meta, {})
                if latest_source_trade_date:
                    latest_dates.append(str(latest_source_trade_date))
        latest_source_trade_date = max(latest_dates) if latest_dates else None
        assets: dict[str, Any] = {}
        for asset_kind in N2_CONDITION_BASIS_ASSET_ORDER:
            data = self.fetch_ui_v1_n2_condition_basis(
                asset_kind=asset_kind,
                filters={"source_trade_date": latest_source_trade_date} if latest_source_trade_date else {},
                limit=N2_CONDITION_BASIS_DEFAULT_LIMIT,
                include_all=True,
            )
            assets[asset_kind] = {
                "asset_kind": data.get("asset_kind"),
                "asset_label": data.get("asset_label"),
                "source_table": data.get("source_table"),
                "row_count": len(list(data.get("items") or [])),
                "items": list(data.get("items") or []),
            }
        return {
            "latest_source_trade_date": latest_source_trade_date,
            "assets": assets,
            "side_effects": dict(READ_ONLY_SIDE_EFFECTS),
        }

    def _n2_condition_basis_effective_filters(
        self,
        filters: dict[str, Any],
        *,
        include_all: bool,
        latest_source_trade_date: str | None,
    ) -> tuple[dict[str, Any], bool]:
        effective_filters = dict(filters)
        if include_all or effective_filters.get("source_trade_date") or not latest_source_trade_date:
            return effective_filters, False
        effective_filters["source_trade_date"] = latest_source_trade_date
        return effective_filters, True

    def _n2_condition_basis_latest_source_trade_date(
        self,
        cur: Any,
        meta: dict[str, str],
        filters: dict[str, Any],
    ) -> str | None:
        source_table = str(meta["source_table"])
        where, params = self._n2_condition_basis_where(
            meta,
            {key: value for key, value in filters.items() if key != "source_trade_date"},
        )
        cur.execute(
            f"""
            SELECT max(source_trade_date) AS latest_source_trade_date
            FROM {source_table} t
            WHERE {" AND ".join(where)}
            """,
            tuple(params),
        )
        row = cur.fetchone() or {}
        return row.get("latest_source_trade_date")

    def _n2_condition_basis_where(
        self,
        meta: dict[str, str],
        filters: dict[str, Any],
    ) -> tuple[list[str], list[Any]]:
        identity_column = str(meta["identity_column"])
        code_column = str(meta["code_column"])
        name_column = str(meta["name_column"])
        where = ["TRUE"]
        params: list[Any] = []
        source_trade_date = filters.get("source_trade_date")
        if source_trade_date:
            where.append("t.source_trade_date = %s")
            params.append(source_trade_date)
        condition_key = filters.get("condition_key")
        if condition_key:
            where.append(
                """
                %s IN (
                    t.buy_necessary_key,
                    t.sell_necessary_key,
                    t.buy_full_necessary_key,
                    t.sell_full_necessary_key,
                    t.oversold_hint_key,
                    t.overbought_hint_key
                )
                """
            )
            params.append(condition_key)
        quality_status = filters.get("quality_status")
        if quality_status:
            where.append("t.quality_status = %s")
            params.append(quality_status)
        keyword = filters.get("q")
        if keyword:
            where.append(
                f"""
                (
                    t.{identity_column} ILIKE %s OR
                    t.{code_column} ILIKE %s OR
                    t.{name_column} ILIKE %s OR
                    t.run_id ILIKE %s
                )
                """
            )
            pattern = f"%{keyword}%"
            params.extend([pattern, pattern, pattern, pattern])
        return where, params

    def _raw_message_effective_filters(
        self,
        filters: dict[str, Any],
        *,
        include_all: bool,
        latest_event_date: str | None,
    ) -> tuple[dict[str, Any], bool]:
        effective_filters = dict(filters)
        if include_all or effective_filters.get("event_date") or not latest_event_date:
            return effective_filters, False
        effective_filters["event_date"] = latest_event_date
        return effective_filters, True

    def _raw_message_active_lineage_filters(
        self,
        source_layer: str,
        filters: dict[str, Any],
        *,
        include_all: bool,
    ) -> dict[str, Any]:
        effective_filters = dict(filters)
        if include_all or effective_filters.get("source_run_id"):
            return effective_filters
        event_date = effective_filters.get("event_date")
        if not event_date:
            return effective_filters
        active_source_run_id = ACTIVE_RAW_MESSAGE_SOURCE_RUN_BY_LAYER_DATE.get((source_layer, str(event_date)))
        if active_source_run_id:
            effective_filters["source_run_id"] = active_source_run_id
        return effective_filters

    def _raw_message_latest_event_date(
        self,
        cur: Any,
        source_layer: str,
        event_types: tuple[str, ...],
        filters: dict[str, Any],
    ) -> str | None:
        where = ["source_layer = %s", "event_type = ANY(%s)"]
        params: list[Any] = [source_layer, list(event_types)]
        event_type = filters.get("event_type")
        if event_type:
            where.append("event_type = %s")
            params.append(event_type)
        status = filters.get("status")
        if status:
            where.append("status = %s")
            params.append(status)
        asset_kind = filters.get("asset_kind")
        if asset_kind:
            where.append("asset_kind = %s")
            params.append(asset_kind)
        source_run_id = filters.get("source_run_id")
        if source_run_id:
            where.append("source_run_id = %s")
            params.append(source_run_id)
        keyword = filters.get("q")
        if keyword:
            keyword_where, keyword_params = raw_message_keyword_predicate(keyword)
            where.append(keyword_where)
            params.extend(keyword_params)
        cur.execute(
            f"""
            SELECT max((event_time AT TIME ZONE 'Asia/Shanghai')::date)::text AS latest_event_date
            FROM common_event_outbox
            WHERE {" AND ".join(where)}
            """,
            tuple(params),
        )
        row = cur.fetchone() or {}
        return row.get("latest_event_date")

    def _raw_message_where(
        self,
        source_layer: str,
        event_types: tuple[str, ...],
        filters: dict[str, Any],
    ) -> tuple[list[str], list[Any]]:
        where = ["source_layer = %s", "event_type = ANY(%s)"]
        params: list[Any] = [source_layer, list(event_types)]
        event_type = filters.get("event_type")
        if event_type:
            where.append("event_type = %s")
            params.append(event_type)
        status = filters.get("status")
        if status:
            where.append("status = %s")
            params.append(status)
        asset_kind = filters.get("asset_kind")
        if asset_kind:
            where.append("asset_kind = %s")
            params.append(asset_kind)
        source_run_id = filters.get("source_run_id")
        if source_run_id:
            where.append("source_run_id = %s")
            params.append(source_run_id)
        keyword = filters.get("q")
        if keyword:
            keyword_where, keyword_params = raw_message_keyword_predicate(keyword)
            where.append(keyword_where)
            params.extend(keyword_params)
        event_date = filters.get("event_date")
        if event_date:
            start, end = raw_message_event_date_bounds(event_date)
            where.append("event_time >= %s AND event_time < %s")
            params.extend([start, end])
        return where, params

    def _raw_message_count(self, cur: Any, where: list[str], params: list[Any]) -> int:
        cur.execute(
            f"""
            SELECT count(*)::int AS count
            FROM common_event_outbox
            WHERE {" AND ".join(where)}
            """,
            tuple(params),
        )
        return int((cur.fetchone() or {}).get("count") or 0)

    def _raw_message_summary(
        self,
        cur: Any,
        where: list[str],
        params: list[Any],
        legacy_event_types: tuple[str, ...],
    ) -> dict[str, Any]:
        cur.execute(
            f"""
            SELECT count(*)::int AS total,
                   count(*) FILTER (WHERE status = 'pending')::int AS pending,
                   count(*) FILTER (WHERE event_type = 'ActionBlocked')::int AS action_blocked,
                   count(*) FILTER (WHERE event_type = 'ActionExecuted')::int AS action_executed,
                   count(*) FILTER (WHERE event_type = ANY(%s))::int AS legacy,
                   max(event_time) AS latest_event_time
            FROM common_event_outbox
            WHERE {" AND ".join(where)}
            """,
            tuple([list(legacy_event_types)] + list(params)),
        )
        row = cur.fetchone() or {}
        return {
            "total": int(row.get("total") or 0),
            "pending": int(row.get("pending") or 0),
            "ActionBlocked": int(row.get("action_blocked") or 0),
            "ActionExecuted": int(row.get("action_executed") or 0),
            "legacy": int(row.get("legacy") or 0),
            "latest_event_time": row.get("latest_event_time"),
        }

    def _n3_message_summary(self, cur: Any, where: list[str], params: list[Any]) -> dict[str, Any]:
        cur.execute(
            f"""
            SELECT count(*)::int AS total,
                   count(*) FILTER (WHERE status = 'pending')::int AS pending,
                   count(*) FILTER (WHERE event_type = 'MarketSnapshotUpdated')::int AS market_snapshot_updated,
                   count(*) FILTER (WHERE event_type = 'MinuteBarClosed')::int AS minute_bar_closed,
                   count(*) FILTER (WHERE event_type = 'MarketDisplaySnapshotUpdated')::int
                       AS market_display_snapshot_updated,
                   max(event_time) AS latest_event_time
            FROM common_event_outbox
            WHERE {" AND ".join(where)}
            """,
            tuple(params),
        )
        row = cur.fetchone() or {}
        return {
            "total": int(row.get("total") or 0),
            "pending": int(row.get("pending") or 0),
            "MarketSnapshotUpdated": int(row.get("market_snapshot_updated") or 0),
            "MinuteBarClosed": int(row.get("minute_bar_closed") or 0),
            "MarketDisplaySnapshotUpdated": int(row.get("market_display_snapshot_updated") or 0),
            "latest_event_time": row.get("latest_event_time"),
        }

    def _input_message_where(self, filters: dict[str, Any]) -> tuple[list[str], list[Any]]:
        where = [
            """(
                (source_layer = %s AND event_type = ANY(%s))
                OR (source_layer = %s AND event_type = ANY(%s))
            )"""
        ]
        params: list[Any] = [
            "N5_action",
            list(N5_ALL_EVENT_TYPES),
            "N3_market_data",
            list(N3_DISPLAY_INPUT_EVENT_TYPES),
        ]
        source_layer = filters.get("source_layer")
        if source_layer:
            where.append("source_layer = %s")
            params.append(source_layer)
        input_group = filters.get("input_group")
        if input_group == "n5_action":
            where.append("source_layer = %s")
            params.append("N5_action")
        elif input_group == "n3_display":
            where.append("(source_layer = %s AND event_type = ANY(%s))")
            params.extend(["N3_market_data", list(N3_DISPLAY_INPUT_EVENT_TYPES)])
        elif input_group == "legacy":
            where.append("(source_layer = %s AND event_type = ANY(%s))")
            params.extend(["N5_action", list(N5_LEGACY_EVENT_TYPES)])
        event_type = filters.get("event_type")
        if event_type:
            where.append("event_type = %s")
            params.append(event_type)
        status = filters.get("status")
        if status:
            where.append("status = %s")
            params.append(status)
        asset_kind = filters.get("asset_kind")
        if asset_kind:
            where.append("asset_kind = %s")
            params.append(asset_kind)
        keyword = filters.get("q")
        if keyword:
            keyword_where, keyword_params = raw_message_keyword_predicate(keyword)
            where.append(keyword_where)
            params.extend(keyword_params)
        event_date = filters.get("event_date")
        if event_date:
            start, end = raw_message_event_date_bounds(event_date)
            where.append("event_time >= %s AND event_time < %s")
            params.extend([start, end])
        return where, params

    def _input_message_latest_event_date(self, cur: Any, filters: dict[str, Any]) -> str | None:
        effective_filters = {key: value for key, value in dict(filters).items() if key != "event_date"}
        where, params = self._input_message_where(effective_filters)
        cur.execute(
            f"""
            SELECT max((event_time AT TIME ZONE 'Asia/Shanghai')::date)::text AS latest_event_date
            FROM common_event_outbox
            WHERE {" AND ".join(where)}
            """,
            tuple(params),
        )
        row = cur.fetchone() or {}
        return row.get("latest_event_date")

    def _input_message_summary(self, cur: Any, where: list[str], params: list[Any]) -> dict[str, Any]:
        cur.execute(
            f"""
            SELECT count(*)::int AS total,
                   count(*) FILTER (WHERE status = 'pending')::int AS pending,
                   count(*) FILTER (
                     WHERE source_layer = 'N5_action'
                       AND event_type = ANY(%s)
                   )::int AS n5_canonical,
                   count(*) FILTER (
                     WHERE source_layer = 'N5_action'
                       AND event_type = ANY(%s)
                   )::int AS n5_legacy,
                   count(*) FILTER (
                     WHERE source_layer = 'N3_market_data'
                       AND event_type = ANY(%s)
                   )::int AS n3_display_input,
                   max(event_time) AS latest_event_time
            FROM common_event_outbox
            WHERE {" AND ".join(where)}
            """,
            tuple(
                [
                    list(N5_STANDARD_EVENT_TYPES),
                    list(N5_LEGACY_EVENT_TYPES),
                    list(N3_DISPLAY_INPUT_EVENT_TYPES),
                ]
                + list(params)
            ),
        )
        row = cur.fetchone() or {}
        return {
            "total": int(row.get("total") or 0),
            "pending": int(row.get("pending") or 0),
            "n5_canonical": int(row.get("n5_canonical") or 0),
            "n5_legacy": int(row.get("n5_legacy") or 0),
            "n3_display_input": int(row.get("n3_display_input") or 0),
            "latest_event_time": row.get("latest_event_time"),
        }

    def fetch_ui_v1_status_monitor(
        self,
        user_id: int,
        filters: dict[str, Any],
        limit: int,
        offset: int = 0,
    ) -> dict[str, Any]:
        del user_id
        source_n4_run_id = str(filters.get("source_n4_run_id") or N6_UI_V1_LINEAGE_N4_RUN_ID)
        source_n5_run_id = str(filters.get("source_n5_run_id") or N6_UI_V1_LINEAGE_N5_RUN_ID)
        with self._readonly_connection() as conn, conn.cursor() as cur:
            n4_counts = self._status_monitor_counts(
                cur,
                source_run_id=source_n4_run_id,
                event_types=("TriggerMatched", "TriggerPendingMarketData", "TriggerStateChanged"),
            )
            n5_counts = self._status_monitor_counts(
                cur,
                source_run_id=source_n5_run_id,
                event_types=("ActionExecuted", "ActionBlocked"),
            )
            rows, filtered_count = self._status_monitor_rows(
                cur,
                filters=filters,
                source_n4_run_id=source_n4_run_id,
                source_n5_run_id=source_n5_run_id,
                limit=limit,
                offset=offset,
            )

        trigger_matched = n4_counts.get("TriggerMatched", 0)
        action_executed = n5_counts.get("ActionExecuted", 0)
        action_blocked = n5_counts.get("ActionBlocked", 0)
        pending_market_data = n4_counts.get("TriggerPendingMarketData", 0)
        trigger_state_changed = n4_counts.get("TriggerStateChanged", 0)
        unmatched = max(trigger_matched - action_executed - action_blocked, 0)
        return {
            "source_runs": {
                "N4": source_n4_run_id,
                "N5": source_n5_run_id,
            },
            "event_summary": {
                "N4": {
                    "TriggerMatched": {"pending": trigger_matched, "action_entry": True},
                    "TriggerPendingMarketData": {
                        "pending": pending_market_data,
                        "action_entry": False,
                    },
                    "TriggerStateChanged": {
                        "pending": trigger_state_changed,
                        "action_entry": False,
                    },
                },
                "N5": {
                    "ActionExecuted": {"pending": action_executed},
                    "ActionBlocked": {"pending": action_blocked},
                },
            },
            "relationship_summary": {
                "matched_to_action": {
                    "TriggerMatched": trigger_matched,
                    "ActionExecuted": action_executed,
                    "ActionBlocked": action_blocked,
                    "unmatched": unmatched,
                    "pass": unmatched == 0,
                },
                "status_only": {
                    "TriggerPendingMarketData_action_entries": 0,
                    "TriggerStateChanged_action_entries": 0,
                },
            },
            "status_summary": {
                "active": {"count": trigger_matched, "trigger_live": True},
                "pending_market_data": {"count": pending_market_data, "trigger_live": False},
                "inactive": {"count": 0, "trigger_live": False},
            },
            "items": rows,
            "pagination": {
                "total_count": trigger_matched + pending_market_data + trigger_state_changed,
                "filtered_count": filtered_count,
                "limit": limit,
                "offset": offset,
            },
        }

    def _status_monitor_counts(
        self,
        cur: Any,
        *,
        source_run_id: str,
        event_types: tuple[str, ...],
    ) -> dict[str, int]:
        cur.execute(
            """
            SELECT event_type, count(*)::int AS count
            FROM common_event_outbox
            WHERE source_run_id = %s
              AND event_type = ANY(%s)
              AND status = 'pending'
            GROUP BY event_type
            """,
            (source_run_id, list(event_types)),
        )
        return {str(row["event_type"]): int(row["count"]) for row in cur.fetchall()}

    def _status_monitor_rows(
        self,
        cur: Any,
        *,
        filters: dict[str, Any],
        source_n4_run_id: str,
        source_n5_run_id: str,
        limit: int,
        offset: int,
    ) -> tuple[list[dict[str, Any]], int]:
        n5_mode = (
            filters.get("source_layer") == "N5_action"
            or filters.get("action_event_type") in {"ActionExecuted", "ActionBlocked"}
            or filters.get("event_type") in {"ActionExecuted", "ActionBlocked"}
        )
        if n5_mode:
            return self._status_monitor_n5_rows(
                cur,
                filters=filters,
                source_run_id=source_n5_run_id,
                limit=limit,
                offset=offset,
            )
        return self._status_monitor_n4_rows(
            cur,
            filters=filters,
            source_run_id=source_n4_run_id,
            limit=limit,
            offset=offset,
        )

    def _status_monitor_n4_rows(
        self,
        cur: Any,
        *,
        filters: dict[str, Any],
        source_run_id: str,
        limit: int,
        offset: int,
    ) -> tuple[list[dict[str, Any]], int]:
        where = [
            "source_run_id = %s",
            "event_type IN ('TriggerMatched', 'TriggerPendingMarketData', 'TriggerStateChanged')",
            "status = 'pending'",
        ]
        params: list[Any] = [source_run_id]
        event_type = filters.get("event_type")
        status_filter = filters.get("status")
        if event_type in {"TriggerMatched", "TriggerPendingMarketData", "TriggerStateChanged"}:
            where.append("event_type = %s")
            params.append(event_type)
        elif status_filter == "active":
            where.append("event_type = 'TriggerMatched'")
        elif status_filter == "pending_market_data":
            where.append("event_type = 'TriggerPendingMarketData'")
        elif status_filter == "inactive":
            where.append("event_type = 'TriggerStateChanged'")
            where.append("payload_json->>'current_status' = 'inactive'")
        self._append_status_monitor_common_filters(where, params, filters)
        return self._execute_status_monitor_row_query(
            cur,
            where=where,
            params=params,
            limit=limit,
            offset=offset,
            source_layer_label="'N4_trigger'",
            select_extra="""
                CASE
                  WHEN event_type = 'TriggerMatched' THEN 'active'
                  WHEN event_type = 'TriggerPendingMarketData' THEN 'pending_market_data'
                  WHEN payload_json->>'current_status' = 'matched' THEN 'active'
                  WHEN payload_json->>'current_status' = 'pending_market_data' THEN 'pending_market_data'
                  WHEN payload_json->>'current_status' = 'inactive' THEN 'inactive'
                  ELSE 'inactive'
                END AS status_key,
                COALESCE(payload_json->>'current_status',
                  CASE WHEN event_type = 'TriggerMatched' THEN 'matched'
                       WHEN event_type = 'TriggerPendingMarketData' THEN 'pending_market_data'
                       ELSE NULL END
                ) AS current_status,
                COALESCE((payload_json->>'trigger_live')::boolean, event_type = 'TriggerMatched') AS trigger_live,
                NULL::text AS action_event_type,
                NULL::text AS action_state,
                NULL::text AS blocked_reason,
                NULL::text AS related_n4_event_id
            """,
        )

    def _status_monitor_n5_rows(
        self,
        cur: Any,
        *,
        filters: dict[str, Any],
        source_run_id: str,
        limit: int,
        offset: int,
    ) -> tuple[list[dict[str, Any]], int]:
        where = [
            "source_run_id = %s",
            "event_type IN ('ActionExecuted', 'ActionBlocked')",
            "status = 'pending'",
        ]
        params: list[Any] = [source_run_id]
        action_event_type = filters.get("action_event_type") or filters.get("event_type")
        if action_event_type in {"ActionExecuted", "ActionBlocked"}:
            where.append("event_type = %s")
            params.append(action_event_type)
        self._append_status_monitor_common_filters(where, params, filters)
        return self._execute_status_monitor_row_query(
            cur,
            where=where,
            params=params,
            limit=limit,
            offset=offset,
            source_layer_label="'N5_action'",
            select_extra="""
                'active'::text AS status_key,
                'matched'::text AS current_status,
                true AS trigger_live,
                event_type AS action_event_type,
                COALESCE(payload_json->>'action_state',
                  CASE WHEN event_type = 'ActionExecuted' THEN 'executed'
                       WHEN event_type = 'ActionBlocked' THEN 'blocked'
                       ELSE NULL END
                ) AS action_state,
                payload_json->>'blocked_reason' AS blocked_reason,
                COALESCE(payload_json->>'source_trigger_event_id',
                         payload_json->>'source_n4_event_id',
                         payload_json->>'source_event_id') AS related_n4_event_id
            """,
        )

    def _append_status_monitor_common_filters(
        self,
        where: list[str],
        params: list[Any],
        filters: dict[str, Any],
    ) -> None:
        for key, expression in (
            ("asset_kind", "asset_kind = %s"),
            ("direction", "payload_json->>'direction' = %s"),
            ("signal_type", "payload_json->>'signal_type' = %s"),
            ("trade_date", "trade_date::text = %s"),
        ):
            value = filters.get(key)
            if value:
                where.append(expression)
                params.append(value)
        keyword = filters.get("q")
        if keyword:
            where.append(
                """
                (
                  event_id ILIKE %s
                  OR identity_key ILIKE %s
                  OR payload_json->>'code' ILIKE %s
                  OR payload_json->>'condition_key' ILIKE %s
                )
                """
            )
            like = f"%{keyword}%"
            params.extend([like, like, like, like])

    def _execute_status_monitor_row_query(
        self,
        cur: Any,
        *,
        where: list[str],
        params: list[Any],
        limit: int,
        offset: int,
        source_layer_label: str,
        select_extra: str,
    ) -> tuple[list[dict[str, Any]], int]:
        where_sql = " AND ".join(where)
        cur.execute(
            f"""
            SELECT count(*)::int AS count
            FROM common_event_outbox
            WHERE {where_sql}
            """,
            tuple(params),
        )
        count_row = cur.fetchone() or {}
        filtered_count = int(count_row.get("count") or 0)
        cur.execute(
            f"""
            SELECT
                {source_layer_label} AS source_layer,
                event_type,
                event_id,
                status AS outbox_status,
                source_run_id,
                event_time,
                asset_kind,
                identity_key,
                payload_json->>'code' AS code,
                payload_json->>'name' AS name,
                payload_json->>'direction' AS direction,
                payload_json->>'signal_type' AS signal_type,
                COALESCE(payload_json->>'condition_key', payload_json->>'original_condition_key') AS condition_key,
                {select_extra}
            FROM common_event_outbox
            WHERE {where_sql}
            ORDER BY event_time DESC, event_id DESC
            LIMIT %s OFFSET %s
            """,
            tuple(params + [limit, offset]),
        )
        return [dict(row) for row in cur.fetchall()], filtered_count

    def fetch_ui_v1_signal_detail(self, user_id: int, user_signal_projection_id: int) -> dict[str, Any] | None:
        where_sql, params = self._ui_v1_signal_where(user_id, {})
        params["user_signal_projection_id"] = int(user_signal_projection_id)
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT {self._ui_v1_signal_select_list()}
                {self._ui_v1_signal_from_sql()}
                WHERE {where_sql}
                  AND p.user_signal_projection_id = %(user_signal_projection_id)s
                LIMIT 1
                """,
                params,
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_ui_v1_dashboard_metrics(self, user_id: int) -> dict[str, Any]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                WITH scoped AS (
                  SELECT p.user_projection_run_id,
                         p.created_at,
                         {self._ui_v1_action_state_expr()} AS action_state,
                         q.queue_status,
                         r.status AS run_status
                  FROM user_signal_projection p
                  LEFT JOIN user_signal_card c
                    ON c.user_signal_projection_id = p.user_signal_projection_id
                   AND c.user_id = p.user_id
                  LEFT JOIN LATERAL (
                    SELECT q.queue_status,
                           q.action_state
                    FROM user_notification_queue q
                    WHERE q.user_signal_projection_id = p.user_signal_projection_id
                      AND q.user_id = p.user_id
                    ORDER BY
                      CASE WHEN q.queue_status = 'queued_only' THEN 0 ELSE 1 END,
                      q.created_at DESC,
                      q.user_notification_queue_id DESC
                    LIMIT 1
                  ) q ON true
                  LEFT JOIN user_projection_run r
                    ON r.user_projection_run_id = p.user_projection_run_id
                  WHERE p.user_id = %s
                )
                SELECT count(*)::int AS today_signal_count,
                       count(*) FILTER (WHERE action_state = 'blocked')::int AS action_blocked,
                       count(*) FILTER (WHERE action_state IN ('executed', 'action_confirmed'))::int AS action_executed,
                       count(*) FILTER (WHERE queue_status = 'queued_only')::int AS queued_only,
                       count(*) FILTER (WHERE queue_status = 'ready_for_future_push')::int AS pending_delivery,
                       COALESCE(bool_and(COALESCE(run_status, 'passed') IN ('passed', 'ready')), true) AS rollback_safe,
                       (array_agg(user_projection_run_id ORDER BY created_at DESC))[1] AS latest_run_id
                FROM scoped
                """,
                (user_id,),
            )
            row = cur.fetchone()
        return dict(row) if row else {}

    def fetch_ui_v1_virtual_account(self, user_id: int) -> dict[str, Any] | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT v.virtual_account_id,
                       v.principal_id,
                       v.principal_type,
                       v.account_name,
                       v.virtual_account_status,
                       v.base_currency,
                       v.initial_cash,
                       v.current_cash_snapshot_id,
                       v.run_id,
                       v.policy_version,
                       v.policy_hash,
                       v.rollback_scope,
                       v.quality_status,
                       v.created_at,
                       v.updated_at
                FROM n6_virtual_account v
                JOIN n6_principal p
                  ON p.principal_id = v.principal_id
                 AND p.principal_type = v.principal_type
                WHERE p.owner_user_id = %s
                  AND p.principal_status = 'active'
                  AND v.virtual_account_status = 'active'
                  AND v.principal_type IN ('admin', 'human_user')
                ORDER BY v.virtual_account_id ASC
                LIMIT 1
                """,
                (user_id,),
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_ui_v1_cash_snapshot(self, user_id: int) -> dict[str, Any] | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                WITH active_account AS (
                  SELECT v.virtual_account_id,
                         v.current_cash_snapshot_id
                  FROM n6_virtual_account v
                  JOIN n6_principal p
                    ON p.principal_id = v.principal_id
                   AND p.principal_type = v.principal_type
                  WHERE p.owner_user_id = %s
                    AND p.principal_status = 'active'
                    AND v.virtual_account_status = 'active'
                    AND v.principal_type IN ('admin', 'human_user')
                  ORDER BY v.virtual_account_id ASC
                  LIMIT 1
                )
                SELECT s.cash_snapshot_id,
                       s.virtual_account_id,
                       s.snapshot_time,
                       s.trade_date,
                       s.available_cash,
                       s.frozen_cash,
                       s.total_cash,
                       s.currency,
                       s.source_ledger_max_id,
                       s.snapshot_status,
                       s.run_id,
                       s.policy_version,
                       s.policy_hash,
                       s.rollback_scope,
                       s.quality_status,
                       s.created_at,
                       (a.current_cash_snapshot_id IS NULL) AS pointer_missing_warning
                FROM active_account a
                JOIN n6_virtual_cash_snapshot s
                  ON s.virtual_account_id = a.virtual_account_id
                WHERE (
                    a.current_cash_snapshot_id IS NOT NULL
                    AND s.cash_snapshot_id = a.current_cash_snapshot_id
                  )
                   OR (
                    a.current_cash_snapshot_id IS NULL
                    AND s.snapshot_status = 'active'
                  )
                ORDER BY
                  CASE WHEN s.cash_snapshot_id = a.current_cash_snapshot_id THEN 0 ELSE 1 END,
                  s.snapshot_time DESC,
                  s.cash_snapshot_id DESC
                LIMIT 1
                """,
                (user_id,),
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_ui_v1_cash_ledger(self, user_id: int, limit: int) -> list[dict[str, Any]]:
        clamped_limit = max(1, min(int(limit), 100))
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                WITH active_account AS (
                  SELECT v.virtual_account_id
                  FROM n6_virtual_account v
                  JOIN n6_principal p
                    ON p.principal_id = v.principal_id
                   AND p.principal_type = v.principal_type
                  WHERE p.owner_user_id = %s
                    AND p.principal_status = 'active'
                    AND v.virtual_account_status = 'active'
                    AND v.principal_type IN ('admin', 'human_user')
                  ORDER BY v.virtual_account_id ASC
                  LIMIT 1
                )
                SELECT l.cash_ledger_id,
                       l.virtual_account_id,
                       l.ledger_type,
                       l.amount,
                       l.currency,
                       l.trade_date,
                       l.event_time,
                       l.source_event_type,
                       l.source_event_id,
                       l.source_virtual_order_id,
                       l.source_virtual_trade_id,
                       l.run_id,
                       l.policy_version,
                       l.policy_hash,
                       l.rollback_scope,
                       l.quality_status,
                       l.created_at
                FROM n6_virtual_cash_ledger l
                JOIN active_account a
                  ON a.virtual_account_id = l.virtual_account_id
                ORDER BY l.event_time DESC, l.cash_ledger_id DESC
                LIMIT %s
                """,
                (user_id, clamped_limit),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_ui_v1_artifacts(self) -> dict[str, Any]:
        paths = [
            "docs/N6_USER_INTERFACE_SPEC_v1.md",
            "docs/N6_USER_INTERFACE_SPEC_v1_TRACEABILITY.md",
            "docs/N6_USER_INTERFACE_SPEC_v1_POST_REVIEW.md",
            "docs/N6_USER_INTERFACE_SPEC_v1_DRY_RUN_PREVIEW.md",
            "docs/N6_USER_INTERFACE_SPEC_v1_DRY_RUN_PREVIEW.json",
        ]
        repo_root = Path(__file__).resolve().parents[3]
        return {
            "artifacts": paths,
            "stale_artifact": any(not (repo_root / path).exists() for path in paths),
        }

    def fetch_ui_v1_rollback_summary(self) -> dict[str, Any]:
        repo_root = Path(__file__).resolve().parents[3]
        rollback_path = "sql/N6_projection_business_rollback.sql"
        return {
            "rollback_sql_path": rollback_path,
            "rollback_safe": (repo_root / rollback_path).exists(),
            "delete_order": [
                "user_notification_queue",
                "user_signal_card",
                "user_signal_projection",
                "user_projection_run",
            ],
            "hard_fail_guards": ["decision", "sim", "voice", "mobile", "position"],
        }

    def fetch_app_principals(self, user_id: int) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT p.principal_id,
                       p.principal_type,
                       p.owner_user_id,
                       p.principal_status,
                       COALESCE(p.principal_label, u.display_name, u.login_name) AS display_name
                FROM n6_principal p
                JOIN user_account u
                  ON u.user_id = p.owner_user_id
                WHERE p.owner_user_id = %s
                  AND p.principal_status = 'active'
                  AND p.principal_type IN ('admin', 'human_user')
                  AND u.status = 'active'
                ORDER BY
                  CASE WHEN p.principal_type = 'admin' THEN 0 ELSE 1 END,
                  p.principal_id ASC
                """,
                (user_id,),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_app_virtual_account(self, principal_id: int, principal_type: str) -> dict[str, Any] | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT virtual_account_id,
                       principal_id,
                       principal_type,
                       account_name,
                       virtual_account_status,
                       base_currency,
                       initial_cash,
                       current_cash_snapshot_id,
                       run_id,
                       policy_version,
                       policy_hash,
                       rollback_scope,
                       quality_status,
                       created_at,
                       updated_at
                FROM n6_virtual_account
                WHERE principal_id = %s
                  AND principal_type = %s
                  AND virtual_account_status = 'active'
                ORDER BY virtual_account_id ASC
                LIMIT 1
                """,
                (principal_id, principal_type),
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_app_cash_snapshot(self, virtual_account_id: int) -> dict[str, Any] | None:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                WITH account AS (
                  SELECT virtual_account_id,
                         current_cash_snapshot_id
                  FROM n6_virtual_account
                  WHERE virtual_account_id = %s
                )
                SELECT s.cash_snapshot_id,
                       s.virtual_account_id,
                       s.snapshot_time,
                       s.trade_date,
                       s.available_cash,
                       s.frozen_cash,
                       s.total_cash,
                       s.currency,
                       s.source_ledger_max_id,
                       s.snapshot_status,
                       s.run_id,
                       s.policy_version,
                       s.policy_hash,
                       s.rollback_scope,
                       s.quality_status,
                       s.created_at,
                       (a.current_cash_snapshot_id IS NULL) AS pointer_missing_warning
                FROM account a
                JOIN n6_virtual_cash_snapshot s
                  ON s.virtual_account_id = a.virtual_account_id
                WHERE (
                    a.current_cash_snapshot_id IS NOT NULL
                    AND s.cash_snapshot_id = a.current_cash_snapshot_id
                  )
                   OR (
                    a.current_cash_snapshot_id IS NULL
                    AND s.snapshot_status = 'active'
                  )
                ORDER BY
                  CASE WHEN s.cash_snapshot_id = a.current_cash_snapshot_id THEN 0 ELSE 1 END,
                  s.snapshot_time DESC,
                  s.cash_snapshot_id DESC
                LIMIT 1
                """,
                (virtual_account_id,),
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_app_positions(self, principal_id: int, principal_type: str) -> list[dict[str, Any]]:
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                WITH active_principal AS (
                  SELECT p.principal_id,
                         p.principal_type
                  FROM n6_principal p
                  WHERE p.principal_id = %s
                    AND p.principal_type = %s
                    AND p.principal_status = 'active'
                ),
                active_accounts AS (
                  SELECT a.virtual_account_id,
                         a.principal_id,
                         a.principal_type
                  FROM n6_virtual_account a
                  JOIN active_principal p
                    ON p.principal_id = a.principal_id
                   AND p.principal_type = a.principal_type
                  WHERE a.virtual_account_status = 'active'
                ),
                active_scope AS (
                  SELECT min(virtual_account_id) AS virtual_account_id,
                         min(principal_id) AS principal_id,
                         min(principal_type) AS principal_type
                  FROM active_accounts
                  HAVING count(*) = 1
                ),
                current_trade_day AS (
                  SELECT min(to_date(c.trade_date, 'YYYYMMDD')) AS trade_date
                  FROM common_trade_calendar c
                  WHERE c.trade_date = to_char(
                          (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Shanghai')::date,
                          'YYYYMMDD'
                        )
                    AND c.is_open IS TRUE
                  HAVING count(*) = 1
                ),
                scoped_positions AS (
                  SELECT p.virtual_position_id,
                         p.virtual_account_id,
                         p.principal_id,
                         p.principal_type,
                         p.asset_kind,
                         p.identity_key,
                         p.position_status,
                         p.quantity,
                         p.average_cost,
                         p.holding_episode_no,
                         p.first_open_trade_date,
                         p.locked_target_price,
                         p.target_price_status,
                         p.stop_loss_price,
                         p.stop_loss_status,
                         p.stop_loss_effective_trade_date,
                         p.stop_loss_source_quote_snapshot_id,
                         p.stop_loss_frozen_at,
                         p.stop_loss_policy_version,
                         p.stop_loss_policy_hash,
                         d.trade_date AS current_trade_date
                  FROM active_scope s
                  JOIN n6_virtual_position p
                    ON p.virtual_account_id = s.virtual_account_id
                   AND p.principal_id = s.principal_id
                   AND p.principal_type = s.principal_type
                  CROSS JOIN current_trade_day d
                  WHERE p.asset_kind = 'stock'
                    AND p.position_status = 'open_virtual'
                    AND p.quantity > 0
                ),
                lot_rollup AS (
                  SELECT p.virtual_position_id,
                         coalesce(sum(l.remaining_quantity) FILTER (
                           WHERE l.virtual_account_id = p.virtual_account_id
                             AND l.principal_id = p.principal_id
                             AND l.principal_type = p.principal_type
                             AND l.identity_key = p.identity_key
                             AND l.holding_episode_no = p.holding_episode_no
                             AND l.remaining_quantity > 0
                         ), 0) AS lot_quantity_total,
                         coalesce(sum(l.remaining_quantity) FILTER (
                           WHERE l.virtual_account_id = p.virtual_account_id
                             AND l.principal_id = p.principal_id
                             AND l.principal_type = p.principal_type
                             AND l.identity_key = p.identity_key
                             AND l.holding_episode_no = p.holding_episode_no
                             AND l.remaining_quantity > 0
                             AND l.available_trade_date <= p.current_trade_date
                             AND l.lot_status IN ('locked_t1', 'available')
                         ), 0) AS sellable_quantity,
                         coalesce(sum(l.remaining_quantity) FILTER (
                           WHERE l.virtual_account_id = p.virtual_account_id
                             AND l.principal_id = p.principal_id
                             AND l.principal_type = p.principal_type
                             AND l.identity_key = p.identity_key
                             AND l.holding_episode_no = p.holding_episode_no
                             AND l.remaining_quantity > 0
                             AND l.available_trade_date > p.current_trade_date
                         ), 0) AS t1_locked_quantity,
                         bool_or(
                           l.remaining_quantity > 0
                           AND (
                             l.virtual_account_id <> p.virtual_account_id
                             OR l.principal_id <> p.principal_id
                             OR l.principal_type <> p.principal_type
                             OR l.identity_key <> p.identity_key
                             OR l.holding_episode_no <> p.holding_episode_no
                           )
                         ) AS lot_scope_mismatch,
                         bool_or(
                           l.remaining_quantity > 0
                           AND l.lot_status NOT IN ('locked_t1', 'available')
                         ) FILTER (
                           WHERE l.virtual_account_id = p.virtual_account_id
                             AND l.principal_id = p.principal_id
                             AND l.principal_type = p.principal_type
                             AND l.identity_key = p.identity_key
                             AND l.holding_episode_no = p.holding_episode_no
                         ) AS lot_status_mismatch
                  FROM scoped_positions p
                  JOIN n6_virtual_position_lot l
                    ON l.virtual_position_id = p.virtual_position_id
                  GROUP BY p.virtual_position_id
                )
                SELECT p.virtual_position_id,
                       p.identity_key,
                       split_part(p.identity_key, ':', 2) AS position_exchange,
                       split_part(p.identity_key, ':', 3) AS stock_code,
                       p.position_status,
                       p.quantity,
                       r.sellable_quantity,
                       r.t1_locked_quantity,
                       r.lot_quantity_total,
                       p.average_cost,
                       p.current_trade_date,
                       p.holding_episode_no,
                       p.first_open_trade_date,
                       p.locked_target_price AS target_price,
                       p.target_price_status,
                       p.stop_loss_price,
                       p.stop_loss_status,
                       p.stop_loss_effective_trade_date,
                       p.stop_loss_source_quote_snapshot_id,
                       p.stop_loss_frozen_at,
                       p.stop_loss_policy_version,
                       p.stop_loss_policy_hash,
                       stock_identity.stock_name,
                       industry_identity.industry_code,
                       industry_identity.industry_name,
                       q.exchange AS quote_exchange,
                       q.current_price,
                       q.quote_minute,
                       q.fetched_at,
                       q.quality_status,
                       q.quality_reason
                FROM scoped_positions p
                JOIN lot_rollup r
                  ON r.virtual_position_id = p.virtual_position_id
                LEFT JOIN v_n6_virtual_quote_latest q
                  ON q.identity_key = p.identity_key
                LEFT JOIN LATERAL (
                  SELECT CASE WHEN count(*) = 1 THEN max(latest_name.stock_name) END AS stock_name
                  FROM (
                    SELECT DISTINCT btrim(basis.display_name::text) AS stock_name
                    FROM v_n6_stock_condition_display_basis basis
                    WHERE basis.identity_key = p.identity_key
                      AND NULLIF(btrim(basis.display_name::text), '') IS NOT NULL
                      AND basis.for_trade_date = (
                        SELECT max(asof_basis.for_trade_date)
                        FROM v_n6_stock_condition_display_basis asof_basis
                        WHERE asof_basis.identity_key = p.identity_key
                          AND asof_basis.for_trade_date <= to_char(p.current_trade_date, 'YYYYMMDD')
                      )
                  ) latest_name
                ) stock_identity ON TRUE
                LEFT JOIN LATERAL (
                  SELECT
                    CASE WHEN count(*) = 1 THEN max(latest_industry.board_code) END AS industry_code,
                    CASE WHEN count(*) = 1 THEN max(latest_industry.board_name) END AS industry_name
                  FROM (
                    SELECT DISTINCT
                           membership.board_identity_key,
                           membership.board_code,
                           membership.board_name
                    FROM v_n6_board_membership_fact membership
                    WHERE membership.stock_identity_key = p.identity_key
                      AND membership.board_type = 'tdx_industry'
                      AND NULLIF(btrim(membership.board_identity_key::text), '') IS NOT NULL
                      AND NULLIF(btrim(membership.board_code::text), '') IS NOT NULL
                      AND NULLIF(btrim(membership.board_name::text), '') IS NOT NULL
                      AND membership.trade_date = (
                        SELECT max(asof_membership.trade_date)
                        FROM v_n6_board_membership_fact asof_membership
                        WHERE asof_membership.stock_identity_key = p.identity_key
                          AND asof_membership.board_type = 'tdx_industry'
                          AND asof_membership.trade_date <= to_char(p.current_trade_date, 'YYYYMMDD')
                      )
                  ) latest_industry
                ) industry_identity ON TRUE
                WHERE r.lot_quantity_total = p.quantity
                  AND coalesce(r.lot_scope_mismatch, false) IS FALSE
                  AND coalesce(r.lot_status_mismatch, false) IS FALSE
                ORDER BY p.identity_key
                """,
                (principal_id, principal_type),
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_app_signals(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        filters: dict[str, Any],
        limit: int,
    ) -> list[dict[str, Any]]:
        if not self._app_v1_signal_scope_relations_ready():
            return []
        where_sql, params = self._app_v1_signal_where(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
            filters=filters,
            include_monitor_scope=False,
        )
        params["limit"] = max(1, min(int(limit), 500))
        historical_projection_mode = bool(filters.get("historical_projection_mode"))
        scope_cte_sql = self._app_v1_web_signal_scope_cte(
            include_expired=historical_projection_mode,
            include_realtime_scope=not historical_projection_mode,
            use_current_approved_batch=not historical_projection_mode,
        )
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                WITH {scope_cte_sql}
                SELECT {self._app_v1_signal_select_list()}
                FROM user_signal_projection p
                JOIN user_projection_run r
                  ON r.user_projection_run_id = p.user_projection_run_id
                 AND r.status IN ('passed', 'ready')
                LEFT JOIN user_signal_card c
                  ON c.user_signal_projection_id = p.user_signal_projection_id
                 AND c.user_projection_run_id = p.user_projection_run_id
                 AND c.user_id = p.user_id
                {self._app_v1_web_signal_scope_join()}
                WHERE {where_sql}
                ORDER BY p.created_at DESC, p.user_signal_projection_id DESC
                LIMIT %(limit)s
                """,
                params,
            )
            return [dict(row) for row in cur.fetchall()]

    def fetch_app_signal_detail(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        user_signal_projection_id: int,
        trade_date: str | None = None,
    ) -> dict[str, Any] | None:
        if not self._app_v1_signal_scope_relations_ready():
            return None
        effective_trade_date = normalize_filter_value(trade_date) or self.fetch_app_current_signal_trade_date()
        if not effective_trade_date:
            return None
        where_sql, params = self._app_v1_signal_where(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
            filters={"trade_date": effective_trade_date},
            include_monitor_scope=False,
        )
        params["user_signal_projection_id"] = int(user_signal_projection_id)
        scope_cte_sql = self._app_v1_web_signal_scope_cte()
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                WITH {scope_cte_sql}
                SELECT {self._app_v1_signal_detail_select_list()}
                FROM user_signal_projection p
                JOIN user_projection_run r
                  ON r.user_projection_run_id = p.user_projection_run_id
                 AND r.status IN ('passed', 'ready')
                LEFT JOIN user_signal_card c
                  ON c.user_signal_projection_id = p.user_signal_projection_id
                 AND c.user_projection_run_id = p.user_projection_run_id
                 AND c.user_id = p.user_id
                {self._app_v1_web_signal_scope_join()}
                WHERE {where_sql}
                  AND p.user_signal_projection_id = %(user_signal_projection_id)s
                LIMIT 1
                """,
                params,
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def fetch_app_current_signal_trade_date(self) -> str | None:
        current_filter_batch = self._app_v2_current_filter_batches(["stock", "index", "board"])
        return current_app_signal_trade_date({"current_filter_batch": current_filter_batch})

    def fetch_app_signal_scope_metadata(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        trade_date: str | None = None,
    ) -> dict[str, Any]:
        metadata = {
            "scope_mode": "effective_monitor",
            "current_filter_batch": {
                asset_kind: {"source_trade_date": "", "for_trade_date": "", "source_run_id": ""}
                for asset_kind in APP_V2_MONITOR_TABLE_BY_ASSET
            },
            "available_trade_dates": [],
            "effective_monitor_count": 0,
            "expired_monitor_count": 0,
            "matched_signal_count": 0,
            "excluded_reason_counts": {
                "message_trade_date_missing": 0,
                "message_trade_date_mismatch": 0,
                "monitor_expired": 0,
            },
        }
        if not self._app_v1_signal_scope_relations_ready():
            return metadata
        current_filter_batch = self._app_v2_current_filter_batches(["stock", "index", "board"])
        current_trade_date = current_app_signal_trade_date({"current_filter_batch": current_filter_batch})
        requested_trade_date = normalize_filter_value(trade_date)
        if requested_trade_date and requested_trade_date != current_trade_date:
            current_trade_date = None
        cache_key = (int(principal_id), str(principal_type), int(user_id), current_trade_date or "")
        now_monotonic = time.monotonic()
        with self._app_user_scope_cache_lock:
            cached = self._app_user_scope_cache.get(cache_key)
            if cached and now_monotonic - cached[0] <= 2.0:
                return copy.deepcopy(cached[1])
        metadata["current_filter_batch"] = current_filter_batch
        if not current_trade_date:
            return metadata
        where_sql, params = self._app_v1_signal_where(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
            filters={"trade_date": current_trade_date},
            include_monitor_scope=False,
        )
        scope_cte_sql = self._app_v1_effective_monitor_scope_cte()
        all_monitor_cte_sql = self._app_v1_all_monitor_scope_cte()
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT pg_catalog.to_char(p.for_trade_date, 'YYYYMMDD') AS trade_date
                FROM user_signal_projection p
                JOIN user_projection_run r
                  ON r.user_projection_run_id = p.user_projection_run_id
                 AND r.status IN ('passed', 'ready')
                WHERE p.projection_status IN ('visible', 'blocked')
                  AND p.for_trade_date IS NOT NULL
                  AND (
                    p.source_action_run_id IS NULL
                    OR NOT (p.source_action_run_id = ANY(%(stale_source_action_run_ids)s))
                  )
                GROUP BY p.for_trade_date
                ORDER BY p.for_trade_date DESC
                """,
                {"stale_source_action_run_ids": params["stale_source_action_run_ids"]},
            )
            metadata["available_trade_dates"] = [
                str(item["trade_date"])
                for item in cur.fetchall()
                if item.get("trade_date")
            ]
            cur.execute(
                f"""
                WITH {scope_cte_sql},
                {all_monitor_cte_sql},
                deduplicated_monitor_scope AS (
                  SELECT DISTINCT principal_id,
                                  principal_type,
                                  asset_kind,
                                  identity_key,
                                  direction,
                                  valid_for_trade_date
                  FROM effective_monitor_scope
                ),
                candidate_messages AS (
                  SELECT p.user_signal_projection_id,
                         p.asset_kind,
                         p.identity_key,
                         p.direction,
                         pg_catalog.to_char(p.for_trade_date, 'YYYYMMDD') AS message_trade_date
                  FROM user_signal_projection p
                  JOIN user_projection_run r
                    ON r.user_projection_run_id = p.user_projection_run_id
                   AND r.status IN ('passed', 'ready')
                  WHERE {where_sql}
                ),
                scoped_candidate_messages AS (
                  SELECT candidate_messages.user_signal_projection_id,
                         candidate_messages.asset_kind,
                         candidate_messages.identity_key,
                         candidate_messages.direction,
                         candidate_messages.message_trade_date
                  FROM candidate_messages
                  WHERE EXISTS (
                    SELECT 1
                    FROM deduplicated_monitor_scope
                    WHERE deduplicated_monitor_scope.asset_kind = candidate_messages.asset_kind
                      AND deduplicated_monitor_scope.identity_key = candidate_messages.identity_key
                      AND deduplicated_monitor_scope.direction = candidate_messages.direction
                  )
                )
                SELECT
                  (SELECT COUNT(*) FROM deduplicated_monitor_scope) AS effective_monitor_count,
                  (
                    SELECT COUNT(*)
                    FROM all_monitor_scope expired_monitor_scope
                    WHERE expired_monitor_scope.status = 'active'
                      AND NOT EXISTS (
                        SELECT 1
                        FROM effective_monitor_scope
                        WHERE effective_monitor_scope.asset_kind = expired_monitor_scope.asset_kind
                          AND effective_monitor_scope.monitor_id = expired_monitor_scope.monitor_id
                      )
                  ) AS expired_monitor_count,
                  COUNT(*) FILTER (
                    WHERE scoped_candidate_messages.message_trade_date IS NULL
                  ) AS message_trade_date_missing,
                  COUNT(*) FILTER (
                    WHERE scoped_candidate_messages.message_trade_date IS NOT NULL
                      AND NOT EXISTS (
                        SELECT 1
                        FROM deduplicated_monitor_scope
                        WHERE deduplicated_monitor_scope.asset_kind = scoped_candidate_messages.asset_kind
                          AND deduplicated_monitor_scope.identity_key = scoped_candidate_messages.identity_key
                          AND deduplicated_monitor_scope.direction = scoped_candidate_messages.direction
                          AND deduplicated_monitor_scope.valid_for_trade_date = scoped_candidate_messages.message_trade_date
                      )
                  ) AS message_trade_date_mismatch
                FROM scoped_candidate_messages
                """,
                params,
            )
            row = cur.fetchone()
        if row:
            metadata["effective_monitor_count"] = int(row.get("effective_monitor_count") or 0)
            metadata["expired_monitor_count"] = int(row.get("expired_monitor_count") or 0)
            metadata["excluded_reason_counts"]["monitor_expired"] = int(row.get("expired_monitor_count") or 0)
            metadata["excluded_reason_counts"]["message_trade_date_missing"] = int(
                row.get("message_trade_date_missing") or 0
            )
            metadata["excluded_reason_counts"]["message_trade_date_mismatch"] = int(
                row.get("message_trade_date_mismatch") or 0
            )
        with self._app_user_scope_cache_lock:
            self._app_user_scope_cache[cache_key] = (time.monotonic(), copy.deepcopy(metadata))
        return metadata

    def fetch_app_realtime_scope(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
    ) -> dict[str, Any]:
        if not self._app_v2_relation_exists(APP_REALTIME_SCOPE_TABLE):
            return {
                "tables_ready": False,
                "items": self._app_v2_default_realtime_scope_items(
                    principal_id=principal_id,
                    principal_type=principal_type,
                    user_id=user_id,
                ),
            }
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT realtime_scope_id,
                       principal_id,
                       principal_type,
                       user_id,
                       asset_kind,
                       identity_key,
                       display_name,
                       source_type,
                       source_snapshot_json,
                       is_default_seed,
                       status,
                       deleted_at,
                       created_at,
                       updated_at
                FROM {APP_REALTIME_SCOPE_TABLE}
                WHERE principal_id = %(principal_id)s
                  AND principal_type = %(principal_type)s
                  AND user_id = %(user_id)s
                ORDER BY
                  CASE WHEN status = 'active' THEN 0 ELSE 1 END,
                  asset_kind ASC,
                  identity_key ASC
                """,
                {
                    "principal_id": principal_id,
                    "principal_type": principal_type,
                    "user_id": user_id,
                },
            )
            rows = [dict(row) for row in cur.fetchall()]
        return {
            "tables_ready": True,
            "items": self._app_v2_merge_realtime_scope_defaults(
                rows,
                principal_id=principal_id,
                principal_type=principal_type,
                user_id=user_id,
            ),
        }

    def _app_v1_signal_select_list(self) -> str:
        actual_trigger_period_expr = self._app_v1_actual_trigger_period_expr()
        return f"""
               p.user_signal_projection_id,
               c.user_signal_card_id,
               NULL::bigint AS user_notification_queue_id,
               p.user_projection_run_id,
               p.user_projection_run_id AS projection_run_id,
               {self._app_v1_event_type_expr()} AS event_type,
               {self._app_v1_trade_date_expr()} AS trade_date,
               {self._app_v1_event_time_expr()} AS event_time,
               p.asset_kind,
               p.identity_key,
               p.code,
               p.code AS display_code,
               p.name,
               p.name AS display_name,
               COALESCE(c.board_code, p.board_code) AS industry_code,
               COALESCE(c.board_name, p.board_name) AS industry_name,
               p.direction,
               p.signal_type,
               {self._app_v1_action_state_expr()} AS action_state,
               COALESCE(NULLIF(p.action_mark, ''), NULLIF(c.action_mark, ''), '—') AS action_mark,
               c.card_status,
               {self._app_v1_blocked_reason_expr()} AS blocked_reason,
               p.list_payload_json->>'trigger_kind' AS trigger_kind,
               COALESCE(NULLIF(p.condition_key, ''), NULLIF(c.condition_key, ''), p.list_payload_json->>'condition_key') AS condition_key,
               COALESCE(NULLIF(p.original_condition_key, ''), NULLIF(c.original_condition_key, ''), p.list_payload_json->>'original_condition_key') AS original_condition_key,
               {actual_trigger_period_expr} AS primary_trigger_period,
               COALESCE(p.list_payload_json->>'trigger_time', p.created_at::text) AS trigger_time,
               'readonly'::text AS queue_status,
               'not_delivered'::text AS delivery_status,
               'b_track_reviewed_projection'::text AS notification_source,
               NULL::text AS channel,
               c.title,
               c.summary AS message,
               NULL::jsonb AS notification_payload_json,
               p.source_action_run_id,
               p.source_action_run_id AS source_run_id,
               p.source_event_id,
               COALESCE(NULLIF(p.source_action_event_id, ''), NULLIF(c.source_action_event_id, ''), p.source_event_id) AS source_action_event_id,
               {self._app_v1_event_type_expr()} AS source_action_event_type,
               p.list_payload_json->>'source_n4_run_id' AS source_n4_run_id,
               p.list_payload_json->>'n4_trigger_event_id' AS n4_trigger_event_id,
               COALESCE(p.list_payload_json->>'source_action_status', p.action_state, c.action_state) AS source_action_status,
               {self._app_v1_trigger_price_expr()} AS trigger_price,
               {self._app_v1_triggered_periods_expr(actual_trigger_period_expr)} AS triggered_periods,
               {self._app_v1_baseline_source_expr(actual_trigger_period_expr)} AS baseline_source,
               p.list_payload_json->'condition_projection_context' AS condition_projection_context,
               p.list_payload_json->>'condition_projection_context_status' AS condition_projection_context_status,
               p.list_payload_json->'condition_projection_context_trace' AS condition_projection_context_trace,
               p.list_payload_json->>'projection_message_contract_version' AS projection_message_contract_version,
               p.list_payload_json->>'projection_message_contract_hash' AS projection_message_contract_hash,
               p.list_payload_json->>'projection_message_status' AS projection_message_status,
               p.list_payload_json->'projection_message_not_ready_reasons' AS projection_message_not_ready_reasons,
               p.list_payload_json->>'trigger_pct' AS trigger_pct,
               p.list_payload_json->>'trigger_pct_status' AS trigger_pct_status,
               p.list_payload_json->>'action_price' AS action_price,
               p.list_payload_json->>'action_pct' AS action_pct,
               p.list_payload_json->>'action_pct_status' AS action_pct_status,
               p.list_payload_json->>'buy_expected_return_pct' AS buy_expected_return_pct,
               p.list_payload_json->>'up_secondary_expected_return_pct' AS up_secondary_expected_return_pct,
               p.list_payload_json->>'sell_expected_return_pct' AS sell_expected_return_pct,
               p.list_payload_json->>'up_reference_period' AS up_reference_period,
               p.list_payload_json->>'down_reference_period' AS down_reference_period,
               p.list_payload_json->'all_trigger_periods' AS all_trigger_periods,
               p.list_payload_json->>'score' AS score,
               p.list_payload_json->>'pe_core' AS pe_core,
               p.list_payload_json->>'industry_status' AS industry_status,
               p.list_payload_json->'industry_provenance' AS industry_provenance,
               COALESCE(c.target_price, p.target_price) AS target_price,
               COALESCE(c.current_price, p.current_price) AS current_price,
               COALESCE(c.expected_return_pct, p.expected_return_pct) AS expected_return_pct,
               COALESCE(c.board_code, p.board_code) AS board_code,
               COALESCE(c.board_name, p.board_name) AS board_name,
               p.source_display_table,
               p.source_condition_display_basis_id,
               p.source_condition_display_run_id,
               monitor_scope.source_type_raw,
               monitor_scope.source_type,
               monitor_scope.source_type_label,
               monitor_scope.source_object_kind,
               monitor_scope.source_object_identity_key,
               monitor_scope.source_object_code,
               monitor_scope.source_object_name,
               monitor_scope.membership_relation_date,
               NULL::text AS condition_display_cache_source,
               NULL::text AS membership_cache_source,
               COALESCE(p.list_payload_json->>'quality_status', 'reviewed') AS quality_status,
               true AS rollback_safe,
               p.created_at
        """

    def _app_v1_signal_detail_select_list(self) -> str:
        return f"""
               {self._app_v1_signal_select_list()},
               p.source_payload_json,
               p.display_payload_json,
               p.trace_json,
               c.card_payload_json
        """

    def _app_v1_signal_sse_select_list(self) -> str:
        actual_trigger_period_expr = self._app_v1_actual_trigger_period_expr()
        return f"""
               p.user_signal_projection_id,
               c.user_signal_card_id,
               p.user_projection_run_id,
               {self._app_v1_event_type_expr()} AS event_type,
               {self._app_v1_trade_date_expr()} AS trade_date,
               {self._app_v1_event_time_expr()} AS event_time,
               p.asset_kind,
               p.identity_key,
               p.code,
               p.code AS display_code,
               p.name,
               p.name AS display_name,
               COALESCE(c.board_code, p.board_code) AS industry_code,
               COALESCE(c.board_name, p.board_name) AS industry_name,
               p.direction,
               p.signal_type,
               {self._app_v1_action_state_expr()} AS action_state,
               COALESCE(NULLIF(p.action_mark, ''), NULLIF(c.action_mark, ''), '—') AS action_mark,
               {self._app_v1_blocked_reason_expr()} AS blocked_reason,
               COALESCE(NULLIF(p.condition_key, ''), NULLIF(c.condition_key, ''), p.list_payload_json->>'condition_key') AS condition_key,
               {actual_trigger_period_expr} AS primary_trigger_period,
               {self._app_v1_triggered_periods_expr(actual_trigger_period_expr)} AS triggered_periods,
               p.list_payload_json->'condition_projection_context' AS condition_projection_context,
               COALESCE(c.target_price, p.target_price) AS target_price,
               COALESCE(c.current_price, p.current_price) AS current_price,
               COALESCE(c.expected_return_pct, p.expected_return_pct) AS expected_return_pct,
               p.list_payload_json->>'buy_expected_return_pct' AS buy_expected_return_pct,
               p.list_payload_json->>'up_secondary_expected_return_pct' AS up_secondary_expected_return_pct,
               p.list_payload_json->>'sell_expected_return_pct' AS sell_expected_return_pct,
               p.list_payload_json->>'up_reference_period' AS up_reference_period,
               p.list_payload_json->>'down_reference_period' AS down_reference_period,
               p.list_payload_json->>'score' AS score,
               p.list_payload_json->>'pe_core' AS pe_core,
               {self._app_v1_trigger_price_expr()} AS trigger_price,
               p.list_payload_json->>'trigger_pct' AS trigger_pct,
               p.list_payload_json->>'action_price' AS action_price,
               p.list_payload_json->>'action_pct' AS action_pct,
               COALESCE(p.list_payload_json->>'projection_message_status', 'not_ready') AS projection_message_status,
               COALESCE(p.list_payload_json->>'quality_status', 'reviewed') AS quality_status,
               p.source_action_run_id AS source_run_id,
               p.user_projection_run_id AS projection_run_id,
               c.title,
               c.summary AS message,
               p.created_at
        """

    def _app_v1_signal_where(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        filters: dict[str, Any],
        include_monitor_scope: bool = True,
    ) -> tuple[str, dict[str, Any]]:
        params: dict[str, Any] = {
            "principal_id": int(principal_id),
            "principal_type": str(principal_type),
            "user_id": int(user_id),
            "trade_date": normalize_filter_value(filters.get("trade_date")) or "",
            "stale_source_action_run_ids": list(stale_source_action_run_ids()),
        }
        where_clauses = [
            """
            (
              p.source_action_run_id IS NULL
              OR NOT (p.source_action_run_id = ANY(%(stale_source_action_run_ids)s))
            )
            """,
            """
            p.projection_status IN ('visible', 'blocked')
            """,
        ]
        if include_monitor_scope:
            where_clauses.append(self._app_v1_effective_monitor_scope_clause())
        filter_specs = (
            ("event_type", self._app_v1_event_type_expr()),
            ("direction", "p.direction"),
            ("signal_type", "p.signal_type"),
            ("action_state", self._app_v1_action_state_expr()),
            ("blocked_reason", self._app_v1_blocked_reason_expr()),
        )
        asset_kinds = (
            app_message_asset_kinds(filters.get("asset_kinds"))
            if "asset_kinds" in filters
            else []
        )
        if asset_kinds:
            params["asset_kinds"] = asset_kinds
            where_clauses.append("p.asset_kind = ANY(%(asset_kinds)s)")
        else:
            asset_kind = normalize_filter_value(filters.get("asset_kind"))
            if asset_kind:
                params["asset_kind"] = asset_kind
                where_clauses.append("p.asset_kind = %(asset_kind)s")
        trade_date = normalize_filter_value(filters.get("trade_date"))
        if trade_date:
            params["trade_date"] = trade_date
            where_clauses.append(
                "p.for_trade_date = pg_catalog.to_date(pg_catalog.replace(%(trade_date)s, '-', ''), 'YYYYMMDD')"
            )
        before_created_at = filters.get("before_created_at")
        before_id = filters.get("before_id")
        if isinstance(before_created_at, datetime) and before_id is not None:
            params["before_created_at"] = before_created_at
            params["before_id"] = int(before_id)
            where_clauses.append(
                """
                (p.created_at, p.user_signal_projection_id)
                  < (%(before_created_at)s, %(before_id)s)
                """
            )
        for key, expression in filter_specs:
            value = normalize_filter_value(filters.get(key))
            if value:
                params[key] = value
                where_clauses.append(f"{expression} = %({key})s")

        time_field = normalize_time_field(filters.get("time_field"))
        time_expression = "p.created_at" if time_field == "created_at" else f"({self._app_v1_event_time_expr()})::timestamptz"
        date_from = normalize_filter_value(filters.get("date_from"))
        if date_from:
            params["date_from"] = date_from
            where_clauses.append(f"(({time_expression}) AT TIME ZONE 'Asia/Shanghai')::date >= %(date_from)s::date")
        date_to = normalize_filter_value(filters.get("date_to"))
        if date_to:
            params["date_to"] = date_to
            where_clauses.append(f"(({time_expression}) AT TIME ZONE 'Asia/Shanghai')::date <= %(date_to)s::date")

        keyword = normalize_filter_value(filters.get("q"))
        if keyword:
            params["q_like"] = f"%{keyword}%"
            condition_expr = "COALESCE(NULLIF(p.condition_key, ''), NULLIF(c.condition_key, ''), p.list_payload_json->>'condition_key')"
            where_clauses.append(
                f"""
                (
                  p.code ILIKE %(q_like)s
                  OR p.name ILIKE %(q_like)s
                  OR p.identity_key ILIKE %(q_like)s
                  OR p.source_event_id ILIKE %(q_like)s
                  OR p.source_action_event_id ILIKE %(q_like)s
                  OR {condition_expr} ILIKE %(q_like)s
                )
                """
            )
        return "\n                  AND ".join(where_clauses), params

    def _app_v1_signal_scope_relations_ready(self) -> bool:
        required_monitor_tables = (
            "user_monitor_stock",
            "user_monitor_index",
            "user_monitor_board",
        )
        required_display_views = (
            "v_n6_stock_condition_display_basis",
            "v_n6_index_condition_display_basis",
            "v_n6_board_condition_display_basis",
        )
        capabilities = self._app_v1_signal_schema_capabilities()
        monitor_lineage_columns = {
            "monitor_id",
            "principal_id",
            "principal_type",
            "user_id",
            "identity_key",
            "direction",
            "source_type",
            "source_run_id",
            "source_snapshot_json",
            "valid_source_trade_date",
            "valid_for_trade_date",
            "valid_source_run_id",
            "status",
        }
        approved_batch_columns = {"identity_key", "source_trade_date", "for_trade_date", "run_id"}
        return all(
            monitor_lineage_columns.issubset(capabilities.get(table_name, frozenset()))
            for table_name in required_monitor_tables
        ) and all(
            approved_batch_columns.issubset(capabilities.get(view_name, frozenset()))
            for view_name in required_display_views
        )

    def _app_v1_signal_schema_capabilities(self) -> dict[str, frozenset[str]]:
        cached = self._app_v1_signal_schema_capability_cache
        if cached is not None:
            return cached
        monitor_tables = tuple(APP_V2_MONITOR_TABLE_BY_ASSET.values())
        display_views = (
            "v_n6_stock_condition_display_basis",
            "v_n6_index_condition_display_basis",
            "v_n6_board_condition_display_basis",
        )
        relation_names = (
            *monitor_tables,
            *display_views,
            APP_REALTIME_SCOPE_TABLE,
            "n6_virtual_account",
            "n6_virtual_position",
        )
        with self._app_v1_signal_schema_capability_lock:
            cached = self._app_v1_signal_schema_capability_cache
            if cached is not None:
                return cached
            columns_by_relation: dict[str, set[str]] = {}
            with self._readonly_connection() as conn, conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT c.relname AS relation_name,
                           a.attname AS column_name
                    FROM pg_catalog.pg_class c
                    JOIN pg_catalog.pg_namespace n
                      ON n.oid = c.relnamespace
                    LEFT JOIN pg_catalog.pg_attribute a
                      ON a.attrelid = c.oid
                     AND a.attnum > 0
                     AND NOT a.attisdropped
                    WHERE n.nspname = current_schema()
                      AND c.relname = ANY(%s)
                      AND c.relkind IN ('r', 'p', 'v', 'm', 'f')
                    ORDER BY c.relname, a.attnum
                    """,
                    (list(relation_names),),
                )
                for row in cur.fetchall():
                    relation_name = str(row.get("relation_name") or "").strip()
                    column_name = str(row.get("column_name") or "").strip()
                    if not relation_name:
                        continue
                    columns_by_relation.setdefault(relation_name, set())
                    if column_name:
                        columns_by_relation[relation_name].add(column_name)
            snapshot = {
                relation_name: frozenset(columns)
                for relation_name, columns in columns_by_relation.items()
            }
            for relation_name in relation_names:
                self._app_v2_relation_existence_cache[relation_name] = relation_name in snapshot
            for table_name in monitor_tables:
                self._app_v2_monitor_column_cache[table_name] = set(snapshot.get(table_name, frozenset()))
            for view_name in display_views:
                self._app_v2_filter_column_cache[view_name] = set(snapshot.get(view_name, frozenset()))
            self._app_v1_signal_schema_capability_cache = snapshot
            return snapshot

    def _app_v1_effective_monitor_scope_clause(self) -> str:
        message_trade_date_expr = self._app_v1_trade_date_expr()
        return f"""
            EXISTS (
              SELECT 1
              FROM effective_monitor_scope
              WHERE effective_monitor_scope.asset_kind = p.asset_kind
                AND effective_monitor_scope.identity_key = p.identity_key
                AND effective_monitor_scope.direction = p.direction
                AND effective_monitor_scope.valid_for_trade_date = ({message_trade_date_expr})
            )
        """

    def _app_v1_effective_monitor_scope_lookup(self, column_name: str) -> str:
        message_trade_date_expr = self._app_v1_trade_date_expr()
        return f"""
            (
              SELECT effective_monitor_scope.{column_name}
              FROM effective_monitor_scope
              WHERE effective_monitor_scope.asset_kind = p.asset_kind
                AND effective_monitor_scope.identity_key = p.identity_key
                AND effective_monitor_scope.direction = p.direction
                AND effective_monitor_scope.valid_for_trade_date = ({message_trade_date_expr})
              ORDER BY effective_monitor_scope.monitor_id DESC
              LIMIT 1
            )
        """

    def _app_v1_effective_monitor_scope_join(self) -> str:
        message_trade_date_expr = self._app_v1_trade_date_expr()
        return f"""
            JOIN LATERAL (
              SELECT effective_monitor_scope.source_type_raw,
                     effective_monitor_scope.source_type,
                     effective_monitor_scope.source_type_label,
                     effective_monitor_scope.source_object_kind,
                     effective_monitor_scope.source_object_identity_key,
                     effective_monitor_scope.source_object_code,
                     effective_monitor_scope.source_object_name,
                     effective_monitor_scope.membership_relation_date
              FROM effective_monitor_scope
              WHERE effective_monitor_scope.asset_kind = p.asset_kind
                AND effective_monitor_scope.identity_key = p.identity_key
                AND effective_monitor_scope.direction = p.direction
                AND effective_monitor_scope.valid_for_trade_date = ({message_trade_date_expr})
              ORDER BY effective_monitor_scope.monitor_id DESC
              LIMIT 1
            ) monitor_scope ON TRUE
        """

    def _app_v1_web_signal_scope_join(self) -> str:
        message_trade_date_expr = self._app_v1_trade_date_expr()
        return f"""
            JOIN deduplicated_monitor_scope monitor_scope
              ON monitor_scope.asset_kind = p.asset_kind
             AND monitor_scope.identity_key = p.identity_key
             AND monitor_scope.direction = p.direction
             AND monitor_scope.valid_for_trade_date = ({message_trade_date_expr})
        """

    def _app_v1_web_signal_scope_cte(
        self,
        *,
        include_expired: bool = False,
        include_realtime_scope: bool = True,
        use_current_approved_batch: bool = True,
    ) -> str:
        effective_scope_cte = self._app_v1_effective_monitor_scope_cte(
            include_expired=include_expired,
            include_realtime_scope=include_realtime_scope,
            use_current_approved_batch=use_current_approved_batch,
        )
        return f"""
        {effective_scope_cte},
        deduplicated_monitor_scope AS MATERIALIZED (
          SELECT DISTINCT ON (
                   asset_kind,
                   identity_key,
                   direction,
                   valid_for_trade_date
                 )
                 monitor_id,
                 principal_id,
                 principal_type,
                 asset_kind,
                 identity_key,
                 direction,
                 valid_source_trade_date,
                 valid_for_trade_date,
                 valid_source_run_id,
                 source_type_raw,
                 source_type,
                 source_type_label,
                 source_object_kind,
                 source_object_identity_key,
                 source_object_code,
                 source_object_name,
                 membership_relation_date
          FROM effective_monitor_scope
          ORDER BY asset_kind,
                   identity_key,
                   direction,
                   valid_for_trade_date,
                   monitor_id DESC
        )
        """

    def _app_v1_effective_monitor_scope_cte(
        self,
        *,
        include_expired: bool = False,
        include_realtime_scope: bool = True,
        use_current_approved_batch: bool = True,
    ) -> str:
        realtime_scope_union = ""
        if include_realtime_scope and not include_expired and self._app_v2_relation_exists(APP_REALTIME_SCOPE_TABLE):
            realtime_scope_union = f"""
          UNION ALL
          {self._app_v1_realtime_scope_select()}
            """
        holding_scope_union = ""
        capabilities = self._app_v1_signal_schema_capabilities()
        if not include_expired and {"n6_virtual_account", "n6_virtual_position"}.issubset(capabilities):
            holding_scope_union = f"""
          UNION ALL
          {self._app_v1_holding_scope_select()}
            """
        current_approved_batch_ctes = ""
        if use_current_approved_batch:
            current_approved_batch_ctes = f"""
        {self._app_v1_current_approved_batch_cte(
            asset_kind="stock",
            view_name="v_n6_stock_condition_display_basis",
        )},
        {self._app_v1_current_approved_batch_cte(
            asset_kind="index",
            view_name="v_n6_index_condition_display_basis",
        )},
        {self._app_v1_current_approved_batch_cte(
            asset_kind="board",
            view_name="v_n6_board_condition_display_basis",
        )},
            """
        return f"""
        {current_approved_batch_ctes}
        {self._app_v1_principal_monitor_cte(
            asset_kind="stock",
            table_name="user_monitor_stock",
            include_expired=include_expired,
        )},
        {self._app_v1_principal_monitor_cte(
            asset_kind="index",
            table_name="user_monitor_index",
            include_expired=include_expired,
        )},
        {self._app_v1_principal_monitor_cte(
            asset_kind="board",
            table_name="user_monitor_board",
            include_expired=include_expired,
        )},
        effective_monitor_scope AS (
          {self._app_v1_effective_monitor_scope_select(
              asset_kind="stock",
              table_name="user_monitor_stock",
              view_name="v_n6_stock_condition_display_basis",
              use_current_approved_batch=use_current_approved_batch,
          )}
          UNION ALL
          {self._app_v1_effective_monitor_scope_select(
              asset_kind="index",
              table_name="user_monitor_index",
              view_name="v_n6_index_condition_display_basis",
              use_current_approved_batch=use_current_approved_batch,
          )}
          UNION ALL
          {self._app_v1_effective_monitor_scope_select(
              asset_kind="board",
              table_name="user_monitor_board",
              view_name="v_n6_board_condition_display_basis",
              use_current_approved_batch=use_current_approved_batch,
          )}
          {realtime_scope_union}
          {holding_scope_union}
        )
        """

    def _app_v1_principal_monitor_cte(
        self,
        *,
        asset_kind: str,
        table_name: str,
        include_expired: bool,
    ) -> str:
        status_clause = "status <> 'removed'" if include_expired else "status = 'active'"
        return f"""
        current_{asset_kind}_principal_monitors AS MATERIALIZED (
          SELECT monitor_id,
                 principal_id,
                 principal_type,
                 user_id,
                 identity_key,
                 direction,
                 source_type,
                 source_run_id,
                 source_snapshot_json,
                 valid_source_trade_date,
                 valid_for_trade_date,
                 valid_source_run_id,
                 status
          FROM {table_name}
          WHERE principal_id = %(principal_id)s
            AND principal_type = %(principal_type)s
            AND user_id = %(user_id)s
            AND {status_clause}
            AND ('{asset_kind}' <> 'stock' OR direction = 'buy')
        )
        """

    def _app_v1_current_approved_batch_cte(self, *, asset_kind: str, view_name: str) -> str:
        return f"""
        current_{asset_kind}_approved_batch AS MATERIALIZED (
          SELECT min(source_trade_date::text) AS source_trade_date,
                 min(for_trade_date::text) AS for_trade_date,
                 min(run_id::text) AS source_run_id
          FROM {view_name}
          WHERE for_trade_date = (SELECT max(for_trade_date) FROM {view_name})
          HAVING count(*) > 0
             AND count(source_trade_date) = count(*)
             AND count(for_trade_date) = count(*)
             AND count(run_id) = count(*)
             AND count(DISTINCT (source_trade_date::text, for_trade_date::text, run_id::text)) = 1
        )
        """

    def _app_v1_all_monitor_scope_cte(self) -> str:
        return f"""
        all_monitor_scope AS (
          {self._app_v1_all_monitor_scope_select(
              asset_kind="stock",
              table_name="user_monitor_stock",
          )}
          UNION ALL
          {self._app_v1_all_monitor_scope_select(
              asset_kind="index",
              table_name="user_monitor_index",
          )}
          UNION ALL
          {self._app_v1_all_monitor_scope_select(
              asset_kind="board",
              table_name="user_monitor_board",
          )}
        )
        """

    def _app_v1_monitor_validity_sql_exprs(
        self,
        table_name: str,
        alias: str = "m",
    ) -> tuple[str, str, str, str]:
        columns = self._app_v2_monitor_columns(table_name)
        expressions = tuple(
            f"{alias}.{column_name}::text" if column_name in columns else "NULL::text"
            for column_name in (
                "source_run_id",
                "valid_source_trade_date",
                "valid_for_trade_date",
                "valid_source_run_id",
            )
        )
        return expressions  # type: ignore[return-value]

    def _app_v1_effective_monitor_scope_select(
        self,
        *,
        asset_kind: str,
        table_name: str,
        view_name: str,
        use_current_approved_batch: bool = True,
    ) -> str:
        source_run_expr, valid_source_trade_expr, valid_for_trade_expr, valid_run_expr = (
            self._app_v1_monitor_validity_sql_exprs(table_name)
        )
        source_type_raw_expr = "COALESCE(NULLIF(m.source_type, ''), 'single_row')"
        source_type_expr = f"""
            CASE {source_type_raw_expr}
              WHEN 'single_row' THEN 'direct'
              WHEN 'direct' THEN 'direct'
              WHEN 'index_linked_stock' THEN 'index_linked_stock'
              WHEN 'board_linked_stock' THEN 'board_linked_stock'
              ELSE {source_type_raw_expr}
            END
        """
        scope_gate_sql = f"""
          WHERE NULLIF({valid_for_trade_expr}, '') IS NOT NULL
        """
        if use_current_approved_batch:
            scope_gate_sql = f"""
          JOIN current_{asset_kind}_approved_batch current_batch
            ON NULLIF({source_run_expr}, '') = current_batch.source_run_id
           AND NULLIF({valid_source_trade_expr}, '') = current_batch.source_trade_date
           AND NULLIF({valid_for_trade_expr}, '') = current_batch.for_trade_date
           AND NULLIF({valid_run_expr}, '') = current_batch.source_run_id
          WHERE EXISTS (
            SELECT 1
            FROM {view_name} approved
            WHERE approved.identity_key = m.identity_key
              AND approved.source_trade_date::text = current_batch.source_trade_date
              AND approved.for_trade_date::text = current_batch.for_trade_date
              AND approved.run_id::text = current_batch.source_run_id
          )
            """
        return f"""
          SELECT m.monitor_id,
                 m.principal_id,
                 m.principal_type,
                 '{asset_kind}'::text AS asset_kind,
                 m.identity_key,
                 m.direction,
                 {valid_source_trade_expr} AS valid_source_trade_date,
                 {valid_for_trade_expr} AS valid_for_trade_date,
                 {valid_run_expr} AS valid_source_run_id,
                 {source_type_raw_expr} AS source_type_raw,
                 {source_type_expr} AS source_type,
                 CASE {source_type_expr}
                   WHEN 'direct' THEN '直接加入'
                   WHEN 'index_linked_stock' THEN '来源指数'
                   WHEN 'board_linked_stock' THEN '来源板块'
                   ELSE {source_type_expr}
                 END AS source_type_label,
                 CASE
                   WHEN {source_type_expr} = 'direct' THEN 'none'
                   ELSE COALESCE(
                     NULLIF(m.source_snapshot_json->>'parent_asset_kind', ''),
                     CASE {source_type_expr}
                       WHEN 'index_linked_stock' THEN 'index'
                       WHEN 'board_linked_stock' THEN 'board'
                       ELSE 'none'
                     END
                   )
                 END AS source_object_kind,
                 CASE WHEN {source_type_expr} = 'direct' THEN NULL ELSE NULLIF(m.source_snapshot_json->>'parent_identity_key', '') END AS source_object_identity_key,
                 CASE WHEN {source_type_expr} = 'direct' THEN NULL ELSE NULLIF(m.source_snapshot_json->>'parent_code', '') END AS source_object_code,
                 CASE WHEN {source_type_expr} = 'direct' THEN NULL ELSE NULLIF(m.source_snapshot_json->>'parent_name', '') END AS source_object_name,
                 CASE WHEN {source_type_expr} = 'direct' THEN NULL ELSE NULLIF(m.source_snapshot_json->>'membership_trade_date', '') END AS membership_relation_date
          FROM current_{asset_kind}_principal_monitors m
          {scope_gate_sql}
        """

    def _app_v1_realtime_scope_select(self) -> str:
        return f"""
          {self._app_v1_explicit_realtime_scope_select()}
          UNION ALL
          {self._app_v1_default_realtime_scope_seed_select()}
        """

    def _app_v1_holding_scope_select(self) -> str:
        return """
          SELECT NULL::bigint AS monitor_id,
                 a.principal_id,
                 a.principal_type,
                 'stock'::text AS asset_kind,
                 p.identity_key,
                 holding_direction.direction,
                 NULL::text AS valid_source_trade_date,
                 %(trade_date)s::text AS valid_for_trade_date,
                 NULL::text AS valid_source_run_id,
                 'virtual_position'::text AS source_type_raw,
                 'virtual_position'::text AS source_type,
                 '虚拟持仓'::text AS source_type_label,
                 'none'::text AS source_object_kind,
                 NULL::text AS source_object_identity_key,
                 NULL::text AS source_object_code,
                 NULL::text AS source_object_name,
                 NULL::text AS membership_relation_date
          FROM n6_virtual_account a
          JOIN n6_virtual_position p
            ON p.virtual_account_id = a.virtual_account_id
           AND p.principal_id = a.principal_id
           AND p.principal_type = a.principal_type
          CROSS JOIN (VALUES ('buy'::text), ('sell'::text)) AS holding_direction(direction)
          WHERE a.principal_id = %(principal_id)s
            AND a.principal_type = %(principal_type)s
            AND a.virtual_account_status = 'active'
            AND p.asset_kind = 'stock'
            AND p.position_status = 'open_virtual'
            AND p.quantity > 0
            AND NULLIF(%(trade_date)s, '') IS NOT NULL
        """

    def _app_v1_explicit_realtime_scope_select(self) -> str:
        return f"""
          SELECT s.realtime_scope_id AS monitor_id,
                 s.principal_id,
                 s.principal_type,
                 s.asset_kind,
                 s.identity_key,
                 scope_direction.direction,
                 NULL::text AS valid_source_trade_date,
                 %(trade_date)s::text AS valid_for_trade_date,
                 NULL::text AS valid_source_run_id,
                 COALESCE(NULLIF(s.source_type, ''), 'realtime_scope') AS source_type_raw,
                 'realtime_scope'::text AS source_type,
                 '实时监控范围'::text AS source_type_label,
                 'none'::text AS source_object_kind,
                 NULL::text AS source_object_identity_key,
                 NULL::text AS source_object_code,
                 NULL::text AS source_object_name,
                 NULL::text AS membership_relation_date
          FROM {APP_REALTIME_SCOPE_TABLE} s
          CROSS JOIN LATERAL unnest(
            CASE
              WHEN s.asset_kind = 'stock' THEN ARRAY['buy'::text]
              ELSE ARRAY['buy'::text, 'sell'::text]
            END
          ) AS scope_direction(direction)
          WHERE s.principal_id = %(principal_id)s
            AND s.principal_type = %(principal_type)s
            AND s.user_id = %(user_id)s
            AND s.status = 'active'
            AND NULLIF(%(trade_date)s, '') IS NOT NULL
        """

    def _app_v1_default_realtime_scope_seed_select(self) -> str:
        values_sql = ",\n                 ".join(
            f"({_sql_text_literal(item['asset_kind'])}, {_sql_text_literal(item['identity_key'])}, {_sql_text_literal(item['display_name'])})"
            for item in DEFAULT_REALTIME_SCOPE_INDEXES
        )
        return f"""
          SELECT NULL::bigint AS monitor_id,
                 %(principal_id)s::bigint AS principal_id,
                 %(principal_type)s::text AS principal_type,
                 d.asset_kind,
                 d.identity_key,
                 scope_direction.direction,
                 NULL::text AS valid_source_trade_date,
                 %(trade_date)s::text AS valid_for_trade_date,
                 NULL::text AS valid_source_run_id,
                 'default_seed'::text AS source_type_raw,
                 'realtime_scope'::text AS source_type,
                 '实时监控范围'::text AS source_type_label,
                 'none'::text AS source_object_kind,
                 NULL::text AS source_object_identity_key,
                 NULL::text AS source_object_code,
                 NULL::text AS source_object_name,
                 NULL::text AS membership_relation_date
          FROM (
                 VALUES
                 {values_sql}
          ) AS d(asset_kind, identity_key, display_name)
          CROSS JOIN LATERAL unnest(
            CASE
              WHEN d.asset_kind = 'stock' THEN ARRAY['buy'::text]
              ELSE ARRAY['buy'::text, 'sell'::text]
            END
          ) AS scope_direction(direction)
          WHERE NULLIF(%(trade_date)s, '') IS NOT NULL
            AND NOT EXISTS (
              SELECT 1
              FROM {APP_REALTIME_SCOPE_TABLE} deleted_scope
              WHERE deleted_scope.principal_id = %(principal_id)s
                AND deleted_scope.principal_type = %(principal_type)s
                AND deleted_scope.user_id = %(user_id)s
                AND deleted_scope.asset_kind = d.asset_kind
                AND deleted_scope.identity_key = d.identity_key
                AND deleted_scope.status = 'deleted'
            )
            AND NOT EXISTS (
              SELECT 1
              FROM {APP_REALTIME_SCOPE_TABLE} active_scope
              WHERE active_scope.principal_id = %(principal_id)s
                AND active_scope.principal_type = %(principal_type)s
                AND active_scope.user_id = %(user_id)s
                AND active_scope.asset_kind = d.asset_kind
                AND active_scope.identity_key = d.identity_key
                AND active_scope.status = 'active'
            )
            -- n6_default_realtime_monitor_scope_v1
        """

    def _app_v1_all_monitor_scope_select(self, *, asset_kind: str, table_name: str) -> str:
        _, valid_source_trade_expr, valid_for_trade_expr, valid_run_expr = self._app_v1_monitor_validity_sql_exprs(
            table_name
        )
        user_id_clause = "AND m.user_id = %(user_id)s" if "user_id" in self._app_v2_monitor_columns(table_name) else ""
        return f"""
          SELECT m.monitor_id,
                 m.principal_id,
                 m.principal_type,
                 '{asset_kind}'::text AS asset_kind,
                 m.identity_key,
                 m.direction,
                 m.status,
                 {valid_source_trade_expr} AS valid_source_trade_date,
                 {valid_for_trade_expr} AS valid_for_trade_date,
                 {valid_run_expr} AS valid_source_run_id
          FROM {table_name} m
          WHERE m.principal_id = %(principal_id)s
            AND m.principal_type = %(principal_type)s
            {user_id_clause}
            AND m.status <> 'removed'
        """

    def _app_v1_trade_date_expr(self) -> str:
        return "pg_catalog.to_char(p.for_trade_date, 'YYYYMMDD')"

    def _app_v1_event_type_expr(self) -> str:
        return "COALESCE(NULLIF(p.source_action_event_type, ''), NULLIF(c.source_action_event_type, ''), p.source_event_type)"

    def _app_v1_event_time_expr(self) -> str:
        return "COALESCE(p.list_payload_json->>'event_time', p.created_at::text)"

    def _app_v1_action_state_expr(self) -> str:
        return """
            COALESCE(
              NULLIF(p.action_state, ''),
              NULLIF(c.action_state, ''),
              CASE
                WHEN p.source_action_event_type = 'ActionBlocked' THEN 'blocked'
                WHEN p.source_action_event_type = 'ActionExecuted' THEN 'executed'
                WHEN p.source_action_event_type = 'ActionEligible' THEN 'eligible'
                WHEN p.source_action_event_type = 'ActionSkipped' THEN 'skipped'
                ELSE NULL
              END,
              p.projection_status,
              c.card_status
            )
        """

    def _app_v1_blocked_reason_expr(self) -> str:
        return "p.list_payload_json->>'blocked_reason'"

    def _app_v1_actual_trigger_period_expr(self) -> str:
        return "NULLIF(p.list_payload_json->>'primary_trigger_period', '')"

    def _app_v1_trigger_price_expr(self) -> str:
        return "p.list_payload_json->>'trigger_price'"

    def _app_v1_triggered_periods_expr(self, actual_trigger_period_expr: str) -> str:
        return f"""
            COALESCE(
              NULLIF(NULLIF(p.list_payload_json->>'triggered_periods', ''), '[]'),
              CASE
                WHEN {actual_trigger_period_expr} IS NOT NULL
                THEN pg_catalog.jsonb_build_array({actual_trigger_period_expr})::text
                ELSE NULL
              END
            )
        """

    def _app_v1_baseline_source_expr(self, actual_trigger_period_expr: str) -> str:
        _ = actual_trigger_period_expr
        return "p.list_payload_json->>'baseline_source'"

    def _app_v2_filter_source_identity_keys(self, filters: dict[str, Any]) -> list[str]:
        return normalize_filter_identity_values(
            [
                *normalize_filter_values(filters.get("source_identity_keys")),
                *normalize_filter_values(filters.get("source_identity_key")),
            ]
        )

    def _app_v2_stock_source_context(
        self,
        cur: Any,
        *,
        filters: dict[str, Any],
        selected_for_trade_date: str,
    ) -> dict[str, Any] | None:
        source_asset_type = normalize_filter_value(filters.get("source_asset_type"))
        source_identity_keys = self._app_v2_filter_source_identity_keys(filters)
        if not source_asset_type and not source_identity_keys:
            return {
                "source_asset_type": None,
                "source_identity_keys": [],
                "source_display_names": [],
                "membership_source_table": None,
                "membership_trade_date": None,
                "membership_count": 0,
                "matched_stock_count": 0,
                "for_trade_date": selected_for_trade_date,
                "empty_state": "",
                "fallback_used": False,
                "source_filter_active": False,
            }
        membership_table_by_type = {
            "index": "v_n6_index_membership_fact",
            "board": "v_n6_board_membership_fact",
        }
        parent_table_by_type = {
            "index": "v_n6_index_condition_display_basis",
            "board": "v_n6_board_condition_display_basis",
        }
        parent_column_by_type = {
            "index": "index_identity_key",
            "board": "board_identity_key",
        }
        parent_name_column_by_type = {
            "index": "index_name",
            "board": "board_name",
        }
        source_table = membership_table_by_type.get(source_asset_type or "")
        parent_table = parent_table_by_type.get(source_asset_type or "")
        parent_column = parent_column_by_type.get(source_asset_type or "")
        parent_name_column = parent_name_column_by_type.get(source_asset_type or "")
        if (
            source_asset_type not in membership_table_by_type
            or not source_identity_keys
            or source_table is None
            or parent_table is None
            or parent_column is None
            or parent_name_column is None
        ):
            return {
                "source_asset_type": source_asset_type,
                "source_identity_keys": source_identity_keys,
                "source_display_names": [],
                "membership_source_table": source_table,
                "membership_trade_date": None,
                "membership_count": 0,
                "matched_stock_count": 0,
                "for_trade_date": selected_for_trade_date,
                "empty_state": "非法来源类型" if source_asset_type not in membership_table_by_type else "暂无来源对象",
                "fallback_used": False,
                "source_filter_active": True,
                "_stock_identity_keys": [],
            }
        if not self._app_v2_relation_exists(source_table) or not self._app_v2_relation_exists(parent_table):
            return {
                "source_asset_type": source_asset_type,
                "source_identity_keys": source_identity_keys,
                "source_display_names": [],
                "membership_source_table": source_table,
                "membership_trade_date": None,
                "membership_count": 0,
                "matched_stock_count": 0,
                "for_trade_date": selected_for_trade_date,
                "empty_state": "暂无成分股",
                "fallback_used": False,
                "source_filter_active": True,
                "_stock_identity_keys": [],
            }
        cur.execute(
            f"""
            SELECT identity_key,
                   COALESCE(NULLIF(display_name, ''), identity_key) AS display_name,
                   max(source_trade_date::text) AS source_trade_date
            FROM {parent_table}
            WHERE identity_key = ANY(%(source_identity_keys)s)
              AND for_trade_date = %(selected_for_trade_date)s
            GROUP BY identity_key, COALESCE(NULLIF(display_name, ''), identity_key)
            """,
            {
                "source_identity_keys": source_identity_keys,
                "selected_for_trade_date": selected_for_trade_date,
            },
        )
        parent_rows = [dict(row) for row in cur.fetchall()]
        parent_display_by_key = {
            str(row.get("identity_key") or "").strip(): str(row.get("display_name") or row.get("identity_key") or "").strip()
            for row in parent_rows
            if str(row.get("identity_key") or "").strip()
        }
        parent_source_dates = sorted(
            {
                str(row.get("source_trade_date") or "").strip()
                for row in parent_rows
                if str(row.get("source_trade_date") or "").strip()
            }
        )
        parent_source_trade_date = parent_source_dates[-1] if parent_source_dates else None
        fallback_used = parent_source_trade_date is None
        if parent_source_trade_date:
            cur.execute(
                f"""
                SELECT max(trade_date::text) AS membership_trade_date
                FROM {source_table}
                WHERE {parent_column} = ANY(%(source_identity_keys)s)
                  AND trade_date::text <= %(parent_source_trade_date)s
                """,
                {
                    "source_identity_keys": source_identity_keys,
                    "parent_source_trade_date": parent_source_trade_date,
                },
            )
        else:
            cur.execute(
                f"""
                SELECT max(trade_date::text) AS membership_trade_date
                FROM {source_table}
                WHERE {parent_column} = ANY(%(source_identity_keys)s)
                """,
                {"source_identity_keys": source_identity_keys},
            )
        membership_trade_date = normalize_filter_value((cur.fetchone() or {}).get("membership_trade_date"))
        if not membership_trade_date and parent_source_trade_date:
            fallback_used = True
            cur.execute(
                f"""
                SELECT max(trade_date::text) AS membership_trade_date
                FROM {source_table}
                WHERE {parent_column} = ANY(%(source_identity_keys)s)
                """,
                {"source_identity_keys": source_identity_keys},
            )
            membership_trade_date = normalize_filter_value((cur.fetchone() or {}).get("membership_trade_date"))
        if membership_trade_date:
            cur.execute(
                f"""
                SELECT {parent_column} AS parent_identity_key,
                       {parent_name_column} AS parent_name,
                       stock_identity_key
                FROM {source_table}
                WHERE {parent_column} = ANY(%(source_identity_keys)s)
                  AND trade_date::text = %(membership_trade_date)s
                ORDER BY {parent_column} ASC, stock_identity_key ASC
                """,
                {
                    "source_identity_keys": source_identity_keys,
                    "membership_trade_date": membership_trade_date,
                },
            )
            membership_rows = [dict(row) for row in cur.fetchall()]
        else:
            membership_rows = []
        for row in membership_rows:
            parent_key = str(row.get("parent_identity_key") or "").strip()
            if parent_key and parent_key not in parent_display_by_key:
                parent_display_by_key[parent_key] = str(row.get("parent_name") or parent_key).strip()
        stock_identity_keys = list(
            dict.fromkeys(
                str(row.get("stock_identity_key") or "").strip()
                for row in membership_rows
                if str(row.get("stock_identity_key") or "").strip()
            )
        )
        return {
            "source_asset_type": source_asset_type,
            "source_identity_keys": source_identity_keys,
            "source_display_names": [
                parent_display_by_key.get(identity_key, identity_key)
                for identity_key in source_identity_keys
            ],
            "membership_source_table": source_table,
            "membership_trade_date": membership_trade_date,
            "membership_count": len(stock_identity_keys),
            "matched_stock_count": 0,
            "for_trade_date": selected_for_trade_date,
            "empty_state": "暂无成分股" if not stock_identity_keys else "",
            "fallback_used": fallback_used,
            "source_filter_active": True,
            "_stock_identity_keys": stock_identity_keys,
        }

    def fetch_app_filter_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        filters: dict[str, Any],
        limit: int,
    ) -> dict[str, Any]:
        table_by_asset = {
            "stock": "v_n6_stock_condition_display_basis",
            "index": "v_n6_index_condition_display_basis",
            "board": "v_n6_board_condition_display_basis",
        }
        table_name = table_by_asset.get(asset_kind)
        if table_name is None or not self._app_v2_relation_exists(table_name):
            return {"cache_ready": False, "items": [], "total_count": 0, "filtered_count": 0, "returned_count": 0}
        where_sql, params = self._app_v2_filter_where(filters, asset_kind=asset_kind)
        limit_cap = APP_V2_FILTER_PAGE_MAX_ROWS
        params["limit"] = max(1, min(int(limit), limit_cap))
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT for_trade_date::text AS for_trade_date,
                       max(run_id::text) AS source_run_id
                FROM {table_name}
                WHERE for_trade_date IS NOT NULL
                GROUP BY for_trade_date
                ORDER BY for_trade_date DESC
                LIMIT 120
                """
            )
            batch_rows = [dict(row) for row in cur.fetchall()]
            available_for_trade_dates = [
                str(row.get("for_trade_date") or "").strip()
                for row in batch_rows
                if str(row.get("for_trade_date") or "").strip()
            ]
            selected_for_trade_date = normalize_filter_value(filters.get("for_trade_date"))
            if not selected_for_trade_date:
                selected_for_trade_date = available_for_trade_dates[0] if available_for_trade_dates else None
            if not selected_for_trade_date:
                return {
                    "cache_ready": True,
                    "items": [],
                    "total_count": 0,
                    "filtered_count": 0,
                    "returned_count": 0,
                    "available_for_trade_dates": available_for_trade_dates,
                    "selected_for_trade_date": "",
                }
            source_run_id = next(
                (
                    str(row.get("source_run_id") or "").strip()
                    for row in batch_rows
                    if str(row.get("for_trade_date") or "").strip() == selected_for_trade_date
                ),
                "",
            )
            cache_key = (
                asset_kind,
                selected_for_trade_date,
                source_run_id,
                json.dumps(filters, ensure_ascii=False, sort_keys=True, default=str),
                int(params["limit"]),
            )
            cached_result = self._app_shared_cache_get(self._app_filter_result_cache, cache_key)
            if cached_result is not None:
                return cached_result
            params["selected_for_trade_date"] = selected_for_trade_date
            source_context = (
                self._app_v2_stock_source_context(
                    cur,
                    filters=filters,
                    selected_for_trade_date=selected_for_trade_date,
                )
                if asset_kind == "stock"
                else None
            )
            source_where_sql = "TRUE"
            if source_context and source_context.get("source_filter_active"):
                source_stock_identity_keys = list(source_context.get("_stock_identity_keys") or [])
                if source_stock_identity_keys:
                    params["source_stock_identity_keys"] = source_stock_identity_keys
                    source_where_sql = "identity_key = ANY(%(source_stock_identity_keys)s)"
                else:
                    source_where_sql = "FALSE"
            level_up_recommendation = self._app_v2_filter_level_up_recommendation_context(
                cur,
                filters=filters,
                asset_kind=asset_kind,
                params=params,
                selected_for_trade_date=selected_for_trade_date,
            )
            recommendation_where_sql = str(level_up_recommendation.get("_where_sql") or "TRUE")
            level_up_recommendation.pop("_where_sql", None)
            stock_industry_join_sql = (
                self._app_v2_stock_industry_join_sql()
                if asset_kind == "stock"
                else ""
            )
            count_stock_industry_join_sql = (
                stock_industry_join_sql
                if asset_kind == "stock" and normalize_filter_value(filters.get("q"))
                else ""
            )
            cur.execute(
                f"""
                SELECT count(*) FILTER (WHERE {source_where_sql})::int AS total_count,
                       count(*) FILTER (WHERE {source_where_sql} AND {where_sql} AND {recommendation_where_sql})::int AS filtered_count,
                       avg(level_up_score) FILTER (
                         WHERE {source_where_sql}
                           AND {where_sql}
                           AND {recommendation_where_sql}
                           AND level_up_score IS NOT NULL
                           AND level_down_score IS NOT NULL
                       ) AS level_up_score_avg,
                       avg(level_down_score) FILTER (
                         WHERE {source_where_sql}
                           AND {where_sql}
                           AND {recommendation_where_sql}
                           AND level_up_score IS NOT NULL
                           AND level_down_score IS NOT NULL
                       ) AS level_down_score_avg,
                       count(*) FILTER (
                         WHERE {source_where_sql}
                           AND {where_sql}
                           AND {recommendation_where_sql}
                           AND level_up_score IS NOT NULL
                           AND level_down_score IS NOT NULL
                       )::int AS score_sample_count
                FROM {table_name} t
                {count_stock_industry_join_sql}
                WHERE for_trade_date = %(selected_for_trade_date)s
                """,
                params,
            )
            count_row = cur.fetchone() or {}
            total_count = int(count_row.get("total_count") or 0)
            filtered_count = int(count_row.get("filtered_count") or 0)
            score_sample_count = int(count_row.get("score_sample_count") or 0)
            level_up_score_avg = count_row.get("level_up_score_avg")
            level_down_score_avg = count_row.get("level_down_score_avg")
            score_regime = "unavailable"
            if score_sample_count:
                score_regime = "bull" if level_up_score_avg > level_down_score_avg else "bear"
            score_comparison = {
                "level_up_score_avg": level_up_score_avg if score_sample_count else None,
                "level_down_score_avg": level_down_score_avg if score_sample_count else None,
                "sample_count": score_sample_count,
                "regime": score_regime,
                "label": (
                    "牛市"
                    if score_regime == "bull"
                    else "熊市"
                    if score_regime == "bear"
                    else "暂无数据"
                ),
            }
            order_sql = self._app_v2_filter_order_sql(
                table_name,
                filters,
                asset_kind=asset_kind,
            )
            select_sql = self._app_v2_filter_select_sql(
                table_name,
                asset_kind,
                include_stock_industry=asset_kind == "stock",
            )
            cur.execute(
                f"""
                SELECT {select_sql}
                FROM {table_name} t
                {stock_industry_join_sql}
                WHERE {where_sql}
                  AND {source_where_sql}
                  AND {recommendation_where_sql}
                  AND for_trade_date = %(selected_for_trade_date)s
                ORDER BY {order_sql}
                LIMIT %(limit)s
                """,
                params,
            )
            rows = [dict(row) for row in cur.fetchall()]
            linked_stock_filter_source_identity_keys: list[str] = []
            if asset_kind in {"index", "board"}:
                cur.execute(
                    f"""
                    SELECT t.identity_key
                    FROM {table_name} t
                    WHERE {where_sql}
                      AND {source_where_sql}
                      AND {recommendation_where_sql}
                      AND for_trade_date = %(selected_for_trade_date)s
                    ORDER BY {order_sql}
                    """,
                    params,
                )
                linked_stock_filter_source_identity_keys = list(
                    dict.fromkeys(
                        str(row.get("identity_key") or "").strip()
                        for row in cur.fetchall()
                        if str(row.get("identity_key") or "").strip()
                    )
                )
            if source_context is not None:
                source_context["matched_stock_count"] = filtered_count
                if (
                    source_context.get("source_filter_active")
                    and source_context.get("membership_count")
                    and not filtered_count
                    and not source_context.get("empty_state")
                ):
                    source_context["empty_state"] = "成分股中无符合当前筛选条件的个股"
                source_context.pop("_stock_identity_keys", None)
        result = {
            "cache_ready": True,
            "items": rows,
            "total_count": total_count,
            "filtered_count": filtered_count,
            "returned_count": len(rows),
            "available_for_trade_dates": available_for_trade_dates,
            "selected_for_trade_date": selected_for_trade_date,
            "source_run_id": source_run_id,
            "source_context": source_context,
            "level_up_recommendation": level_up_recommendation,
            "score_comparison": score_comparison,
            "linked_stock_filter_source_identity_keys": linked_stock_filter_source_identity_keys,
        }
        self._app_shared_cache_set(self._app_filter_result_cache, cache_key, result)
        return result

    def fetch_app_current_filter_identity(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
    ) -> dict[str, Any]:
        del principal_id, principal_type, user_id
        table_by_asset = {
            "stock": "v_n6_stock_condition_display_basis",
            "index": "v_n6_index_condition_display_basis",
            "board": "v_n6_board_condition_display_basis",
        }
        table_name = table_by_asset.get(asset_kind)
        if table_name is None or not self._app_v2_relation_exists(table_name):
            return {"approved": False, "for_trade_date": ""}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                WITH current_batch AS (
                  SELECT max(for_trade_date) AS for_trade_date
                  FROM {table_name}
                )
                SELECT current_batch.for_trade_date::text AS for_trade_date,
                       EXISTS (
                         SELECT 1
                         FROM {table_name} approved
                         WHERE approved.identity_key = %(identity_key)s
                           AND approved.for_trade_date = current_batch.for_trade_date
                       ) AS approved
                FROM current_batch
                """,
                {"identity_key": identity_key},
            )
            row = cur.fetchone() or {}
        return {
            "approved": bool(row.get("approved")),
            "for_trade_date": str(row.get("for_trade_date") or "").strip(),
        }

    def _app_v2_filter_level_up_recommendation_context(
        self,
        cur: Any,
        *,
        filters: dict[str, Any],
        asset_kind: str,
        params: dict[str, Any],
        selected_for_trade_date: str,
    ) -> dict[str, Any]:
        value = app_v2_level_up_recommendation_value(filters.get("level_up_score_recommendation"))
        if asset_kind not in {"board", "stock"} or value != "index_max":
            return {"active": False, "value": "", "available": True, "_where_sql": "TRUE"}
        context: dict[str, Any] = {
            "active": True,
            "value": "index_max",
            "source_asset_kind": "index",
            "source_field": "level_up_score",
            "available": False,
            "threshold": None,
            "blocker": "",
            "_where_sql": "FALSE",
        }
        if not self._app_v2_relation_exists("v_n6_index_condition_display_basis"):
            context["blocker"] = "index_level_up_score_max_unavailable"
            return context
        cur.execute(
            """
            SELECT max(level_up_score) AS threshold
            FROM v_n6_index_condition_display_basis
            WHERE for_trade_date = %(selected_for_trade_date)s
              AND level_up_score IS NOT NULL
            """,
            {"selected_for_trade_date": selected_for_trade_date},
        )
        row = cur.fetchone() or {}
        threshold = row.get("threshold")
        if threshold is None:
            context["blocker"] = "index_level_up_score_max_unavailable"
            return context
        params["level_up_score_recommendation_min"] = threshold
        context.update(
            {
                "available": True,
                "threshold": threshold,
                "blocker": "",
                "_where_sql": "level_up_score >= %(level_up_score_recommendation_min)s",
            }
        )
        return context

    def fetch_app_filter_members(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        membership_kind: str,
        parent_identity_key: str,
        limit: int,
    ) -> dict[str, Any]:
        table_by_kind = {
            "index": "v_n6_index_membership_fact",
            "board": "v_n6_board_membership_fact",
        }
        table_name = table_by_kind.get(membership_kind)
        if table_name is None or not self._app_v2_relation_exists(table_name):
            return {"cache_ready": False, "items": []}
        membership_select_by_kind = {
            "index": """
                       'index'::text AS membership_kind,
                       trade_date,
                       index_identity_key,
                       index_identity_key AS parent_identity_key,
                       index_code,
                       index_code AS parent_code,
                       index_name,
                       index_name AS parent_name,
                       stock_identity_key,
                       stock_code,
                       stock_name,
                       NULL::text AS board_type,
                       source_version,
                       source_batch_id
            """,
            "board": """
                       'board'::text AS membership_kind,
                       trade_date,
                       board_identity_key,
                       board_identity_key AS parent_identity_key,
                       board_code,
                       board_code AS parent_code,
                       board_name,
                       board_name AS parent_name,
                       stock_identity_key,
                       stock_code,
                       stock_name,
                       board_type,
                       source_version,
                       source_batch_id
            """,
        }
        membership_select = membership_select_by_kind[membership_kind]
        parent_column = "index_identity_key" if membership_kind == "index" else "board_identity_key"
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT max(trade_date)::text AS trade_date,
                       max(source_batch_id::text) AS source_batch_id
                FROM {table_name}
                """
            )
            batch_row = dict(cur.fetchone() or {})
            cache_key = (
                membership_kind,
                parent_identity_key,
                str(batch_row.get("trade_date") or ""),
                str(batch_row.get("source_batch_id") or ""),
                max(1, min(int(limit), 500)),
            )
            cached_result = self._app_shared_cache_get(self._app_membership_result_cache, cache_key)
            if cached_result is not None:
                return cached_result
            cur.execute(
                f"""
                SELECT {membership_select}
                FROM {table_name}
                WHERE {parent_column} = %(parent_identity_key)s
                  AND trade_date = (SELECT max(trade_date) FROM {table_name})
                ORDER BY stock_code ASC, stock_identity_key ASC
                LIMIT %(limit)s
                """,
                {
                    "parent_identity_key": parent_identity_key,
                    "limit": max(1, min(int(limit), 500)),
                },
            )
            rows = [dict(row) for row in cur.fetchall()]
        result = {"cache_ready": True, "items": rows}
        self._app_shared_cache_set(self._app_membership_result_cache, cache_key, result)
        return result

    def fetch_app_filter_linked_stocks(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        membership_kind: str,
        parent_identity_key: str,
        limit: int,
        view: str = "matched",
    ) -> dict[str, Any]:
        del principal_id, principal_type, user_id
        view = "all" if view == "all" else "matched"
        table_by_kind = {
            "index": "v_n6_index_membership_fact",
            "board": "v_n6_board_membership_fact",
        }
        membership_table = table_by_kind.get(membership_kind)
        stock_table = "v_n6_stock_condition_display_basis"
        if (
            membership_table is None
            or not self._app_v2_relation_exists(membership_table)
            or not self._app_v2_relation_exists(stock_table)
        ):
            return {
                "cache_ready": False,
                "items": [],
                "membership_count": 0,
                "linked_count": 0,
                "missing_count": 0,
                "view": view,
                "current_view_count": 0,
            }
        membership_select_by_kind = {
            "index": """
                       'index'::text AS membership_kind,
                       m.trade_date,
                       m.index_identity_key,
                       m.index_identity_key AS parent_identity_key,
                       m.index_code,
                       m.index_code AS parent_code,
                       m.index_name,
                       m.index_name AS parent_name,
                       m.stock_identity_key,
                       m.stock_code,
                       m.stock_name,
                       NULL::text AS board_type,
                       m.source_version,
                       m.source_batch_id
            """,
            "board": """
                       'board'::text AS membership_kind,
                       m.trade_date,
                       m.board_identity_key,
                       m.board_identity_key AS parent_identity_key,
                       m.board_code,
                       m.board_code AS parent_code,
                       m.board_name,
                       m.board_name AS parent_name,
                       m.stock_identity_key,
                       m.stock_code,
                       m.stock_name,
                       m.board_type,
                       m.source_version,
                       m.source_batch_id
            """,
        }
        membership_select = membership_select_by_kind[membership_kind]
        parent_column = "index_identity_key" if membership_kind == "index" else "board_identity_key"
        query_params = {
            "parent_identity_key": parent_identity_key,
            "limit": max(1, min(int(limit), APP_V2_FILTER_PAGE_MAX_ROWS)),
        }
        join_type = "LEFT JOIN" if view == "all" else "JOIN"
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                WITH latest_membership AS (
                  SELECT max(trade_date) AS trade_date
                  FROM {membership_table}
                ),
                latest_stock AS (
                  SELECT max(source_trade_date) AS source_trade_date
                  FROM v_n6_stock_condition_display_basis
                ),
                membership_rows AS (
                  SELECT {membership_select}
                  FROM {membership_table} m, latest_membership lm
                  WHERE m.{parent_column} = %(parent_identity_key)s
                    AND m.trade_date = lm.trade_date
                ),
                stock_rows AS (
                  SELECT identity_key
                  FROM v_n6_stock_condition_display_basis s, latest_stock ls
                  WHERE s.source_trade_date = ls.source_trade_date
                )
                SELECT (SELECT count(*) FROM membership_rows)::int AS membership_count,
                       (
                         SELECT count(*)
                         FROM membership_rows mr
                         JOIN stock_rows s ON s.identity_key = mr.stock_identity_key
                       )::int AS linked_count
                """,
                query_params,
            )
            count_row = cur.fetchone() or {}
            membership_count = int(count_row.get("membership_count") or 0)
            linked_count = int(count_row.get("linked_count") or 0)
            missing_count = max(0, membership_count - linked_count)
            cur.execute(
                f"""
                WITH latest_membership AS (
                  SELECT max(trade_date) AS trade_date
                  FROM {membership_table}
                ),
                latest_stock AS (
                  SELECT max(source_trade_date) AS source_trade_date
                  FROM v_n6_stock_condition_display_basis
                ),
                membership_rows AS (
                  SELECT {membership_select}
                  FROM {membership_table} m, latest_membership lm
                  WHERE m.{parent_column} = %(parent_identity_key)s
                    AND m.trade_date = lm.trade_date
                )
                SELECT NULL::text AS cache_run_id,
                       'stock'::text AS asset_kind,
                       COALESCE(s.identity_key, mr.stock_identity_key) AS identity_key,
                       s.source_display_basis_id,
                       s.run_id,
                       s.for_trade_date,
                       s.source_trade_date,
                       COALESCE(s.stock_identity_key, mr.stock_identity_key) AS stock_identity_key,
                       COALESCE(s.code, mr.stock_code) AS code,
                       s.exchange,
                       COALESCE(s.name, mr.stock_name) AS name,
                       COALESCE(s.display_title, CONCAT(mr.stock_code, ' ', mr.stock_name)) AS display_title,
                       COALESCE(s.display_summary, 'not_in_stock_filter') AS display_summary,
                       s.period_transition_y,
                       s.period_transition_q,
                       s.period_transition_m,
                       s.period_transition_w,
                       s.period_transition_d,
                       s.buy_target_price,
                       s.sell_target_price,
                       s.up_sell_reference_period,
                       s.down_buy_reference_period,
                       s.total_mv,
                       s.circ_mv,
                       s.score,
                       s.recommendation_level,
                       s.main_index_code,
                       s.main_index_name,
                       s.preferred_board_code,
                       s.preferred_board_name,
                       s.is_st,
                       s.stock_status,
                       s.display_status,
                       s.quality_reason,
                       COALESCE(s.display_code, mr.stock_code) AS display_code,
                       COALESCE(s.display_name, mr.stock_name) AS display_name,
                       NULL::text AS direction,
                       NULL::text AS condition_key,
                       s.selected_directions,
                       s.selected_condition_keys,
                       s.selected_signal_types,
                       s.selected_lanes,
                       s.selected_monitor_types,
                       s.period_grade_y,
                       s.period_grade_q,
                       s.period_grade_m,
                       s.period_grade_w,
                       s.period_grade_d,
                       s.period_grade_y AS year_overheat_level,
                       s.period_grade_q AS quarter_overheat_level,
                       s.period_grade_m AS month_overheat_level,
                       s.period_grade_w AS week_overheat_level,
                       s.period_grade_d AS day_overheat_level,
                       NULL::text AS last_signal_state,
                       s.quality_status,
                       s.run_id AS source_run_id,
                       NULL::text AS projection_run_id,
                       NULL::text AS board_type,
                       s.updated_at AS source_updated_at,
                       CASE WHEN s.identity_key IS NULL THEN FALSE ELSE TRUE END AS in_stock_filter,
                       CASE WHEN s.identity_key IS NULL THEN 'not_in_filter' ELSE 'in_filter' END AS stock_filter_status,
                       mr.membership_kind,
                       mr.trade_date AS membership_trade_date,
                       mr.parent_identity_key,
                       mr.parent_code,
                       mr.parent_name,
                       mr.stock_code,
                       mr.stock_name,
                       mr.board_type AS membership_board_type,
                       mr.source_version AS membership_source_version,
                       mr.source_batch_id AS membership_source_batch_id
                FROM membership_rows mr
                {join_type} v_n6_stock_condition_display_basis s
                  ON s.identity_key = mr.stock_identity_key
                 AND s.source_trade_date = (SELECT source_trade_date FROM latest_stock)
                ORDER BY s.score DESC NULLS LAST, COALESCE(s.identity_key, mr.stock_identity_key) ASC
                LIMIT %(limit)s
                """,
                query_params,
            )
            rows = [dict(row) for row in cur.fetchall()]
        return {
            "cache_ready": True,
            "items": rows,
            "membership_count": membership_count,
            "linked_count": linked_count,
            "missing_count": missing_count,
            "view": view,
            "current_view_count": len(rows),
        }

    def fetch_app_membership_stocks(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        entity_type: str,
        identity_key: str,
        limit: int,
    ) -> dict[str, Any]:
        del principal_id, principal_type, user_id
        table_by_type = {
            "index": "v_n6_index_membership_fact",
            "board": "v_n6_board_membership_fact",
        }
        table_name = table_by_type.get(entity_type)
        if table_name is None or not self._app_v2_relation_exists(table_name):
            return {"members": [], "member_count": 0}
        if entity_type == "index":
            parent_column = "index_identity_key"
            select_sql = """
                index_identity_key,
                stock_identity_key,
                stock_name,
                stock_code,
                trade_date,
                NULL::numeric AS weight
            """
        else:
            parent_column = "board_identity_key"
            select_sql = """
                board_identity_key,
                stock_identity_key,
                stock_name,
                stock_code,
                trade_date,
                NULL::numeric AS weight
            """
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT {select_sql}
                FROM {table_name}
                WHERE {parent_column} = %(identity_key)s
                  AND trade_date = (SELECT max(trade_date) FROM {table_name})
                ORDER BY stock_code ASC, stock_identity_key ASC
                LIMIT %(limit)s
                """,
                {
                    "identity_key": identity_key,
                    "limit": max(1, min(int(limit), 500)),
                },
            )
            rows = [dict(row) for row in cur.fetchall()]
        members = (
            get_index_membership_stocks(identity_key, rows)
            if entity_type == "index"
            else get_board_membership_stocks(identity_key, rows)
        )
        return {"members": members, "member_count": len(members)}

    def fetch_app_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str | None = None,
        limit: int = 500,
        monitor_status: str = "active",
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        table_items = list(APP_V2_MONITOR_TABLE_BY_ASSET.items())
        if asset_kind is not None:
            table_name = APP_V2_MONITOR_TABLE_BY_ASSET.get(asset_kind)
            table_items = [(asset_kind, table_name)] if table_name else []
        if not table_items:
            return {"tables_ready": False, "items": []}
        monitor_status = app_v2_monitor_status_filter(monitor_status)
        existing_tables = [
            (kind, table_name)
            for kind, table_name in table_items
            if table_name and self._app_v2_monitor_relation_exists(table_name)
        ]
        if len(existing_tables) != len(table_items):
            return {"tables_ready": False, "items": []}
        requested_for_trade_date = self._batch_text(for_trade_date)
        selected_for_trade_date = requested_for_trade_date
        available_for_trade_dates = self._app_v2_monitor_available_for_trade_dates(
            [kind for kind, _ in existing_tables]
        )
        if not selected_for_trade_date and available_for_trade_dates:
            selected_for_trade_date = available_for_trade_dates[0]
        current_filter_batch = self._app_v2_current_filter_batches(
            [kind for kind, _ in existing_tables],
            for_trade_date=selected_for_trade_date,
        )
        query_parts = []
        params: dict[str, Any] = {
            "principal_id": principal_id,
            "principal_type": principal_type,
            "limit": max(1, min(max(int(limit) * 5, int(limit)), 5000)),
            "selected_for_trade_date": selected_for_trade_date,
        }
        for kind, table_name in existing_tables:
            columns = self._app_v2_monitor_columns(table_name)
            display_table = APP_V2_FILTER_DISPLAY_TABLE_BY_ASSET.get(kind, "")
            display_join = ""
            display_json_expr = "'{}'::jsonb"
            if (
                selected_for_trade_date
                and display_table
                and self._app_v2_relation_exists(display_table)
            ):
                approved_display_json_expr = self._app_v2_monitor_display_json_expr(
                    kind,
                    display_table,
                    row_alias="approved_display",
                )
                display_join = f"""
                LEFT JOIN LATERAL (
                  SELECT {approved_display_json_expr} AS current_display_row_json
                  FROM {display_table} approved_display
                  WHERE %(current_{kind}_source_trade_date)s <> ''
                    AND %(current_{kind}_for_trade_date)s <> ''
                    AND %(current_{kind}_source_run_id)s <> ''
                    AND approved_display.identity_key = monitor_base.identity_key
                    AND approved_display.source_trade_date::text = %(current_{kind}_source_trade_date)s
                    AND approved_display.for_trade_date::text = %(current_{kind}_for_trade_date)s
                    AND approved_display.run_id::text = %(current_{kind}_source_run_id)s
                  LIMIT 1
                ) display_row ON true
                """
                display_json_expr = (
                    "COALESCE(display_row.current_display_row_json, '{}'::jsonb)"
                )
            source_run_expr = (
                "monitor_base.source_run_id::text" if "source_run_id" in columns else "NULL::text"
            )
            valid_source_trade_expr = (
                "monitor_base.valid_source_trade_date::text"
                if "valid_source_trade_date" in columns
                else "NULL::text"
            )
            valid_for_trade_expr = (
                "monitor_base.valid_for_trade_date::text"
                if "valid_for_trade_date" in columns
                else "NULL::text"
            )
            valid_run_expr = (
                "monitor_base.valid_source_run_id::text"
                if "valid_source_run_id" in columns
                else "NULL::text"
            )
            current_batch = current_filter_batch.get(kind) or {}
            params[f"current_{kind}_source_trade_date"] = self._batch_text(
                current_batch.get("source_trade_date")
            )
            params[f"current_{kind}_for_trade_date"] = self._batch_text(
                current_batch.get("for_trade_date")
            )
            params[f"current_{kind}_source_run_id"] = self._batch_text(
                current_batch.get("source_run_id")
            )
            identity_approved_expr = "false"
            if display_table and self._app_v2_relation_exists(display_table):
                identity_approved_expr = f"""
                EXISTS (
                  SELECT 1
                  FROM {display_table} approved_identity
                  WHERE approved_identity.identity_key = monitor_base.identity_key
                    AND approved_identity.source_trade_date::text = %(current_{kind}_source_trade_date)s
                    AND approved_identity.for_trade_date::text = %(current_{kind}_for_trade_date)s
                    AND approved_identity.run_id::text = %(current_{kind}_source_run_id)s
                )
                """
            expired_at_expr = "monitor_base.expired_at" if "expired_at" in columns else "NULL::timestamptz"
            expired_reason_expr = "monitor_base.expired_reason" if "expired_reason" in columns else "NULL::text"
            user_id_expr = "monitor_base.user_id" if "user_id" in columns else "NULL::text"
            user_filter_clause = "AND monitor_base.user_id = %(user_id)s" if "user_id" in columns else ""
            params["user_id"] = str(user_id)
            query_parts.append(
                f"""
                SELECT monitor_base.monitor_id,
                       monitor_base.principal_id,
                       monitor_base.principal_type,
                       {user_id_expr} AS user_id,
                       '{kind}'::text AS asset_kind,
                       monitor_base.identity_key,
                       monitor_base.direction,
                       monitor_base.source_type AS source,
                       monitor_base.condition_key,
                       {source_run_expr} AS source_run_id,
                       monitor_base.projection_run_id,
                       monitor_base.status,
                       monitor_base.quality_status,
                       monitor_base.last_signal_state,
                       {valid_source_trade_expr} AS valid_source_trade_date,
                       {valid_for_trade_expr} AS valid_for_trade_date,
                       {valid_run_expr} AS valid_source_run_id,
                       {identity_approved_expr} AS identity_in_current_batch,
                       {expired_at_expr} AS expired_at,
                       {expired_reason_expr} AS expired_reason,
                       monitor_base.source_snapshot_json,
                       monitor_base.source_snapshot_json->>'display_name' AS display_name,
                       monitor_base.source_snapshot_json->>'display_code' AS display_code,
                       monitor_base.source_snapshot_json->>'parent_asset_kind' AS source_parent_asset_kind,
                       monitor_base.source_snapshot_json->>'parent_identity_key' AS source_parent_identity_key,
                       monitor_base.source_snapshot_json->>'parent_code' AS source_parent_code,
                       monitor_base.source_snapshot_json->>'parent_name' AS source_parent_name,
                       monitor_base.source_snapshot_json->>'membership_trade_date' AS source_parent_trade_date,
                       monitor_base.source_snapshot_json->>'membership_source_version' AS source_parent_source_version,
                       monitor_base.source_snapshot_json->>'membership_source_batch_id' AS source_parent_source_batch_id,
                       monitor_base.source_snapshot_json->>'linked_mode' AS source_linked_mode,
                       {display_json_expr} AS current_display_row_json,
                       monitor_base.created_at,
                       monitor_base.updated_at,
                       monitor_base.removed_at
                FROM {table_name} monitor_base
                {display_join}
                WHERE monitor_base.principal_id = %(principal_id)s
                  AND monitor_base.principal_type = %(principal_type)s
                  {user_filter_clause}
                  AND monitor_base.status <> 'removed'
                """
            )
        union_sql = "\nUNION ALL\n".join(query_parts)
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT monitor_rows.monitor_id,
                       monitor_rows.principal_id,
                       monitor_rows.principal_type,
                       monitor_rows.user_id,
                       monitor_rows.asset_kind,
                       monitor_rows.identity_key,
                       monitor_rows.direction,
                       monitor_rows.source,
                       monitor_rows.condition_key,
                       monitor_rows.source_run_id,
                       monitor_rows.projection_run_id,
                       monitor_rows.status,
                       monitor_rows.quality_status,
                       monitor_rows.last_signal_state,
                       monitor_rows.valid_source_trade_date,
                       monitor_rows.valid_for_trade_date,
                       monitor_rows.valid_source_run_id,
                       monitor_rows.identity_in_current_batch,
                       monitor_rows.expired_at,
                       monitor_rows.expired_reason,
                       monitor_rows.source_snapshot_json,
                       monitor_rows.display_name,
                       monitor_rows.display_code,
                       monitor_rows.source_parent_asset_kind,
                       monitor_rows.source_parent_identity_key,
                       monitor_rows.source_parent_code,
                       monitor_rows.source_parent_name,
                       monitor_rows.source_parent_trade_date,
                       monitor_rows.source_parent_source_version,
                       monitor_rows.source_parent_source_batch_id,
                       monitor_rows.source_linked_mode,
                       monitor_rows.current_display_row_json,
                       monitor_rows.created_at,
                       monitor_rows.updated_at,
                       monitor_rows.removed_at
                FROM (
                  {union_sql}
                ) monitor_rows
                ORDER BY created_at DESC NULLS LAST, monitor_id DESC
                LIMIT %(limit)s
                """,
                params,
            )
            rows = [dict(row) for row in cur.fetchall()]
        self._app_v2_apply_monitor_validity(rows, current_filter_batch)
        if requested_for_trade_date:
            rows = [
                row
                for row in rows
                if self._batch_text(row.get("valid_for_trade_date")) == requested_for_trade_date
            ]
        status_counts = self._app_v2_monitor_status_counts(rows)
        if monitor_status != "all":
            rows = [row for row in rows if row.get("effective_status") == monitor_status]
        rows = rows[: max(1, min(int(limit), 1000))]
        self._app_v2_enrich_monitor_memberships(rows)
        return {
            "tables_ready": True,
            "items": rows,
            "monitor_status_filter": monitor_status,
            "current_filter_batch": current_filter_batch,
            "selected_for_trade_date": selected_for_trade_date,
            "available_for_trade_dates": available_for_trade_dates,
            "status_counts": status_counts,
        }

    def _app_v2_monitor_display_json_expr(
        self,
        asset_kind: str,
        display_table: str,
        *,
        row_alias: str = "display_row",
    ) -> str:
        columns = self._app_v2_filter_columns(display_table)
        fields: list[str] = []
        for field in V2_FILTER_VISIBLE_FIELDS_BY_ASSET.get(asset_kind, ()):
            if field in columns and re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", field):
                fields.append(field)
        for field in (
            "asset_kind",
            "source_display_basis_id",
            "run_id",
            "for_trade_date",
            "source_trade_date",
            "identity_key",
            "display_code",
            "display_name",
            "code",
            "name",
        ):
            if field in columns and field not in fields and re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", field):
                fields.append(field)
        if not fields:
            return "'{}'::jsonb"
        pairs = ", ".join(f"'{field}', {row_alias}.\"{field}\"" for field in fields)
        return f"COALESCE(jsonb_strip_nulls(jsonb_build_object({pairs})), '{{}}'::jsonb)"

    def _app_v2_monitor_columns(self, table_name: str) -> set[str]:
        if table_name not in set(APP_V2_MONITOR_TABLE_BY_ASSET.values()):
            return set()
        cached = self._app_v2_monitor_column_cache.get(table_name)
        if cached is not None:
            return set(cached)
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = current_schema()
                  AND table_name = %s
                """,
                (table_name,),
            )
            columns = {str(row["column_name"]) for row in cur.fetchall()}
        self._app_v2_monitor_column_cache[table_name] = set(columns)
        return columns

    def _app_v2_monitor_available_for_trade_dates(self, asset_kinds: list[str]) -> list[str]:
        dates: set[str] = set()
        with self._readonly_connection() as conn, conn.cursor() as cur:
            for asset_kind in asset_kinds:
                table_name = APP_V2_FILTER_DISPLAY_TABLE_BY_ASSET.get(asset_kind)
                if table_name is None or not self._app_v2_relation_exists(table_name):
                    continue
                cur.execute(
                    f"""
                    SELECT DISTINCT for_trade_date::text AS for_trade_date
                    FROM {table_name}
                    WHERE for_trade_date IS NOT NULL
                    """
                )
                dates.update(str(row.get("for_trade_date") or "").strip() for row in cur.fetchall())
        return sorted((date for date in dates if date), reverse=True)

    def _app_v2_current_filter_batches(
        self,
        asset_kinds: list[str],
        *,
        for_trade_date: str = "",
    ) -> dict[str, dict[str, str]]:
        table_by_asset = {
            "stock": "v_n6_stock_condition_display_basis",
            "index": "v_n6_index_condition_display_basis",
            "board": "v_n6_board_condition_display_basis",
        }
        batches = {
            asset_kind: {
                "source_trade_date": "",
                "for_trade_date": "",
                "source_run_id": "",
            }
            for asset_kind in APP_V2_MONITOR_TABLE_BY_ASSET
        }
        with self._readonly_connection() as conn, conn.cursor() as cur:
            for asset_kind in asset_kinds:
                table_name = table_by_asset.get(asset_kind)
                if table_name is None or not self._app_v2_relation_exists(table_name):
                    continue
                params: dict[str, Any] = {}
                where_clause = "for_trade_date = (SELECT max(for_trade_date) FROM {table_name})"
                if for_trade_date:
                    where_clause = "for_trade_date::text = %(for_trade_date)s"
                    params["for_trade_date"] = for_trade_date
                cur.execute(
                    f"""
                    SELECT min(source_trade_date::text) AS source_trade_date,
                           min(for_trade_date::text) AS for_trade_date,
                           min(run_id::text) AS source_run_id
                    FROM {table_name}
                    WHERE {where_clause.format(table_name=table_name)}
                    HAVING count(*) > 0
                       AND count(source_trade_date) = count(*)
                       AND count(for_trade_date) = count(*)
                       AND count(run_id) = count(*)
                       AND count(DISTINCT (source_trade_date::text, for_trade_date::text, run_id::text)) = 1
                    """,
                    params,
                )
                row = cur.fetchone()
                if row:
                    batches[asset_kind] = {
                        "source_trade_date": str(row.get("source_trade_date") or "").strip(),
                        "for_trade_date": str(row.get("for_trade_date") or "").strip(),
                        "source_run_id": str(row.get("source_run_id") or "").strip(),
                    }
        return batches

    def _app_v2_apply_monitor_validity(
        self,
        rows: list[dict[str, Any]],
        current_filter_batch: dict[str, dict[str, str]],
    ) -> None:
        for row in rows:
            asset_kind = str(row.get("asset_kind") or "").strip()
            current_batch = current_filter_batch.get(asset_kind) or {}
            self._app_v2_apply_monitor_validity_to_row(row, current_batch)

    def _app_v2_apply_monitor_validity_to_row(self, row: dict[str, Any], current_batch: dict[str, Any]) -> None:
        original_batch = {
            "source_trade_date": self._batch_text(row.get("valid_source_trade_date")),
            "for_trade_date": self._batch_text(row.get("valid_for_trade_date")),
            "source_run_id": self._batch_text(row.get("source_run_id")),
            "valid_source_run_id": self._batch_text(row.get("valid_source_run_id")),
        }
        normalized_current = {
            "source_trade_date": self._batch_text(current_batch.get("source_trade_date")),
            "for_trade_date": self._batch_text(current_batch.get("for_trade_date")),
            "source_run_id": self._batch_text(current_batch.get("source_run_id")),
        }
        row["valid_source_trade_date"] = original_batch["source_trade_date"]
        row["valid_for_trade_date"] = original_batch["for_trade_date"]
        row["valid_source_run_id"] = original_batch["valid_source_run_id"]

        status = str(row.get("status") or "active").strip().lower()
        effective_status = status
        effective_active = False
        expired_reason = self._batch_text(row.get("expired_reason"))
        if status == "removed":
            effective_status = "removed"
        elif status == "expired":
            effective_status = "expired"
        elif status == "active":
            source_trade_matches = (
                original_batch["source_trade_date"]
                and original_batch["source_trade_date"] == normalized_current["source_trade_date"]
            )
            for_trade_matches = (
                original_batch["for_trade_date"]
                and original_batch["for_trade_date"] == normalized_current["for_trade_date"]
            )
            run_matches = (
                original_batch["source_run_id"]
                and original_batch["source_run_id"] == normalized_current["source_run_id"]
                and original_batch["valid_source_run_id"]
                and original_batch["valid_source_run_id"] == normalized_current["source_run_id"]
            )
            identity_matches = row.get("identity_in_current_batch") is True
            effective_active = bool(
                source_trade_matches and for_trade_matches and run_matches and identity_matches
            )
            if not effective_active:
                effective_status = "expired"
                expired_reason = expired_reason or "filter_batch_changed"

        row["effective_status"] = effective_status
        row["effective_active"] = bool(effective_active)
        row["expired_reason"] = expired_reason
        row["expired_reason_label"] = {
            "filter_batch_changed": "筛选中心已更新",
            "current_filter_batch_missing": "当前筛选批次不可用",
            "source_batch_missing": "来源批次缺失",
        }.get(expired_reason, "筛选中心已更新" if effective_status == "expired" else "")
        row["effective_status_label"] = {
            "active": "有效",
            "expired": "已失效",
            "removed": "已删除",
        }.get(effective_status, effective_status)
        row["validity"] = {
            "original_batch": original_batch,
            "current_batch": normalized_current,
            "identity_in_current_batch": row.get("identity_in_current_batch") is True,
        }

    def _app_v2_monitor_status_counts(self, rows: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
        counts = {
            asset_kind: {"active": 0, "expired": 0, "removed": 0, "all": 0}
            for asset_kind in APP_V2_MONITOR_TABLE_BY_ASSET
        }
        for row in rows:
            asset_kind = str(row.get("asset_kind") or "").strip()
            if asset_kind not in counts:
                continue
            effective_status = str(row.get("effective_status") or "expired").strip()
            counts[asset_kind]["all"] += 1
            if effective_status in counts[asset_kind]:
                counts[asset_kind][effective_status] += 1
        return counts

    @staticmethod
    def _batch_text(value: Any) -> str:
        text = str(value or "").strip()
        return text

    def _app_v2_enrich_monitor_memberships(self, rows: list[dict[str, Any]]) -> None:
        stock_identity_keys = sorted(
            {
                str(row.get("identity_key") or "").strip()
                for row in rows
                if row.get("asset_kind") == "stock" and str(row.get("identity_key") or "").strip()
            }
        )
        empty_memberships = {
            "indexes": [],
            "boards": [],
            "index_count": 0,
            "board_count": 0,
            "summary_label": "所属指数 0 个 / 所属板块 0 个",
        }
        for row in rows:
            row["current_memberships"] = dict(empty_memberships)
        if not stock_identity_keys:
            return

        memberships_by_stock = {
            identity_key: {"indexes": [], "boards": []}
            for identity_key in stock_identity_keys
        }
        for membership in self._app_v2_fetch_stock_membership_context("index", stock_identity_keys):
            stock_identity_key = str(membership.get("stock_identity_key") or "").strip()
            if stock_identity_key in memberships_by_stock:
                memberships_by_stock[stock_identity_key]["indexes"].append(membership)
        for membership in self._app_v2_fetch_stock_membership_context("board", stock_identity_keys):
            stock_identity_key = str(membership.get("stock_identity_key") or "").strip()
            if stock_identity_key in memberships_by_stock:
                memberships_by_stock[stock_identity_key]["boards"].append(membership)

        for row in rows:
            if row.get("asset_kind") != "stock":
                continue
            identity_key = str(row.get("identity_key") or "").strip()
            current = memberships_by_stock.get(identity_key, {"indexes": [], "boards": []})
            index_count = len(current["indexes"])
            board_count = len(current["boards"])
            display_row = row.get("current_display_row_json")
            display_row = display_row if isinstance(display_row, dict) else {}
            industry = next(
                (
                    membership
                    for membership in current["boards"]
                    if str(membership.get("board_type") or "").strip() == "tdx_industry"
                ),
                None,
            )
            if industry:
                display_row = dict(display_row)
                display_row.setdefault("industry_code", industry.get("display_code"))
                display_row.setdefault("industry_name", industry.get("display_name"))
                row["current_display_row_json"] = display_row
            row["current_memberships"] = {
                "indexes": current["indexes"],
                "boards": current["boards"],
                "index_count": index_count,
                "board_count": board_count,
                "summary_label": f"所属指数 {index_count} 个 / 所属板块 {board_count} 个",
            }

    def _app_v2_fetch_stock_membership_context(
        self,
        membership_kind: str,
        stock_identity_keys: list[str],
    ) -> list[dict[str, Any]]:
        table_name = {
            "index": "v_n6_index_membership_fact",
            "board": "v_n6_board_membership_fact",
        }.get(membership_kind)
        if table_name is None or not stock_identity_keys or not self._app_v2_relation_exists(table_name):
            return []
        select_sql = {
            "index": """
                m.stock_identity_key,
                'index'::text AS asset_kind,
                m.index_identity_key AS identity_key,
                m.index_code AS display_code,
                m.index_name AS display_name,
                NULL::text AS board_type,
                m.trade_date,
                m.source_version,
                m.source_batch_id
            """,
            "board": """
                m.stock_identity_key,
                'board'::text AS asset_kind,
                m.board_identity_key AS identity_key,
                m.board_code AS display_code,
                m.board_name AS display_name,
                m.board_type,
                m.trade_date,
                m.source_version,
                m.source_batch_id
            """,
        }[membership_kind]
        order_sql = (
            "m.index_code ASC NULLS LAST, m.index_identity_key ASC"
            if membership_kind == "index"
            else "m.board_type ASC NULLS LAST, m.board_code ASC NULLS LAST, m.board_identity_key ASC"
        )
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT {select_sql}
                FROM {table_name} m
                WHERE m.stock_identity_key = ANY(%(stock_identity_keys)s)
                  AND m.trade_date = (SELECT max(trade_date) FROM {table_name})
                ORDER BY m.stock_identity_key ASC, {order_sql}
                """,
                {"stock_identity_keys": stock_identity_keys},
            )
            return [dict(row) for row in cur.fetchall()]

    def add_app_monitor_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
        direction: str,
        source: str = "single_row",
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        table_name = APP_V2_MONITOR_TABLE_BY_ASSET.get(asset_kind)
        if table_name is None or direction not in APP_V2_VALID_DIRECTIONS:
            return {"ok": False, "status": "invalid_request", "error": "invalid_monitor_request"}
        if not self._app_v2_monitor_relation_exists(table_name):
            return {"ok": False, "status": "data_not_ready", "error": "monitor_table_not_ready"}
        source_row = self._app_v2_fetch_filter_source_row(
            asset_kind=asset_kind,
            identity_key=identity_key,
            for_trade_date=for_trade_date,
        )
        if source_row is None:
            return {"ok": False, "status": "not_found", "error": "source_not_found"}
        snapshot = self._app_v2_monitor_snapshot(source_row, asset_kind=asset_kind)
        condition_key = self._app_v2_monitor_condition_key(source_row)
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                existing = self._app_v2_fetch_existing_monitor(
                    cur,
                    table_name=table_name,
                    principal_id=principal_id,
                    principal_type=principal_type,
                    user_id=user_id,
                    identity_key=identity_key,
                    direction=direction,
                    valid_source_trade_date=snapshot.get("source_trade_date"),
                    valid_for_trade_date=snapshot.get("for_trade_date"),
                    valid_source_run_id=snapshot.get("source_run_id"),
                    require_matching_batch=self._app_v2_monitor_batch_identity_enabled(table_name),
                )
                if existing:
                    return {
                        "ok": True,
                        "status": "already_exists",
                        "added_count": 0,
                        "skipped_count": 1,
                        "item": dict(existing),
                    }
                lifecycle_columns, lifecycle_values, lifecycle_params = self._app_v2_monitor_lifecycle_insert_parts(
                    table_name,
                    snapshot,
                )
                user_id_columns, user_id_values, user_id_params = self._app_v2_monitor_user_insert_parts(
                    table_name,
                    user_id,
                )
                insert_params = {
                    "principal_id": principal_id,
                    "principal_type": principal_type,
                    "asset_kind": asset_kind,
                    "identity_key": identity_key,
                    "direction": direction,
                    "source_type": source,
                    "source_run_id": snapshot.get("source_run_id"),
                    "projection_run_id": snapshot.get("projection_run_id"),
                    "condition_key": condition_key,
                    "quality_status": snapshot.get("quality_status") or "reviewed",
                    "last_signal_state": snapshot.get("last_signal_state"),
                    "source_snapshot_json": Jsonb(snapshot),
                }
                insert_params.update(lifecycle_params)
                insert_params.update(user_id_params)
                cur.execute(
                    f"""
                    INSERT INTO {table_name} (
                      principal_id,
                      principal_type,
                      asset_kind
                      {user_id_columns},
                      identity_key,
                      direction,
                      source_type,
                      source_run_id,
                      projection_run_id,
                      condition_key,
                      status,
                      quality_status,
                      last_signal_state,
                      source_snapshot_json
                      {lifecycle_columns}
                    )
                    VALUES (
                      %(principal_id)s,
                      %(principal_type)s,
                      %(asset_kind)s
                      {user_id_values},
                      %(identity_key)s,
                      %(direction)s,
                      %(source_type)s,
                      %(source_run_id)s,
                      %(projection_run_id)s,
                      %(condition_key)s,
                      'active',
                      %(quality_status)s,
                      %(last_signal_state)s,
                      %(source_snapshot_json)s
                      {lifecycle_values}
                    )
                    RETURNING monitor_id,
                              principal_id,
                              principal_type,
                              asset_kind,
                              identity_key,
                              direction,
                              source_type AS source,
                              condition_key,
                              source_run_id,
                              projection_run_id,
                              status,
                              quality_status,
                              last_signal_state,
                              created_at,
                              updated_at
                    """,
                    insert_params,
                )
                item = dict(cur.fetchone())
        return {"ok": True, "status": "added", "added_count": 1, "skipped_count": 0, "item": item}

    def add_app_monitor_directions(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
        directions: tuple[str, ...],
        source: str = "single_row",
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        table_name = APP_V2_MONITOR_TABLE_BY_ASSET.get(asset_kind)
        normalized_directions = tuple(dict.fromkeys(str(item).strip() for item in directions))
        if (
            table_name is None
            or not normalized_directions
            or any(direction not in APP_V2_VALID_DIRECTIONS for direction in normalized_directions)
        ):
            return {"ok": False, "status": "invalid_request", "error": "invalid_monitor_request"}
        if not self._app_v2_monitor_relation_exists(table_name):
            return {"ok": False, "status": "data_not_ready", "error": "monitor_table_not_ready"}
        source_row = self._app_v2_fetch_filter_source_row(
            asset_kind=asset_kind,
            identity_key=identity_key,
            for_trade_date=for_trade_date,
        )
        if source_row is None:
            return {"ok": False, "status": "not_found", "error": "source_not_found"}
        snapshot = self._app_v2_monitor_snapshot(source_row, asset_kind=asset_kind)
        condition_key = self._app_v2_monitor_condition_key(source_row)
        results: list[dict[str, Any]] = []
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                for direction in normalized_directions:
                    existing = self._app_v2_fetch_existing_monitor(
                        cur,
                        table_name=table_name,
                        principal_id=principal_id,
                        principal_type=principal_type,
                        user_id=user_id,
                        identity_key=identity_key,
                        direction=direction,
                        valid_source_trade_date=snapshot.get("source_trade_date"),
                        valid_for_trade_date=snapshot.get("for_trade_date"),
                        valid_source_run_id=snapshot.get("source_run_id"),
                        require_matching_batch=self._app_v2_monitor_batch_identity_enabled(table_name),
                    )
                    if existing:
                        results.append(
                            {
                                "ok": True,
                                "status": "already_exists",
                                "added_count": 0,
                                "skipped_count": 1,
                                "item": dict(existing),
                            }
                        )
                        continue
                    lifecycle_columns, lifecycle_values, lifecycle_params = (
                        self._app_v2_monitor_lifecycle_insert_parts(table_name, snapshot)
                    )
                    user_id_columns, user_id_values, user_id_params = self._app_v2_monitor_user_insert_parts(
                        table_name,
                        user_id,
                    )
                    insert_params = {
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                        "asset_kind": asset_kind,
                        "identity_key": identity_key,
                        "direction": direction,
                        "source_type": source,
                        "source_run_id": snapshot.get("source_run_id"),
                        "projection_run_id": snapshot.get("projection_run_id"),
                        "condition_key": condition_key,
                        "quality_status": snapshot.get("quality_status") or "reviewed",
                        "last_signal_state": snapshot.get("last_signal_state"),
                        "source_snapshot_json": Jsonb(snapshot),
                    }
                    insert_params.update(lifecycle_params)
                    insert_params.update(user_id_params)
                    cur.execute(
                        f"""
                        INSERT INTO {table_name} (
                          principal_id, principal_type, asset_kind {user_id_columns},
                          identity_key, direction, source_type, source_run_id,
                          projection_run_id, condition_key, status, quality_status,
                          last_signal_state, source_snapshot_json {lifecycle_columns}
                        )
                        VALUES (
                          %(principal_id)s, %(principal_type)s, %(asset_kind)s {user_id_values},
                          %(identity_key)s, %(direction)s, %(source_type)s, %(source_run_id)s,
                          %(projection_run_id)s, %(condition_key)s, 'active', %(quality_status)s,
                          %(last_signal_state)s, %(source_snapshot_json)s {lifecycle_values}
                        )
                        RETURNING monitor_id, principal_id, principal_type, asset_kind,
                                  identity_key, direction, source_type AS source, condition_key,
                                  source_run_id, projection_run_id, status, quality_status,
                                  last_signal_state, created_at, updated_at
                        """,
                        insert_params,
                    )
                    results.append(
                        {
                            "ok": True,
                            "status": "added",
                            "added_count": 1,
                            "skipped_count": 0,
                            "item": dict(cur.fetchone()),
                        }
                    )
        return {
            "ok": True,
            "status": "added" if any(item["added_count"] for item in results) else "already_exists",
            "added_count": sum(int(item["added_count"]) for item in results),
            "skipped_count": sum(int(item["skipped_count"]) for item in results),
            "results": results,
        }

    def bulk_add_app_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        direction: str,
        filters: dict[str, Any],
    ) -> dict[str, Any]:
        table_name = APP_V2_MONITOR_TABLE_BY_ASSET.get(asset_kind)
        if table_name is None or direction not in APP_V2_VALID_DIRECTIONS:
            return {"ok": False, "status": "invalid_request", "error": "invalid_monitor_request"}
        if not self._app_v2_monitor_relation_exists(table_name):
            return {"ok": False, "status": "data_not_ready", "error": "monitor_table_not_ready"}
        result = self.fetch_app_filter_items(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
            asset_kind=asset_kind,
            filters=filters,
            limit=APP_V2_FILTER_PAGE_MAX_ROWS,
        )
        if not result.get("cache_ready"):
            return {"ok": False, "status": "data_not_ready", "error": "filter_source_not_ready"}
        added = 0
        skipped = 0
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                for source_row in result.get("items") or []:
                    identity_key = str(source_row.get("identity_key") or "").strip()
                    if not identity_key:
                        continue
                    snapshot = self._app_v2_monitor_snapshot(source_row, asset_kind=asset_kind)
                    existing = self._app_v2_fetch_existing_monitor(
                        cur,
                        table_name=table_name,
                        principal_id=principal_id,
                        principal_type=principal_type,
                        user_id=user_id,
                        identity_key=identity_key,
                        direction=direction,
                        valid_source_trade_date=snapshot.get("source_trade_date"),
                        valid_for_trade_date=snapshot.get("for_trade_date"),
                        valid_source_run_id=snapshot.get("source_run_id"),
                        require_matching_batch=self._app_v2_monitor_batch_identity_enabled(table_name),
                    )
                    if existing:
                        skipped += 1
                        continue
                    lifecycle_columns, lifecycle_values, lifecycle_params = self._app_v2_monitor_lifecycle_insert_parts(
                        table_name,
                        snapshot,
                    )
                    user_id_columns, user_id_values, user_id_params = self._app_v2_monitor_user_insert_parts(
                        table_name,
                        user_id,
                    )
                    insert_params = {
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                        "asset_kind": asset_kind,
                        "identity_key": identity_key,
                        "direction": direction,
                        "source_run_id": snapshot.get("source_run_id"),
                        "projection_run_id": snapshot.get("projection_run_id"),
                        "condition_key": self._app_v2_monitor_condition_key(source_row),
                        "quality_status": snapshot.get("quality_status") or "reviewed",
                        "last_signal_state": snapshot.get("last_signal_state"),
                        "source_snapshot_json": Jsonb(snapshot),
                    }
                    insert_params.update(lifecycle_params)
                    insert_params.update(user_id_params)
                    cur.execute(
                        f"""
                        INSERT INTO {table_name} (
                          principal_id,
                          principal_type,
                          asset_kind
                          {user_id_columns},
                          identity_key,
                          direction,
                          source_type,
                          source_run_id,
                          projection_run_id,
                          condition_key,
                          status,
                          quality_status,
                          last_signal_state,
                          source_snapshot_json
                          {lifecycle_columns}
                        )
                        VALUES (
                          %(principal_id)s,
                          %(principal_type)s,
                          %(asset_kind)s
                          {user_id_values},
                          %(identity_key)s,
                          %(direction)s,
                          'filter_result',
                          %(source_run_id)s,
                          %(projection_run_id)s,
                          %(condition_key)s,
                          'active',
                          %(quality_status)s,
                          %(last_signal_state)s,
                          %(source_snapshot_json)s
                          {lifecycle_values}
                        )
                        """,
                        insert_params,
                    )
                    added += 1
        return {
            "ok": True,
            "status": "completed",
            "asset_kind": asset_kind,
            "direction": direction,
            "filtered_count": int(result.get("filtered_count") or 0),
            "added_count": added,
            "skipped_count": skipped,
            "failed_count": 0,
        }

    def selected_add_app_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        direction: str,
        identity_keys: list[str],
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        table_name = APP_V2_MONITOR_TABLE_BY_ASSET.get(asset_kind)
        if table_name is None or direction not in APP_V2_VALID_DIRECTIONS:
            return {"ok": False, "status": "invalid_request", "error": "invalid_monitor_request"}
        if not self._app_v2_monitor_relation_exists(table_name):
            return {"ok": False, "status": "data_not_ready", "error": "monitor_table_not_ready"}
        clean_identity_keys = list(dict.fromkeys(str(key).strip() for key in identity_keys if str(key or "").strip()))
        if not clean_identity_keys:
            return {"ok": False, "status": "invalid_request", "error": "empty_identity_keys"}

        added = 0
        skipped = 0
        failed = 0
        details: list[dict[str, Any]] = []
        for identity_key in clean_identity_keys:
            try:
                result = self.add_app_monitor_item(
                    principal_id=principal_id,
                    principal_type=principal_type,
                    user_id=user_id,
                    asset_kind=asset_kind,
                    identity_key=identity_key,
                    direction=direction,
                    source="selected_rows",
                    for_trade_date=for_trade_date,
                )
            except Exception as exc:
                failed += 1
                details.append({"identity_key": identity_key, "status": "failed", "error": type(exc).__name__})
                continue
            status = str(result.get("status") or "")
            if status == "added":
                added += int(result.get("added_count") or 1)
            elif status == "already_exists":
                skipped += int(result.get("skipped_count") or 1)
            else:
                failed += 1
            details.append({"identity_key": identity_key, "status": status or "unknown", "error": result.get("error")})
        return {
            "ok": failed == 0,
            "status": "completed",
            "asset_kind": asset_kind,
            "direction": direction,
            "requested_count": len(clean_identity_keys),
            "added_count": added,
            "skipped_count": skipped,
            "failed_count": failed,
            "items": details,
        }

    def add_app_linked_stock_monitor_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        parent_asset_kind: str,
        parent_identity_key: str,
        mode: str,
        stock_identity_keys: list[str],
        direction: str,
    ) -> dict[str, Any]:
        table_name = APP_V2_MONITOR_TABLE_BY_ASSET["stock"]
        if (
            parent_asset_kind not in {"index", "board"}
            or mode not in {"selected", "matched_stock_filter"}
            or direction not in APP_V2_VALID_DIRECTIONS
            or not parent_identity_key
        ):
            return {"ok": False, "status": "invalid_request", "error": "invalid_monitor_request"}
        if not self._app_v2_monitor_relation_exists(table_name):
            return {"ok": False, "status": "data_not_ready", "error": "monitor_table_not_ready"}
        result = self.fetch_app_filter_linked_stocks(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
            membership_kind=parent_asset_kind,
            parent_identity_key=parent_identity_key,
            limit=APP_V2_FILTER_PAGE_MAX_ROWS,
            view="matched",
        )
        if not result.get("cache_ready"):
            return {"ok": False, "status": "data_not_ready", "error": "linked_stock_source_not_ready"}

        linked_rows = list(result.get("items") or [])
        selected_keys = {str(key).strip() for key in stock_identity_keys if str(key or "").strip()}
        if mode == "selected":
            if not selected_keys:
                return {"ok": False, "status": "invalid_request", "error": "empty_stock_identity_keys"}
            linked_rows = [
                row
                for row in linked_rows
                if str(row.get("stock_identity_key") or row.get("identity_key") or "").strip() in selected_keys
            ]

        added = 0
        skipped = 0
        source_type = f"{parent_asset_kind}_linked_stock"
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                for source_row in linked_rows:
                    identity_key = str(source_row.get("stock_identity_key") or source_row.get("identity_key") or "").strip()
                    if not identity_key:
                        continue
                    snapshot = self._app_v2_monitor_snapshot(source_row, asset_kind="stock")
                    snapshot.update(
                        {
                            "source": source_type,
                            "linked_mode": mode,
                            "parent_asset_kind": parent_asset_kind,
                            "parent_identity_key": parent_identity_key,
                            "parent_code": source_row.get("parent_code"),
                            "parent_name": source_row.get("parent_name"),
                            "membership_trade_date": source_row.get("membership_trade_date"),
                            "membership_source_version": source_row.get("membership_source_version"),
                            "membership_source_batch_id": source_row.get("membership_source_batch_id"),
                            "stock_filter_source_trade_date": source_row.get("source_trade_date"),
                            "stock_filter_run_id": source_row.get("run_id") or source_row.get("source_run_id"),
                            "in_stock_filter": True,
                        }
                    )
                    existing = self._app_v2_fetch_existing_monitor(
                        cur,
                        table_name=table_name,
                        principal_id=principal_id,
                        principal_type=principal_type,
                        user_id=user_id,
                        identity_key=identity_key,
                        direction=direction,
                        valid_source_trade_date=snapshot.get("source_trade_date"),
                        valid_for_trade_date=snapshot.get("for_trade_date"),
                        valid_source_run_id=snapshot.get("source_run_id"),
                        require_matching_batch=self._app_v2_monitor_batch_identity_enabled(table_name),
                    )
                    if existing:
                        skipped += 1
                        continue
                    lifecycle_columns, lifecycle_values, lifecycle_params = self._app_v2_monitor_lifecycle_insert_parts(
                        table_name,
                        snapshot,
                    )
                    user_id_columns, user_id_values, user_id_params = self._app_v2_monitor_user_insert_parts(
                        table_name,
                        user_id,
                    )
                    insert_params = {
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                        "identity_key": identity_key,
                        "direction": direction,
                        "source_type": source_type,
                        "source_run_id": snapshot.get("source_run_id"),
                        "projection_run_id": snapshot.get("projection_run_id"),
                        "condition_key": self._app_v2_monitor_condition_key(source_row),
                        "quality_status": snapshot.get("quality_status") or "reviewed",
                        "last_signal_state": snapshot.get("last_signal_state"),
                        "source_snapshot_json": Jsonb(snapshot),
                    }
                    insert_params.update(lifecycle_params)
                    insert_params.update(user_id_params)
                    cur.execute(
                        f"""
                        INSERT INTO {table_name} (
                          principal_id,
                          principal_type,
                          asset_kind
                          {user_id_columns},
                          identity_key,
                          direction,
                          source_type,
                          source_run_id,
                          projection_run_id,
                          condition_key,
                          status,
                          quality_status,
                          last_signal_state,
                          source_snapshot_json
                          {lifecycle_columns}
                        )
                        VALUES (
                          %(principal_id)s,
                          %(principal_type)s,
                          'stock'
                          {user_id_values},
                          %(identity_key)s,
                          %(direction)s,
                          %(source_type)s,
                          %(source_run_id)s,
                          %(projection_run_id)s,
                          %(condition_key)s,
                          'active',
                          %(quality_status)s,
                          %(last_signal_state)s,
                          %(source_snapshot_json)s
                          {lifecycle_values}
                        )
                        """,
                        insert_params,
                    )
                    added += 1
        return {
            "ok": True,
            "status": "completed",
            "asset_kind": "stock",
            "direction": direction,
            "parent_asset_kind": parent_asset_kind,
            "parent_identity_key": parent_identity_key,
            "mode": mode,
            "membership_count": int(result.get("membership_count") or 0),
            "candidate_count": len(linked_rows),
            "added_count": added,
            "skipped_count": skipped,
            "failed_count": 0,
        }

    def remove_app_monitor_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        monitor_id: int,
    ) -> dict[str, Any]:
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                for table_name in APP_V2_MONITOR_TABLE_BY_ASSET.values():
                    if not self._app_v2_monitor_relation_exists(table_name):
                        continue
                    user_id_clause = ""
                    params = {
                        "monitor_id": monitor_id,
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                    }
                    if "user_id" in self._app_v2_monitor_columns(table_name):
                        user_id_clause = "AND user_id = %(user_id)s"
                        params["user_id"] = int(user_id)
                    cur.execute(
                        f"""
                        UPDATE {table_name}
                        SET status = 'removed',
                            removed_at = now(),
                            updated_at = now()
                        WHERE monitor_id = %(monitor_id)s
                          AND principal_id = %(principal_id)s
                          AND principal_type = %(principal_type)s
                          {user_id_clause}
                          AND status <> 'removed'
                        RETURNING monitor_id
                        """,
                        params,
                    )
                    row = cur.fetchone()
                    if row:
                        return {"ok": True, "status": "removed", "monitor_id": int(row["monitor_id"])}
        return {"ok": False, "status": "not_found", "error": "monitor_not_found"}

    def add_app_realtime_scope_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
        for_trade_date: str = "",
        source: str = "single_row",
    ) -> dict[str, Any]:
        if asset_kind not in APP_V2_MONITOR_TABLE_BY_ASSET or not identity_key:
            return {"ok": False, "status": "invalid_request", "error": "invalid_realtime_scope_request"}
        if not self._app_v2_relation_exists(APP_REALTIME_SCOPE_TABLE):
            return {"ok": False, "status": "data_not_ready", "error": "realtime_scope_table_not_ready"}
        source_row = self._app_v2_fetch_filter_source_row(
            asset_kind=asset_kind,
            identity_key=identity_key,
            for_trade_date=for_trade_date,
        )
        if source_row is None:
            return {"ok": False, "status": "not_found", "error": "source_not_found"}
        snapshot = self._app_v2_monitor_snapshot(source_row, asset_kind=asset_kind)
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                item = self._app_v2_upsert_realtime_scope_item(
                    cur,
                    principal_id=principal_id,
                    principal_type=principal_type,
                    user_id=user_id,
                    asset_kind=asset_kind,
                    identity_key=identity_key,
                    display_name=str(snapshot.get("display_name") or identity_key),
                    source_type=source,
                    source_snapshot=snapshot,
                    is_default_seed=False,
                )
        return {"ok": True, "status": "added", "added_count": 1, "skipped_count": 0, "item": item}

    def selected_add_app_realtime_scope_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_keys: list[str],
        for_trade_date: str = "",
    ) -> dict[str, Any]:
        if asset_kind not in APP_V2_MONITOR_TABLE_BY_ASSET or not identity_keys:
            return {"ok": False, "status": "invalid_request", "error": "invalid_realtime_scope_request"}
        if not self._app_v2_relation_exists(APP_REALTIME_SCOPE_TABLE):
            return {"ok": False, "status": "data_not_ready", "error": "realtime_scope_table_not_ready"}
        added = 0
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                for identity_key in identity_keys:
                    identity_key = str(identity_key or "").strip()
                    if not identity_key:
                        continue
                    source_row = self._app_v2_fetch_filter_source_row(
                        asset_kind=asset_kind,
                        identity_key=identity_key,
                        for_trade_date=for_trade_date,
                    )
                    snapshot = (
                        self._app_v2_monitor_snapshot(source_row, asset_kind=asset_kind)
                        if source_row is not None
                        else {
                            "asset_kind": asset_kind,
                            "identity_key": identity_key,
                            "display_name": identity_key,
                            "for_trade_date": self._batch_text(for_trade_date),
                        }
                    )
                    self._app_v2_upsert_realtime_scope_item(
                        cur,
                        principal_id=principal_id,
                        principal_type=principal_type,
                        user_id=user_id,
                        asset_kind=asset_kind,
                        identity_key=identity_key,
                        display_name=str(snapshot.get("display_name") or identity_key),
                        source_type="selected_rows",
                        source_snapshot=snapshot,
                        is_default_seed=False,
                    )
                    added += 1
        return {
            "ok": True,
            "status": "completed",
            "asset_kind": asset_kind,
            "requested_count": len(identity_keys),
            "added_count": added,
            "skipped_count": 0,
            "failed_count": 0,
        }

    def bulk_add_app_realtime_scope_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        filters: dict[str, Any],
    ) -> dict[str, Any]:
        if asset_kind not in APP_V2_MONITOR_TABLE_BY_ASSET:
            return {"ok": False, "status": "invalid_request", "error": "invalid_realtime_scope_request"}
        if not self._app_v2_relation_exists(APP_REALTIME_SCOPE_TABLE):
            return {"ok": False, "status": "data_not_ready", "error": "realtime_scope_table_not_ready"}
        result = self.fetch_app_filter_items(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
            asset_kind=asset_kind,
            filters=filters,
            limit=APP_V2_FILTER_PAGE_MAX_ROWS,
        )
        if not result.get("cache_ready"):
            return {"ok": False, "status": "data_not_ready", "error": "filter_source_not_ready"}
        added = 0
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                for source_row in result.get("items") or []:
                    identity_key = str(source_row.get("identity_key") or "").strip()
                    if not identity_key:
                        continue
                    snapshot = self._app_v2_monitor_snapshot(source_row, asset_kind=asset_kind)
                    self._app_v2_upsert_realtime_scope_item(
                        cur,
                        principal_id=principal_id,
                        principal_type=principal_type,
                        user_id=user_id,
                        asset_kind=asset_kind,
                        identity_key=identity_key,
                        display_name=str(snapshot.get("display_name") or identity_key),
                        source_type="filter_result",
                        source_snapshot=snapshot,
                        is_default_seed=False,
                    )
                    added += 1
        return {
            "ok": True,
            "status": "completed",
            "asset_kind": asset_kind,
            "filtered_count": int(result.get("filtered_count") or 0),
            "added_count": added,
            "skipped_count": 0,
            "failed_count": 0,
        }

    def remove_app_realtime_scope_item(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        realtime_scope_id: int,
    ) -> dict[str, Any]:
        if not self._app_v2_relation_exists(APP_REALTIME_SCOPE_TABLE):
            return {"ok": False, "status": "data_not_ready", "error": "realtime_scope_table_not_ready"}
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                cur.execute(
                    f"""
                    UPDATE {APP_REALTIME_SCOPE_TABLE}
                    SET status = 'deleted',
                        deleted_at = now(),
                        updated_at = now()
                    WHERE realtime_scope_id = %(realtime_scope_id)s
                      AND principal_id = %(principal_id)s
                      AND principal_type = %(principal_type)s
                      AND user_id = %(user_id)s
                      AND status = 'active'
                    RETURNING realtime_scope_id
                    """,
                    {
                        "realtime_scope_id": realtime_scope_id,
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                        "user_id": user_id,
                    },
                )
                row = cur.fetchone()
        if not row:
            return {"ok": False, "status": "not_found", "error": "realtime_scope_not_found"}
        return {"ok": True, "status": "deleted", "realtime_scope_id": int(row["realtime_scope_id"])}

    def fetch_app_trade_proposals(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        limit: int = 100,
    ) -> dict[str, Any]:
        if not self._app_v2_relation_exists("n6_virtual_trade_proposal"):
            return {"tables_ready": False, "items": []}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT proposal_id,
                       source_type,
                       source_id,
                       source_signal_projection_id,
                       source_virtual_position_id,
                       holding_episode_no,
                       asset_kind,
                       identity_key,
                       proposal_side,
                       signal_reference_kind,
                       signal_reference_price,
                       proposal_status,
                       expires_at,
                       confirmed_at,
                       executed_virtual_order_id,
                       executed_virtual_trade_id,
                       failure_reason,
                       created_at,
                       updated_at
                FROM n6_virtual_trade_proposal
                WHERE principal_id = %(principal_id)s
                  AND principal_type = %(principal_type)s
                  AND user_id = %(user_id)s
                ORDER BY created_at DESC, proposal_id DESC
                LIMIT %(limit)s
                """,
                {
                    "principal_id": principal_id,
                    "principal_type": principal_type,
                    "user_id": user_id,
                    "limit": max(1, min(int(limit), 500)),
                },
            )
            rows = [dict(row) for row in cur.fetchall()]
        return {"tables_ready": True, "items": rows}

    def create_app_trade_proposal(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        source_type: str,
        source_id: str,
    ) -> dict[str, Any]:
        if principal_type not in {"admin", "human_user"}:
            return {"ok": False, "status": "forbidden", "error": "principal_type_not_supported"}
        if not self._app_v2_relation_exists("n6_virtual_trade_proposal"):
            return {"ok": False, "status": "data_not_ready", "error": "proposal_table_not_ready"}
        source_id = str(source_id or "").strip()
        if not re.fullmatch(r"[1-9][0-9]*", source_id):
            return {"ok": False, "status": "invalid_request", "error": "invalid_source_id"}
        source_signal_projection_id: int | None = None
        source_virtual_position_id: int | None = None
        holding_episode_no: int | None = None
        signal_reference_kind = "manual"
        signal_reference_price: Decimal | None = None
        locked_target_price: Decimal | None = None
        if source_type == "signal":
            source_signal_projection_id = int(source_id)
            signal_row = self.fetch_app_signal_detail(
                principal_id=principal_id,
                principal_type=principal_type,
                user_id=user_id,
                user_signal_projection_id=source_signal_projection_id,
            )
            if not signal_row:
                return {"ok": False, "status": "not_found", "error": "signal_not_in_effective_scope"}
            signal = app_signal_item(signal_row)
            asset_kind = str(signal.get("asset_kind") or "")
            identity_key = str(signal.get("identity_key") or "")
            proposal_side = str(signal.get("direction") or "")
            action_state = str(signal.get("action_state") or "")
            signal_trade_date = normalize_filter_value(signal.get("trade_date"))
            current_stock_batch = self._app_v2_current_filter_batches(["stock"]).get("stock") or {}
            current_trade_date = normalize_filter_value(current_stock_batch.get("for_trade_date"))
            if not current_trade_date or signal_trade_date != current_trade_date:
                return {
                    "ok": False,
                    "status": "conflict",
                    "error": "current_for_trade_date_signal_required",
                }
            if asset_kind != "stock" or proposal_side not in {"buy", "sell"}:
                return {"ok": False, "status": "invalid_request", "error": "stock_signal_required"}
            if action_state == "executed":
                signal_reference_kind = "action_price"
                signal_reference_price = _decimal_or_none(signal.get("action_price"))
            elif action_state == "eligible":
                signal_reference_kind = "trigger_price"
                signal_reference_price = _decimal_or_none(signal.get("trigger_price"))
            else:
                return {"ok": False, "status": "invalid_request", "error": "signal_action_state_not_actionable"}
            if signal_reference_price is None or not signal_reference_price.is_finite() or signal_reference_price <= 0:
                return {"ok": False, "status": "not_ready", "error": "signal_reference_price_not_ready"}
            locked_target_price = _decimal_or_none(signal.get("target_price"))
            if locked_target_price is not None and (
                not locked_target_price.is_finite() or locked_target_price <= 0
            ):
                locked_target_price = None
        elif source_type == "manual_position":
            source_virtual_position_id = int(source_id)
            with self._readonly_connection() as conn, conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT p.virtual_position_id,
                           p.asset_kind,
                           p.identity_key,
                           p.available_quantity,
                           p.holding_episode_no
                    FROM n6_virtual_position p
                    JOIN n6_virtual_account a
                      ON a.virtual_account_id = p.virtual_account_id
                     AND a.principal_id = p.principal_id
                     AND a.principal_type = p.principal_type
                    WHERE p.virtual_position_id = %(source_id)s
                      AND p.principal_id = %(principal_id)s
                      AND p.principal_type = %(principal_type)s
                      AND a.virtual_account_status = 'active'
                      AND p.position_status = 'open_virtual'
                      AND p.asset_kind = 'stock'
                      AND p.available_quantity > 0
                    """,
                    {
                        "source_id": source_virtual_position_id,
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                    },
                )
                position = cur.fetchone()
            if not position:
                return {"ok": False, "status": "not_found", "error": "sellable_position_not_found"}
            asset_kind = "stock"
            identity_key = str(position["identity_key"])
            proposal_side = "sell"
            holding_episode_no = int(position.get("holding_episode_no") or 1)
        else:
            return {"ok": False, "status": "invalid_request", "error": "invalid_proposal_source_type"}
        policy_version = "n6_virtual_trade_proposal_v1"
        policy_hash = hashlib.sha256(
            b"source-only-request|server-derived-side|60s-expiry|fresh-n6-quote-at-execution"
        ).hexdigest()
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT virtual_account_id
                    FROM n6_virtual_account
                    WHERE principal_id = %(principal_id)s
                      AND principal_type = %(principal_type)s
                      AND virtual_account_status = 'active'
                    FOR SHARE
                    """,
                    {"principal_id": principal_id, "principal_type": principal_type},
                )
                accounts = cur.fetchall()
                if len(accounts) != 1:
                    return {"ok": False, "status": "not_ready", "error": "exactly_one_active_virtual_account_required"}
                cur.execute(
                    """
                    INSERT INTO n6_virtual_trade_proposal (
                      principal_id, principal_type, user_id, virtual_account_id,
                      source_type, source_id, source_signal_projection_id,
                      source_virtual_position_id, holding_episode_no,
                      asset_kind, identity_key, proposal_side,
                      signal_reference_kind, signal_reference_price,
                      locked_target_price,
                      proposal_status, expires_at, policy_version, policy_hash,
                      source_lineage_json
                    ) VALUES (
                      %(principal_id)s, %(principal_type)s, %(user_id)s, %(virtual_account_id)s,
                      %(source_type)s, %(source_id)s, %(source_signal_projection_id)s,
                      %(source_virtual_position_id)s, %(holding_episode_no)s,
                      %(asset_kind)s, %(identity_key)s, %(proposal_side)s,
                      %(signal_reference_kind)s, %(signal_reference_price)s,
                      %(locked_target_price)s,
                      'pending', now() + interval '60 seconds', %(policy_version)s, %(policy_hash)s,
                      %(source_lineage_json)s
                    )
                    ON CONFLICT DO NOTHING
                    RETURNING proposal_id, proposal_status, expires_at, proposal_side,
                              identity_key, signal_reference_kind, signal_reference_price
                    """,
                    {
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                        "user_id": user_id,
                        "virtual_account_id": int(accounts[0]["virtual_account_id"]),
                        "source_type": source_type,
                        "source_id": source_id,
                        "source_signal_projection_id": source_signal_projection_id,
                        "source_virtual_position_id": source_virtual_position_id,
                        "holding_episode_no": holding_episode_no,
                        "asset_kind": asset_kind,
                        "identity_key": identity_key,
                        "proposal_side": proposal_side,
                        "signal_reference_kind": signal_reference_kind,
                        "signal_reference_price": signal_reference_price,
                        "locked_target_price": locked_target_price,
                        "policy_version": policy_version,
                        "policy_hash": policy_hash,
                        "source_lineage_json": Jsonb({"source_type": source_type, "source_id": source_id}),
                    },
                )
                inserted = cur.fetchone()
                if not inserted:
                    return {"ok": False, "status": "conflict", "error": "proposal_already_exists"}
                row = dict(inserted)
        return {"ok": True, "status": "created", "item": row}

    def confirm_app_trade_proposal(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        proposal_id: int,
        idempotency_key: str,
    ) -> dict[str, Any]:
        idempotency_key = str(idempotency_key or "").strip()
        if not re.fullmatch(r"[A-Za-z0-9._:-]{8,128}", idempotency_key):
            return {"ok": False, "status": "invalid_request", "error": "invalid_idempotency_key"}
        if not self._app_v2_relation_exists("n6_virtual_trade_proposal"):
            return {"ok": False, "status": "data_not_ready", "error": "proposal_table_not_ready"}
        with psycopg.connect(self.dsn, connect_timeout=10, row_factory=dict_row) as conn:
            with conn.transaction(), conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT proposal_id, proposal_status, expires_at, confirm_idempotency_key
                    FROM n6_virtual_trade_proposal
                    WHERE proposal_id = %(proposal_id)s
                      AND principal_id = %(principal_id)s
                      AND principal_type = %(principal_type)s
                      AND user_id = %(user_id)s
                    FOR UPDATE
                    """,
                    {
                        "proposal_id": proposal_id,
                        "principal_id": principal_id,
                        "principal_type": principal_type,
                        "user_id": user_id,
                    },
                )
                row = cur.fetchone()
                if not row:
                    return {"ok": False, "status": "not_found", "error": "proposal_not_found"}
                if row["proposal_status"] == "confirmed" and row.get("confirm_idempotency_key") == idempotency_key:
                    return {"ok": True, "status": "confirmed", "proposal_id": proposal_id, "idempotent": True}
                if row["proposal_status"] != "pending":
                    return {"ok": False, "status": "conflict", "error": "proposal_not_pending"}
                if ensure_aware(row["expires_at"]) <= utc_now():
                    cur.execute(
                        "UPDATE n6_virtual_trade_proposal SET proposal_status = 'expired', updated_at = now() WHERE proposal_id = %s",
                        (proposal_id,),
                    )
                    return {"ok": False, "status": "expired", "error": "proposal_expired"}
                cur.execute(
                    """
                    UPDATE n6_virtual_trade_proposal
                    SET proposal_status = 'confirmed',
                        confirmed_at = now(),
                        confirm_idempotency_key = %(idempotency_key)s,
                        updated_at = now()
                    WHERE proposal_id = %(proposal_id)s
                    RETURNING proposal_id, proposal_status, confirmed_at, expires_at
                    """,
                    {"proposal_id": proposal_id, "idempotency_key": idempotency_key},
                )
                confirmed = dict(cur.fetchone())
        return {"ok": True, "status": "confirmed", "item": confirmed, "idempotent": False}

    def fetch_app_virtual_trades(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        limit: int = 200,
    ) -> dict[str, Any]:
        if not self._app_v2_relation_exists("n6_virtual_trade"):
            return {"tables_ready": False, "items": []}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT t.virtual_trade_id,
                       t.virtual_order_id,
                       t.virtual_account_id,
                       t.identity_key,
                       split_part(t.identity_key, ':', 3) AS stock_code,
                       stock_identity.stock_name,
                       industry_identity.industry_code,
                       industry_identity.industry_name,
                       t.trade_side,
                       t.filled_quantity,
                       t.filled_price,
                       t.gross_amount,
                       t.total_fee_amount,
                       t.net_amount,
                       t.trade_status,
                       t.trade_time,
                       t.source_proposal_id,
                       t.signal_reference_kind,
                       t.signal_reference_price,
                       t.fill_quote_snapshot_id
                FROM n6_virtual_trade t
                LEFT JOIN LATERAL (
                  SELECT CASE WHEN count(*) = 1 THEN max(latest_name.stock_name) END AS stock_name
                  FROM (
                    SELECT DISTINCT btrim(basis.display_name::text) AS stock_name
                    FROM v_n6_stock_condition_display_basis basis
                    WHERE basis.identity_key = t.identity_key
                      AND NULLIF(btrim(basis.display_name::text), '') IS NOT NULL
                      AND basis.for_trade_date = (
                        SELECT max(asof_basis.for_trade_date)
                        FROM v_n6_stock_condition_display_basis asof_basis
                        WHERE asof_basis.identity_key = t.identity_key
                          AND asof_basis.for_trade_date <= to_char(
                                t.trade_time AT TIME ZONE 'Asia/Shanghai',
                                'YYYYMMDD'
                              )
                      )
                  ) latest_name
                ) stock_identity ON TRUE
                LEFT JOIN LATERAL (
                  SELECT
                    CASE WHEN count(*) = 1 THEN max(latest_industry.board_code) END AS industry_code,
                    CASE WHEN count(*) = 1 THEN max(latest_industry.board_name) END AS industry_name
                  FROM (
                    SELECT DISTINCT
                           membership.board_identity_key,
                           membership.board_code,
                           membership.board_name
                    FROM v_n6_board_membership_fact membership
                    WHERE membership.stock_identity_key = t.identity_key
                      AND membership.board_type = 'tdx_industry'
                      AND NULLIF(btrim(membership.board_identity_key::text), '') IS NOT NULL
                      AND NULLIF(btrim(membership.board_code::text), '') IS NOT NULL
                      AND NULLIF(btrim(membership.board_name::text), '') IS NOT NULL
                      AND membership.trade_date = (
                        SELECT max(asof_membership.trade_date)
                        FROM v_n6_board_membership_fact asof_membership
                        WHERE asof_membership.stock_identity_key = t.identity_key
                          AND asof_membership.board_type = 'tdx_industry'
                          AND asof_membership.trade_date <= to_char(
                                t.trade_time AT TIME ZONE 'Asia/Shanghai',
                                'YYYYMMDD'
                              )
                      )
                  ) latest_industry
                ) industry_identity ON TRUE
                WHERE t.principal_id = %(principal_id)s
                  AND t.principal_type = %(principal_type)s
                  AND EXISTS (
                    SELECT 1
                    FROM n6_principal p
                    WHERE p.principal_id = %(principal_id)s
                      AND p.principal_type = %(principal_type)s
                      AND p.owner_user_id = %(user_id)s
                      AND p.principal_status = 'active'
                  )
                ORDER BY t.trade_time DESC, t.virtual_trade_id DESC
                LIMIT %(limit)s
                """,
                {
                    "principal_id": principal_id,
                    "principal_type": principal_type,
                    "user_id": user_id,
                    "limit": max(1, min(int(limit), 500)),
                },
            )
            rows = [dict(row) for row in cur.fetchall()]
        return {"tables_ready": True, "items": rows}

    def _app_v2_upsert_realtime_scope_item(
        self,
        cur: Any,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
        asset_kind: str,
        identity_key: str,
        display_name: str,
        source_type: str,
        source_snapshot: dict[str, Any],
        is_default_seed: bool,
    ) -> dict[str, Any]:
        cur.execute(
            f"""
            INSERT INTO {APP_REALTIME_SCOPE_TABLE} (
              principal_id,
              principal_type,
              user_id,
              asset_kind,
              identity_key,
              display_name,
              source_type,
              source_snapshot_json,
              is_default_seed,
              status,
              deleted_at
            )
            VALUES (
              %(principal_id)s,
              %(principal_type)s,
              %(user_id)s,
              %(asset_kind)s,
              %(identity_key)s,
              %(display_name)s,
              %(source_type)s,
              %(source_snapshot_json)s,
              %(is_default_seed)s,
              'active',
              NULL
            )
            ON CONFLICT (principal_id, principal_type, user_id, asset_kind, identity_key)
            DO UPDATE SET
              display_name = EXCLUDED.display_name,
              source_type = EXCLUDED.source_type,
              source_snapshot_json = EXCLUDED.source_snapshot_json,
              is_default_seed = {APP_REALTIME_SCOPE_TABLE}.is_default_seed OR EXCLUDED.is_default_seed,
              status = 'active',
              deleted_at = NULL,
              updated_at = now()
            RETURNING realtime_scope_id,
                      principal_id,
                      principal_type,
                      user_id,
                      asset_kind,
                      identity_key,
                      display_name,
                      source_type,
                      source_snapshot_json,
                      is_default_seed,
                      status,
                      deleted_at,
                      created_at,
                      updated_at
            """,
            {
                "principal_id": principal_id,
                "principal_type": principal_type,
                "user_id": user_id,
                "asset_kind": asset_kind,
                "identity_key": identity_key,
                "display_name": display_name,
                "source_type": source_type,
                "source_snapshot_json": Jsonb(source_snapshot),
                "is_default_seed": is_default_seed,
            },
        )
        return dict(cur.fetchone())

    def _app_v2_relation_exists(self, relation_name: str) -> bool:
        allowed_relations = {
            "v_n6_stock_condition_display_basis",
            "v_n6_index_condition_display_basis",
            "v_n6_board_condition_display_basis",
            "n6_index_membership_display_cache",
            "n6_board_membership_display_cache",
            "v_n6_index_membership_fact",
            "v_n6_board_membership_fact",
            APP_REALTIME_SCOPE_TABLE,
            "n6_virtual_trade_proposal",
            "n6_virtual_trade",
        }
        if relation_name not in allowed_relations:
            return False
        cached = self._app_v2_relation_existence_cache.get(relation_name)
        if cached is not None:
            return cached
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute("SELECT to_regclass(%s) AS relation_name", (relation_name,))
            row = cur.fetchone()
        return bool(row and row.get("relation_name"))

    def _app_v2_default_realtime_scope_items(
        self,
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
    ) -> list[dict[str, Any]]:
        return [
            {
                "realtime_scope_id": None,
                "principal_id": principal_id,
                "principal_type": principal_type,
                "user_id": user_id,
                "asset_kind": str(item["asset_kind"]),
                "identity_key": str(item["identity_key"]),
                "display_name": str(item["display_name"]),
                "source_type": "default_seed",
                "source_snapshot_json": {
                    "asset_kind": str(item["asset_kind"]),
                    "identity_key": str(item["identity_key"]),
                    "display_name": str(item["display_name"]),
                    "seed_policy": "n6_default_realtime_monitor_scope_v1",
                },
                "is_default_seed": True,
                "status": "active",
                "deleted_at": None,
                "created_at": None,
                "updated_at": None,
            }
            for item in DEFAULT_REALTIME_SCOPE_INDEXES
        ]

    def _app_v2_merge_realtime_scope_defaults(
        self,
        rows: list[dict[str, Any]],
        *,
        principal_id: int,
        principal_type: str,
        user_id: int,
    ) -> list[dict[str, Any]]:
        deleted_keys = {
            (str(row.get("asset_kind") or ""), str(row.get("identity_key") or ""))
            for row in rows
            if str(row.get("status") or "") == "deleted"
        }
        active_rows: list[dict[str, Any]] = []
        active_keys: set[tuple[str, str]] = set()
        for row in rows:
            if str(row.get("status") or "") != "active":
                continue
            key = (str(row.get("asset_kind") or ""), str(row.get("identity_key") or ""))
            if key in active_keys:
                continue
            active_rows.append(dict(row))
            active_keys.add(key)
        defaults = []
        for row in self._app_v2_default_realtime_scope_items(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=user_id,
        ):
            key = (str(row.get("asset_kind") or ""), str(row.get("identity_key") or ""))
            if key in deleted_keys or key in active_keys:
                continue
            defaults.append(row)
        return active_rows + defaults

    def _app_v2_filter_columns(self, table_name: str) -> set[str]:
        allowed_relations = {
            "v_n6_stock_condition_display_basis",
            "v_n6_index_condition_display_basis",
            "v_n6_board_condition_display_basis",
        }
        if table_name not in allowed_relations:
            return set()
        cached = self._app_v2_filter_column_cache.get(table_name)
        if cached is not None:
            return set(cached)
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = current_schema()
                  AND table_name = %s
                """,
                (table_name,),
            )
            columns = {
                str(row.get("column_name") or "").strip()
                for row in cur.fetchall()
                if str(row.get("column_name") or "").strip()
            }
        self._app_v2_filter_column_cache[table_name] = set(columns)
        return columns

    def _app_v2_filter_select_sql(
        self,
        table_name: str,
        asset_kind: str,
        *,
        alias: str = "t",
        include_stock_industry: bool = False,
    ) -> str:
        available_columns = self._app_v2_filter_columns(table_name)
        select_expressions: list[str] = []
        for field in V2_FILTER_VISIBLE_FIELDS_BY_ASSET.get(asset_kind, ()):
            if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", field):
                continue
            if include_stock_industry and field in {"industry_code", "industry_name"}:
                select_expressions.append(f'industry."{field}" AS "{field}"')
            elif field in available_columns:
                select_expressions.append(f'{alias}."{field}" AS "{field}"')
        if not select_expressions:
            raise RuntimeError("approved N6 filter view exposes no approved visible columns")
        return ",\n                       ".join(select_expressions)

    @staticmethod
    def _app_v2_stock_industry_join_sql() -> str:
        return """
                LEFT JOIN LATERAL (
                  SELECT
                    CASE WHEN count(*) = 1 THEN max(industry_identity.board_code) END AS industry_code,
                    CASE WHEN count(*) = 1 THEN max(industry_identity.board_name) END AS industry_name
                  FROM (
                    SELECT DISTINCT
                           membership.board_identity_key,
                           membership.board_code,
                           membership.board_name
                    FROM v_n6_board_membership_fact membership
                    WHERE membership.stock_identity_key = t.identity_key
                      AND membership.board_type = 'tdx_industry'
                      AND NULLIF(btrim(membership.board_identity_key::text), '') IS NOT NULL
                      AND NULLIF(btrim(membership.board_code::text), '') IS NOT NULL
                      AND NULLIF(btrim(membership.board_name::text), '') IS NOT NULL
                      AND membership.trade_date = (
                        SELECT max(asof_membership.trade_date)
                        FROM v_n6_board_membership_fact asof_membership
                        WHERE asof_membership.stock_identity_key = t.identity_key
                          AND asof_membership.board_type = 'tdx_industry'
                          AND asof_membership.trade_date <= t.source_trade_date
                      )
                  ) industry_identity
                ) industry ON TRUE
        """

    def _app_v2_filter_order_sql(
        self,
        table_name: str,
        filters: dict[str, Any],
        *,
        asset_kind: str = "",
    ) -> str:
        default_order = "t.updated_at DESC NULLS LAST, t.identity_key ASC"
        sort_key = normalize_filter_value(filters.get("sort"))
        if not sort_key or not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", sort_key):
            return default_order
        if asset_kind == "stock" and sort_key in {"industry_code", "industry_name"}:
            sort_expression = f'industry."{sort_key}"'
        elif sort_key in self._app_v2_filter_columns(table_name):
            sort_expression = f't."{sort_key}"'
        else:
            return default_order
        sort_dir = str(filters.get("sort_dir") or "asc").strip().lower()
        direction = "DESC" if sort_dir == "desc" else "ASC"
        return f"{sort_expression} {direction} NULLS LAST, t.identity_key ASC"

    def _app_v2_monitor_relation_exists(self, relation_name: str) -> bool:
        if relation_name not in set(APP_V2_MONITOR_TABLE_BY_ASSET.values()):
            return False
        cached = self._app_v2_relation_existence_cache.get(relation_name)
        if cached is not None:
            return cached
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute("SELECT to_regclass(%s) AS relation_name", (relation_name,))
            row = cur.fetchone()
        return bool(row and row.get("relation_name"))

    def _app_v2_monitor_lifecycle_insert_parts(
        self,
        table_name: str,
        snapshot: dict[str, Any],
    ) -> tuple[str, str, dict[str, Any]]:
        columns = self._app_v2_monitor_columns(table_name)
        mapping = (
            ("valid_source_trade_date", "source_trade_date"),
            ("valid_for_trade_date", "for_trade_date"),
            ("valid_source_run_id", "source_run_id"),
        )
        insert_columns: list[str] = []
        value_placeholders: list[str] = []
        params: dict[str, Any] = {}
        for column_name, snapshot_key in mapping:
            if column_name not in columns:
                continue
            param_name = f"insert_{column_name}"
            insert_columns.append(f",\n                      {column_name}")
            value_placeholders.append(f",\n                      %({param_name})s")
            params[param_name] = snapshot.get(snapshot_key)
        return "".join(insert_columns), "".join(value_placeholders), params

    def _app_v2_monitor_batch_identity_enabled(self, table_name: str) -> bool:
        columns = self._app_v2_monitor_columns(table_name)
        return {
            "valid_source_trade_date",
            "valid_for_trade_date",
            "valid_source_run_id",
        }.issubset(columns)

    def _app_v2_monitor_user_insert_parts(self, table_name: str, user_id: int) -> tuple[str, str, dict[str, Any]]:
        if "user_id" not in self._app_v2_monitor_columns(table_name):
            return "", "", {}
        return ",\n                      user_id", ",\n                      %(insert_user_id)s", {"insert_user_id": int(user_id)}

    def _app_v2_fetch_filter_source_row(
        self,
        *,
        asset_kind: str,
        identity_key: str,
        for_trade_date: str = "",
    ) -> dict[str, Any] | None:
        table_name = {
            "stock": "v_n6_stock_condition_display_basis",
            "index": "v_n6_index_condition_display_basis",
            "board": "v_n6_board_condition_display_basis",
        }.get(asset_kind)
        if table_name is None or not self._app_v2_relation_exists(table_name):
            return None
        selected_for_trade_date = self._batch_text(for_trade_date)
        date_filter_sql = ""
        params: dict[str, Any] = {"identity_key": identity_key}
        if selected_for_trade_date:
            date_filter_sql = "AND t.for_trade_date::text = %(for_trade_date)s"
            params["for_trade_date"] = selected_for_trade_date
        select_sql = self._app_v2_filter_select_sql(table_name, asset_kind)
        with self._readonly_connection() as conn, conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT {select_sql}
                FROM {table_name} t
                WHERE t.identity_key = %(identity_key)s
                  {date_filter_sql}
                  AND t.source_trade_date = (
                    SELECT max(source_trade_date)
                    FROM {table_name}
                    WHERE identity_key = %(identity_key)s
                      {date_filter_sql.replace("t.", "")}
                  )
                ORDER BY t.updated_at DESC NULLS LAST, t.identity_key ASC
                LIMIT 1
                """,
                params,
            )
            row = cur.fetchone()
        return dict(row) if row else None

    def _app_v2_fetch_existing_monitor(
        self,
        cur: Any,
        *,
        table_name: str,
        principal_id: int,
        principal_type: str,
        identity_key: str,
        direction: str,
        user_id: int = 0,
        valid_source_trade_date: Any = None,
        valid_for_trade_date: Any = None,
        valid_source_run_id: Any = None,
        require_matching_batch: bool = True,
    ) -> dict[str, Any] | None:
        columns = self._app_v2_monitor_columns(table_name)
        valid_source_trade_expr = (
            "COALESCE(valid_source_trade_date, source_snapshot_json->>'source_trade_date')"
            if "valid_source_trade_date" in columns
            else "source_snapshot_json->>'source_trade_date'"
        )
        valid_for_trade_expr = (
            "COALESCE(valid_for_trade_date, source_snapshot_json->>'for_trade_date')"
            if "valid_for_trade_date" in columns
            else "source_snapshot_json->>'for_trade_date'"
        )
        valid_run_expr = (
            "COALESCE(valid_source_run_id, source_snapshot_json->>'source_run_id')"
            if "valid_source_run_id" in columns
            else "source_snapshot_json->>'source_run_id'"
        )
        params: dict[str, Any] = {
            "principal_id": principal_id,
            "principal_type": principal_type,
            "identity_key": identity_key,
            "direction": direction,
        }
        user_filter_sql = ""
        if "user_id" in columns:
            user_filter_sql = "AND user_id = %(user_id)s"
            params["user_id"] = int(user_id)
        batch_filters: list[str] = []
        valid_source_trade_date_text = self._batch_text(valid_source_trade_date)
        valid_for_trade_date_text = self._batch_text(valid_for_trade_date)
        valid_source_run_id_text = self._batch_text(valid_source_run_id)
        if require_matching_batch and valid_source_trade_date_text:
            params["valid_source_trade_date"] = valid_source_trade_date_text
            batch_filters.append(f"{valid_source_trade_expr} = %(valid_source_trade_date)s")
        if require_matching_batch and valid_for_trade_date_text:
            params["valid_for_trade_date"] = valid_for_trade_date_text
            batch_filters.append(f"{valid_for_trade_expr} = %(valid_for_trade_date)s")
        if require_matching_batch and valid_source_run_id_text:
            params["valid_source_run_id"] = valid_source_run_id_text
            batch_filters.append(f"{valid_run_expr} = %(valid_source_run_id)s")
        batch_filter_sql = ""
        if batch_filters:
            batch_filter_sql = "\n              AND " + "\n              AND ".join(batch_filters)
        cur.execute(
            f"""
            SELECT monitor_id,
                   principal_id,
                   principal_type,
                   asset_kind,
                   identity_key,
                   direction,
                   source_type AS source,
                   status,
                   quality_status,
                   last_signal_state,
                   {valid_source_trade_expr} AS valid_source_trade_date,
                   {valid_for_trade_expr} AS valid_for_trade_date,
                   {valid_run_expr} AS valid_source_run_id,
                   source_snapshot_json->>'display_name' AS display_name,
                   source_snapshot_json->>'display_code' AS display_code,
                   created_at,
                   updated_at
            FROM {table_name}
            WHERE principal_id = %(principal_id)s
              AND principal_type = %(principal_type)s
              {user_filter_sql}
              AND identity_key = %(identity_key)s
              AND direction = %(direction)s
              AND status <> 'removed'
              {batch_filter_sql}
            LIMIT 1
            """,
            params,
        )
        row = cur.fetchone()
        if not row:
            return None
        existing = dict(row)
        requested_batch = {
            "source_trade_date": valid_source_trade_date_text,
            "for_trade_date": valid_for_trade_date_text,
            "source_run_id": valid_source_run_id_text,
        }
        if require_matching_batch and any(requested_batch.values()):
            existing_batch = {
                "source_trade_date": self._batch_text(existing.get("valid_source_trade_date")),
                "for_trade_date": self._batch_text(existing.get("valid_for_trade_date")),
                "source_run_id": self._batch_text(existing.get("valid_source_run_id")),
            }
            if existing_batch != requested_batch:
                return None
        return existing

    def _app_v2_monitor_condition_key(self, row: dict[str, Any]) -> str | None:
        value = row.get("condition_key")
        if value is not None and str(value).strip():
            return str(value)
        keys = row.get("selected_condition_keys")
        if isinstance(keys, (list, tuple)):
            return ",".join(str(item) for item in keys if str(item or "").strip()) or None
        if keys is not None and str(keys).strip():
            return str(keys)
        return None

    def _app_v2_monitor_snapshot(self, row: dict[str, Any], *, asset_kind: str) -> dict[str, Any]:
        return {
            "asset_kind": asset_kind,
            "identity_key": row.get("identity_key"),
            "display_name": row.get("display_name") or row.get("name") or row.get("board_name"),
            "display_code": row.get("display_code") or row.get("code") or row.get("board_code"),
            "display_title": row.get("display_title"),
            "condition_key": self._app_v2_monitor_condition_key(row),
            "source_run_id": row.get("source_run_id") or row.get("run_id"),
            "source_display_basis_id": (
                row.get("source_display_basis_id")
                or row.get("stock_condition_display_basis_id")
                or row.get("index_condition_display_basis_id")
                or row.get("board_condition_display_basis_id")
            ),
            "source_trade_date": row.get("source_trade_date"),
            "for_trade_date": row.get("for_trade_date"),
            "projection_run_id": row.get("projection_run_id"),
            "quality_status": row.get("quality_status"),
            "last_signal_state": row.get("last_signal_state"),
            "period_grade_y": row.get("period_grade_y") or row.get("year_overheat_level"),
            "period_grade_q": row.get("period_grade_q") or row.get("quarter_overheat_level"),
            "period_grade_m": row.get("period_grade_m") or row.get("month_overheat_level"),
            "period_grade_w": row.get("period_grade_w") or row.get("week_overheat_level"),
            "period_grade_d": row.get("period_grade_d") or row.get("day_overheat_level"),
        }

    def _app_v2_filter_where(self, filters: dict[str, Any], *, asset_kind: str) -> tuple[str, dict[str, Any]]:
        where_clauses = ["TRUE"]
        params: dict[str, Any] = {}
        direct_filters = {
            "asset_kind": "asset_kind",
            "quality_status": "quality_status",
            "source_run_id": "run_id",
        }
        if asset_kind == "board":
            direct_filters["board_type"] = "board_type"
        for key, column in direct_filters.items():
            value = normalize_filter_value(filters.get(key))
            if value:
                params[key] = value
                where_clauses.append(f"{column} = %({key})s")
        direction = normalize_filter_value(filters.get("direction"))
        if direction:
            params["direction"] = direction
            where_clauses.append("%(direction)s = ANY(selected_directions)")
        condition_key = normalize_filter_value(filters.get("condition_key"))
        if condition_key:
            params["condition_key"] = condition_key
            where_clauses.append("%(condition_key)s = ANY(selected_condition_keys)")
        period_filters = {
            "year_overheat_level": "period_transition_y",
            "quarter_overheat_level": "period_transition_q",
            "month_overheat_level": "period_transition_m",
            "week_overheat_level": "period_transition_w",
            "day_overheat_level": "period_transition_d",
        }
        for key, expression in period_filters.items():
            values = normalize_filter_values(filters.get(key))
            if len(values) == 1:
                params[key] = values[0]
                where_clauses.append(f"{expression} = %({key})s")
            elif values:
                params[key] = values
                where_clauses.append(f"{expression} = ANY(%({key})s)")
        keyword = normalize_filter_value(filters.get("q"))
        if keyword:
            params["q_like"] = f"%{keyword}%"
            if asset_kind == "board":
                search_columns = ("board_code", "board_name", "identity_key")
            elif asset_kind == "stock":
                search_columns = ("code", "name", "identity_key", "industry.industry_code")
            else:
                search_columns = ("code", "name", "identity_key")
            where_clauses.append(
                "("
                + " OR ".join(f"{column} ILIKE %(q_like)s" for column in search_columns)
                + ")"
            )
        expected_return_min = app_v2_expected_return_threshold(filters.get("buy_expected_return_pct_min"))
        if expected_return_min is not None:
            params["buy_expected_return_pct_min"] = expected_return_min
            where_clauses.append("buy_expected_return_pct >= %(buy_expected_return_pct_min)s")
        return " AND ".join(where_clauses), params

    def _ui_v1_signal_select_list(self) -> str:
        actual_trigger_period_expr = self._ui_v1_actual_trigger_period_expr()
        return f"""
               p.user_signal_projection_id,
               c.user_signal_card_id,
               q.user_notification_queue_id,
               p.user_projection_run_id,
               {self._ui_v1_event_type_expr()} AS event_type,
               {self._ui_v1_trade_date_expr()} AS trade_date,
               {self._ui_v1_event_time_expr()} AS event_time,
               p.asset_kind,
               p.identity_key,
               p.code,
               p.name,
               p.direction,
               p.signal_type,
               {self._ui_v1_action_state_expr()} AS action_state,
               COALESCE(NULLIF(p.action_mark, ''), NULLIF(c.action_mark, ''), NULLIF(q.action_mark, ''), '—') AS action_mark,
               c.card_status,
               {self._ui_v1_blocked_reason_expr()} AS blocked_reason,
               COALESCE(e.payload_json->>'trigger_kind', p.source_payload_json->'payload_json'->>'trigger_kind', c.card_payload_json->>'trigger_kind', p.display_payload_json->>'trigger_kind', p.trace_json->>'trigger_kind', p.source_payload_json->>'trigger_kind') AS trigger_kind,
               COALESCE(NULLIF(p.condition_key, ''), NULLIF(c.condition_key, ''), NULLIF(q.condition_key, ''), p.display_payload_json->>'condition_key', p.trace_json->>'condition_key') AS condition_key,
               COALESCE(NULLIF(p.original_condition_key, ''), NULLIF(c.original_condition_key, ''), NULLIF(q.original_condition_key, ''), p.display_payload_json->>'original_condition_key', p.trace_json->>'original_condition_key') AS original_condition_key,
               {actual_trigger_period_expr} AS primary_trigger_period,
               COALESCE(p.trace_json->>'trigger_time', p.source_payload_json->>'event_time', p.created_at::text) AS trigger_time,
               q.queue_status,
               CASE
                 WHEN q.queue_status = 'ready_for_future_push' THEN 'preview'
                 WHEN q.queue_status = 'queued_only' THEN 'not_delivered'
                 ELSE COALESCE(q.queue_status, 'not_delivered')
               END AS delivery_status,
               q.notification_source,
               q.channel,
               q.title,
               q.message,
               q.notification_payload_json,
               p.source_action_run_id,
               p.source_event_id,
               COALESCE(NULLIF(p.source_action_event_id, ''), NULLIF(c.source_action_event_id, ''), NULLIF(q.source_action_event_id, ''), e.event_id) AS source_action_event_id,
               {self._ui_v1_event_type_expr()} AS source_action_event_type,
               COALESCE(e.payload_json->>'source_trigger_run_id', c.card_payload_json->>'source_n4_run_id', p.trace_json->>'source_n4_run_id', p.source_payload_json->>'source_n4_run_id', p.trace_json->>'source_trigger_run_id') AS source_n4_run_id,
               COALESCE(e.payload_json->>'source_trigger_event_id', p.source_payload_json->'payload_json'->>'source_trigger_event_id', p.trace_json#>>'{{condition_provenance,source_trigger_event_ids,0}}') AS n4_trigger_event_id,
               COALESCE(e.payload_json->>'confirmation_status', p.source_payload_json->'payload_json'->>'confirmation_status', p.trace_json->>'confirmation_status', p.action_state, c.action_state) AS source_action_status,
               {self._ui_v1_trigger_price_expr()} AS trigger_price,
               {self._ui_v1_triggered_periods_expr(actual_trigger_period_expr)} AS triggered_periods,
               {self._ui_v1_baseline_source_expr(actual_trigger_period_expr)} AS baseline_source,
               COALESCE(c.target_price, p.target_price) AS target_price,
               COALESCE(c.current_price, p.current_price) AS current_price,
               COALESCE(c.expected_return_pct, p.expected_return_pct) AS expected_return_pct,
               COALESCE(c.board_code, p.board_code) AS board_code,
               COALESCE(c.board_name, p.board_name) AS board_name,
               c.card_payload_json,
               p.display_payload_json,
               true AS rollback_safe,
               p.created_at
        """

    def _ui_v1_signal_from_sql(self) -> str:
        return """
                FROM user_signal_projection p
                LEFT JOIN common_event_outbox e
                  ON e.event_id = p.source_event_id
                LEFT JOIN user_signal_card c
                  ON c.user_signal_projection_id = p.user_signal_projection_id
                 AND c.user_id = p.user_id
                LEFT JOIN LATERAL (
                  SELECT q.user_notification_queue_id,
                         q.notification_source,
                         q.queue_status,
                         q.channel,
                         q.title,
                         q.message,
                         q.notification_payload_json,
                         q.source_action_event_id,
                         q.source_action_event_type,
                         q.action_state,
                         q.action_mark,
                         q.condition_key,
                         q.original_condition_key
                  FROM user_notification_queue q
                  WHERE q.user_signal_projection_id = p.user_signal_projection_id
                    AND q.user_id = p.user_id
                  ORDER BY
                    CASE WHEN q.queue_status = 'queued_only' THEN 0 ELSE 1 END,
                    q.created_at DESC,
                    q.user_notification_queue_id DESC
                  LIMIT 1
                ) q ON true
        """

    def _ui_v1_signal_where(self, user_id: int, filters: dict[str, Any]) -> tuple[str, dict[str, Any]]:
        params: dict[str, Any] = {
            "user_id": user_id,
            "stale_source_action_run_ids": list(stale_source_action_run_ids()),
        }
        where_clauses = [
            "p.user_id = %(user_id)s",
            """
            (
              p.source_action_run_id IS NULL
              OR NOT (p.source_action_run_id = ANY(%(stale_source_action_run_ids)s))
            )
            """,
            """
            p.user_projection_run_id = (
              SELECT upr.user_projection_run_id
              FROM user_projection_run upr
              WHERE upr.status = 'passed'
              ORDER BY COALESCE(upr.finished_at, upr.updated_at, upr.created_at) DESC,
                       upr.user_projection_run_id DESC
              LIMIT 1
            )
            """,
        ]
        filter_specs = (
            ("trade_date", self._ui_v1_trade_date_expr()),
            ("event_type", self._ui_v1_event_type_expr()),
            ("asset_kind", "p.asset_kind"),
            ("direction", "p.direction"),
            ("signal_type", "p.signal_type"),
            ("action_state", self._ui_v1_action_state_expr()),
            ("blocked_reason", self._ui_v1_blocked_reason_expr()),
        )
        for key, expression in filter_specs:
            value = normalize_filter_value(filters.get(key))
            if value:
                params[key] = value
                where_clauses.append(f"{expression} = %({key})s")

        time_field = normalize_time_field(filters.get("time_field"))
        time_expression = "p.created_at" if time_field == "created_at" else f"({self._ui_v1_event_time_expr()})::timestamptz"
        date_from = normalize_filter_value(filters.get("date_from"))
        if date_from:
            params["date_from"] = date_from
            where_clauses.append(f"(({time_expression}) AT TIME ZONE 'Asia/Shanghai')::date >= %(date_from)s::date")
        date_to = normalize_filter_value(filters.get("date_to"))
        if date_to:
            params["date_to"] = date_to
            where_clauses.append(f"(({time_expression}) AT TIME ZONE 'Asia/Shanghai')::date <= %(date_to)s::date")

        keyword = normalize_filter_value(filters.get("q"))
        if keyword:
            params["q_like"] = f"%{keyword}%"
            condition_expr = "COALESCE(NULLIF(p.condition_key, ''), NULLIF(c.condition_key, ''), NULLIF(q.condition_key, ''), p.display_payload_json->>'condition_key', p.trace_json->>'condition_key')"
            where_clauses.append(
                f"""
                (
                  p.code ILIKE %(q_like)s
                  OR p.name ILIKE %(q_like)s
                  OR p.identity_key ILIKE %(q_like)s
                  OR p.source_event_id ILIKE %(q_like)s
                  OR p.source_action_event_id ILIKE %(q_like)s
                  OR {condition_expr} ILIKE %(q_like)s
                )
                """
            )
        return "\n                  AND ".join(where_clauses), params

    def _ui_v1_trade_date_expr(self) -> str:
        return "COALESCE(p.display_payload_json->>'trade_date', p.source_payload_json->>'trade_date', c.card_payload_json->>'trade_date', p.trace_json->>'trade_date')"

    def _ui_v1_event_type_expr(self) -> str:
        return "COALESCE(NULLIF(e.event_type, ''), NULLIF(p.source_action_event_type, ''), NULLIF(c.source_action_event_type, ''), NULLIF(q.source_action_event_type, ''), p.source_event_type)"

    def _ui_v1_event_time_expr(self) -> str:
        return "COALESCE(e.payload_json->>'event_time', p.source_payload_json->>'event_time', p.source_payload_json->'payload_json'->>'event_time', p.trace_json->>'event_time', p.display_payload_json->>'event_time', c.card_payload_json->>'event_time', p.created_at::text)"

    def _ui_v1_action_state_expr(self) -> str:
        return """
            COALESCE(
              NULLIF(p.action_state, ''),
              NULLIF(c.action_state, ''),
              NULLIF(q.action_state, ''),
              CASE
                WHEN p.source_action_event_type = 'ActionBlocked' THEN 'blocked'
                WHEN p.source_action_event_type = 'ActionExecuted' THEN 'executed'
                WHEN p.source_action_event_type = 'ActionEligible' THEN 'eligible'
                WHEN p.source_action_event_type = 'ActionSkipped' THEN 'skipped'
                ELSE NULL
              END,
              p.projection_status,
              c.card_status
            )
        """

    def _ui_v1_blocked_reason_expr(self) -> str:
        return "COALESCE(e.payload_json->>'blocked_reason', c.card_payload_json->>'blocked_reason', p.display_payload_json->>'blocked_reason', p.source_payload_json->'payload_json'->>'blocked_reason', p.trace_json->>'blocked_reason', p.source_payload_json->>'blocked_reason')"

    def _ui_v1_actual_trigger_period_expr(self) -> str:
        return """
            COALESCE(
              e.payload_json->>'primary_trigger_period',
              e.payload_json->>'trigger_period',
              p.source_payload_json->'payload_json'->>'primary_trigger_period',
              p.source_payload_json->'payload_json'->>'trigger_period',
              c.card_payload_json->>'primary_trigger_period',
              c.card_payload_json->>'trigger_period',
              p.display_payload_json->>'primary_trigger_period'
            )
        """

    def _ui_v1_trigger_price_expr(self) -> str:
        return """
            COALESCE(
              e.payload_json->>'trigger_price',
              p.source_payload_json->'payload_json'->>'trigger_price',
              p.source_payload_json->'payload_json'->'trace_json'->>'trigger_price',
              p.trace_json->>'trigger_price',
              c.card_payload_json->>'trigger_price'
            )
        """

    def _ui_v1_triggered_periods_expr(self, actual_trigger_period_expr: str) -> str:
        return f"""
            COALESCE(
              e.payload_json->>'all_trigger_periods',
              p.source_payload_json->'payload_json'->>'all_trigger_periods',
              p.source_payload_json->'payload_json'->'trace_json'->>'all_trigger_periods',
              p.trace_json->>'all_trigger_periods',
              c.card_payload_json->>'all_trigger_periods',
              e.payload_json->>'triggered_periods',
              p.source_payload_json->'payload_json'->>'triggered_periods',
              p.source_payload_json->'payload_json'->'trace_json'->>'triggered_periods',
              p.trace_json->>'triggered_periods',
              c.card_payload_json->>'triggered_periods',
              CASE
                WHEN {actual_trigger_period_expr} IS NOT NULL
                THEN jsonb_build_array({actual_trigger_period_expr})::text
                ELSE NULL
              END
            )
        """

    def _ui_v1_baseline_source_expr(self, actual_trigger_period_expr: str) -> str:
        period_baseline_cases = "\n".join(
            f"""
              WHEN '{period}' THEN COALESCE(
                e.payload_json#>>'{{period_trigger_baseline_trace,traced_periods,{period},baseline_source}}',
                e.payload_json#>>'{{trace_json,period_trigger_baseline_trace,traced_periods,{period},baseline_source}}',
                e.payload_json#>>'{{source_market_trace,period_trigger_baseline_trace,traced_periods,{period},baseline_source}}',
                p.source_payload_json#>>'{{payload_json,period_trigger_baseline_trace,traced_periods,{period},baseline_source}}',
                p.source_payload_json#>>'{{payload_json,trace_json,period_trigger_baseline_trace,traced_periods,{period},baseline_source}}',
                p.source_payload_json#>>'{{payload_json,source_market_trace,period_trigger_baseline_trace,traced_periods,{period},baseline_source}}',
                p.trace_json#>>'{{period_trigger_baseline_trace,traced_periods,{period},baseline_source}}',
                c.card_payload_json#>>'{{period_trigger_baseline_trace,traced_periods,{period},baseline_source}}'
              )
            """
            for period in ("Y", "Q", "M", "W", "D", "30m", "120m", "5m", "1m")
        )
        return f"""
            COALESCE(
              e.payload_json->>'baseline_source',
              p.source_payload_json->'payload_json'->>'baseline_source',
              p.source_payload_json->'payload_json'->'trace_json'->>'baseline_source',
              p.trace_json->>'baseline_source',
              c.card_payload_json->>'baseline_source',
              CASE {actual_trigger_period_expr}
                {period_baseline_cases}
                ELSE NULL
              END
            )
        """

    def fetch_index_board_c1_minute_rows(self, trade_date: str) -> dict[str, list[dict[str, Any]]]:
        normalized_trade_date = str(trade_date or "").replace("-", "")
        output: dict[str, list[dict[str, Any]]] = {"index": [], "board": []}
        with self._readonly_connection() as conn, conn.cursor() as cur:
            for asset_kind, table_name, identity_column in (
                ("index", "index_minute_bar_1m", "index_identity_key"),
                ("board", "board_minute_bar_1m", "board_identity_key"),
            ):
                cur.execute(
                    f"""
                    SELECT
                      %s AS asset_kind,
                      {identity_column} AS identity_key,
                      exchange,
                      code,
                      COALESCE(display_code, code) AS display_code,
                      bar_id,
                      bar_time,
                      open,
                      high,
                      low,
                      close,
                      volume,
                      amount,
                      quality_status,
                      trade_date
                    FROM {table_name}
                    WHERE trade_date = %s
                    ORDER BY identity_key, bar_time
                    """,
                    (asset_kind, normalized_trade_date),
                )
                output[asset_kind] = [dict(row) for row in cur.fetchall()]
        return output

    def _readonly_connection(self) -> psycopg.Connection[Any]:
        return psycopg.connect(
            self.dsn,
            connect_timeout=10,
            options="-c default_transaction_read_only=on",
            row_factory=dict_row,
        )


def config_from_env() -> N6UserWebConfig:
    csrf_secret_file = (
        ""
        if "ASHARE_V3_N6_CSRF_SECRET" in os.environ
        else os.environ.get("ASHARE_V3_N6_CSRF_SECRET_FILE", "")
    )
    return N6UserWebConfig(
        dsn=os.environ.get("ASHARE_V3_POSTGRES_DSN", DEFAULT_DSN),
        cookie_secure=os.environ.get("ASHARE_V3_N6_COOKIE_SECURE", "0") == "1",
        session_ttl_seconds=int(os.environ.get("ASHARE_V3_N6_SESSION_TTL_SECONDS", str(8 * 60 * 60))),
        card_limit=int(os.environ.get("ASHARE_V3_N6_CARD_LIMIT", "500")),
        notification_limit=int(os.environ.get("ASHARE_V3_N6_NOTIFICATION_LIMIT", "500")),
        action_event_limit=int(os.environ.get("ASHARE_V3_N6_ACTION_EVENT_LIMIT", "500")),
        ui_signal_limit=int(os.environ.get("ASHARE_V3_N6_UI_SIGNAL_LIMIT", "500")),
        signal_source_user_id=int(os.environ.get("ASHARE_V3_N6_SIGNAL_SOURCE_USER_ID", "1")),
        post_close_fastlane_docs_root=os.environ.get(
            "ASHARE_V3_POST_CLOSE_FASTLANE_DOCS_ROOT",
            DEFAULT_POST_CLOSE_FASTLANE_DOCS_ROOT,
        ),
        runtime_archive_docs_root=os.environ.get(
            "ASHARE_V3_RUNTIME_ARCHIVE_DOCS_ROOT",
            DEFAULT_RUNTIME_ARCHIVE_DOCS_ROOT,
        ),
        runtime_archive_root=os.environ.get("ASHARE_V3_RUNTIME_ARCHIVE_ROOT", DEFAULT_RUNTIME_ARCHIVE_ROOT),
        rag_docs_root=os.environ.get("ASHARE_V3_RAG_DOCS_ROOT", DEFAULT_RAG_DOCS_ROOT),
        rag_sql_root=os.environ.get("ASHARE_V3_RAG_SQL_ROOT", DEFAULT_RAG_SQL_ROOT),
        scope_write_enabled=os.environ.get("ASHARE_V3_N6_SCOPE_WRITE_ENABLED", "0") == "1",
        scope_bulk_write_enabled=os.environ.get(
            "ASHARE_V3_N6_SCOPE_BULK_WRITE_ENABLED",
            "0",
        )
        == "1",
        proposal_write_enabled=os.environ.get("ASHARE_V3_N6_PROPOSAL_WRITE_ENABLED", "0") == "1",
        csrf_secret_file=csrf_secret_file,
    )


def load_n6_csrf_secret_file(path: str) -> str:
    normalized_path = str(path or "").strip()
    if not normalized_path:
        return ""
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(normalized_path, flags)
    except (OSError, ValueError):
        return ""
    try:
        file_stat = os.fstat(descriptor)
        if not stat.S_ISREG(file_stat.st_mode):
            return ""
        if file_stat.st_uid != os.geteuid():
            return ""
        if stat.S_IMODE(file_stat.st_mode) != 0o600:
            return ""
        if file_stat.st_size <= 0 or file_stat.st_size > N6_CSRF_SECRET_MAX_BYTES:
            return ""
        raw_secret = os.read(descriptor, N6_CSRF_SECRET_MAX_BYTES + 1)
        if len(raw_secret) > N6_CSRF_SECRET_MAX_BYTES or b"\x00" in raw_secret:
            return ""
        try:
            secret = raw_secret.decode("utf-8").strip()
        except UnicodeDecodeError:
            return ""
        if not secret or "\n" in secret or "\r" in secret:
            return ""
        return secret
    except OSError:
        return ""
    finally:
        os.close(descriptor)


def build_runtime_btrack_authority_repository(
    environ: Mapping[str, str] | None = None,
) -> N6BTrackAuthorityRepository | None:
    source = os.environ if environ is None else environ
    if source.get("PGSERVICE") != N6_BTRACK_WEB_DB_SERVICE:
        return None
    if "PGPASSWORD" in source:
        return None
    if any(
        key in source
        for key in (
            "ASHARE_V3_N6_BTRACK_DSN",
            "ASHARE_V3_N6_BTRACK_PASSWORD",
        )
    ):
        return None
    try:
        return PostgresN6BTrackAuthorityRepository(
            f"service={N6_BTRACK_WEB_DB_SERVICE}"
        )
    except Exception:
        return None


def strategy_center_retired_response() -> JSONResponse:
    response = JSONResponse(
        {"ok": False, "code": "strategy_center_retired"},
        status_code=410,
    )
    response.headers["Cache-Control"] = "no-store"
    return response


def create_app(
    *,
    repository: N6UserRepository | None = None,
    btrack_authority_repository: N6BTrackAuthorityRepository | None = None,
    btrack_authority_required: bool = False,
    config: N6UserWebConfig | None = None,
    password_verifier: PasswordVerifier | None = None,
    password_hasher: PasswordHasher | None = None,
) -> FastAPI:
    web_config = config or config_from_env()
    csrf_secret = (
        load_n6_csrf_secret_file(web_config.csrf_secret_file)
        if (
            web_config.scope_write_enabled
            or web_config.scope_bulk_write_enabled
            or web_config.proposal_write_enabled
        )
        else ""
    )
    repo = repository or PostgresN6UserRepository(web_config.dsn)
    verifier = password_verifier or verify_password
    hasher = password_hasher or hash_password
    templates = Jinja2Templates(directory=str(TEMPLATE_DIR))
    app = FastAPI(
        title="Ashare v3 N6 User MVP",
        description="N6 login and read-only user projection pages.",
    )
    scope_write_active = bool(
        web_config.scope_write_enabled
        and csrf_secret
        and btrack_authority_repository is not None
    )
    scope_bulk_write_active = bool(
        scope_write_active and web_config.scope_bulk_write_enabled
    )
    proposal_write_active = bool(
        web_config.proposal_write_enabled
        and csrf_secret
        and btrack_authority_repository is not None
    )

    @app.get("/n6/login", response_class=HTMLResponse)
    async def login_page(request: Request) -> HTMLResponse:
        next_target = safe_b_track_next(request.query_params.get("next"))
        return templates.TemplateResponse(request, "n6_login.html", {"request": request, "next": next_target or ""})

    @app.post("/api/n6/auth/login")
    async def login(request: Request) -> Response:
        payload = await read_login_payload(request)
        login_name = str(payload.get("login_name") or "").strip()
        password = str(payload.get("password") or "")
        requested_next = payload.get("next") or request.query_params.get("next")
        try:
            user = repo.fetch_user_for_login(login_name) if login_name else None
        except psycopg.OperationalError:
            return JSONResponse(
                {"ok": False, "error": "authentication_service_unavailable"},
                status_code=503,
                headers={"Retry-After": "5"},
            )
        if user is None or user.status != "active" or not verifier(password, user.password_hash, user.password_hash_algo):
            return JSONResponse({"ok": False, "error": "invalid_login"}, status_code=401)

        raw_token = generate_session_token()
        expires_at = utc_now() + timedelta(seconds=web_config.session_ttl_seconds)
        try:
            session = repo.create_session(
                user_id=user.user_id,
                session_token_hash=hash_session_token(raw_token),
                session_token_hash_algo=SESSION_HASH_ALGO,
                expires_at=expires_at,
                client_info=client_info_from_request(request),
            )
        except psycopg.OperationalError:
            return JSONResponse(
                {"ok": False, "error": "authentication_service_unavailable"},
                status_code=503,
                headers={"Retry-After": "5"},
            )
        next_target = login_success_location(user, requested_next)
        response = RedirectResponse(next_target, status_code=302)
        response.set_cookie(
            COOKIE_NAME,
            raw_token,
            httponly=True,
            secure=web_config.cookie_secure,
            samesite="lax",
            path="/",
            max_age=web_config.session_ttl_seconds,
        )
        return response

    @app.get("/api/n6/me")
    async def me(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        return JSONResponse({"ok": True, "user": session_user_payload(session)})

    @app.get("/api/n6/app/v1/me")
    async def app_v1_me(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        return JSONResponse(app_me_model(principal, user=session_user_payload(session)))

    @app.get("/api/n6/app/v1/account")
    async def app_v1_account(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        account = repo.fetch_app_virtual_account(
            int(principal["principal_id"]),
            str(principal["principal_type"]),
        )
        cash_snapshot = (
            repo.fetch_app_cash_snapshot(int(account["virtual_account_id"]))
            if account and account.get("virtual_account_id")
            else None
        )
        return JSONResponse(
            app_account_model(
                principal,
                user=session_user_payload(session),
                account=account,
                cash_snapshot=cash_snapshot,
            )
        )

    @app.get("/api/n6/app/v1/dashboard")
    async def app_v1_dashboard(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        return JSONResponse(build_app_dashboard_data(session, principal))

    @app.get("/api/n6/app/v1/watchlist")
    async def app_v1_watchlist(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        current_trade_date = repo.fetch_app_current_signal_trade_date()
        rows = repo.fetch_app_signals(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            filters={"trade_date": current_trade_date},
            limit=web_config.ui_signal_limit,
        ) if current_trade_date else []
        return JSONResponse(app_watchlist_model(principal, user=session_user_payload(session), rows=rows))

    @app.get("/api/n6/app/v1/signals")
    async def app_v1_signals(request: Request) -> Response:
        raw_session_token = request.cookies.get(COOKIE_NAME)
        filters = ui_v1_filters_from_request(request)
        page_limit = ui_v1_limit_from_request(
            request,
            N6_SIGNAL_PAGE_DEFAULT_LIMIT,
            max_limit=N6_SIGNAL_PAGE_MAX_LIMIT,
        )
        try:
            n6_signal_filters_with_cursor(filters)
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid_signal_page_cursor"}, status_code=400)
        session = await asyncio.to_thread(current_session_from_token, raw_session_token, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = await asyncio.to_thread(resolve_app_principal, session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        current_trade_date = await asyncio.to_thread(repo.fetch_app_current_signal_trade_date)
        if not current_trade_date:
            return JSONResponse({"ok": False, "error": "signal_current_trade_date_unavailable"}, status_code=409)
        date_policy = n6_trade_date_access_policy(
            current_trade_date=current_trade_date,
            requested_trade_date=filters.get("trade_date"),
        )
        if date_policy["blocked"]:
            return n6_trading_session_blocker_response(date_policy)
        scope_metadata = await asyncio.to_thread(
            repo.fetch_app_signal_scope_metadata,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            trade_date=current_trade_date,
        )
        filters = app_signal_filters_with_trade_date_defaults(filters, scope_metadata)
        scope_metadata = app_signal_scope_metadata_for_filters(scope_metadata, filters)
        query_filters = n6_signal_filters_with_cursor(filters)
        rows = await asyncio.to_thread(
            repo.fetch_app_signals,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            filters=query_filters,
            limit=page_limit + 1,
        )
        page_rows, pagination = n6_signal_keyset_page(rows, limit=page_limit)
        payload = app_signals_model(
            principal,
            user=session_user_payload(session),
            rows=page_rows,
            filters=filters,
            scope_metadata=scope_metadata,
        )
        payload = n6_compact_signal_payload(payload)
        payload["pagination"] = pagination
        etag = n6_signal_response_etag(
            principal_id=int(principal["principal_id"]),
            filters=filters,
            pagination=pagination,
        )
        return n6_signal_json_response(
            request,
            payload,
            etag=etag,
            watermark=str(pagination["watermark"]),
        )

    @app.get("/api/n6/app/v1/signals/stream")
    async def app_v1_signal_stream(request: Request) -> Response:
        forbidden_scope_params = {"principal_id", "principal_type", "user_id"}
        query_param_names = frozenset(request.query_params.keys())
        raw_session_token = request.cookies.get(COOKIE_NAME)
        last_event_id = request.headers.get("last-event-id")
        requested_after_id = request.query_params.get("after_id")
        filters = ui_v1_filters_from_request(request)
        if len(request.query_params.getlist("asset_kind")) > 1:
            selected_asset_kinds = app_message_asset_kinds(
                request.query_params.getlist("asset_kind")
            )
            filters["asset_kinds"] = selected_asset_kinds
            filters["asset_kind"] = (
                selected_asset_kinds[0]
                if len(selected_asset_kinds) == 1
                else None
            )
        if forbidden_scope_params.intersection(query_param_names):
            return JSONResponse({"ok": False, "error": "client_scope_not_allowed"}, status_code=400)
        session = await asyncio.to_thread(current_session_from_token, raw_session_token, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = await asyncio.to_thread(resolve_app_principal, session, repo)
        if principal is None or principal.get("principal_source") == "session_scope":
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        try:
            after_id = parse_n6_signal_sse_cursor(
                last_event_id=last_event_id,
                after_id=requested_after_id,
            )
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid_signal_sse_cursor"}, status_code=400)
        current_trade_date = await asyncio.to_thread(repo.fetch_app_current_signal_trade_date)
        requested_trade_date = normalize_filter_value(filters.get("trade_date"))
        if not current_trade_date or (requested_trade_date and requested_trade_date != current_trade_date):
            return JSONResponse(
                {
                    "ok": False,
                    "error": "signal_sse_current_trade_date_required",
                    "current_trade_date": current_trade_date or "",
                },
                status_code=409,
            )
        scope_metadata = await asyncio.to_thread(
            repo.fetch_app_signal_scope_metadata,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            trade_date=current_trade_date,
        )
        filters = app_signal_filters_with_trade_date_defaults(filters, scope_metadata)
        principal_id = int(principal["principal_id"])
        principal_type = str(principal["principal_type"])
        user_id = int(session.user_id)

        async def read_batch(cursor: int, limit: int) -> list[dict[str, Any]]:
            return await asyncio.to_thread(
                repo.fetch_app_signal_events,
                principal_id=principal_id,
                principal_type=principal_type,
                user_id=user_id,
                filters=filters,
                after_id=cursor,
                limit=limit,
            )

        stream = iter_n6_signal_sse(
            after_id=after_id,
            read_batch=read_batch,
            is_disconnected=request.is_disconnected,
        )
        return StreamingResponse(
            stream,
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache, no-store",
                "X-Accel-Buffering": "no",
            },
        )

    @app.get("/api/n6/app/v1/signals/{user_signal_projection_id}")
    async def app_v1_signal_detail(request: Request, user_signal_projection_id: int) -> JSONResponse:
        raw_session_token = request.cookies.get(COOKIE_NAME)
        session = await asyncio.to_thread(current_session_from_token, raw_session_token, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = await asyncio.to_thread(resolve_app_principal, session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        current_trade_date = await asyncio.to_thread(repo.fetch_app_current_signal_trade_date)
        if not current_trade_date:
            return JSONResponse({"ok": False, "error": "signal_current_trade_date_unavailable"}, status_code=409)
        row = await asyncio.to_thread(
            repo.fetch_app_signal_detail,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            user_signal_projection_id=user_signal_projection_id,
            trade_date=current_trade_date,
        )
        if row is None:
            return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
        return JSONResponse(app_signal_detail_model(principal, user=session_user_payload(session), row=row))

    @app.get("/api/n6/app/v1/status-monitor")
    async def app_v1_status_monitor(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        filters = ui_v1_filters_from_request(request)
        current_trade_date = repo.fetch_app_current_signal_trade_date()
        if not current_trade_date:
            return JSONResponse({"ok": False, "error": "signal_current_trade_date_unavailable"}, status_code=409)
        date_policy = n6_trade_date_access_policy(
            current_trade_date=current_trade_date,
            requested_trade_date=filters.get("trade_date"),
            live_only=True,
        )
        if date_policy["blocked"]:
            return n6_trading_session_blocker_response(date_policy)
        filters["trade_date"] = current_trade_date
        rows = repo.fetch_app_signals(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            filters=filters,
            limit=web_config.ui_signal_limit,
        ) if current_trade_date else []
        return JSONResponse(
            app_status_monitor_model(principal, user=session_user_payload(session), rows=rows, filters=filters)
        )

    @app.get("/api/n6/app/v1/proposals")
    async def app_v1_proposals(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        return JSONResponse(
            app_locked_future_module_model(
                principal, user=session_user_payload(session), module_key="proposals", component="B Track Proposals"
            )
        )

    @app.get("/api/n6/app/v1/portfolio")
    async def app_v1_portfolio(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        positions = repo.fetch_app_positions(
            int(principal["principal_id"]),
            str(principal["principal_type"]),
        )
        return JSONResponse(
            app_portfolio_model(
                principal,
                user=session_user_payload(session),
                positions=positions,
                now=n6_trading_session_now(),
                proposal_write_enabled=proposal_write_active,
            )
        )

    @app.get("/api/n6/app/v1/pnl")
    async def app_v1_pnl(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        return JSONResponse(
            app_locked_future_module_model(
                principal, user=session_user_payload(session), module_key="pnl", component="B Track PnL"
            )
        )

    @app.get("/api/n6/app/v1/ai-users")
    async def app_v1_ai_users(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        return JSONResponse(app_ai_users_model(principal, user=session_user_payload(session)))

    @app.get("/api/n6/app/v1/leaderboard")
    async def app_v1_leaderboard(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        return JSONResponse(app_leaderboard_model(principal, user=session_user_payload(session)))

    async def app_v2_message_context(
        request: Request,
    ) -> (
        tuple[
            AuthSession,
            dict[str, Any],
            list[dict[str, Any]],
            dict[str, Any],
            dict[str, Any],
            dict[str, Any],
        ]
        | JSONResponse
    ):
        raw_session_token = request.cookies.get(COOKIE_NAME)
        filters = app_v2_message_filters_from_request(request)
        page_limit = ui_v1_limit_from_request(
            request,
            N6_SIGNAL_PAGE_DEFAULT_LIMIT,
            max_limit=N6_SIGNAL_PAGE_MAX_LIMIT,
        )
        try:
            n6_signal_filters_with_cursor(filters)
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid_signal_page_cursor"}, status_code=400)
        session = await asyncio.to_thread(current_session_from_token, raw_session_token, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = await asyncio.to_thread(resolve_app_principal, session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        current_trade_date = await asyncio.to_thread(repo.fetch_app_current_signal_trade_date)
        if not current_trade_date:
            return JSONResponse({"ok": False, "error": "signal_current_trade_date_unavailable"}, status_code=409)
        date_policy = n6_trade_date_access_policy(
            current_trade_date=current_trade_date,
            requested_trade_date=filters.get("trade_date"),
        )
        if date_policy["blocked"]:
            return n6_trading_session_blocker_response(date_policy)
        scope_metadata = await asyncio.to_thread(
            repo.fetch_app_signal_scope_metadata,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            trade_date=current_trade_date,
        )
        filters = app_signal_filters_with_trade_date_defaults(filters, scope_metadata)
        scope_metadata = app_signal_scope_metadata_for_filters(scope_metadata, filters)
        query_filters = n6_signal_filters_with_cursor(filters)
        rows = await asyncio.to_thread(
            repo.fetch_app_signals,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            filters=query_filters,
            limit=page_limit + 1,
        )
        page_rows, pagination = n6_signal_keyset_page(rows, limit=page_limit)
        return session, principal, page_rows, scope_metadata, filters, pagination

    def build_app_v2_buy_messages_data(
        session: AuthSession,
        principal: dict[str, Any],
        *,
        selected_asset_kind: str | None = None,
    ) -> dict[str, Any]:
        current_trade_date = repo.fetch_app_current_signal_trade_date()
        scope_metadata = repo.fetch_app_signal_scope_metadata(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            trade_date=current_trade_date,
        )
        rows = repo.fetch_app_signals(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            filters={"trade_date": current_trade_date},
            limit=web_config.ui_signal_limit,
        ) if current_trade_date else []
        return app_v2_buy_messages_model(
            principal,
            user=session_user_payload(session),
            rows=rows,
            scope_metadata=scope_metadata,
            selected_asset_kind=selected_asset_kind,
        )

    @app.get("/api/n6/app/v2/message-dashboard")
    async def app_v2_message_dashboard(request: Request) -> Response:
        context = await app_v2_message_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, rows, scope_metadata, filters, pagination = context
        payload = build_app_v2_message_dashboard(
            principal,
            user=session_user_payload(session),
            rows=rows,
            filters=filters,
            scope_metadata=scope_metadata,
            limit=int(pagination["limit"]),
        )
        payload = n6_compact_message_dashboard_payload(payload)
        payload["pagination"] = pagination
        etag = n6_signal_response_etag(
            principal_id=int(principal["principal_id"]),
            filters=filters,
            pagination=pagination,
        )
        return n6_signal_json_response(
            request,
            payload,
            etag=etag,
            watermark=str(pagination["watermark"]),
        )

    @app.get("/api/n6/app/v2/message-dashboard/groups")
    async def app_v2_message_dashboard_groups(request: Request) -> JSONResponse:
        context = await app_v2_message_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, rows, scope_metadata, _filters, _pagination = context
        return JSONResponse(
            n6_compact_message_dashboard_payload(
                build_app_v2_message_groups(
                    principal,
                    user=session_user_payload(session),
                    rows=rows,
                    scope_metadata=scope_metadata,
                )
            )
        )

    @app.get("/api/n6/app/v2/message-dashboard/projection-status")
    async def app_v2_message_dashboard_projection_status(request: Request) -> JSONResponse:
        context = await app_v2_message_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, rows, scope_metadata, _filters, _pagination = context
        return JSONResponse(
            build_app_v2_projection_status(
                principal,
                user=session_user_payload(session),
                rows=rows,
                scope_metadata=scope_metadata,
            )
        )

    @app.get("/api/n6/app/v2/buy-messages")
    async def app_v2_buy_messages(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        asset_kind = normalize_filter_value(request.query_params.get("asset_kind"))
        if asset_kind and asset_kind not in {"index", "board", "stock"}:
            return JSONResponse({"ok": False, "error": "invalid_asset_kind"}, status_code=400)
        return JSONResponse(
            build_app_v2_buy_messages_data(
                session,
                principal,
                selected_asset_kind=asset_kind or None,
            )
        )

    @app.get("/api/n6/app/v2/filter/stocks")
    async def app_v2_filter_stocks(request: Request) -> JSONResponse:
        return app_v2_filter_response(request, "stock")

    @app.get("/api/n6/app/v2/filter/boards")
    async def app_v2_filter_boards(request: Request) -> JSONResponse:
        return app_v2_filter_response(request, "board")

    @app.get("/api/n6/app/v2/filter/indexes")
    async def app_v2_filter_indexes(request: Request) -> JSONResponse:
        return app_v2_filter_response(request, "index")

    @app.get("/api/n6/app/v2/filter/board-members")
    async def app_v2_filter_board_members(request: Request) -> JSONResponse:
        return app_v2_filter_members_response(request, "board")

    @app.get("/api/n6/app/v2/filter/index-members")
    async def app_v2_filter_index_members(request: Request) -> JSONResponse:
        return app_v2_filter_members_response(request, "index")

    @app.get("/api/n6/app/v2/filter/board-linked-stocks")
    async def app_v2_filter_board_linked_stocks(request: Request) -> JSONResponse:
        return app_v2_filter_linked_stocks_response(request, "board")

    @app.get("/api/n6/app/v2/filter/index-linked-stocks")
    async def app_v2_filter_index_linked_stocks(request: Request) -> JSONResponse:
        return app_v2_filter_linked_stocks_response(request, "index")

    @app.get("/api/n6/app/v2/membership/index/{index_identity_key:path}")
    async def app_v2_membership_index(request: Request, index_identity_key: str) -> JSONResponse:
        return app_v2_membership_response(request, "index", index_identity_key)

    @app.get("/api/n6/app/v2/membership/board/{board_identity_key:path}")
    async def app_v2_membership_board(request: Request, board_identity_key: str) -> JSONResponse:
        return app_v2_membership_response(request, "board", board_identity_key)

    @app.get("/api/n6/app/v2/monitor")
    async def app_v2_monitor(request: Request) -> JSONResponse:
        return app_v2_monitor_response(request, None)

    @app.get("/api/n6/app/v2/monitor/stocks")
    async def app_v2_monitor_stocks(request: Request) -> JSONResponse:
        return app_v2_monitor_response(request, "stock")

    @app.get("/api/n6/app/v2/monitor/boards")
    async def app_v2_monitor_boards(request: Request) -> JSONResponse:
        return app_v2_monitor_response(request, "board")

    @app.get("/api/n6/app/v2/monitor/indexes")
    async def app_v2_monitor_indexes(request: Request) -> JSONResponse:
        return app_v2_monitor_response(request, "index")

    async def app_v3_authority_context(
        request: Request,
    ) -> tuple[AuthSession, dict[str, Any], str] | JSONResponse:
        raw_session_token = request.cookies.get(COOKIE_NAME)
        session = await asyncio.to_thread(current_session_from_token, raw_session_token, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        session_token_hash = hash_session_token(raw_session_token or "")
        if btrack_authority_required and btrack_authority_repository is None:
            return JSONResponse(
                {"ok": False, "error": "btrack_db_authority_unavailable"},
                status_code=503,
            )
        if btrack_authority_repository is not None:
            try:
                authority = await asyncio.to_thread(
                    btrack_authority_repository.resolve_authority,
                    session_token_hash,
                )
            except (AttributeError, TypeError, ValueError, psycopg.Error):
                return JSONResponse(
                    {"ok": False, "error": "btrack_db_authority_unavailable"},
                    status_code=503,
                )
            if authority is None or authority.user_id != session.user_id:
                return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
            principal = authority.principal_payload()
        else:
            principal = await asyncio.to_thread(resolve_app_principal, session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        return session, principal, session_token_hash

    @app.get("/api/n6/app/v3/strategy-center")
    async def app_v3_strategy_center(_request: Request) -> JSONResponse:
        return strategy_center_retired_response()

    @app.put("/api/n6/app/v3/strategy-center/selection")
    async def app_v3_strategy_center_selection(_request: Request) -> JSONResponse:
        return strategy_center_retired_response()

    @app.get("/api/n6/app/v3/strategy-center/stream")
    async def app_v3_strategy_center_stream(_request: Request) -> JSONResponse:
        return strategy_center_retired_response()

    async def app_v3_ai_agent_public_snapshot(
        request: Request,
    ) -> tuple[AuthSession, dict[str, Any], dict[str, Any]] | JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        if btrack_authority_repository is None:
            return JSONResponse(
                {"ok": False, "error": "ai_agent_public_service_unavailable"},
                status_code=503,
            )
        session, principal, session_token_hash = context
        try:
            result = await asyncio.to_thread(
                btrack_authority_repository.fetch_public_ai_agent_dashboard,
                session_token_hash,
                decision_limit=50,
                trade_limit=50,
                summary_limit=30,
            )
        except (AttributeError, TypeError, ValueError, psycopg.Error):
            return JSONResponse(
                {"ok": False, "error": "ai_agent_public_service_unavailable"},
                status_code=503,
            )
        if not isinstance(result, dict):
            return JSONResponse(
                {"ok": False, "error": "principal_scope_unavailable"},
                status_code=403,
            )
        return session, principal, app_ai_agent_public_model(result)

    @app.get("/api/n6/app/v3/ai-agent")
    async def app_v3_ai_agent(request: Request) -> JSONResponse:
        context = await app_v3_ai_agent_public_snapshot(request)
        if isinstance(context, JSONResponse):
            return context
        _session, _principal, payload = context
        response = n6_json_response(payload)
        response.headers["Cache-Control"] = "no-store"
        return response

    @app.get("/api/n6/app/v3/ai-agent/decisions/{decision_id}")
    async def app_v3_ai_agent_decision_detail(
        request: Request,
        decision_id: int,
    ) -> JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        if btrack_authority_repository is None:
            return JSONResponse(
                {"ok": False, "error": "ai_agent_public_service_unavailable"},
                status_code=503,
            )
        _session, _principal, session_token_hash = context
        try:
            decision_id_text = canonical_bigint_id(
                decision_id,
                field_name="ai_decision_id",
                required=True,
            )
            result = await asyncio.to_thread(
                btrack_authority_repository.fetch_public_ai_decision_detail,
                session_token_hash,
                decision_id=int(decision_id_text),
            )
        except ValueError:
            return JSONResponse(
                {"ok": False, "error": "invalid_ai_decision_id"},
                status_code=400,
            )
        except (AttributeError, TypeError, psycopg.Error):
            return JSONResponse(
                {"ok": False, "error": "ai_agent_public_service_unavailable"},
                status_code=503,
            )
        payload = app_ai_agent_public_decision_detail_model(result)
        response = n6_json_response(
            payload,
            status_code=200 if payload["ok"] else 404,
        )
        response.headers["Cache-Control"] = "no-store"
        return response

    @app.get("/api/n6/app/v3/ai-agent/{section}")
    async def app_v3_ai_agent_section(
        request: Request,
        section: str,
    ) -> JSONResponse:
        if section not in {
            "overview",
            "positions",
            "trades",
            "decisions",
            "daily-summaries",
            "performance",
        }:
            return JSONResponse(
                {"ok": False, "error": "not_found"},
                status_code=404,
            )
        context = await app_v3_ai_agent_public_snapshot(request)
        if isinstance(context, JSONResponse):
            return context
        _session, _principal, payload = context
        response = n6_json_response(
            app_ai_agent_public_section_model(payload, section)
        )
        response.headers["Cache-Control"] = "no-store"
        return response

    async def app_v3_scope_write_context(
        request: Request,
    ) -> tuple[AuthSession, dict[str, Any], str] | JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        if not web_config.scope_write_enabled:
            return JSONResponse({"ok": False, "error": "scope_write_disabled"}, status_code=403)
        if not csrf_secret:
            return JSONResponse({"ok": False, "error": "csrf_not_configured"}, status_code=503)
        if btrack_authority_repository is None:
            return JSONResponse({"ok": False, "error": "btrack_db_authority_unavailable"}, status_code=503)
        session, principal, session_token_hash = context
        if not n6_csrf_valid(request, session, csrf_secret):
            return JSONResponse({"ok": False, "error": "csrf_rejected"}, status_code=403)
        return session, principal, session_token_hash

    async def app_v3_current_filter_identity(
        *,
        session: AuthSession,
        principal: dict[str, Any],
        asset_kind: str,
        identity_key: str,
    ) -> dict[str, Any]:
        return await asyncio.to_thread(
            repo.fetch_app_current_filter_identity,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            asset_kind=asset_kind,
            identity_key=identity_key,
        )

    def invalidate_app_user_scope(session: AuthSession, principal: dict[str, Any]) -> None:
        invalidator = getattr(repo, "invalidate_app_user_scope_cache", None)
        if callable(invalidator):
            invalidator(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
            )

    async def app_v3_bulk_filter_selection(
        *,
        session: AuthSession,
        principal: dict[str, Any],
        asset_kind: str,
        filters: Mapping[str, Any],
    ) -> tuple[dict[str, Any] | None, JSONResponse | None]:
        if asset_kind not in APP_V2_MONITOR_TABLE_BY_ASSET:
            return None, JSONResponse({"ok": False, "error": "invalid_asset_kind"}, status_code=400)
        canonical_filters = n6_scope_bulk_canonical_filters(filters)
        current_filters = {
            key: value
            for key, value in canonical_filters.items()
            if key != "for_trade_date"
        }
        current_result = await asyncio.to_thread(
            repo.fetch_app_filter_items,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            asset_kind=asset_kind,
            filters=current_filters,
            limit=1,
        )
        current_trade_date = next(
            (
                str(value).strip()
                for value in current_result.get("available_for_trade_dates") or []
                if str(value).strip()
            ),
            str(current_result.get("selected_for_trade_date") or "").strip(),
        )
        requested_trade_date = str(canonical_filters.get("for_trade_date") or "").strip()
        if not current_trade_date or (
            requested_trade_date and requested_trade_date != current_trade_date
        ):
            return None, JSONResponse(
                {
                    "ok": False,
                    "error": "current_for_trade_date_required",
                    "current_trade_date": current_trade_date,
                },
                status_code=409,
            )
        canonical_filters["for_trade_date"] = current_trade_date
        result = await asyncio.to_thread(
            repo.fetch_app_filter_items,
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            asset_kind=asset_kind,
            filters=canonical_filters,
            limit=N6_SCOPE_BULK_MAX_IDENTITIES,
        )
        filtered_count = int(result.get("filtered_count") or 0)
        rows = list(result.get("items") or [])
        identity_keys = sorted(
            {
                str(row.get("identity_key") or "").strip()
                for row in rows
                if str(row.get("identity_key") or "").strip()
            }
        )
        if filtered_count > N6_SCOPE_BULK_MAX_IDENTITIES:
            return None, JSONResponse(
                {
                    "ok": False,
                    "error": "bulk_scope_too_large",
                    "matched_count": filtered_count,
                    "max_count": N6_SCOPE_BULK_MAX_IDENTITIES,
                },
                status_code=409,
            )
        if (
            filtered_count < 1
            or len(rows) != filtered_count
            or len(identity_keys) != filtered_count
            or any(not value.startswith(f"{asset_kind}:") for value in identity_keys)
        ):
            return None, JSONResponse(
                {"ok": False, "error": "filter_snapshot_stale"},
                status_code=409,
            )
        source_run_id = str(result.get("source_run_id") or "").strip()
        selected_trade_date = str(result.get("selected_for_trade_date") or "").strip()
        if not source_run_id or selected_trade_date != current_trade_date:
            return None, JSONResponse(
                {"ok": False, "error": "filter_snapshot_stale"},
                status_code=409,
            )
        return {
            "asset_kind": asset_kind,
            "filters": canonical_filters,
            "for_trade_date": current_trade_date,
            "source_run_id": source_run_id,
            "identity_keys": identity_keys,
            "identity_count": len(identity_keys),
            "selection_sha256": n6_scope_bulk_selection_sha256(identity_keys),
        }, None

    @app.get("/api/n6/app/v3/filter-scope-bulk-preview")
    async def app_v3_filter_scope_bulk_preview(request: Request) -> JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        if not scope_bulk_write_active:
            return JSONResponse({"ok": False, "error": "scope_bulk_write_disabled"}, status_code=403)
        session, principal, session_token_hash = context
        target_scope = str(request.query_params.get("target") or "").strip()
        asset_kind = str(request.query_params.get("asset_kind") or "").strip()
        if target_scope not in {"monitor", "realtime"}:
            return JSONResponse({"ok": False, "error": "invalid_target_scope"}, status_code=400)
        filters = app_v2_filter_filters_from_request(request)
        selection, blocker = await app_v3_bulk_filter_selection(
            session=session,
            principal=principal,
            asset_kind=asset_kind,
            filters=filters,
        )
        if blocker is not None:
            return blocker
        assert selection is not None
        preview = await asyncio.to_thread(
            btrack_authority_repository.preview_bulk_scope,
            session_token_hash,
            target_scope=target_scope,
            asset_kind=asset_kind,
            identity_keys=selection["identity_keys"],
            for_trade_date=selection["for_trade_date"],
            source_run_id=selection["source_run_id"],
            selection_sha256=selection["selection_sha256"],
        )
        if not preview.get("ok"):
            return n6_json_response(app_v3_public_payload(preview), status_code=409)
        selection_token = n6_scope_bulk_selection_token(
            session=session,
            principal=principal,
            secret=csrf_secret,
            target_scope=target_scope,
            asset_kind=asset_kind,
            filters=selection["filters"],
            for_trade_date=selection["for_trade_date"],
            source_run_id=selection["source_run_id"],
            identity_count=selection["identity_count"],
            selection_sha256=selection["selection_sha256"],
        )
        response = n6_json_response(
            app_v3_public_payload(
                {
                    **preview,
                    "selection_token": selection_token,
                    "selection_expires_in_seconds": N6_SCOPE_BULK_SELECTION_TTL_SECONDS,
                }
            )
        )
        response.headers["Cache-Control"] = "no-store"
        return response

    async def app_v3_execute_scope_bulk(
        request: Request,
        *,
        target_scope: str,
    ) -> JSONResponse:
        context = await app_v3_scope_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        if not scope_bulk_write_active:
            return JSONResponse({"ok": False, "error": "scope_bulk_write_disabled"}, status_code=403)
        session, principal, session_token_hash = context
        payload = await read_json_object(request)
        if set(payload) != {"selection_token"}:
            return JSONResponse({"ok": False, "error": "client_scope_not_allowed"}, status_code=400)
        token_payload, token_error = n6_scope_bulk_selection_payload(
            str(payload.get("selection_token") or ""),
            session=session,
            principal=principal,
            secret=csrf_secret,
            expected_target_scope=target_scope,
        )
        if token_payload is None:
            return JSONResponse({"ok": False, "error": token_error}, status_code=409)
        selection, blocker = await app_v3_bulk_filter_selection(
            session=session,
            principal=principal,
            asset_kind=str(token_payload["asset_kind"]),
            filters=dict(token_payload["filters"]),
        )
        if blocker is not None:
            return blocker
        assert selection is not None
        if any(
            (
                selection["for_trade_date"] != str(token_payload["for_trade_date"]),
                selection["source_run_id"] != str(token_payload["source_run_id"]),
                selection["identity_count"] != int(token_payload["identity_count"]),
                selection["selection_sha256"] != str(token_payload["selection_sha256"]),
            )
        ):
            return JSONResponse({"ok": False, "error": "filter_snapshot_stale"}, status_code=409)
        writer = (
            btrack_authority_repository.bulk_upsert_monitor_items
            if target_scope == "monitor"
            else btrack_authority_repository.bulk_upsert_realtime_scope_items
        )
        result = await asyncio.to_thread(
            writer,
            session_token_hash,
            asset_kind=selection["asset_kind"],
            identity_keys=selection["identity_keys"],
            for_trade_date=selection["for_trade_date"],
            source_run_id=selection["source_run_id"],
            selection_sha256=selection["selection_sha256"],
        )
        if not result.get("ok"):
            return n6_json_response(app_v3_public_payload(result), status_code=409)
        invalidate_app_user_scope(session, principal)
        return n6_json_response(app_v3_public_payload(result))

    @app.post("/api/n6/app/v3/monitor-items/bulk")
    async def app_v3_add_monitor_items_bulk(request: Request) -> JSONResponse:
        return await app_v3_execute_scope_bulk(request, target_scope="monitor")

    @app.post("/api/n6/app/v3/realtime-scope-items/bulk")
    async def app_v3_add_realtime_scope_items_bulk(request: Request) -> JSONResponse:
        return await app_v3_execute_scope_bulk(request, target_scope="realtime")

    @app.get("/api/n6/app/v3/realtime-scope-items")
    async def app_v3_realtime_scope_items(request: Request) -> JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, session_token_hash = context
        if btrack_authority_repository is not None:
            result = await asyncio.to_thread(
                btrack_authority_repository.list_realtime_scope_items,
                session_token_hash,
                limit=500,
            )
        else:
            result = await asyncio.to_thread(
                repo.fetch_app_realtime_scope,
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
            )
        return n6_json_response(
            app_realtime_scope_model(
                principal,
                user=session_user_payload(session),
                result=result,
                write_enabled=scope_write_active,
            )
        )

    @app.post("/api/n6/app/v3/monitor-items")
    async def app_v3_add_monitor_item(request: Request) -> JSONResponse:
        context = await app_v3_scope_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, session_token_hash = context
        payload = await read_json_object(request)
        if set(payload) - {"asset_kind", "identity_key", "for_trade_date"}:
            return JSONResponse({"ok": False, "error": "client_scope_not_allowed"}, status_code=400)
        asset_kind = str(payload.get("asset_kind") or "").strip()
        identity_key = str(payload.get("identity_key") or "").strip()
        requested_trade_date = str(payload.get("for_trade_date") or "").strip()
        if asset_kind not in APP_V2_MONITOR_TABLE_BY_ASSET or not identity_key:
            return JSONResponse({"ok": False, "error": "invalid_monitor_request"}, status_code=400)
        current_filter_identity = await app_v3_current_filter_identity(
            session=session,
            principal=principal,
            asset_kind=asset_kind,
            identity_key=identity_key,
        )
        current_trade_date = str(current_filter_identity.get("for_trade_date") or "").strip()
        if not current_trade_date or (requested_trade_date and requested_trade_date != current_trade_date):
            return JSONResponse(
                {"ok": False, "error": "current_for_trade_date_required", "current_trade_date": current_trade_date},
                status_code=409,
            )
        if not current_filter_identity.get("approved"):
            return JSONResponse({"ok": False, "error": "source_not_found"}, status_code=409)
        directions = ("buy",) if asset_kind == "stock" else ("buy", "sell")
        results = [
            await asyncio.to_thread(
                btrack_authority_repository.upsert_monitor_item,
                session_token_hash,
                asset_kind=asset_kind,
                identity_key=identity_key,
                direction=direction,
                for_trade_date=current_trade_date,
            )
            for direction in directions
        ]
        result = next((item for item in results if not item.get("ok")), None) or {
            "ok": True,
            "status": "active",
            "added_count": len(results),
            "results": results,
        }
        if not result.get("ok"):
            return n6_json_response(app_v3_public_payload(result), status_code=409)
        invalidate_app_user_scope(session, principal)
        return n6_json_response(
            app_v3_public_payload(
                {"ok": True, "asset_kind": asset_kind, "identity_key": identity_key, **result}
            )
        )

    @app.delete("/api/n6/app/v3/monitor-items/{monitor_id}")
    async def app_v3_remove_monitor_item(request: Request, monitor_id: int) -> JSONResponse:
        context = await app_v3_scope_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, session_token_hash = context
        try:
            monitor_id = int(canonical_bigint_id(monitor_id, field_name="monitor_id", required=True))
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid_monitor_id"}, status_code=400)
        result = await asyncio.to_thread(
            btrack_authority_repository.remove_monitor_item,
            session_token_hash,
            monitor_id=monitor_id,
        )
        if result.get("ok"):
            invalidate_app_user_scope(session, principal)
        return n6_json_response(
            app_v3_public_payload(result),
            status_code=200 if result.get("ok") else 404,
        )

    @app.post("/api/n6/app/v3/realtime-scope-items")
    async def app_v3_add_realtime_scope_item(request: Request) -> JSONResponse:
        context = await app_v3_scope_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, session_token_hash = context
        payload = await read_json_object(request)
        if set(payload) - {"asset_kind", "identity_key"}:
            return JSONResponse({"ok": False, "error": "client_scope_not_allowed"}, status_code=400)
        asset_kind = str(payload.get("asset_kind") or "").strip()
        identity_key = str(payload.get("identity_key") or "").strip()
        if asset_kind not in APP_V2_MONITOR_TABLE_BY_ASSET or not identity_key:
            return JSONResponse({"ok": False, "error": "invalid_realtime_scope_request"}, status_code=400)
        current_filter_identity = await app_v3_current_filter_identity(
            session=session,
            principal=principal,
            asset_kind=asset_kind,
            identity_key=identity_key,
        )
        current_trade_date = str(current_filter_identity.get("for_trade_date") or "").strip()
        if not current_trade_date:
            return JSONResponse(
                {"ok": False, "error": "current_for_trade_date_required"},
                status_code=409,
            )
        if not current_filter_identity.get("approved"):
            return JSONResponse({"ok": False, "error": "source_not_found"}, status_code=409)
        result = await asyncio.to_thread(
            btrack_authority_repository.upsert_realtime_scope_item,
            session_token_hash,
            asset_kind=asset_kind,
            identity_key=identity_key,
            for_trade_date=current_trade_date,
        )
        if result.get("ok"):
            invalidate_app_user_scope(session, principal)
        return n6_json_response(
            app_v3_public_payload(result),
            status_code=200 if result.get("ok") else 409,
        )

    @app.delete("/api/n6/app/v3/realtime-scope-items/{scope_id}")
    async def app_v3_remove_realtime_scope_item(request: Request, scope_id: int) -> JSONResponse:
        context = await app_v3_scope_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, session_token_hash = context
        try:
            scope_id = int(canonical_bigint_id(scope_id, field_name="realtime_scope_id", required=True))
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid_realtime_scope_id"}, status_code=400)
        result = await asyncio.to_thread(
            btrack_authority_repository.remove_realtime_scope_item,
            session_token_hash,
            realtime_scope_id=scope_id,
        )
        if result.get("ok"):
            invalidate_app_user_scope(session, principal)
        return n6_json_response(
            app_v3_public_payload(result),
            status_code=200 if result.get("ok") else 404,
        )

    async def app_v3_proposal_write_context(
        request: Request,
    ) -> tuple[AuthSession, dict[str, Any], str] | JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        if not web_config.proposal_write_enabled:
            return JSONResponse({"ok": False, "error": "proposal_write_disabled"}, status_code=403)
        if not csrf_secret:
            return JSONResponse({"ok": False, "error": "csrf_not_configured"}, status_code=503)
        if btrack_authority_repository is None:
            return JSONResponse({"ok": False, "error": "btrack_db_authority_unavailable"}, status_code=503)
        session, principal, session_token_hash = context
        if not n6_csrf_valid(request, session, csrf_secret):
            return JSONResponse({"ok": False, "error": "csrf_rejected"}, status_code=403)
        return session, principal, session_token_hash

    @app.get("/api/n6/app/v3/virtual-account/proposals")
    async def app_v3_virtual_account_proposals(request: Request) -> JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, session_token_hash = context
        if btrack_authority_repository is not None:
            result = await asyncio.to_thread(
                btrack_authority_repository.list_trade_proposals,
                session_token_hash,
                limit=100,
            )
        else:
            result = await asyncio.to_thread(
                repo.fetch_app_trade_proposals,
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                limit=100,
            )
        return n6_json_response(
            app_trade_proposals_model(
                principal,
                user=session_user_payload(session),
                result=result,
                write_enabled=proposal_write_active,
            )
        )

    @app.post("/api/n6/app/v3/virtual-account/proposals")
    async def app_v3_create_virtual_account_proposal(request: Request) -> JSONResponse:
        context = await app_v3_proposal_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        _session, _principal, session_token_hash = context
        payload = await read_json_object(request)
        if set(payload) != {"source_type", "source_id"}:
            return JSONResponse({"ok": False, "error": "source_only_request_required"}, status_code=400)
        source_type = str(payload.get("source_type") or "")
        source_id_field = {
            "signal": "user_signal_projection_id",
            "manual_position": "virtual_position_id",
        }.get(source_type)
        if source_id_field is None or not isinstance(payload.get("source_id"), str):
            return JSONResponse({"ok": False, "error": "canonical_source_id_required"}, status_code=400)
        try:
            source_id = canonical_bigint_id(
                payload.get("source_id"),
                field_name=source_id_field,
                required=True,
            )
        except ValueError:
            return JSONResponse({"ok": False, "error": "canonical_source_id_required"}, status_code=400)
        try:
            result = await asyncio.to_thread(
                btrack_authority_repository.create_trade_proposal,
                session_token_hash,
                source_type=source_type,
                source_id=int(source_id),
            )
        except psycopg.Error:
            return JSONResponse(
                {"ok": False, "error": "proposal_service_unavailable"},
                status_code=503,
            )
        return n6_json_response(
            app_v3_public_payload(result),
            status_code=201 if result.get("ok") else 409,
        )

    @app.post("/api/n6/app/v3/virtual-account/proposals/{proposal_id}/confirm")
    async def app_v3_confirm_virtual_account_proposal(request: Request, proposal_id: int) -> JSONResponse:
        context = await app_v3_proposal_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        _session, _principal, session_token_hash = context
        try:
            proposal_id = int(canonical_bigint_id(proposal_id, field_name="proposal_id", required=True))
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid_proposal_id"}, status_code=400)
        payload = await read_json_object(request)
        if payload:
            return JSONResponse({"ok": False, "error": "empty_confirm_body_required"}, status_code=400)
        try:
            result = await asyncio.to_thread(
                btrack_authority_repository.confirm_trade_proposal,
                session_token_hash,
                proposal_id=proposal_id,
                idempotency_key=str(request.headers.get("idempotency-key") or ""),
            )
        except psycopg.Error:
            return JSONResponse(
                {"ok": False, "error": "proposal_service_unavailable"},
                status_code=503,
            )
        return n6_json_response(
            app_v3_public_payload(result),
            status_code=200 if result.get("ok") else 409,
        )

    @app.post("/api/n6/app/v3/virtual-account/proposals/cancel")
    async def app_v3_cancel_virtual_account_proposals(request: Request) -> JSONResponse:
        context = await app_v3_proposal_write_context(request)
        if isinstance(context, JSONResponse):
            return context
        _session, _principal, session_token_hash = context
        payload = await read_json_object(request)
        if set(payload) != {"proposal_ids"}:
            return JSONResponse(
                {"ok": False, "error": "proposal_ids_only_request_required"},
                status_code=400,
            )
        raw_ids = payload.get("proposal_ids")
        if not isinstance(raw_ids, list) or not raw_ids or len(raw_ids) > 100:
            return JSONResponse(
                {"ok": False, "error": "invalid_proposal_ids"},
                status_code=400,
            )
        try:
            proposal_ids = [
                int(
                    canonical_bigint_id(
                        value,
                        field_name="proposal_id",
                        required=True,
                    )
                )
                for value in raw_ids
            ]
        except (TypeError, ValueError):
            return JSONResponse(
                {"ok": False, "error": "invalid_proposal_ids"},
                status_code=400,
            )
        if len(set(proposal_ids)) != len(proposal_ids):
            return JSONResponse(
                {"ok": False, "error": "duplicate_proposal_ids"},
                status_code=400,
            )
        try:
            result = await asyncio.to_thread(
                btrack_authority_repository.cancel_trade_proposals,
                session_token_hash,
                proposal_ids=proposal_ids,
            )
        except psycopg.Error:
            return JSONResponse(
                {"ok": False, "error": "proposal_service_unavailable"},
                status_code=503,
            )
        return n6_json_response(
            app_v3_public_payload(result),
            status_code=200 if result.get("ok") else 409,
        )

    @app.get("/api/n6/app/v3/virtual-account/trades")
    async def app_v3_virtual_account_trades(request: Request) -> JSONResponse:
        context = await app_v3_authority_context(request)
        if isinstance(context, JSONResponse):
            return context
        session, principal, session_token_hash = context
        if btrack_authority_repository is not None:
            result = await asyncio.to_thread(
                btrack_authority_repository.list_virtual_trades,
                session_token_hash,
                limit=200,
            )
        else:
            result = await asyncio.to_thread(
                repo.fetch_app_virtual_trades,
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                limit=200,
            )
        return n6_json_response(
            app_virtual_trades_model(principal, user=session_user_payload(session), result=result)
        )

    @app.get("/api/n6/ui/v1/signals")
    async def ui_v1_signals(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        filters = ui_v1_filters_from_request(request)
        limit = ui_v1_limit_from_request(request, 100)
        offset = ui_v1_offset_from_request(request)
        source_user_id = signal_source_user_id(session, web_config)
        rows = repo.fetch_ui_v1_signals(source_user_id, filters, limit, offset)
        pagination = {
            "total_count": repo.count_ui_v1_signals(source_user_id, {}),
            "filtered_count": repo.count_ui_v1_signals(source_user_id, filters),
            "limit": limit,
            "offset": offset,
        }
        statistics = repo.fetch_ui_v1_signal_statistics(source_user_id)
        return JSONResponse(signal_list_model(rows, filters=filters, pagination=pagination, statistics=statistics))

    @app.get("/api/n6/ui/v1/signals/{user_signal_projection_id}")
    async def ui_v1_signal_detail(request: Request, user_signal_projection_id: int) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        row = repo.fetch_ui_v1_signal_detail(signal_source_user_id(session, web_config), user_signal_projection_id)
        if row is None:
            return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
        artifacts = repo.fetch_ui_v1_artifacts()
        rollback_summary = repo.fetch_ui_v1_rollback_summary()
        return JSONResponse(signal_detail_model(row, artifacts=artifacts, rollback_summary=rollback_summary))

    @app.get("/api/n6/ui/v1/dashboard/metrics")
    async def ui_v1_dashboard_metrics(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        metrics = repo.fetch_ui_v1_dashboard_metrics(signal_source_user_id(session, web_config))
        return JSONResponse(dashboard_metrics_model(metrics))

    @app.get("/api/n6/ui/v1/virtual-account")
    async def ui_v1_virtual_account(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        account = repo.fetch_ui_v1_virtual_account(session.user_id)
        snapshot = repo.fetch_ui_v1_cash_snapshot(session.user_id)
        return JSONResponse(virtual_account_summary_model(account, cash_snapshot=snapshot))

    @app.get("/api/n6/ui/v1/cash-snapshot")
    async def ui_v1_cash_snapshot(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        snapshot = repo.fetch_ui_v1_cash_snapshot(session.user_id)
        return JSONResponse(cash_snapshot_model(snapshot))

    @app.get("/api/n6/ui/v1/cash-ledger")
    async def ui_v1_cash_ledger(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        limit = cash_ledger_limit_from_request(request)
        rows = repo.fetch_ui_v1_cash_ledger(session.user_id, limit)
        return JSONResponse(cash_ledger_model(rows))

    @app.get("/api/n6/ui/v1/message-dashboard")
    async def ui_v1_message_dashboard(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        dashboard = repo.fetch_message_dashboard(web_config.action_event_limit)
        return JSONResponse(message_dashboard_model(dashboard))

    @app.get("/api/n6/ui/v1/lineage-stats")
    async def ui_v1_lineage_stats(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        stats = repo.fetch_ui_v1_lineage_stats(signal_source_user_id(session, web_config))
        return JSONResponse(lineage_stats_model(stats))

    @app.get("/api/n6/ui/v1/n3-messages")
    async def ui_v1_n3_messages(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N3_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        data = repo.fetch_ui_v1_n3_messages(
            filters=raw_message_filters_from_request(request),
            limit=limit,
            include_all=include_all,
        )
        return JSONResponse(n3_messages_model(data))

    @app.get("/api/n6/ui/v1/n4-messages")
    async def ui_v1_n4_messages(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N4_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        data = repo.fetch_ui_v1_n4_messages(
            filters=raw_message_filters_from_request(request),
            limit=limit,
            include_all=include_all,
        )
        return JSONResponse(n4_messages_model(data))

    @app.get("/api/n6/ui/v1/n5-messages")
    async def ui_v1_n5_messages(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N5_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        data = repo.fetch_ui_v1_n5_messages(
            filters=raw_message_filters_from_request(request),
            limit=limit,
            include_all=include_all,
        )
        return JSONResponse(n5_messages_model(data))

    @app.get("/api/n6/ui/v1/n5-actions")
    async def ui_v1_n5_actions(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        limit = ui_v1_limit_from_request(request, N5_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        data = repo.fetch_ui_v1_n5_actions(
            filters=n5_action_filters_from_request(request),
            limit=limit,
        )
        return JSONResponse(data)

    @app.get("/api/n6/ui/v1/input-messages")
    async def ui_v1_input_messages(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N6_INPUT_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        data = repo.fetch_ui_v1_input_messages(
            filters=raw_message_filters_from_request(request),
            limit=limit,
            include_all=include_all,
        )
        return JSONResponse(input_messages_model(data))

    @app.get("/api/n6/ui/v1/post-close-fastlane-status")
    async def ui_v1_post_close_fastlane_status(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        data = read_post_close_fastlane_status(
            docs_root=web_config.post_close_fastlane_docs_root,
            for_trade_date=request.query_params.get("for_trade_date"),
        )
        return JSONResponse(post_close_fastlane_status_model(data))

    @app.get("/api/n6/ui/v1/rag-search")
    async def ui_v1_rag_search(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        limit = ui_v1_limit_from_request(request, 8, max_limit=20)
        data = read_rag_status_answer(
            docs_root=web_config.rag_docs_root,
            sql_root=web_config.rag_sql_root,
            query=str(request.query_params.get("q") or ""),
            limit=limit,
            layer_role=request.query_params.get("layer_role"),
            trade_date=request.query_params.get("trade_date"),
            artifact_type=request.query_params.get("artifact_type"),
        )
        return JSONResponse(rag_search_model(data))

    @app.get("/api/n6/ui/v1/archive-status")
    async def ui_v1_archive_status(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        data = read_runtime_archive_status(
            docs_root=web_config.runtime_archive_docs_root,
            archive_root=web_config.runtime_archive_root,
            trade_date=request.query_params.get("trade_date"),
        )
        return JSONResponse(runtime_archive_status_model(data))

    @app.get("/api/n6/ui/v1/n2-condition-basis")
    async def ui_v1_n2_condition_basis(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        asset_kind = str(request.query_params.get("asset_kind") or "index").strip().lower()
        if asset_kind not in N2_CONDITION_BASIS_ASSET_META:
            return JSONResponse({"ok": False, "error": "invalid_asset_kind"}, status_code=400)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N2_CONDITION_BASIS_DEFAULT_LIMIT, max_limit=5000)
        data = repo.fetch_ui_v1_n2_condition_basis(
            asset_kind=asset_kind,
            filters=n2_condition_basis_filters_from_request(request),
            limit=limit,
            include_all=include_all,
        )
        return JSONResponse(n2_condition_basis_model(data))

    @app.get("/api/n6/ui/v1/status-monitor")
    async def ui_v1_status_monitor(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        filters = status_monitor_filters_from_request(request)
        limit = ui_v1_limit_from_request(request, 100)
        offset = ui_v1_offset_from_request(request)
        data = repo.fetch_ui_v1_status_monitor(
            signal_source_user_id(session, web_config),
            filters,
            limit,
            offset,
        )
        return JSONResponse(status_monitor_model(data, filters=filters))

    @app.get("/api/n6/ui/v1/artifacts")
    async def ui_v1_artifacts(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        return JSONResponse(artifacts_model(repo.fetch_ui_v1_artifacts()))

    @app.get("/api/n6/ui/v1/rollback-summary")
    async def ui_v1_rollback_summary(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        return JSONResponse(rollback_summary_model(repo.fetch_ui_v1_rollback_summary()))

    @app.post("/api/n6/auth/logout")
    async def logout(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        response = JSONResponse({"ok": True})
        response.delete_cookie(COOKIE_NAME, path="/")
        if session is None:
            return response
        repo.revoke_session(session.session_token_hash, utc_now())
        return response

    @app.post("/api/n6/admin/users")
    async def admin_create_user(request: Request) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        payload = await read_login_payload(request)
        login_name = normalize_login_name(payload.get("login_name"))
        display_name = normalize_optional_text(payload.get("display_name"))
        role = str(payload.get("role") or "user").strip()
        password = str(payload.get("password") or "")
        blockers = validate_user_create_payload(login_name=login_name, role=role)
        blockers.extend(validate_password(password))
        if blockers:
            return JSONResponse({"ok": False, "error": "invalid_user_payload", "blockers": blockers}, status_code=400)
        try:
            hash_result = hasher(password)
            user_row = repo.create_user_with_defaults(
                login_name=login_name,
                display_name=display_name,
                role=role,
                password_hash=hash_result.password_hash,
                password_hash_algo=hash_result.password_hash_algo,
                created_by_user_id=session.user_id,
            )
        except UserManagementError as exc:
            return JSONResponse({"ok": False, "error": exc.code, "blockers": [exc.code]}, status_code=409)
        except ValueError as exc:
            code = str(exc) or "user_create_failed"
            status = 409 if code == "login_name_exists" else 400
            return JSONResponse({"ok": False, "error": code, "blockers": [code]}, status_code=status)
        except RuntimeError:
            return JSONResponse({"ok": False, "error": "password_hash_unavailable"}, status_code=500)
        return JSONResponse(
            {
                "ok": True,
                "user": user_view(user_row),
                "created": {
                    "user_account": 1,
                    "user_filter_profile": 1,
                    "user_sim_account": 1,
                    "n6_principal": 1,
                    "n6_virtual_account": 1
                    if user_row.get("virtual_account_provisioning_status") == "created"
                    else 0,
                    "n6_principal_account": 1
                    if user_row.get("virtual_account_provisioning_status") == "created"
                    else 0,
                    "n6_virtual_cash_ledger": 1
                    if user_row.get("virtual_account_provisioning_status") == "created"
                    else 0,
                    "n6_virtual_cash_snapshot": 1
                    if user_row.get("virtual_account_provisioning_status") == "created"
                    else 0,
                },
                "password_value_logged": False,
                "password_hash_logged": False,
            }
        )

    @app.post("/api/n6/admin/users/{target_user_id}/delete")
    async def admin_delete_user(request: Request, target_user_id: int) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        if session.role != "admin":
            return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
        if target_user_id == session.user_id:
            return JSONResponse({"ok": False, "error": "cannot_delete_self"}, status_code=400)
        try:
            user_row = repo.delete_user(target_user_id=target_user_id, deleted_by_user_id=session.user_id)
        except UserManagementError as exc:
            return JSONResponse({"ok": False, "error": exc.code, "blockers": [exc.code]}, status_code=404)
        except ValueError as exc:
            code = str(exc) or "user_not_found"
            return JSONResponse({"ok": False, "error": code, "blockers": [code]}, status_code=404)
        return JSONResponse({"ok": True, "user": user_view(user_row), "soft_deleted": True, "sessions_revoked": True})

    def build_app_dashboard_data(session: AuthSession, principal: dict[str, Any]) -> dict[str, Any]:
        user = session_user_payload(session)
        principal_id = int(principal["principal_id"])
        principal_type = str(principal["principal_type"])
        current_trade_date = repo.fetch_app_current_signal_trade_date()
        signals = repo.fetch_app_signals(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=session.user_id,
            filters={"trade_date": current_trade_date},
            limit=web_config.ui_signal_limit,
        ) if current_trade_date else []
        account = repo.fetch_app_virtual_account(principal_id, principal_type)
        cash_snapshot = (
            repo.fetch_app_cash_snapshot(int(account["virtual_account_id"]))
            if account and account.get("virtual_account_id")
            else None
        )
        monitor_result = repo.fetch_app_monitor_items(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=session.user_id,
            limit=500,
            monitor_status="all",
            for_trade_date=current_trade_date or "",
        )
        realtime_scope_result = repo.fetch_app_realtime_scope(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=session.user_id,
        )
        positions = repo.fetch_app_positions(principal_id, principal_type)
        proposal_result = repo.fetch_app_trade_proposals(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=session.user_id,
            limit=100,
        )
        trade_result = repo.fetch_app_virtual_trades(
            principal_id=principal_id,
            principal_type=principal_type,
            user_id=session.user_id,
            limit=200,
        )
        return app_dashboard_model(
            principal,
            user=user,
            account=account,
            cash_snapshot=cash_snapshot,
            signal_rows=signals,
            current_trade_date=current_trade_date or "",
            trading_session_active=n6_is_trading_session_for_trade_date(current_trade_date),
            monitor_result=monitor_result,
            realtime_scope_result=realtime_scope_result,
            positions=positions,
            proposal_result=proposal_result,
            trade_result=trade_result,
            scope_write_enabled=scope_write_active,
            scope_bulk_write_enabled=scope_bulk_write_active,
            proposal_write_enabled=proposal_write_active,
        )

    def app_v2_filter_response(request: Request, asset_kind: str) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        filters = {key: value for key, value in app_v2_filter_filters_from_request(request).items() if value}
        date_policy = app_v2_filter_trade_date_policy(session, principal, asset_kind, filters)
        if date_policy["blocked"]:
            return n6_trading_session_blocker_response(date_policy)
        result = repo.fetch_app_filter_items(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            asset_kind=asset_kind,
            filters=filters,
            limit=ui_v1_limit_from_request(request, 100, max_limit=200),
        )
        return JSONResponse(
            app_v2_filter_api_model(
                app_v2_filter_model(
                    principal,
                    user=session_user_payload(session),
                    asset_kind=asset_kind,
                    result=result,
                    filters=filters,
                    base_href=f"/n6/app/filter-center/{APP_FILTER_CENTER_PAGE_BY_ASSET[asset_kind]}",
                    write_enabled=scope_write_active,
                ),
                asset_kind=asset_kind,
            )
        )

    def app_v2_filter_trade_date_policy(
        session: AuthSession,
        principal: dict[str, Any],
        asset_kind: str,
        filters: dict[str, Any],
    ) -> dict[str, Any]:
        requested_trade_date = normalize_filter_value(filters.get("for_trade_date"))
        if not requested_trade_date:
            return n6_trade_date_access_policy(current_trade_date=None, requested_trade_date=None)
        current_filters = {key: value for key, value in filters.items() if key != "for_trade_date" and value}
        current_result = repo.fetch_app_filter_items(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            asset_kind=asset_kind,
            filters=current_filters,
            limit=1,
        )
        current_trade_date = ""
        for value in current_result.get("available_for_trade_dates") or []:
            text = normalize_filter_value(value)
            if text:
                current_trade_date = text
                break
        if not current_trade_date:
            current_trade_date = normalize_filter_value(current_result.get("selected_for_trade_date"))
        return n6_trade_date_access_policy(
            current_trade_date=current_trade_date,
            requested_trade_date=requested_trade_date,
        )

    def app_v2_filter_members_response(request: Request, membership_kind: str) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        parent_identity_key = normalize_filter_value(request.query_params.get("identity_key")) or ""
        result = repo.fetch_app_filter_members(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            membership_kind=membership_kind,
            parent_identity_key=parent_identity_key,
            limit=ui_v1_limit_from_request(request, 100, max_limit=500),
        )
        return JSONResponse(
            app_v2_filter_members_model(
                principal,
                user=session_user_payload(session),
                membership_kind=membership_kind,
                parent_identity_key=parent_identity_key,
                result=result,
            )
        )

    def app_v2_filter_linked_stocks_response(request: Request, membership_kind: str) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        identity_param = f"{membership_kind}_identity_key"
        parent_identity_key = (
            normalize_filter_value(request.query_params.get(identity_param))
            or normalize_filter_value(request.query_params.get("identity_key"))
            or ""
        )
        view = normalize_filter_value(request.query_params.get("view")) or "matched"
        if view not in {"matched", "all"}:
            view = "matched"
        result = repo.fetch_app_filter_linked_stocks(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            membership_kind=membership_kind,
            parent_identity_key=parent_identity_key,
            limit=ui_v1_limit_from_request(request, 100, max_limit=500),
            view=view,
        )
        return JSONResponse(
            app_v2_filter_linked_stocks_model(
                principal,
                user=session_user_payload(session),
                membership_kind=membership_kind,
                parent_identity_key=parent_identity_key,
                result=result,
            )
        )

    def app_v2_membership_response(request: Request, entity_type: str, identity_key: str) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        result = repo.fetch_app_membership_stocks(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            entity_type=entity_type,
            identity_key=identity_key,
            limit=ui_v1_limit_from_request(request, 500, max_limit=500),
        )
        return JSONResponse(
            app_v2_membership_drilldown_model(
                principal,
                user=session_user_payload(session),
                entity_type=entity_type,
                identity_key=identity_key,
                result=result,
            )
        )

    def app_v2_monitor_response(request: Request, asset_kind: str | None) -> JSONResponse:
        session = current_session(request, repo)
        if session is None:
            return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
        principal = resolve_app_principal(session, repo)
        if principal is None:
            return JSONResponse({"ok": False, "error": "principal_scope_unavailable"}, status_code=403)
        requested_trade_date = normalize_filter_value(request.query_params.get("for_trade_date"))
        current_trade_date = repo.fetch_app_current_signal_trade_date()
        if not current_trade_date:
            return JSONResponse({"ok": False, "error": "signal_current_trade_date_unavailable"}, status_code=409)
        date_policy = n6_trade_date_access_policy(
            current_trade_date=current_trade_date,
            requested_trade_date=requested_trade_date,
        )
        if date_policy["blocked"]:
            return n6_trading_session_blocker_response(date_policy)
        historical_readonly = bool(requested_trade_date and requested_trade_date != current_trade_date)
        result = repo.fetch_app_monitor_items(
            principal_id=int(principal["principal_id"]),
            principal_type=str(principal["principal_type"]),
            user_id=session.user_id,
            asset_kind=asset_kind,
            limit=500,
            monitor_status=app_v2_monitor_status_from_request(request),
            for_trade_date=requested_trade_date or "",
        )
        return JSONResponse(
            app_v2_monitor_model(
                principal,
                user=session_user_payload(session),
                result=result,
                selected_asset_kind=asset_kind,
                write_enabled=scope_write_active and not historical_readonly,
            )
        )

    def build_app_page_data(
        page_key: str,
        session: AuthSession,
        principal: dict[str, Any],
        *,
        app_filters: dict[str, Any] | None = None,
        app_show_all: bool = False,
    ) -> dict[str, Any]:
        user = session_user_payload(session)
        if page_key in {"home", "dashboard"}:
            return build_app_dashboard_data(session, principal)
        if page_key == "guide":
            return app_user_guide_model(principal, user=user)
        if page_key == "account":
            account = repo.fetch_app_virtual_account(
                int(principal["principal_id"]),
                str(principal["principal_type"]),
            )
            cash_snapshot = (
                repo.fetch_app_cash_snapshot(int(account["virtual_account_id"]))
                if account and account.get("virtual_account_id")
                else None
            )
            return app_account_model(principal, user=user, account=account, cash_snapshot=cash_snapshot)
        if page_key == "realtime-scope":
            result = repo.fetch_app_realtime_scope(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
            )
            return app_realtime_scope_model(
                principal,
                user=user,
                result=result,
                write_enabled=scope_write_active,
            )
        if page_key == "trade-log":
            result = repo.fetch_app_virtual_trades(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                limit=200,
            )
            return app_virtual_trades_model(principal, user=user, result=result)
        if page_key == "signals":
            page_limit = max(
                1,
                min(
                    int((app_filters or {}).get("page_limit") or N6_SIGNAL_PAGE_DEFAULT_LIMIT),
                    N6_SIGNAL_PAGE_MAX_LIMIT,
                ),
            )
            current_trade_date = normalize_filter_value((app_filters or {}).get("trade_date"))
            scope_metadata = repo.fetch_app_signal_scope_metadata(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                trade_date=current_trade_date,
            )
            filters = app_signal_filters_with_trade_date_defaults(
                {
                    key: value
                    for key, value in dict(app_filters or {}).items()
                    if key != "page_limit"
                },
                scope_metadata,
                enforce_trading_session=True,
            )
            scope_metadata = app_signal_scope_metadata_for_filters(scope_metadata, filters)
            query_filters = n6_signal_filters_with_cursor(filters)
            rows = repo.fetch_app_signals(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                filters=query_filters,
                limit=page_limit + 1,
            )
            page_rows, pagination = n6_signal_keyset_page(rows, limit=page_limit)
            payload = app_signals_model(
                principal,
                user=user,
                rows=page_rows,
                filters=filters,
                scope_metadata=scope_metadata,
            )
            payload = n6_compact_signal_payload(payload)
            payload["pagination"] = pagination
            return payload
        if page_key == "messages":
            page_limit = max(
                1,
                min(
                    int((app_filters or {}).get("page_limit") or N6_SIGNAL_PAGE_DEFAULT_LIMIT),
                    N6_SIGNAL_PAGE_MAX_LIMIT,
                ),
            )
            current_trade_date = normalize_filter_value((app_filters or {}).get("trade_date"))
            scope_metadata = repo.fetch_app_signal_scope_metadata(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                trade_date=current_trade_date,
            )
            filters = app_signal_filters_with_trade_date_defaults(
                {
                    key: value
                    for key, value in dict(app_filters or {}).items()
                    if key != "page_limit"
                },
                scope_metadata,
                enforce_trading_session=True,
            )
            scope_metadata = app_signal_scope_metadata_for_filters(scope_metadata, filters)
            query_filters = n6_signal_filters_with_cursor(filters)
            rows = repo.fetch_app_signals(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                filters=query_filters,
                limit=page_limit + 1,
            )
            page_rows, pagination = n6_signal_keyset_page(rows, limit=page_limit)
            payload = build_app_v2_message_dashboard(
                principal,
                user=user,
                rows=page_rows,
                filters=filters,
                scope_metadata=scope_metadata,
                limit=page_limit,
            )
            payload = n6_compact_message_dashboard_payload(payload)
            payload["pagination"] = pagination
            return payload
        if page_key == "buy-messages":
            return build_app_v2_buy_messages_data(session, principal)
        if page_key in APP_FILTER_CENTER_ASSET_BY_PAGE_KEY:
            asset_kind = APP_FILTER_CENTER_ASSET_BY_PAGE_KEY[page_key]
            filters = {key: value for key, value in (app_filters or {}).items() if value}
            date_policy = app_v2_filter_trade_date_policy(session, principal, asset_kind, filters)
            if date_policy["blocked"]:
                filters["for_trade_date"] = date_policy["effective_trade_date"]
            selected_result = repo.fetch_app_filter_items(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                asset_kind=asset_kind,
                filters=filters,
                limit=(
                    APP_V2_FILTER_PAGE_MAX_ROWS
                    if app_show_all
                    else app_v2_filter_default_limit(asset_kind)
                ),
            )
            if date_policy["blocked"]:
                selected_result["date_policy_blocker"] = date_policy["blocker"]
                selected_result["date_policy_message"] = date_policy["message"]
            empty_result = {"cache_ready": False, "items": []}
            filter_results = {
                "index": empty_result,
                "board": empty_result,
                "stock": empty_result,
                asset_kind: selected_result,
            }
            return app_v2_filter_center_model(
                principal,
                user=user,
                stock_result=filter_results["stock"],
                board_result=filter_results["board"],
                index_result=filter_results["index"],
                selected_asset_kind=asset_kind,
                filters=filters,
                show_all=app_show_all,
                write_enabled=scope_write_active,
                bulk_write_enabled=scope_bulk_write_active,
            )
        if page_key in APP_MONITOR_ASSET_BY_PAGE_KEY:
            asset_kind = APP_MONITOR_ASSET_BY_PAGE_KEY[page_key]
            monitor_status = app_v2_monitor_status_filter((app_filters or {}).get("monitor_status"))
            historical_readonly = bool((app_filters or {}).get("historical_readonly"))
            result = repo.fetch_app_monitor_items(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                asset_kind=asset_kind,
                limit=500,
                monitor_status=monitor_status,
                for_trade_date=str((app_filters or {}).get("for_trade_date") or "").strip(),
            )
            return app_v2_monitor_model(
                principal,
                user=user,
                result=result,
                selected_asset_kind=asset_kind,
                write_enabled=scope_write_active and not historical_readonly,
            )
        if page_key == "status-monitor":
            current_trade_date = repo.fetch_app_current_signal_trade_date()
            rows = repo.fetch_app_signals(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                filters={"trade_date": current_trade_date},
                limit=web_config.ui_signal_limit,
            ) if current_trade_date else []
            return app_status_monitor_model(principal, user=user, rows=rows, filters={})
        if page_key == "watchlist":
            current_trade_date = repo.fetch_app_current_signal_trade_date()
            rows = repo.fetch_app_signals(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                filters={"trade_date": current_trade_date},
                limit=web_config.ui_signal_limit,
            ) if current_trade_date else []
            return app_watchlist_model(principal, user=user, rows=rows)
        if page_key == "proposals":
            result = repo.fetch_app_trade_proposals(
                principal_id=int(principal["principal_id"]),
                principal_type=str(principal["principal_type"]),
                user_id=session.user_id,
                limit=100,
            )
            return app_trade_proposals_model(
                principal,
                user=user,
                result=result,
                write_enabled=proposal_write_active,
            )
        if page_key == "portfolio":
            positions = repo.fetch_app_positions(
                int(principal["principal_id"]),
                str(principal["principal_type"]),
            )
            return app_portfolio_model(
                principal,
                user=user,
                positions=positions,
                now=n6_trading_session_now(),
                proposal_write_enabled=proposal_write_active,
            )
        if page_key == "pnl":
            return app_locked_future_module_model(principal, user=user, module_key="pnl", component="B Track PnL")
        if page_key == "leaderboard":
            return app_leaderboard_model(principal, user=user)
        if page_key == "ai-users":
            return app_ai_users_model(principal, user=user)
        component = {
            "watchlist": "B Track Watchlist",
            "proposals": "B Track Proposals",
            "strategy-center": "B Track Strategy Center",
        }.get(page_key, "B Track Dashboard")
        return app_empty_planned_model(principal, user=user, component=component)

    @app.get("/n6/app", response_class=HTMLResponse)
    async def app_shell_home(request: Request) -> Response:
        return await app_shell_page(request, "home")

    @app.get("/n6/app/filter-center/{filter_page}", response_class=HTMLResponse)
    async def app_filter_center_page(request: Request, filter_page: str) -> Response:
        page_key = APP_FILTER_CENTER_PAGE_KEY_BY_SLUG.get(filter_page)
        if page_key is None:
            return HTMLResponse("not found", status_code=404)
        return await app_shell_page(request, page_key)

    @app.get("/n6/app/my-monitor/{monitor_page}", response_class=HTMLResponse)
    async def app_my_monitor_page(request: Request, monitor_page: str) -> Response:
        page_key = APP_MONITOR_PAGE_KEY_BY_SLUG.get(monitor_page)
        if page_key is None:
            return HTMLResponse("not found", status_code=404)
        return await app_shell_page(request, page_key)

    @app.get("/n6/app/{page_key}", response_class=HTMLResponse)
    async def app_shell_page(request: Request, page_key: str) -> Response:
        if page_key == "strategy-center":
            return RedirectResponse(
                "/n6/app/signals?notice=strategy_center_retired",
                status_code=307,
            )
        if page_key == "watchlist":
            return RedirectResponse("/n6/app/my-monitor", status_code=307)
        if page_key == "buy-messages":
            return RedirectResponse("/n6/app/signals", status_code=307)
        if page_key not in {
            "dashboard",
            "guide",
            "account",
            "signals",
            "messages",
            "realtime-scope",
            "trade-log",
            "filter-center",
            "filter-center:indexes",
            "filter-center:boards",
            "filter-center:stocks",
            "my-monitor",
            "my-monitor:stocks",
            "my-monitor:boards",
            "my-monitor:indexes",
            "status-monitor",
            "proposals",
            "portfolio",
            "pnl",
            "ai-agent",
            "ai-users",
            "leaderboard",
            "home",
        }:
            return HTMLResponse("not found", status_code=404)
        raw_session_token = request.cookies.get(COOKIE_NAME)
        is_filter_center_page = page_key in APP_FILTER_CENTER_ASSET_BY_PAGE_KEY
        is_monitor_page = page_key in APP_MONITOR_ASSET_BY_PAGE_KEY
        app_filters = (
            app_v2_filter_filters_from_request(request)
            if is_filter_center_page
            else app_v2_monitor_filters_from_request(request)
            if is_monitor_page
            else app_v2_message_filters_from_request(request)
            if page_key == "messages"
            else ui_v1_filters_from_request(request)
            if page_key == "signals"
            else None
        )
        if page_key in {"signals", "messages"}:
            try:
                n6_signal_filters_with_cursor(app_filters or {})
            except ValueError:
                return HTMLResponse("invalid_signal_page_cursor", status_code=400)
            app_filters = dict(app_filters or {})
            app_filters["page_limit"] = ui_v1_limit_from_request(
                request,
                N6_SIGNAL_PAGE_DEFAULT_LIMIT,
                max_limit=N6_SIGNAL_PAGE_MAX_LIMIT,
            )
        app_show_all = query_param_enabled(request, "show_all") if is_filter_center_page else False
        session = await asyncio.to_thread(current_session_from_token, raw_session_token, repo)
        if session is None:
            return RedirectResponse(b_track_login_location(request), status_code=302)
        if page_key == "ai-agent":
            if btrack_authority_repository is None:
                return HTMLResponse("AI模拟投资员服务暂不可用", status_code=503)
            session_token_hash = hash_session_token(raw_session_token or "")
            try:
                authority = await asyncio.to_thread(
                    btrack_authority_repository.resolve_authority,
                    session_token_hash,
                )
                result = await asyncio.to_thread(
                    btrack_authority_repository.fetch_public_ai_agent_dashboard,
                    session_token_hash,
                    decision_limit=50,
                    trade_limit=50,
                    summary_limit=30,
                )
            except (AttributeError, TypeError, ValueError, psycopg.Error):
                return HTMLResponse("AI模拟投资员服务暂不可用", status_code=503)
            if (
                authority is None
                or authority.user_id != session.user_id
                or not isinstance(result, dict)
            ):
                return HTMLResponse("principal_scope_unavailable", status_code=403)
            principal = authority.principal_payload()
            data = app_ai_agent_public_model(result)
        else:
            principal = await asyncio.to_thread(resolve_app_principal, session, repo)
            if principal is None:
                return HTMLResponse("principal_scope_unavailable", status_code=403)
            if page_key in {"signals", "messages"}:
                current_trade_date = await asyncio.to_thread(repo.fetch_app_current_signal_trade_date)
                if not current_trade_date:
                    return HTMLResponse("signal_current_trade_date_unavailable", status_code=409)
                date_policy = n6_trade_date_access_policy(
                    current_trade_date=current_trade_date,
                    requested_trade_date=(app_filters or {}).get("trade_date"),
                )
                if date_policy["blocked"]:
                    return HTMLResponse(N6_TRADING_SESSION_HISTORY_MESSAGE, status_code=409)
                app_filters = dict(app_filters or {})
                app_filters["trade_date"] = date_policy["effective_trade_date"]
            if is_monitor_page:
                current_trade_date = await asyncio.to_thread(repo.fetch_app_current_signal_trade_date)
                if not current_trade_date:
                    return HTMLResponse("signal_current_trade_date_unavailable", status_code=409)
                requested_trade_date = normalize_filter_value((app_filters or {}).get("for_trade_date"))
                date_policy = n6_trade_date_access_policy(
                    current_trade_date=current_trade_date,
                    requested_trade_date=requested_trade_date,
                )
                if date_policy["blocked"]:
                    return HTMLResponse(N6_TRADING_SESSION_HISTORY_MESSAGE, status_code=409)
                app_filters = dict(app_filters or {})
                app_filters["historical_readonly"] = bool(
                    requested_trade_date and requested_trade_date != current_trade_date
                )
            data = await asyncio.to_thread(
                build_app_page_data,
                page_key,
                session,
                principal,
                app_filters=app_filters,
                app_show_all=app_show_all,
            )
        display_page_key = (
            "filter-center"
            if page_key.startswith("filter-center:")
            else "my-monitor"
            if is_monitor_page
            else page_key
        )
        page_model = app_page_model(
            display_page_key,
            principal,
            user=session_user_payload(session),
            data=data,
        )
        page_model["csrf_token"] = (
            n6_csrf_token(session, csrf_secret)
            if scope_write_active or proposal_write_active
            else ""
        )
        page_model["scope_write_enabled"] = scope_write_active
        page_model["scope_bulk_write_enabled"] = scope_bulk_write_active
        page_model["proposal_write_enabled"] = proposal_write_active
        page_model["ux_safety_status"] = (
            "投影只读模式；"
            f"本人监控范围{'可管理' if scope_write_active else '管理未启用'}；"
            + (
                "交易申请可用；申请由受限 Web 创建，成交由独立 N6 executor 处理（不代表 executor 正在运行）"
                if proposal_write_active
                else "交易申请未启用"
            )
            + "；不连接真实券商；不执行真实下单"
        )
        if display_page_key in {"home", "dashboard", "guide"}:
            page_model["ux_safety_status"] = (
                "投影只读模式工作台；交易申请状态按页面能力显示；"
                "executor 本页未验证；不连接真实券商；不执行真实下单"
            )
        return templates.TemplateResponse(
            request,
            "n6_app_shell.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "page": page_model,
            },
        )

    @app.get("/n6/", response_class=HTMLResponse)
    async def home(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        return RedirectResponse(a_track_default_location(session), status_code=303)

    @app.get("/n6/portfolio", response_class=HTMLResponse)
    async def portfolio(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        return RedirectResponse(a_track_default_location(session), status_code=303)

    @app.get("/n6/notifications", response_class=HTMLResponse)
    async def notifications(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        return RedirectResponse(a_track_default_location(session), status_code=303)

    @app.get("/n6/action-events", response_class=HTMLResponse)
    async def action_events(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        dashboard = repo.fetch_message_dashboard(web_config.action_event_limit)
        rows = dashboard.get("messages") or []
        filters = ui_v1_filters_from_request(request)
        limit = ui_v1_limit_from_request(request, 100)
        offset = ui_v1_offset_from_request(request)
        source_user_id = signal_source_user_id(session, web_config)
        signal_rows = repo.fetch_ui_v1_signals(source_user_id, filters, limit, offset)
        signal_page = signal_list_model(
            signal_rows,
            filters=filters,
            pagination={
                "total_count": repo.count_ui_v1_signals(source_user_id, {}),
                "filtered_count": repo.count_ui_v1_signals(source_user_id, filters),
                "limit": limit,
                "offset": offset,
            },
            statistics=repo.fetch_ui_v1_signal_statistics(source_user_id),
        )
        lineage_stats = lineage_stats_model(repo.fetch_ui_v1_lineage_stats(source_user_id))
        virtual_account = repo.fetch_ui_v1_virtual_account(session.user_id)
        cash_snapshot = repo.fetch_ui_v1_cash_snapshot(session.user_id)
        return templates.TemplateResponse(
            request,
            "n6_action_events.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "action_events"),
                "events": [action_event_view(row) for row in rows],
                "summary": summarize_action_events(rows),
                "dashboard": message_dashboard_view(dashboard),
                "signal_page": signal_page,
                "lineage_stats": lineage_stats,
                "virtual_account_summary": virtual_account_summary_view(virtual_account, cash_snapshot),
            },
        )

    @app.get("/n6/status-monitor", response_class=HTMLResponse)
    async def status_monitor(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        filters = status_monitor_filters_from_request(request)
        limit = ui_v1_limit_from_request(request, 100)
        offset = ui_v1_offset_from_request(request)
        monitor = status_monitor_model(
            repo.fetch_ui_v1_status_monitor(
                signal_source_user_id(session, web_config),
                filters,
                limit,
                offset,
            ),
            filters=filters,
        )
        return templates.TemplateResponse(
            request,
            "n6_status_monitor.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "status_monitor"),
                "monitor": monitor,
            },
        )

    @app.get("/n6/n3-messages", response_class=HTMLResponse)
    async def n3_messages(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N3_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        n3_page = n3_messages_model(
            repo.fetch_ui_v1_n3_messages(
                filters=raw_message_filters_from_request(request),
                limit=limit,
                include_all=include_all,
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_n3_messages.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "n3_messages"),
                "n3_page": n3_page,
            },
        )

    @app.get("/n6/n4-messages", response_class=HTMLResponse)
    async def n4_messages(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N4_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        n4_page = n4_messages_model(
            repo.fetch_ui_v1_n4_messages(
                filters=raw_message_filters_from_request(request),
                limit=limit,
                include_all=include_all,
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_n4_messages.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "n4_messages"),
                "n4_page": n4_page,
            },
        )

    @app.get("/n6/n5-messages", response_class=HTMLResponse)
    async def n5_messages(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N5_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        n5_page = n5_messages_model(
            repo.fetch_ui_v1_n5_messages(
                filters=raw_message_filters_from_request(request),
                limit=limit,
                include_all=include_all,
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_n5_messages.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "n5_messages"),
                "n5_page": n5_page,
            },
        )

    @app.get("/n6/n5-actions", response_class=HTMLResponse)
    async def n5_actions(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        limit = ui_v1_limit_from_request(request, N5_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        n5_actions_page = repo.fetch_ui_v1_n5_actions(
            filters=n5_action_filters_from_request(request),
            limit=limit,
        )
        return templates.TemplateResponse(
            request,
            "n6_n5_actions.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "n5_actions"),
                "n5_actions_page": n5_actions_page,
            },
        )

    @app.get("/n6/input-messages", response_class=HTMLResponse)
    async def input_messages(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N6_INPUT_MESSAGE_DEFAULT_LIMIT, max_limit=5000)
        input_page = input_messages_model(
            repo.fetch_ui_v1_input_messages(
                filters=raw_message_filters_from_request(request),
                limit=limit,
                include_all=include_all,
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_input_messages.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "input_messages"),
                "input_page": input_page,
            },
        )

    @app.get("/n6/post-close-fastlane-status", response_class=HTMLResponse)
    async def post_close_fastlane_status(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        status_page = post_close_fastlane_status_model(
            read_post_close_fastlane_status(
                docs_root=web_config.post_close_fastlane_docs_root,
                for_trade_date=request.query_params.get("for_trade_date"),
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_post_close_fastlane_status.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "post_close_fastlane_status"),
                "status_page": status_page,
            },
        )

    @app.get("/n6/rag", response_class=HTMLResponse)
    async def rag_search(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        limit = ui_v1_limit_from_request(request, 8, max_limit=20)
        rag_page = rag_search_model(
            read_rag_status_answer(
                docs_root=web_config.rag_docs_root,
                sql_root=web_config.rag_sql_root,
                query=str(request.query_params.get("q") or ""),
                limit=limit,
                layer_role=request.query_params.get("layer_role"),
                trade_date=request.query_params.get("trade_date"),
                artifact_type=request.query_params.get("artifact_type"),
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_rag.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "rag"),
                "rag_page": rag_page,
            },
        )

    @app.get("/n6/archive-status", response_class=HTMLResponse)
    async def archive_status(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        archive_page = runtime_archive_status_model(
            read_runtime_archive_status(
                docs_root=web_config.runtime_archive_docs_root,
                archive_root=web_config.runtime_archive_root,
                trade_date=request.query_params.get("trade_date"),
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_archive_status.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "archive_status"),
                "archive_page": archive_page,
            },
        )

    @app.get("/n6/fastlane-status", response_class=HTMLResponse)
    async def post_close_fastlane_status_alias(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        return RedirectResponse("/n6/post-close-fastlane-status", status_code=303)

    @app.get("/n6/admin/n5-messages", response_class=HTMLResponse)
    async def admin_n5_messages_alias(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        return RedirectResponse("/n6/n5-messages", status_code=303)

    @app.get("/n6/n2-condition-basis", response_class=HTMLResponse)
    async def n2_condition_basis_default(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        return RedirectResponse("/n6/n2-condition-basis/index", status_code=303)

    @app.get("/n6/n2-condition-basis/export-latest.xlsx")
    async def n2_condition_basis_export_latest(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        data = repo.fetch_ui_v1_n2_condition_basis_latest_export()
        return n2_condition_basis_latest_export_response(data)

    @app.get("/n6/n2-condition-basis/{asset_kind}", response_class=HTMLResponse)
    async def n2_condition_basis(request: Request, asset_kind: str) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        asset_kind = str(asset_kind or "").strip().lower()
        if asset_kind not in N2_CONDITION_BASIS_ASSET_META:
            return HTMLResponse("not found", status_code=404)
        include_all = query_param_enabled(request, "show_all")
        limit = ui_v1_limit_from_request(request, N2_CONDITION_BASIS_DEFAULT_LIMIT, max_limit=5000)
        n2_page = n2_condition_basis_model(
            repo.fetch_ui_v1_n2_condition_basis(
                asset_kind=asset_kind,
                filters=n2_condition_basis_filters_from_request(request),
                limit=limit,
                include_all=include_all,
            )
        )
        return templates.TemplateResponse(
            request,
            "n6_n2_condition_basis.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "n2_condition_basis"),
                "n2_page": n2_page,
            },
        )

    @app.get("/n6/admin/account", response_class=HTMLResponse)
    async def admin_account(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        account = repo.fetch_ui_v1_virtual_account(session.user_id)
        snapshot = repo.fetch_ui_v1_cash_snapshot(session.user_id)
        ledger = repo.fetch_ui_v1_cash_ledger(session.user_id, 20)
        return templates.TemplateResponse(
            request,
            "n6_admin_account.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "admin_account"),
                "account": virtual_account_view(account),
                "cash_snapshot": virtual_cash_snapshot_view(snapshot),
                "cash_ledger": [virtual_cash_ledger_view(row) for row in ledger],
                "safety": [
                    "READ ONLY",
                    "NO ORDER",
                    "NO TRADE",
                    "NO POSITION UPDATE",
                    "NO REAL TRADE",
                    "NOT INVESTMENT ADVICE",
                ],
            },
        )

    @app.get("/n6/admin/users", response_class=HTMLResponse)
    async def admin_users(request: Request) -> Response:
        session = current_session(request, repo)
        if session is None:
            return RedirectResponse("/n6/login", status_code=303)
        if session.role != "admin":
            return HTMLResponse("forbidden", status_code=403)
        return templates.TemplateResponse(
            request,
            "n6_admin_users.html",
            {
                "request": request,
                "user": session_user_payload(session),
                "nav": nav_context(session, "admin_users"),
                "users": [user_view(row) for row in repo.fetch_users_for_admin()],
            },
        )

    allowed_b_track_write_routes = {
        ("/api/n6/app/v3/monitor-items", frozenset({"POST"})),
        ("/api/n6/app/v3/monitor-items/bulk", frozenset({"POST"})),
        ("/api/n6/app/v3/monitor-items/{monitor_id}", frozenset({"DELETE"})),
        ("/api/n6/app/v3/realtime-scope-items", frozenset({"POST"})),
        ("/api/n6/app/v3/realtime-scope-items/bulk", frozenset({"POST"})),
        ("/api/n6/app/v3/realtime-scope-items/{scope_id}", frozenset({"DELETE"})),
        ("/api/n6/app/v3/virtual-account/proposals", frozenset({"POST"})),
        ("/api/n6/app/v3/virtual-account/proposals/cancel", frozenset({"POST"})),
        ("/api/n6/app/v3/virtual-account/proposals/{proposal_id}/confirm", frozenset({"POST"})),
        ("/api/n6/app/v3/strategy-center/selection", frozenset({"PUT"})),
    }
    # Keep A-track/admin routes intact and fail closed if a B-track mutation
    # outside the explicitly frozen v3 command surface is ever registered.
    app.router.routes[:] = [
        route
        for route in app.router.routes
        if not (
            getattr(route, "path", "").startswith("/api/n6/app/")
            and set(getattr(route, "methods", set()) or set()) != {"GET"}
            and (
                str(getattr(route, "path", "")),
                frozenset(getattr(route, "methods", set()) or set()),
            )
            not in allowed_b_track_write_routes
        )
    ]

    return app


def generate_session_token() -> str:
    return secrets.token_urlsafe(48)


def hash_session_token(raw_token: str) -> str:
    return hashlib.sha256(raw_token.encode("utf-8")).hexdigest()


def safe_b_track_next(value: Any) -> str | None:
    raw_value = str(value or "").strip()
    if not raw_value:
        return None
    parsed = urlsplit(raw_value)
    if parsed.scheme or parsed.netloc:
        return None
    path = parsed.path or ""
    if path != DEFAULT_B_TRACK_ENTRY and not path.startswith(f"{DEFAULT_B_TRACK_ENTRY}/"):
        return None
    return urlunsplit(("", "", path, parsed.query, ""))


def login_success_location(user: UserAccount, requested_next: Any) -> str:
    safe_next = safe_b_track_next(requested_next)
    if safe_next:
        return safe_next
    if user.role == "admin":
        return DEFAULT_A_TRACK_ADMIN_ENTRY
    return DEFAULT_B_TRACK_ENTRY


def a_track_default_location(session: AuthSession) -> str:
    if session.role == "admin":
        return DEFAULT_A_TRACK_ADMIN_ENTRY
    return DEFAULT_B_TRACK_ENTRY


def b_track_login_location(request: Request) -> str:
    target = request.url.path
    if request.url.query:
        target = f"{target}?{request.url.query}"
    safe_next = safe_b_track_next(target) or DEFAULT_B_TRACK_ENTRY
    return f"/n6/login?next={quote(safe_next, safe='/')}"


def n6_csrf_token(session: AuthSession, secret: str) -> str:
    if not secret:
        return ""
    return hmac.new(
        secret.encode("utf-8"),
        session.session_token_hash.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def n6_csrf_valid(request: Request, session: AuthSession, secret: str) -> bool:
    expected = n6_csrf_token(session, secret)
    supplied = str(request.headers.get("x-csrf-token") or "").strip()
    return bool(expected and supplied and secrets.compare_digest(expected, supplied))


def n6_scope_bulk_canonical_filters(filters: Mapping[str, Any]) -> dict[str, Any]:
    excluded = {"asset_kind", "sort", "sort_dir", "date_policy_blocker", "date_policy_message"}
    output: dict[str, Any] = {}
    for key in sorted(filters):
        if key in excluded:
            continue
        value = filters.get(key)
        if isinstance(value, (list, tuple, set)):
            normalized = sorted(
                {
                    str(item).strip()
                    for item in value
                    if str(item).strip()
                }
            )
            if normalized:
                output[key] = normalized
            continue
        normalized = str(value or "").strip()
        if normalized:
            output[key] = normalized
    return output


def n6_scope_bulk_selection_sha256(identity_keys: list[str]) -> str:
    return hashlib.sha256("\n".join(sorted(identity_keys)).encode("utf-8")).hexdigest()


def _n6_scope_bulk_b64encode(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode("ascii").rstrip("=")


def _n6_scope_bulk_b64decode(value: str) -> bytes:
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode(f"{value}{padding}".encode("ascii"))


def n6_scope_bulk_selection_token(
    *,
    session: AuthSession,
    principal: Mapping[str, Any],
    secret: str,
    target_scope: str,
    asset_kind: str,
    filters: Mapping[str, Any],
    for_trade_date: str,
    source_run_id: str,
    identity_count: int,
    selection_sha256: str,
    issued_at: int | None = None,
) -> str:
    issued = int(time.time()) if issued_at is None else int(issued_at)
    payload = {
        "v": 1,
        "iat": issued,
        "exp": issued + N6_SCOPE_BULK_SELECTION_TTL_SECONDS,
        "target_scope": target_scope,
        "asset_kind": asset_kind,
        "filters": n6_scope_bulk_canonical_filters(filters),
        "for_trade_date": for_trade_date,
        "source_run_id": source_run_id,
        "identity_count": int(identity_count),
        "selection_sha256": selection_sha256,
        "principal_id": str(principal.get("principal_id") or ""),
        "principal_type": str(principal.get("principal_type") or ""),
        "user_id": str(session.user_id),
    }
    encoded = _n6_scope_bulk_b64encode(
        json.dumps(payload, ensure_ascii=True, separators=(",", ":"), sort_keys=True).encode("utf-8")
    )
    signature = hmac.new(
        secret.encode("utf-8"),
        f"{encoded}.{session.session_token_hash}".encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"{encoded}.{signature}"


def n6_scope_bulk_selection_payload(
    token: str,
    *,
    session: AuthSession,
    principal: Mapping[str, Any],
    secret: str,
    expected_target_scope: str,
    now: int | None = None,
) -> tuple[dict[str, Any] | None, str]:
    try:
        encoded, signature = str(token or "").split(".", 1)
        expected_signature = hmac.new(
            secret.encode("utf-8"),
            f"{encoded}.{session.session_token_hash}".encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        if not secrets.compare_digest(expected_signature, signature):
            return None, "selection_token_invalid"
        payload = json.loads(_n6_scope_bulk_b64decode(encoded))
    except (ValueError, UnicodeDecodeError, json.JSONDecodeError, binascii.Error):
        return None, "selection_token_invalid"
    current_time = int(time.time()) if now is None else int(now)
    if (
        not isinstance(payload, dict)
        or payload.get("v") != 1
        or payload.get("target_scope") != expected_target_scope
        or payload.get("asset_kind") not in {"stock", "index", "board"}
        or str(payload.get("principal_id") or "") != str(principal.get("principal_id") or "")
        or str(payload.get("principal_type") or "") != str(principal.get("principal_type") or "")
        or str(payload.get("user_id") or "") != str(session.user_id)
    ):
        return None, "selection_token_invalid"
    try:
        expires_at = int(payload.get("exp"))
        issued_at = int(payload.get("iat"))
        identity_count = int(payload.get("identity_count"))
    except (TypeError, ValueError):
        return None, "selection_token_invalid"
    if issued_at > current_time + 30 or expires_at < current_time:
        return None, "selection_token_expired"
    if (
        identity_count < 1
        or identity_count > N6_SCOPE_BULK_MAX_IDENTITIES
        or not re.fullmatch(r"[0-9a-f]{64}", str(payload.get("selection_sha256") or ""))
        or not re.fullmatch(r"[0-9]{8}", str(payload.get("for_trade_date") or ""))
        or not str(payload.get("source_run_id") or "").strip()
        or not isinstance(payload.get("filters"), dict)
    ):
        return None, "selection_token_invalid"
    return payload, ""


async def read_json_object(request: Request) -> dict[str, Any]:
    if "application/json" not in request.headers.get("content-type", ""):
        raise HTTPException(status_code=415, detail="application_json_required")
    try:
        payload = await request.json()
    except (json.JSONDecodeError, UnicodeDecodeError):
        raise HTTPException(status_code=400, detail="json_object_required")
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="json_object_required")
    return payload


async def read_login_payload(request: Request) -> dict[str, str]:
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        try:
            payload = await request.json()
        except json.JSONDecodeError:
            return {}
        return payload if isinstance(payload, dict) else {}
    body = (await request.body()).decode("utf-8")
    parsed = parse_qs(body, keep_blank_values=True)
    return {key: values[0] if values else "" for key, values in parsed.items()}


def verify_password(password: str, password_hash: str, password_hash_algo: str) -> bool:
    if not password or not password_hash:
        return False
    if password_hash_algo == "argon2id":
        try:
            from argon2 import PasswordHasher
            from argon2.exceptions import InvalidHashError, VerifyMismatchError, VerificationError
            from argon2.low_level import Type
        except ImportError:
            return False
        try:
            return PasswordHasher(type=Type.ID).verify(password_hash, password)
        except (InvalidHashError, VerificationError, VerifyMismatchError):
            return False
    if password_hash_algo == "bcrypt":
        try:
            import bcrypt
        except ImportError:
            return False
        try:
            return bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8"))
        except ValueError:
            return False
    return False


def current_session(request: Request, repository: N6UserRepository) -> AuthSession | None:
    return current_session_from_token(request.cookies.get(COOKIE_NAME), repository)


def current_session_from_token(raw_token: str | None, repository: N6UserRepository) -> AuthSession | None:
    if not raw_token:
        return None
    session = repository.fetch_session(hash_session_token(raw_token))
    if session is None:
        return None
    now = utc_now()
    if session.status != "active" or session.revoked_at is not None or ensure_aware(session.expires_at) <= now:
        return None
    return session


def session_user_payload(session: AuthSession) -> dict[str, Any]:
    return {
        "user_id": session.user_id,
        "login_name": session.login_name,
        "display_name": session.display_name,
        "role": session.role,
        "status": session.status,
        "expires_at": ensure_aware(session.expires_at).isoformat(),
    }


def resolve_app_principal(session: AuthSession, repository: N6UserRepository) -> dict[str, Any] | None:
    principals = repository.fetch_app_principals(session.user_id)
    if len(principals) != 1:
        return None
    principal = principals[0]
    if str(principal.get("principal_status") or "") != "active":
        return None
    if str(principal.get("principal_type") or "") not in {"admin", "human_user", "ai_user"}:
        return None
    return principal


def normalize_login_name(value: Any) -> str:
    return str(value or "").strip().lower()


def normalize_optional_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def normalize_filter_value(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def first_virtual_buy_value(source: dict[str, Any], *keys: str) -> str | None:
    for key in keys:
        value = source.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text and text != "—":
            return text
    return None


def normalize_virtual_buy_date(value: Any) -> str | None:
    text = normalize_filter_value(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def virtual_buy_rejected_response(reason: str, idempotency_key: str | None = None) -> dict[str, Any]:
    return {
        "ok": False,
        "status": "rejected",
        "order_id": None,
        "trade_id": None,
        "position_id": None,
        "idempotency_key": idempotency_key,
        "rejected_reason": reason,
    }


def virtual_buy_result_response(result: Any) -> dict[str, Any]:
    return {
        "ok": True,
        "status": result.status,
        "order_id": result.virtual_order_id,
        "trade_id": result.virtual_trade_id,
        "position_id": result.virtual_position_id,
        "idempotency_key": result.idempotency_key,
        "rejected_reason": None,
    }


def normalize_filter_values(values: Any) -> list[str]:
    if values is None:
        return []
    raw_values = values if isinstance(values, (list, tuple, set)) else [values]
    deduped: list[str] = []
    for raw_value in raw_values:
        value = normalize_filter_value(raw_value)
        if value and value not in deduped:
            deduped.append(value)
    return deduped


def normalize_filter_identity_values(values: Any) -> list[str]:
    if values is None:
        return []
    raw_values = values if isinstance(values, (list, tuple, set)) else [values]
    deduped: list[str] = []
    for raw_value in raw_values:
        for part in str(raw_value or "").split(","):
            value = normalize_filter_value(part)
            if value and value not in deduped:
                deduped.append(value)
    return deduped


def normalize_period_grade_filter_values(values: Any) -> list[str]:
    return [value for value in normalize_filter_values(values) if value in APP_V2_PERIOD_GRADE_FILTER_VALUES]


def normalize_time_field(value: Any) -> str:
    text = str(value or "").strip()
    return text if text in {"event_time", "created_at"} else "event_time"


def normalize_event_date(value: Any) -> str | None:
    text = normalize_filter_value(value)
    if not text:
        return None
    try:
        datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return None
    return text


def ui_v1_filters_from_request(request: Request) -> dict[str, str | None]:
    return {
        "date_from": normalize_filter_value(request.query_params.get("date_from")),
        "date_to": normalize_filter_value(request.query_params.get("date_to")),
        "time_field": normalize_time_field(request.query_params.get("time_field")),
        "event_type": normalize_filter_value(request.query_params.get("event_type")),
        "trade_date": normalize_filter_value(request.query_params.get("trade_date")),
        "asset_kind": normalize_filter_value(request.query_params.get("asset_kind")),
        "direction": normalize_filter_value(request.query_params.get("direction")),
        "signal_type": normalize_filter_value(request.query_params.get("signal_type")),
        "action_state": normalize_filter_value(request.query_params.get("action_state")),
        "blocked_reason": normalize_filter_value(request.query_params.get("blocked_reason")),
        "q": normalize_filter_value(request.query_params.get("q")),
        "cursor": normalize_filter_value(request.query_params.get("cursor")),
    }


def app_v2_message_filters_from_request(request: Request) -> dict[str, Any]:
    filters: dict[str, Any] = dict(ui_v1_filters_from_request(request))
    selected_asset_kinds = app_message_asset_kinds(
        request.query_params.getlist("asset_kind")
    )
    filters["asset_kinds"] = selected_asset_kinds
    filters["asset_kind"] = (
        selected_asset_kinds[0] if len(selected_asset_kinds) == 1 else None
    )
    return filters


def n6_signal_filters_with_cursor(filters: Mapping[str, Any]) -> dict[str, Any]:
    output = dict(filters)
    parsed_cursor = parse_n6_signal_page_cursor(output.pop("cursor", None))
    if parsed_cursor is not None:
        output["before_created_at"], output["before_id"] = parsed_cursor
    return output


def app_signal_filters_with_trade_date_defaults(
    filters: dict[str, Any],
    scope_metadata: dict[str, Any] | None,
    *,
    enforce_trading_session: bool = False,
) -> dict[str, Any]:
    output = dict(filters)
    if "asset_kinds" in output:
        selected_asset_kinds = app_message_asset_kinds(output.get("asset_kinds"))
        output["asset_kinds"] = selected_asset_kinds
        output["asset_kind"] = (
            selected_asset_kinds[0] if len(selected_asset_kinds) == 1 else None
        )
    else:
        output["asset_kind"] = app_message_asset_kind(output.get("asset_kind"))
    current_trade_date = current_app_signal_trade_date(scope_metadata)
    explicit_trade_date = normalize_filter_value(output.get("trade_date"))
    if explicit_trade_date:
        if enforce_trading_session:
            policy = n6_trade_date_access_policy(
                current_trade_date=current_trade_date,
                requested_trade_date=explicit_trade_date,
            )
            if policy["blocked"]:
                output["trade_date"] = policy["effective_trade_date"]
                output.pop("historical_projection_mode", None)
                output["date_policy_blocker"] = policy["blocker"]
                output["date_policy_message"] = policy["message"]
                return output
        output["trade_date"] = explicit_trade_date
        if current_trade_date and explicit_trade_date != current_trade_date:
            output["historical_projection_mode"] = True
        return output
    if current_trade_date:
        output["trade_date"] = current_trade_date
    return output


def app_signal_available_trade_dates(
    *,
    current_trade_date: str | None,
    projection_trade_dates: list[str] | tuple[str, ...] | None,
    selected_trade_date: str | None = None,
) -> list[str]:
    dates = {
        str(value)
        for value in list(projection_trade_dates or []) + [current_trade_date or "", selected_trade_date or ""]
        if str(value or "").isdigit() and len(str(value or "")) == 8
    }
    ordered = sorted(dates, reverse=True)
    if current_trade_date in ordered:
        ordered.remove(current_trade_date)
        ordered.insert(0, current_trade_date)
    return ordered


def app_signal_scope_metadata_for_filters(
    scope_metadata: dict[str, Any] | None,
    filters: dict[str, Any],
) -> dict[str, Any]:
    output = dict(scope_metadata or {})
    date_policy_message = normalize_filter_value(filters.get("date_policy_message"))
    date_policy_blocker = normalize_filter_value(filters.get("date_policy_blocker"))
    if filters.get("historical_projection_mode"):
        output["scope_mode"] = "historical_projection"
    if date_policy_blocker:
        output["date_policy_blocker"] = date_policy_blocker
    if date_policy_message:
        output["date_policy_message"] = date_policy_message
    output["available_trade_dates"] = app_signal_available_trade_dates(
        current_trade_date=current_app_signal_trade_date(output),
        projection_trade_dates=list(output.get("available_trade_dates") or []),
        selected_trade_date=normalize_filter_value(filters.get("trade_date")),
    )
    if date_policy_blocker:
        current_trade_date = current_app_signal_trade_date(output)
        output["available_trade_dates"] = [current_trade_date] if current_trade_date else []
    return output


def current_app_signal_trade_date(scope_metadata: dict[str, Any] | None) -> str | None:
    current_filter_batch = (scope_metadata or {}).get("current_filter_batch")
    if not isinstance(current_filter_batch, dict):
        return None
    for asset_kind in ("stock", "index", "board"):
        batch = current_filter_batch.get(asset_kind)
        if isinstance(batch, dict):
            trade_date = normalize_filter_value(batch.get("for_trade_date"))
            if trade_date:
                return trade_date
    return None


def raw_message_filters_from_request(request: Request) -> dict[str, str | None]:
    return {
        "event_date": normalize_event_date(request.query_params.get("event_date")),
        "source_layer": normalize_filter_value(request.query_params.get("source_layer")),
        "input_group": normalize_filter_value(request.query_params.get("input_group")),
        "event_type": normalize_filter_value(request.query_params.get("event_type")),
        "status": normalize_filter_value(request.query_params.get("status")),
        "asset_kind": normalize_filter_value(request.query_params.get("asset_kind")),
        "source_run_id": normalize_filter_value(request.query_params.get("source_run_id")),
        "q": normalize_filter_value(request.query_params.get("q")),
    }


def n5_action_filters_from_request(request: Request) -> dict[str, str | None]:
    return {
        "action_run_id": normalize_filter_value(request.query_params.get("action_run_id")),
        "source_run_id": normalize_filter_value(request.query_params.get("source_run_id")),
        "event_type": normalize_filter_value(request.query_params.get("event_type")),
        "status": normalize_filter_value(request.query_params.get("status")),
        "asset_kind": normalize_filter_value(request.query_params.get("asset_kind")),
        "action_state": normalize_filter_value(request.query_params.get("action_state")),
        "q": normalize_filter_value(request.query_params.get("q")),
    }


def n2_source_trade_date_from_request(value: Any) -> str | None:
    text = normalize_filter_value(value)
    if not text:
        return None
    if len(text) == 8 and text.isdigit():
        return text
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%Y%m%d")
    except ValueError:
        return None


def n2_condition_basis_filters_from_request(request: Request) -> dict[str, str | None]:
    return {
        "source_trade_date": n2_source_trade_date_from_request(request.query_params.get("source_trade_date")),
        "condition_key": normalize_filter_value(request.query_params.get("condition_key")),
        "quality_status": normalize_filter_value(request.query_params.get("quality_status")),
        "q": normalize_filter_value(request.query_params.get("q")),
    }


def app_v2_filter_filters_from_request(request: Request) -> dict[str, Any]:
    period_filter_keys = (
        "year_overheat_level",
        "quarter_overheat_level",
        "month_overheat_level",
        "week_overheat_level",
        "day_overheat_level",
    )
    period_filters = {
        key: normalize_period_grade_filter_values(request.query_params.getlist(key))
        for key in period_filter_keys
    }
    source_identity_values = [
        *request.query_params.getlist("source_identity_keys"),
        *request.query_params.getlist("source_identity_key"),
    ]
    direction = normalize_filter_value(request.query_params.get("direction"))
    return {
        "asset_kind": normalize_filter_value(request.query_params.get("asset_kind")),
        "for_trade_date": normalize_filter_value(request.query_params.get("for_trade_date")),
        "source_asset_type": normalize_filter_value(request.query_params.get("source_asset_type")),
        "source_identity_keys": normalize_filter_identity_values(source_identity_values),
        "direction": direction if direction in APP_V2_VALID_DIRECTIONS else None,
        "board_type": normalize_filter_value(request.query_params.get("board_type")),
        "year_overheat_level": period_filters["year_overheat_level"],
        "quarter_overheat_level": period_filters["quarter_overheat_level"],
        "month_overheat_level": period_filters["month_overheat_level"],
        "week_overheat_level": period_filters["week_overheat_level"],
        "day_overheat_level": period_filters["day_overheat_level"],
        "condition_key": normalize_filter_value(request.query_params.get("condition_key")),
        "quality_status": normalize_filter_value(request.query_params.get("quality_status")),
        "last_signal_state": normalize_filter_value(request.query_params.get("last_signal_state")),
        "source_run_id": normalize_filter_value(request.query_params.get("source_run_id")),
        "projection_run_id": normalize_filter_value(request.query_params.get("projection_run_id")),
        "cache_run_id": normalize_filter_value(request.query_params.get("cache_run_id")),
        "q": normalize_filter_value(request.query_params.get("q")),
        "buy_expected_return_pct_min": app_v2_expected_return_value_text(
            request.query_params.get("buy_expected_return_pct_min")
        ),
        "level_up_score_recommendation": app_v2_level_up_recommendation_value(
            request.query_params.get("level_up_score_recommendation")
        ),
        "sort": normalize_filter_value(request.query_params.get("sort")),
        "sort_dir": normalize_filter_value(request.query_params.get("sort_dir")),
    }


def app_v2_monitor_status_filter(value: Any) -> str:
    text = str(value or "active").strip().lower()
    return text if text in APP_V2_MONITOR_STATUS_FILTERS else "active"


def app_v2_monitor_status_from_request(request: Request) -> str:
    return app_v2_monitor_status_filter(request.query_params.get("monitor_status"))


def app_v2_monitor_filters_from_request(request: Request) -> dict[str, Any]:
    return {
        "monitor_status": app_v2_monitor_status_from_request(request),
        "for_trade_date": str(request.query_params.get("for_trade_date") or "").strip(),
    }


def status_monitor_filters_from_request(request: Request) -> dict[str, str | None]:
    return {
        "trade_date": normalize_filter_value(request.query_params.get("trade_date")),
        "source_n4_run_id": normalize_filter_value(request.query_params.get("source_n4_run_id")),
        "source_n5_run_id": normalize_filter_value(request.query_params.get("source_n5_run_id")),
        "source_layer": normalize_filter_value(request.query_params.get("source_layer")),
        "status": normalize_filter_value(request.query_params.get("status")),
        "event_type": normalize_filter_value(request.query_params.get("event_type")),
        "action_event_type": normalize_filter_value(request.query_params.get("action_event_type")),
        "asset_kind": normalize_filter_value(request.query_params.get("asset_kind")),
        "direction": normalize_filter_value(request.query_params.get("direction")),
        "signal_type": normalize_filter_value(request.query_params.get("signal_type")),
        "outbox_status": normalize_filter_value(request.query_params.get("outbox_status")),
        "q": normalize_filter_value(request.query_params.get("q")),
    }


def ui_v1_limit_from_request(request: Request, default_limit: int, *, max_limit: int = 500) -> int:
    try:
        requested = int(request.query_params.get("limit") or default_limit)
    except ValueError:
        requested = default_limit
    return max(1, min(requested, max_limit))


def query_param_enabled(request: Request, key: str) -> bool:
    return str(request.query_params.get(key) or "").strip().lower() in {"1", "true", "yes", "all"}


def ui_v1_offset_from_request(request: Request) -> int:
    try:
        requested = int(request.query_params.get("offset") or 0)
    except ValueError:
        requested = 0
    return max(0, requested)


def cash_ledger_limit_from_request(request: Request) -> int:
    try:
        requested = int(request.query_params.get("limit") or 20)
    except ValueError:
        requested = 20
    return max(1, min(requested, 100))


def validate_user_create_payload(*, login_name: str, role: str) -> list[str]:
    blockers: list[str] = []
    if len(login_name) < 3:
        blockers.append("login_name_too_short")
    if not login_name.replace("_", "").replace("-", "").isalnum():
        blockers.append("login_name_invalid_chars")
    if role not in {"admin", "user"}:
        blockers.append("invalid_role")
    return blockers


def client_info_from_request(request: Request) -> dict[str, Any]:
    return {
        "user_agent": request.headers.get("user-agent", ""),
        "client_host": request.client.host if request.client else "",
        "n6_web_mvp": True,
    }


def ensure_aware(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def nav_context(session: AuthSession, active: str) -> dict[str, Any]:
    return {
        "active": active,
        "is_admin": session.role == "admin",
        "links": [
            {"key": "post_close_fastlane_status", "label": "收盘状态", "href": "/n6/post-close-fastlane-status"},
            {"key": "archive_status", "label": "归档状态", "href": "/n6/archive-status"},
            {"key": "n2_condition_basis", "label": "N2条件基础表", "href": "/n6/n2-condition-basis/index"},
            {"key": "n3_messages", "label": "N3消息", "href": "/n6/n3-messages"},
            {"key": "n4_messages", "label": "N4消息", "href": "/n6/n4-messages"},
            {"key": "input_messages", "label": "N6输入消息", "href": "/n6/input-messages"},
            {"key": "n5_messages", "label": "N5消息", "href": "/n6/n5-messages"},
            {"key": "n5_actions", "label": "N5动作", "href": "/n6/n5-actions"},
            {"key": "rag", "label": "RAG问答", "href": "/n6/rag"},
        ],
    }


def n2_condition_basis_asset_meta(asset_kind: str) -> dict[str, str]:
    key = str(asset_kind or "").strip().lower()
    meta = N2_CONDITION_BASIS_ASSET_META.get(key)
    if meta is None:
        raise ValueError("invalid_n2_condition_basis_asset_kind")
    return dict(meta)


def n2_condition_basis_asset_tabs() -> list[dict[str, str]]:
    return [
        {
            "asset_kind": asset_kind,
            "label": str(N2_CONDITION_BASIS_ASSET_META[asset_kind]["label"]),
            "href": f"/n6/n2-condition-basis/{asset_kind}",
            "source_table": str(N2_CONDITION_BASIS_ASSET_META[asset_kind]["source_table"]),
        }
        for asset_kind in N2_CONDITION_BASIS_ASSET_ORDER
    ]


def _n2_condition_basis_export_cell(item: dict[str, Any], path: str) -> Any:
    value: Any = item
    for part in path.split("."):
        if not isinstance(value, dict):
            return ""
        value = value.get(part)
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return value


def build_n2_condition_basis_latest_export(data: dict[str, Any]) -> bytes:
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="E6F3F1")
    header_font = Font(bold=True, color="172124")
    latest_source_trade_date = str(data.get("latest_source_trade_date") or "unknown")
    assets = dict(data.get("assets") or {})
    for sheet_index, asset_kind in enumerate(N2_CONDITION_BASIS_ASSET_ORDER):
        meta = n2_condition_basis_asset_meta(asset_kind)
        worksheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        worksheet.title = str(meta["label"])
        headers = [label for _, label in N2_CONDITION_BASIS_EXPORT_COLUMNS]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        asset_data = dict(assets.get(asset_kind) or {})
        for row in list(asset_data.get("items") or []):
            item = n2_condition_basis_item(row)
            worksheet.append([
                _n2_condition_basis_export_cell(item, key)
                for key, _ in N2_CONDITION_BASIS_EXPORT_COLUMNS
            ])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_number, (_, label) in enumerate(N2_CONDITION_BASIS_EXPORT_COLUMNS, start=1):
            width = 16
            if label in {"运行批次", "标识", "条件键", "原始行JSON", "原始JSON", "周期触发基准JSON", "目标价追溯JSON"}:
                width = 28
            if label in {"名称", "来源表", "质量原因"}:
                width = 20
            worksheet.column_dimensions[get_column_letter(column_number)].width = width
        worksheet.sheet_view.showGridLines = False
    properties = workbook.properties
    properties.title = f"{N2_CONDITION_BASIS_EXPORT_FILENAME_PREFIX}_{latest_source_trade_date}"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def n2_condition_basis_latest_export_response(data: dict[str, Any]) -> Response:
    latest_source_trade_date = str(data.get("latest_source_trade_date") or "unknown")
    filename = f"{N2_CONDITION_BASIS_EXPORT_FILENAME_PREFIX}_{latest_source_trade_date}.xlsx"
    return Response(
        build_n2_condition_basis_latest_export(data),
        media_type=N2_CONDITION_BASIS_EXPORT_MEDIA_TYPE,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
            "Cache-Control": "no-store",
        },
    )


def signal_source_user_id(session: AuthSession, config: N6UserWebConfig) -> int:
    # The current MVP has one shadow projection owner; per-user projection execute is a later gate.
    return config.signal_source_user_id or session.user_id


def filter_profile_view(profile: dict[str, Any] | None) -> dict[str, bool]:
    source = profile or {}
    return {
        "enable_chase": bool(source.get("enable_chase", True)),
        "enable_ultra_short": bool(source.get("enable_ultra_short", True)),
        "enable_short": bool(source.get("enable_short", True)),
        "enable_mid": bool(source.get("enable_mid", True)),
        "enable_long": bool(source.get("enable_long", True)),
    }


def top_strategy_view(row: dict[str, Any] | None) -> dict[str, Any]:
    if not row:
        return {
            "code": "—",
            "name": "暂无指数策略",
            "display_title": "暂无推荐",
            "display_summary": "",
            "recommendation_key": "unclassified",
            "recommendation_label": "暂无推荐策略",
            "signal_types": "—",
            "transition_summary": "—",
            "buy_target_price": "—",
            "sell_target_price": "—",
        }
    strategy_type = strategy_type_from_transitions(row)
    return {
        "code": row.get("code") or "—",
        "name": row.get("name") or "—",
        "display_title": row.get("display_title") or "顶部推荐策略",
        "display_summary": row.get("display_summary") or "",
        "recommendation_key": strategy_type["key"],
        "recommendation_label": f"推荐{strategy_type['label']}",
        "signal_types": join_list(row.get("selected_signal_types")),
        "transition_summary": transition_summary(row),
        "buy_target_price": format_price(row.get("buy_target_price")),
        "sell_target_price": format_price(row.get("sell_target_price")),
    }


def board_view(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "board_code": row.get("board_code") or "—",
        "board_name": row.get("board_name") or "—",
        "display_title": row.get("display_title") or row.get("board_name") or "强势板块",
        "signal_types": join_list(row.get("selected_signal_types")),
        "transition_summary": transition_summary(row),
        "buy_target_price": format_price(row.get("buy_target_price")),
    }


def strong_board_views(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [board_view(row) for row in rows if is_strong_industry_board(row)]


def is_strong_industry_board(row: dict[str, Any]) -> bool:
    return (
        str(row.get("board_code") or "").startswith("881")
        and row.get("period_transition_y") == "volume_up"
        and row.get("period_transition_q") == "volume_up"
        and row.get("period_transition_m") == "volume_up"
    )


def card_view(row: dict[str, Any]) -> dict[str, Any]:
    strategy_type = strategy_type_from_transitions(row)
    return {
        "user_signal_card_id": row.get("user_signal_card_id"),
        "code": row.get("code") or "—",
        "name": row.get("name") or "—",
        "direction": row.get("direction") or "—",
        "direction_label": "买入" if row.get("direction") == "buy" else "卖出" if row.get("direction") == "sell" else "—",
        "signal_type": row.get("signal_type") or "—",
        "target_price": format_price(row.get("target_price")),
        "current_price": format_price(row.get("current_price")),
        "expected_return_pct": format_pct(row.get("expected_return_pct")),
        "board_code": row.get("board_code") or "—",
        "board_name": row.get("board_name") or "—",
        "title": row.get("title") or "",
        "strategy_key": strategy_type["key"],
        "strategy_label": strategy_type["label"],
    }


def resolve_selected_strategy_keys(
    request: Request,
    strategy: dict[str, Any],
    profile: dict[str, bool],
) -> set[str]:
    if request.query_params.get("filter_submitted") == "1":
        return {
            key
            for key in request.query_params.getlist("strategy")
            if key in VALID_STRATEGY_FILTER_KEYS
        }
    recommendation_key = str(strategy.get("recommendation_key") or "")
    if recommendation_key in VALID_STRATEGY_FILTER_KEYS:
        return {recommendation_key}
    return {
        str(option["key"])
        for option in STRATEGY_FILTER_OPTIONS
        if profile.get(str(option["profile_key"]), True)
    }


def strategy_filter_view(selected_keys: set[str]) -> list[dict[str, Any]]:
    return [
        {
            "key": option["key"],
            "label": option["label"],
            "checked": str(option["key"]) in selected_keys,
        }
        for option in STRATEGY_FILTER_OPTIONS
    ]


def filter_stock_card_views(cards: list[dict[str, Any]], selected_keys: set[str]) -> list[dict[str, Any]]:
    rows = []
    for row in cards:
        if row.get("asset_kind") != "stock":
            continue
        view = card_view(row)
        if view["strategy_key"] in selected_keys:
            rows.append(view)
    return rows


def notification_groups(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = [
        ("index", "指数信号", {"index_signal"}),
        ("board", "板块信号", {"board_signal"}),
        ("stock", "个股信号", {"stock_filter_signal"}),
        ("n5", "N5 action/hint 信号", {"n5_action_event", "n5_hint_event"}),
    ]
    grouped = []
    for key, label, sources in buckets:
        group_rows = [notification_view(row) for row in rows if row.get("notification_source") in sources]
        grouped.append({"key": key, "label": label, "rows": group_rows, "count": len(group_rows)})
    return grouped


def notification_view(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "title": row.get("title") or "—",
        "message": row.get("message") or "",
        "queue_status": row.get("queue_status") or "—",
        "channel": row.get("channel") or "—",
        "notification_source": row.get("notification_source") or "—",
        "identity": row.get("identity_key") or row.get("code") or "—",
        "name": row.get("name") or "",
        "queued_at": format_datetime(row.get("queued_at")),
    }


def action_event_view(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "outbox_id": row.get("outbox_id"),
        "event_id": row.get("event_id") or "—",
        "event_type": row.get("event_type") or "—",
        "source_layer": row.get("source_layer") or "—",
        "status": row.get("status") or "—",
        "trade_date": row.get("trade_date") or "—",
        "asset": f"{row.get('asset_kind') or '—'} / {row.get('identity_key') or '—'}",
        "direction": row.get("direction") or "—",
        "signal_type": row.get("signal_type") or "—",
        "action_state": row.get("action_state") or "—",
        "action_mark": row.get("action_mark") or "—",
        "condition_key": row.get("condition_key") or row.get("original_condition_key") or "—",
        "source_run_id": row.get("source_run_id") or "—",
        "event_time": format_datetime(row.get("event_time")),
        "created_at": format_datetime(row.get("created_at")),
    }


def summarize_action_events(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_event_type: dict[str, int] = {}
    by_status: dict[str, int] = {}
    for row in rows:
        event_type = str(row.get("event_type") or "unknown")
        status = str(row.get("status") or "unknown")
        by_event_type[event_type] = by_event_type.get(event_type, 0) + 1
        by_status[status] = by_status.get(status, 0) + 1
    return {
        "total": len(rows),
        "by_event_type": [{"label": key, "count": value} for key, value in sorted(by_event_type.items())],
        "by_status": [{"label": key, "count": value} for key, value in sorted(by_status.items())],
    }


def message_dashboard_view(dashboard: dict[str, Any]) -> dict[str, Any]:
    admin = dashboard.get("admin_dashboard") or {}
    message = dashboard.get("message_dashboard") or {}
    return {
        "safety": [
            "只读预览",
            "不消费 outbox",
            "不触发真实通知/交易",
            "READ ONLY",
            "NO ORDER",
            "NO TRADE",
            "NO POSITION UPDATE",
            "NO REAL TRADE",
            "NOT INVESTMENT ADVICE",
        ],
        "metrics": [
            {
                "label": "今日 N4 TriggerMatched pending",
                "value": int(admin.get("today_n4_trigger_matched_pending") or 0),
            },
            {
                "label": "今日 N5 ActionBlocked",
                "value": int(admin.get("today_n5_action_blocked") or 0),
            },
            {
                "label": "今日 N5 ActionExecuted",
                "value": int(admin.get("today_n5_action_executed") or 0),
            },
            {
                "label": "N5 outbox pending",
                "value": int(admin.get("n5_outbox_pending") or 0),
            },
            {
                "label": "N6 queued_only",
                "value": int(admin.get("n6_queued_only") or 0),
            },
            {
                "label": "N6 ready preview",
                "value": int(admin.get("n6_ready_for_future_push") or 0),
            },
        ],
        "n6_shadow": [
            {"label": "projection", "value": int(admin.get("n6_shadow_projection_count") or 0)},
            {"label": "card", "value": int(admin.get("n6_shadow_card_count") or 0)},
            {"label": "queue", "value": int(admin.get("n6_shadow_queue_count") or 0)},
            {"label": "latest run", "value": admin.get("latest_projection_run_id") or "—"},
        ],
        "event_distribution": [
            {
                "label": f"{row.get('event_type') or '—'} / {row.get('status') or '—'}",
                "count": int(row.get("count") or 0),
            }
            for row in message.get("event_distribution", [])
        ],
        "blocked_reasons": [
            {
                "label": row.get("blocked_reason") or "unknown",
                "count": int(row.get("count") or 0),
            }
            for row in dashboard.get("blocked_reason_distribution", [])
        ],
        "recent_runs": [
            {
                "layer": row.get("layer") or "—",
                "run_id": row.get("run_id") or "—",
                "status": row.get("status") or "—",
                "created_at": format_datetime(row.get("created_at")),
                "finished_at": format_datetime(row.get("finished_at")),
            }
            for row in dashboard.get("recent_runs", [])
        ],
    }


def user_view(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "user_id": row.get("user_id"),
        "login_name": row.get("login_name") or "—",
        "display_name": row.get("display_name") or "—",
        "role": row.get("role") or "—",
        "status": row.get("status") or "—",
        "created_at": format_datetime(row.get("created_at")),
        "last_login_at": format_datetime(row.get("last_login_at")),
        "filter_profile_count": int(row.get("filter_profile_count") or 0),
        "sim_account_count": int(row.get("sim_account_count") or 0),
        "active_position_count": int(row.get("active_position_count") or 0),
    }


def virtual_account_summary_view(
    account: dict[str, Any] | None,
    cash_snapshot: dict[str, Any] | None,
) -> dict[str, Any]:
    account_view = virtual_account_view(account)
    snapshot_view = virtual_cash_snapshot_view(cash_snapshot)
    return {
        "available": account is not None,
        "safety": [
            "READ ONLY",
            "NO ORDER",
            "NO TRADE",
            "NO POSITION UPDATE",
            "NO REAL TRADE",
            "NOT INVESTMENT ADVICE",
        ],
        "account_name": account_view["account_name"],
        "base_currency": account_view["base_currency"],
        "initial_cash": account_view["initial_cash"],
        "available_cash": snapshot_view["available_cash"],
        "frozen_cash": snapshot_view["frozen_cash"],
        "total_cash": snapshot_view["total_cash"],
        "quality_status": account_view["quality_status"],
        "seed_run_id": account_view["seed_run_id"],
    }


def virtual_account_view(row: dict[str, Any] | None) -> dict[str, Any]:
    source = row or {}
    return {
        "virtual_account_id": source.get("virtual_account_id") or "—",
        "principal_id": source.get("principal_id") or "—",
        "principal_type": source.get("principal_type") or "—",
        "account_name": source.get("account_name") or "—",
        "virtual_account_status": source.get("virtual_account_status") or "missing",
        "base_currency": source.get("base_currency") or "—",
        "initial_cash": format_money(source.get("initial_cash")),
        "current_cash_snapshot_id": source.get("current_cash_snapshot_id") or "—",
        "quality_status": source.get("quality_status") or "missing",
        "seed_run_id": source.get("run_id") or "—",
        "policy_version": source.get("policy_version") or "—",
        "policy_hash": source.get("policy_hash") or "—",
        "rollback_scope": source.get("rollback_scope") or "—",
        "created_at": format_datetime(source.get("created_at")),
        "updated_at": format_datetime(source.get("updated_at")),
    }


def virtual_cash_snapshot_view(row: dict[str, Any] | None) -> dict[str, Any]:
    source = row or {}
    return {
        "cash_snapshot_id": source.get("cash_snapshot_id") or "—",
        "virtual_account_id": source.get("virtual_account_id") or "—",
        "snapshot_time": format_datetime(source.get("snapshot_time")),
        "trade_date": source.get("trade_date") or "—",
        "available_cash": format_money(source.get("available_cash")),
        "frozen_cash": format_money(source.get("frozen_cash")),
        "total_cash": format_money(source.get("total_cash")),
        "currency": source.get("currency") or "—",
        "source_ledger_max_id": source.get("source_ledger_max_id") or "—",
        "snapshot_status": source.get("snapshot_status") or "missing",
        "quality_status": source.get("quality_status") or "missing",
        "run_id": source.get("run_id") or "—",
    }


def virtual_cash_ledger_view(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "cash_ledger_id": row.get("cash_ledger_id") or "—",
        "event_time": format_datetime(row.get("event_time")),
        "ledger_type": row.get("ledger_type") or "—",
        "amount": format_money(row.get("amount")),
        "currency": row.get("currency") or "—",
        "source_event_type": row.get("source_event_type") or "—",
        "source_event_id": row.get("source_event_id") or "—",
        "source_virtual_order_id": row.get("source_virtual_order_id") or "—",
        "source_virtual_trade_id": row.get("source_virtual_trade_id") or "—",
        "run_id": row.get("run_id") or "—",
        "quality_status": row.get("quality_status") or "—",
    }


def sim_account_view(row: dict[str, Any] | None) -> dict[str, Any]:
    source = row or {}
    return {
        "account_name": source.get("account_name") or DEFAULT_SIM_ACCOUNT_NAME,
        "initial_cash": format_decimal(source.get("initial_cash", DEFAULT_INITIAL_CASH), places=0),
        "cash_balance": format_decimal(source.get("cash_balance", DEFAULT_INITIAL_CASH), places=0),
        "frozen_cash": format_decimal(source.get("frozen_cash", 0), places=0),
        "settlement_mode": source.get("settlement_mode") or "T_PLUS_1",
        "account_status": source.get("account_status") or "shadow",
    }


def sim_position_view(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "code": row.get("code") or "—",
        "name": row.get("name") or "—",
        "total_qty": format_decimal(row.get("total_qty"), places=0),
        "available_qty": format_decimal(row.get("available_qty"), places=0),
        "t_plus_one_locked_qty": format_decimal(row.get("t_plus_one_locked_qty"), places=0),
        "avg_cost": format_price(row.get("avg_cost")),
        "last_price": format_price(row.get("last_price")),
        "market_value": format_decimal(row.get("market_value"), places=2),
        "unrealized_pnl": format_decimal(row.get("unrealized_pnl"), places=2),
    }


def summarize_card_counts(cards: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "total": len(cards),
        "stock": sum(1 for row in cards if row.get("asset_kind") == "stock"),
        "buy": sum(1 for row in cards if row.get("direction") == "buy"),
        "sell": sum(1 for row in cards if row.get("direction") == "sell"),
    }


def summarize_stock_row_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "total": len(rows),
        "stock": len(rows),
        "buy": sum(1 for row in rows if row.get("direction") == "buy"),
        "sell": sum(1 for row in rows if row.get("direction") == "sell"),
    }


def format_price(value: Any) -> str:
    return format_decimal(value, places=2)


def format_pct(value: Any) -> str:
    formatted = format_decimal(value, places=2)
    return "—" if formatted == "—" else f"{formatted}%"


def format_money(value: Any) -> str:
    if value is None:
        return "—"
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return "—"
    quantized = decimal_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{quantized:,.2f}"


def format_decimal(value: Any, *, places: int) -> str:
    if value is None:
        return "—"
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return "—"
    quantizer = Decimal("1").scaleb(-places)
    return f"{decimal_value.quantize(quantizer, rounding=ROUND_HALF_UP):f}"


def format_datetime(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, datetime):
        return ensure_aware(value).astimezone(DISPLAY_TIMEZONE).strftime("%Y-%m-%d %H:%M")
    return str(value)


def join_list(value: Any) -> str:
    if not value:
        return "—"
    if isinstance(value, str):
        return value
    return " / ".join(str(item) for item in value)


def strategy_type_from_transitions(row: dict[str, Any]) -> dict[str, Any]:
    transitions = [
        row.get("period_transition_y"),
        row.get("period_transition_q"),
        row.get("period_transition_m"),
        row.get("period_transition_w"),
        row.get("period_transition_d"),
    ]
    volume_up = [value == "volume_up" for value in transitions]
    if volume_up == [True, True, True, True, True]:
        return {"key": "chase", "label": "追涨策略", "rank": 60}
    if volume_up == [True, True, True, True, False]:
        return {"key": "ultra_short", "label": "超短策略", "rank": 50}
    if volume_up == [True, True, True, False, False]:
        return {"key": "short", "label": "短线策略", "rank": 40}
    if volume_up == [True, True, False, False, False]:
        return {"key": "mid", "label": "中线策略", "rank": 30}
    if volume_up == [True, False, False, False, False] or volume_up == [False, False, False, False, False]:
        return {"key": "long", "label": "长线策略", "rank": 20}
    return {"key": "unclassified", "label": "未分级策略", "rank": 0}


def transition_summary(row: dict[str, Any]) -> str:
    values = [
        ("年", row.get("period_transition_y")),
        ("季", row.get("period_transition_q")),
        ("月", row.get("period_transition_m")),
        ("周", row.get("period_transition_w")),
        ("日", row.get("period_transition_d")),
    ]
    labels = {"volume_up": "放量上涨", "low_volume_up": "缩量上涨", "volume_down": "放量下跌", "low_volume_down": "缩量下跌", "flat": "平稳", "unknown": "未知"}
    return " / ".join(f"{name}:{labels.get(value, value or '—')}" for name, value in values)


def create_runtime_app() -> FastAPI:
    return create_app(
        btrack_authority_repository=build_runtime_btrack_authority_repository(),
        btrack_authority_required=True,
    )


app = create_runtime_app()


def main() -> None:
    import uvicorn

    host = os.environ.get("ASHARE_V3_N6_WEB_HOST", "127.0.0.1")
    port = int(os.environ.get("ASHARE_V3_N6_WEB_PORT", "8786"))
    uvicorn.run("ashare_v3.web.n6_user_app:app", host=host, port=port, reload=False)


if __name__ == "__main__":
    main()
