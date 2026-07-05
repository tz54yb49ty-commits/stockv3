import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from ashare_v3.web.n6_replay import (
    N6ReplayBlocked,
    _asset_unit_fix_delta_n4_message,
    _active_candidates_for_minute,
    _active_state_reduction_skip_keys,
    _build_asset_unit_fix_n5_delta_only_messages,
    _build_canonical_b2_input,
    _build_closed_confirmation_metric_rows_for_minute,
    _build_n5_evaluator_demand_inputs,
    _canonical_fixture_source_bundle,
    _n4_message_to_outbox_row,
    _N3PNegativeProofCache,
    _n3p_negative_cache_proof_context,
    _prefilter_decision_for_context,
    _suppress_duplicate_action_executed_messages,
    build_c1_index_board_readiness_report,
    build_b2_plan_only_replay_artifact,
    build_n3p_plan_only_replay_artifact,
    create_local_replay_job,
    export_c1_index_board_full_day_source_bundle,
    list_replay_dates,
    read_replay_job,
    read_replay_messages,
    read_replay_timeline,
)
from tests.test_realtime_projection_execute import (
    SHANGHAI,
    sample_calculation_config,
    sample_current_bars,
    sample_previous_bars,
    sample_snapshot,
)
from tests.test_v3_realtime_virtual_metric_writer_runner import (
    LIVE_CURRENT_1M_SOURCE_RUN_ID,
    SOURCE_PREVIOUS_DAY_MINUTE_RUN_ID,
    SOURCE_SNAPSHOT_RUN_ID,
    clean_target_counts,
    live_688596_payload,
)


class N6LocalReplayTests(unittest.TestCase):
    def _c1_rows(self, *, asset_kind: str, identity_key: str, code: str, trade_date: str = "2026-06-26") -> list[dict[str, object]]:
        minutes = [
            *[f"{hour:02d}:{minute:02d}" for hour in range(9, 12) for minute in range(60) if "09:31" <= f"{hour:02d}:{minute:02d}" <= "11:30"],
            *[f"{hour:02d}:{minute:02d}" for hour in range(13, 16) for minute in range(60) if "13:01" <= f"{hour:02d}:{minute:02d}" <= "15:00"],
        ]
        return [
            {
                "asset_kind": asset_kind,
                "identity_key": identity_key,
                "code": code,
                "display_code": code,
                "datetime": f"{trade_date} {minute}",
                "bar_time": f"{trade_date}T{minute}:00+08:00",
                "open": "10",
                "high": "11",
                "low": "9",
                "close": "10",
                "volume": "100",
                "amount": "1000",
                "quality_status": "passed",
            }
            for minute in minutes
        ]

    def _seed_explicit_source_bundle(self, replay_root: Path) -> None:
        target = replay_root / "_sources" / "20260626" / "source_bundle.json"
        target.parent.mkdir(parents=True, exist_ok=True)
        bundle = _canonical_fixture_source_bundle(job_id="local_replay_20260626_explicit_seed")
        bundle["source_meta"] = {
            "historical_source_type": "explicit_source_bundle",
            "historical_source_path": str(target),
            "historical_source_hash": "test-explicit-source-bundle-hash",
            "source_row_count": 960,
            "candidate_count": 4,
            "context_count": 4,
            "b2_snapshot_row_count": 2,
            "b2_live_current_row_count": 2,
            "b2_previous_day_row_count": 2,
            "upstream_source_mode": "live_current_1m",
            "bundle_contract_version": "historical_replay_source_bundle_v1",
        }
        target.write_text(json.dumps(bundle, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")

    def _seed_mixed_asset_scope_bundle(self, replay_root: Path) -> Path:
        target = replay_root / "_sources" / "20260626" / "source_bundle.json"
        target.parent.mkdir(parents=True, exist_ok=True)
        bundle = {
            "source_records": {
                "600000": [
                    {"asset_kind": "stock", "code": "600000", "datetime": "2026-06-25 09:31", "open": 10, "high": 11, "low": 9, "close": 10, "amount": 100},
                    {"asset_kind": "stock", "code": "600000", "datetime": "2026-06-26 10:00", "open": 10, "high": 11, "low": 9, "close": 10, "amount": 100},
                ],
                "000300": [
                    {"asset_kind": "index", "code": "000300", "datetime": "2026-06-25 09:31", "open": 20, "high": 21, "low": 19, "close": 20, "amount": 200},
                    {"asset_kind": "index", "code": "000300", "datetime": "2026-06-26 10:00", "open": 20, "high": 21, "low": 19, "close": 20, "amount": 200},
                ],
                "881001": [
                    {"asset_kind": "board", "code": "881001", "datetime": "2026-06-25 09:31", "open": 30, "high": 31, "low": 29, "close": 30, "amount": 300},
                    {"asset_kind": "board", "code": "881001", "datetime": "2026-06-26 10:00", "open": 30, "high": 31, "low": 29, "close": 30, "amount": 300},
                ],
            },
            "candidates": [
                {"asset_kind": "stock", "identity_key": "stock:SH:600000", "exchange": "SH", "code": "600000", "display_code": "600000", "name": "浦发银行", "signal_type": "B_BUY", "condition_key": "BUY:D", "minute_label": "2026-06-26 10:00", "observed_at": "2026-06-26 10:00:30"},
                {"asset_kind": "index", "identity_key": "index:SH:000300", "exchange": "SH", "code": "000300", "display_code": "000300", "name": "沪深300", "signal_type": "B_BUY", "condition_key": "BUY:D", "minute_label": "2026-06-26 10:00", "observed_at": "2026-06-26 10:00:30"},
                {"asset_kind": "board", "identity_key": "board:TDX:881001", "code": "881001", "display_code": "881001", "name": "银行", "signal_type": "B_BUY", "condition_key": "BUY_HINT", "minute_label": "2026-06-26 10:15", "observed_at": "2026-06-26 10:15:30"},
            ],
            "n4_context_snapshot_rows": [
                {"run_id": "trigger_context_snapshot_20260626_test", "for_trade_date": "20260626", "source_trade_date": "20260625", "prev_trade_date": "20260625", "asset_kind": "stock", "identity_key": "stock:SH:600000", "direction": "buy", "condition_key": "BUY:D", "condition_periods": ["D"], "allowed_signal_types": ["BUY"], "is_hint_scope": False, "quality_status": "passed", "period_trigger_baseline_json": {"periods": {"D": {"previous_transition": "flat"}}}, "raw_json": {"condition_key": "BUY:D", "original_condition_key": "BUY:D"}},
                {"run_id": "trigger_context_snapshot_20260626_test", "for_trade_date": "20260626", "source_trade_date": "20260625", "prev_trade_date": "20260625", "asset_kind": "index", "identity_key": "index:SH:000300", "direction": "buy", "condition_key": "BUY:D", "condition_periods": ["D"], "allowed_signal_types": ["BUY"], "is_hint_scope": False, "quality_status": "passed", "period_trigger_baseline_json": {"periods": {"D": {"previous_transition": "flat"}}}, "raw_json": {"condition_key": "BUY:D", "original_condition_key": "BUY:D"}},
                {"run_id": "trigger_context_snapshot_20260626_test", "for_trade_date": "20260626", "source_trade_date": "20260625", "prev_trade_date": "20260625", "asset_kind": "board", "identity_key": "board:TDX:881001", "direction": "buy", "condition_key": "BUY_HINT", "condition_periods": [], "allowed_signal_types": ["BUY_HINT"], "is_hint_scope": True, "quality_status": "passed", "period_trigger_baseline_json": {"periods": {"D": {"previous_transition": "flat"}}}, "raw_json": {"condition_key": "BUY_HINT", "original_condition_key": "BUY_HINT"}},
            ],
            "replay_config": {
                "replay_run_id": "local_replay_20260626_explicit_seed",
                "job_id": "local_replay_20260626_explicit_seed",
                "trade_date": "20260626",
                "source_trade_date": "20260625",
                "prev_trade_date": "20260625",
                "for_trade_date": "20260626",
                "until_hhmm": "1500",
                "source_mode": "replay",
                "source_condition_run_id": "condition_layer_20260625_test",
                "source_subscription_run_id": "market_data_subscription_20260626_test",
                "source_snapshot_run_id": "realtime_daily_snapshot_20260626_test",
                "source_previous_day_minute_run_id": "previous_day_minute_preload_20260625_test",
                "source_live_minute_run_id": "live_current_1m_source_20260626_test",
                "latest_closed_minute": "2026-06-26T10:01:00+08:00",
                "calculation_config": {
                    "completion_ratio_min_ready": "0.1",
                    "amount_projection_expand_threshold": "1.2",
                    "amount_projection_shrink_threshold": "0.8",
                    "price_flat_abs_pct_threshold": "0.001",
                    "window_total_seconds": 1800,
                    "calculation_method": "active_30m_bucket_projection_v1_strict_current_lineage",
                    "calculation_config_hash": "local-replay-asset-scope-test",
                },
            },
            "b2_input": {
                "snapshot_rows": [
                    {"asset_kind": "stock", "snapshot_id": 1, "subscription_id": 11, "identity_key": "stock:SH:600000", "exchange": "SH", "code": "600000", "display_code": "600000", "name": "浦发银行", "snapshot_time": "2026-06-26T10:00:00+08:00", "current_price": "10.0", "close": "10.0", "source_adapter": "LocalReplaySnapshotAdapter"},
                    {"asset_kind": "index", "snapshot_id": 2, "subscription_id": 12, "identity_key": "index:SH:000300", "exchange": "SH", "code": "000300", "display_code": "000300", "name": "沪深300", "snapshot_time": "2026-06-26T10:00:00+08:00", "current_price": "20.0", "close": "20.0", "source_adapter": "LocalReplaySnapshotAdapter"},
                    {"asset_kind": "board", "snapshot_id": 3, "subscription_id": 13, "identity_key": "board:TDX:881001", "exchange": "TDX", "code": "881001", "display_code": "881001", "name": "银行", "snapshot_time": "2026-06-26T10:15:00+08:00", "current_price": "30.0", "close": "30.0", "source_adapter": "LocalReplaySnapshotAdapter"},
                ],
                "live_current_rows_by_asset": {
                    "stock": [{"bar_id": 1, "bar_time": "2026-06-26T10:00:00+08:00", "open": "10", "high": "11", "low": "9", "close": "10", "volume": "100", "amount": "1000", "quality_status": "passed"}],
                    "index": [{"bar_id": 2, "bar_time": "2026-06-26T10:00:00+08:00", "open": "20", "high": "21", "low": "19", "close": "20", "volume": "200", "amount": "2000", "quality_status": "passed"}],
                    "board": [{"bar_id": 3, "bar_time": "2026-06-26T10:15:00+08:00", "open": "30", "high": "31", "low": "29", "close": "30", "volume": "300", "amount": "3000", "quality_status": "passed"}],
                },
                "previous_day_rows_by_asset": {
                    "stock": [{"bar_id": 11, "bar_time": "2026-06-25T10:00:00+08:00", "open": "10", "high": "11", "low": "9", "close": "10", "volume": "100", "amount": "1000", "quality_status": "passed"}],
                    "index": [{"bar_id": 12, "bar_time": "2026-06-25T10:00:00+08:00", "open": "20", "high": "21", "low": "19", "close": "20", "volume": "200", "amount": "2000", "quality_status": "passed"}],
                    "board": [{"bar_id": 13, "bar_time": "2026-06-25T10:15:00+08:00", "open": "30", "high": "31", "low": "29", "close": "30", "volume": "300", "amount": "3000", "quality_status": "passed"}],
                },
            },
            "source_meta": {
                "historical_source_type": "explicit_source_bundle",
                "historical_source_path": str(target),
                "historical_source_hash": "test-mixed-asset-scope-hash",
                "source_row_count": 6,
                "candidate_count": 3,
                "context_count": 3,
                "b2_snapshot_row_count": 3,
                "b2_live_current_row_count": 3,
                "b2_previous_day_row_count": 3,
                "upstream_source_mode": "live_current_1m",
                "bundle_contract_version": "historical_replay_source_bundle_v1",
            },
        }
        original = json.dumps(bundle, ensure_ascii=False, indent=2, default=str) + "\n"
        target.write_text(original, encoding="utf-8")
        return target

    def _seed_index_board_source_bundle(self, replay_root: Path) -> Path:
        full_scope_path = self._seed_mixed_asset_scope_bundle(replay_root)
        bundle = json.loads(full_scope_path.read_text(encoding="utf-8"))
        target = replay_root / "_sources" / "20260626_index_board_only" / "source_bundle.json"
        target.parent.mkdir(parents=True, exist_ok=True)
        bundle["source_records"] = {
            key: rows
            for key, rows in bundle["source_records"].items()
            if rows and rows[0].get("asset_kind") in {"index", "board"}
        }
        bundle["candidates"] = [row for row in bundle["candidates"] if row.get("asset_kind") in {"index", "board"}]
        bundle["n4_context_snapshot_rows"] = [
            row for row in bundle["n4_context_snapshot_rows"] if row.get("asset_kind") in {"index", "board"}
        ]
        bundle["b2_input"]["snapshot_rows"] = [
            row for row in bundle["b2_input"]["snapshot_rows"] if row.get("asset_kind") in {"index", "board"}
        ]
        bundle["b2_input"]["live_current_rows_by_asset"] = {
            key: rows
            for key, rows in bundle["b2_input"]["live_current_rows_by_asset"].items()
            if key in {"index", "board"}
        }
        bundle["b2_input"]["previous_day_rows_by_asset"] = {
            key: rows
            for key, rows in bundle["b2_input"]["previous_day_rows_by_asset"].items()
            if key in {"index", "board"}
        }
        bundle["source_meta"].update({
            "historical_source_type": "explicit_source_bundle_index_board_only",
            "historical_source_path": str(target),
            "historical_source_hash": "test-index-board-source-bundle-hash",
            "source_row_count": 4,
            "candidate_count": 2,
            "context_count": 2,
            "b2_snapshot_row_count": 2,
            "b2_live_current_row_count": 2,
            "b2_previous_day_row_count": 2,
        })
        target.write_text(json.dumps(bundle, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
        return target

    def _seed_index_board_full_day_source_bundle(self, replay_root: Path) -> Path:
        self._seed_index_board_source_bundle(replay_root)
        export = export_c1_index_board_full_day_source_bundle(
            replay_root=replay_root,
            trade_date="2026-06-26",
            index_rows=self._c1_rows(asset_kind="index", identity_key="index:SH:000300", code="000300"),
            board_rows=self._c1_rows(asset_kind="board", identity_key="board:TDX:881001", code="881001"),
            template_source_bundle_key="20260626_index_board_only",
        )
        return Path(export["source_bundle_path"])

    def _seed_asset_unit_fix_delta_source_bundle(self, replay_root: Path) -> Path:
        self._seed_explicit_source_bundle(replay_root)
        source = replay_root / "_sources" / "20260626" / "source_bundle.json"
        bundle = json.loads(source.read_text(encoding="utf-8"))
        bundle["replay_config"].pop("source_condition_run_id", None)
        target = replay_root / "_sources" / "20260626_asset_unit_fix_delta" / "source_bundle.json"
        target.parent.mkdir(parents=True, exist_ok=True)

        def trigger_row(asset_kind: str, identity_key: str, condition_key: str = "BUY:D") -> dict[str, object]:
            return {
                "for_trade_date": "20260626",
                "asset_kind": asset_kind,
                "identity_key": identity_key,
                "direction": "buy",
                "signal_type": "B_BUY",
                "condition_key": condition_key,
                "trigger_period": "D",
                "trigger_mark_candidate": "normal",
                "trigger_time": "2026-06-26T14:47:00+08:00",
                "trigger_price": "10.00",
                "source_condition_run_id": "condition_layer_20260625_source_20260625_for_20260626_v1",
                "source_run_id": "production_lineage_trace_only",
                "source_metric_run_id": "realtime_action_confirmation_metric_20260626_until_1447__asset_unit_fix_v1_trace_only",
                "source_n3p_live_target_run_id": "realtime_action_confirmation_metric_20260626_until_1447__asset_unit_fix_v1_trace_only",
                "source_mode": "live_current_1m",
                "c1_dependency": False,
                "n5_entry_allowed": True,
                "event_id": f"prod_evt_{identity_key.replace(':', '_')}_{condition_key.replace(':', '_')}",
                "dedup_key": f"production_dedup_{identity_key}",
            }

        old_rows = [
            trigger_row("stock", f"stock:SH:{600000 + index:06d}")
            for index in range(295)
        ]
        corrected_rows = [
            *old_rows,
            *[
                trigger_row("stock", f"stock:SZ:{300000 + index:06d}", "BUY:Q,M,W,D")
                for index in range(90)
            ],
            *[
                trigger_row("index", f"index:SH:{index:06d}", "BUY:D")
                for index in range(1, 6)
            ],
            *[
                trigger_row("board", f"board:TDX:{881000 + index}", "BUY:Q,M,W,D")
                for index in range(1, 17)
            ],
        ]
        bundle["asset_unit_fix_delta_validation"] = {
            "old_unified_n4_run_id": "trigger_provisional_ordinary_20260626_until_1447__old_unified_trace_only",
            "corrected_n4_run_id": "trigger_provisional_ordinary_20260626_until_1447__asset_unit_fix_v1_trace_only",
            "old_unified_trigger_matched": old_rows,
            "corrected_trigger_matched": corrected_rows,
        }
        bundle["source_meta"].update({
            "historical_source_type": "explicit_source_bundle_asset_unit_fix_delta",
            "historical_source_path": str(target),
            "historical_source_hash": "test-asset-unit-fix-delta-source-bundle-hash",
        })
        target.write_text(json.dumps(bundle, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
        return target

    def test_default_replay_engine_is_canonical_plan_v1(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            job = create_local_replay_job(
                replay_root=Path(tmp),
                trade_date="2026-06-26",
            )
            self.assertEqual(job["replay_engine_version"], "canonical_plan_v1")

    def test_canonical_plan_v1_without_historical_source_fails_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SOURCE_UNAVAILABLE"):
                create_local_replay_job(
                    replay_root=Path(tmp),
                    trade_date="2026-06-27",
                )

    def test_create_local_replay_job_writes_only_local_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
            )

            self.assertTrue(job["job_id"].startswith("local_replay_20260626_"))
            self.assertEqual(job["trade_date"], "2026-06-26")
            self.assertEqual(job["safety_flags"]["replay_mode"], "local_only")
            self.assertFalse(job["safety_flags"]["database_write"])
            self.assertFalse(job["safety_flags"]["consume_outbox"])
            self.assertFalse(job["safety_flags"]["update_checkpoint"])
            self.assertFalse(job["safety_flags"]["worker_started"])
            self.assertFalse(job["safety_flags"]["production_run_id_used"])
            self.assertTrue(Path(job["artifact_dir"]).is_relative_to(root.resolve()))

            artifact_dir = Path(job["artifact_dir"])
            expected_files = {
                "replay_config.json",
                "replay_status.json",
                "replay_timeline.jsonl",
                "n3_messages.jsonl",
                "n4_messages.jsonl",
                "n5_messages.jsonl",
                "replay_summary.json",
                "replay_summary.md",
                "n3_n5_full_day_replay.xlsx",
            }
            self.assertTrue(expected_files.issubset({path.name for path in artifact_dir.iterdir()}))

            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            self.assertEqual(summary["replay_engine_version"], "canonical_plan_v1")
            self.assertEqual(summary["source_policy"], "historical_minute_local_replay")
            self.assertEqual(summary["historical_source_status"], "available")
            self.assertEqual(summary["timeline_minutes"], 240)
            self.assertEqual(summary["n5"]["ActionEligible"], 3)
            self.assertEqual(summary["n5"]["ActionExecuted"], 0)
            self.assertEqual(summary["n5"]["BUY_HINT_ActionEligible"], 1)
            self.assertEqual(summary["n5"]["SELL_HINT_ActionEligible"], 1)
            self.assertEqual(summary["n5"]["b2_hint_final_proof_rows"], 0)
            self.assertEqual(summary["canonical_planner_trace"]["n4_ordinary_matcher"], "provisional_ordinary_matcher_v1")
            self.assertEqual(summary["canonical_planner_trace"]["n4_hint_matcher"], "provisional_projection_matcher_v1")
            self.assertEqual(summary["canonical_planner_trace"]["n5_actioneligible"], "provisional_actioneligible_v1")
            self.assertEqual(summary["canonical_planner_trace"]["n5_actionexecuted"], "provisional_action_executed_dry_run_v1")
            self.assertIn("Local replay only", (artifact_dir / "replay_summary.md").read_text(encoding="utf-8"))

            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "summary",
                    "minute_timeline",
                    "n4_ordinary_messages",
                    "n4_hint_messages",
                    "n5_action_eligible",
                    "n5_action_executed",
                    "n5_action_skipped",
                    "quality_blockers",
                    "lineage_and_safety",
                ],
            )
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("replay_engine_version", "canonical_plan_v1"), lineage_rows)
            self.assertIn(
                ("canonical_planner_trace.n4_ordinary_matcher", "provisional_ordinary_matcher_v1"),
                lineage_rows,
            )
            self.assertIn(("source_policy", "historical_minute_local_replay"), lineage_rows)
            n4_headers = [cell.value for cell in next(workbook["n4_ordinary_messages"].iter_rows(min_row=1, max_row=1))]
            n5_headers = [cell.value for cell in next(workbook["n5_action_executed"].iter_rows(min_row=1, max_row=1))]
            self.assertEqual(
                n4_headers,
                [
                    "minute",
                    "event_type",
                    "asset_kind",
                    "identity_key",
                    "signal_type",
                    "condition_key",
                    "original_condition_key",
                    "trigger_type",
                    "trigger_price",
                    "trigger_mark_candidate",
                    "source_mode",
                    "source",
                    "source_policy",
                    "replay_engine_version",
                    "trace_summary",
                    "lineage_summary",
                ],
            )
            self.assertEqual(
                n5_headers,
                [
                    "minute",
                    "event_type",
                    "asset_kind",
                    "identity_key",
                    "signal_type",
                    "condition_key",
                    "original_condition_key",
                    "action_state",
                    "action_mark",
                    "confirmation_metric_id",
                    "source_trigger_run_id",
                    "source_trigger_event_type",
                    "final_proof_source",
                    "replay_engine_version",
                ],
            )

    def test_explicit_source_bundle_metadata_is_exposed_in_summary_and_status(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_explicit_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
            )

            loaded = read_replay_job(root, job["job_id"])
            summary = loaded["summary"]
            self.assertEqual(summary["historical_source_kind"], "local_json")
            self.assertEqual(summary["source_meta"]["historical_source_type"], "explicit_source_bundle")
            self.assertTrue(str(summary["source_meta"]["historical_source_path"]).endswith("source_bundle.json"))
            self.assertEqual(summary["source_meta"]["source_row_count"], 960)
            self.assertEqual(summary["source_meta"]["candidate_count"], 4)
            self.assertEqual(summary["source_meta"]["context_count"], 4)
            self.assertEqual(summary["source_meta"]["b2_snapshot_row_count"], 2)
            self.assertEqual(summary["source_meta"]["b2_live_current_row_count"], 2)
            self.assertEqual(summary["source_meta"]["b2_previous_day_row_count"], 2)
            self.assertEqual(summary["source_meta"]["upstream_source_mode"], "live_current_1m")
            self.assertEqual(
                loaded["historical_source_kind"],
                "local_json",
            )

    def test_source_bundle_key_selects_prefiltered_bundle_and_records_lineage(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_path = self._seed_index_board_source_bundle(root)
            original_text = source_path.read_text(encoding="utf-8")

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only",
            )

            artifact_dir = Path(job["artifact_dir"])
            loaded = read_replay_job(root, job["job_id"])
            summary = loaded["summary"]
            config = json.loads((artifact_dir / "replay_config.json").read_text(encoding="utf-8"))
            status = json.loads((artifact_dir / "replay_status.json").read_text(encoding="utf-8"))
            n4_messages = read_replay_messages(artifact_dir, "n4")
            n5_messages = read_replay_messages(artifact_dir, "n5")
            self.assertEqual(job["source_bundle_key"], "20260626_index_board_only")
            self.assertEqual(summary["source_bundle_key"], "20260626_index_board_only")
            self.assertEqual(summary["source_bundle_selector_mode"], "explicit")
            self.assertTrue(summary["resolved_source_bundle_path"].endswith("20260626_index_board_only/source_bundle.json"))
            self.assertEqual(summary["source_meta"]["historical_source_type"], "explicit_source_bundle_index_board_only")
            self.assertEqual(summary["asset_scope"], "index_board_only")
            self.assertTrue(summary["asset_scope_filter_applied"])
            self.assertEqual(summary["asset_scope_source_counts_after"]["candidates"]["stock"], 0)
            self.assertNotIn("stock", {row.get("asset_kind") for row in n4_messages})
            self.assertNotIn("stock", {row.get("asset_kind") for row in n5_messages})
            self.assertEqual(config["source_bundle_key"], "20260626_index_board_only")
            self.assertEqual(status["source_bundle_selector_mode"], "explicit")
            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("source_bundle_key", "20260626_index_board_only"), lineage_rows)
            self.assertIn(("source_bundle_selector_mode", "explicit"), lineage_rows)
            self.assertIn(("source_meta.historical_source_type", "explicit_source_bundle_index_board_only"), lineage_rows)
            self.assertEqual(source_path.read_text(encoding="utf-8"), original_text)

    def test_source_bundle_key_invalid_or_missing_fails_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_source_bundle(root)
            for source_bundle_key in (
                "missing_bundle_key",
                "../20260626",
                "/tmp/source_bundle",
                "trigger_provisional_ordinary_20260626_until_1447",
            ):
                with self.subTest(source_bundle_key=source_bundle_key):
                    with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SOURCE_BUNDLE_NOT_FOUND"):
                        create_local_replay_job(
                            replay_root=root,
                            trade_date="2026-06-26",
                            source_bundle_key=source_bundle_key,
                        )

    def test_c1_readiness_requires_240_minutes_per_index_board_object(self) -> None:
        index_rows = self._c1_rows(asset_kind="index", identity_key="index:SH:000300", code="000300")
        board_rows = self._c1_rows(asset_kind="board", identity_key="board:TDX:881001", code="881001")

        passed = build_c1_index_board_readiness_report(
            trade_date="2026-06-26",
            index_rows=index_rows,
            board_rows=board_rows,
        )
        self.assertEqual(passed["status"], "passed")
        self.assertEqual(passed["expected_minutes_per_object"], 240)
        self.assertEqual(passed["index_object_count"], 1)
        self.assertEqual(passed["board_object_count"], 1)
        self.assertEqual(passed["actual_minutes_distribution"], {"240": 2})

        blocked = build_c1_index_board_readiness_report(
            trade_date="2026-06-26",
            index_rows=index_rows[:-1],
            board_rows=board_rows,
        )
        self.assertEqual(blocked["status"], "blocked")
        self.assertEqual(blocked["blocked_reason"], "BLOCKED_REPLAY_C1_SOURCE_INCOMPLETE")
        self.assertEqual(blocked["missing_asset_count"], 1)

    def test_c1_export_writes_full_day_source_bundle_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_source_bundle(root)

            export = export_c1_index_board_full_day_source_bundle(
                replay_root=root,
                trade_date="2026-06-26",
                index_rows=self._c1_rows(asset_kind="index", identity_key="index:SH:000300", code="000300"),
                board_rows=self._c1_rows(asset_kind="board", identity_key="board:TDX:881001", code="881001"),
                template_source_bundle_key="20260626_index_board_only",
            )

            self.assertEqual(export["source_bundle_key"], "20260626_index_board_only_full_day")
            source_path = Path(export["source_bundle_path"])
            self.assertTrue(source_path.is_relative_to(root.resolve() / "_sources"))
            bundle = json.loads(source_path.read_text(encoding="utf-8"))
            source_meta = bundle["source_meta"]
            self.assertEqual(source_meta["historical_source_type"], "explicit_source_bundle_index_board_only_full_day")
            self.assertEqual(source_meta["source_origin"], "c1_read_only_export")
            self.assertEqual(source_meta["c1_readiness_status"], "passed")
            self.assertEqual(source_meta["stock_count"], 0)
            self.assertEqual(source_meta["expected_minutes_per_object"], 240)
            self.assertEqual(source_meta["actual_minutes_distribution"], {"240": 2})
            self.assertEqual(sum(len(rows) for rows in bundle["source_records"].values()), 482)
            self.assertEqual(source_meta["c1_current_day_row_count"], 480)
            source_record_keys = set(bundle["source_records"])
            for candidate in bundle["candidates"]:
                self.assertIn(candidate["source_record_key"], source_record_keys)

    def test_c1_export_fails_closed_when_readiness_blocked(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_source_bundle(root)

            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_C1_SOURCE_INCOMPLETE"):
                export_c1_index_board_full_day_source_bundle(
                    replay_root=root,
                    trade_date="2026-06-26",
                    index_rows=self._c1_rows(asset_kind="index", identity_key="index:SH:000300", code="000300")[:-1],
                    board_rows=self._c1_rows(asset_kind="board", identity_key="board:TDX:881001", code="881001"),
                    template_source_bundle_key="20260626_index_board_only",
                )

    def test_asset_unit_fix_delta_validation_mode_writes_delta_artifacts_and_n5_delta_only_messages(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_asset_unit_fix_delta_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="14:47",
                end_hhmm="14:47",
                source_bundle_key="20260626_asset_unit_fix_delta",
                validation_mode="asset_unit_fix_delta_v1",
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = read_replay_job(root, job["job_id"])["summary"]
            config = json.loads((artifact_dir / "replay_config.json").read_text(encoding="utf-8"))
            status = json.loads((artifact_dir / "replay_status.json").read_text(encoding="utf-8"))
            delta_rows = [
                json.loads(line)
                for line in (artifact_dir / "n4_delta_attribution.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            n5_delta_rows = [
                json.loads(line)
                for line in (artifact_dir / "n5_delta_only_messages.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]

            self.assertEqual(job["validation_mode"], "asset_unit_fix_delta_v1")
            self.assertEqual(config["validation_mode"], "asset_unit_fix_delta_v1")
            self.assertEqual(status["validation_mode"], "asset_unit_fix_delta_v1")
            self.assertTrue(summary["asset_unit_fix_delta_validation"])
            self.assertEqual(summary["validation_mode"], "asset_unit_fix_delta_v1")
            self.assertEqual(summary["asset_unit_fix_delta"]["corrected_full_trigger_matched"], 406)
            self.assertEqual(summary["asset_unit_fix_delta"]["old_unified_trigger_matched"], 295)
            self.assertEqual(summary["asset_unit_fix_delta"]["common_old_and_corrected"], 295)
            self.assertEqual(summary["asset_unit_fix_delta"]["corrected_only"], 111)
            self.assertEqual(summary["asset_unit_fix_delta"]["old_only"], 0)
            self.assertEqual(summary["asset_unit_fix_delta"]["index_board_delta"], 21)
            self.assertEqual(summary["asset_unit_fix_delta"]["excluded_stock_replay"], 90)
            self.assertEqual(summary["asset_unit_fix_delta"]["other_requires_review"], 0)
            self.assertEqual(len(delta_rows), 406)
            classifications = {}
            for row in delta_rows:
                classifications[row["delta_classification"]] = classifications.get(row["delta_classification"], 0) + 1
            self.assertEqual(classifications["common_old_and_corrected"], 295)
            self.assertEqual(classifications["index_board_unit_fix_new_signal"], 21)
            self.assertEqual(classifications["stock_replayed_due_no_previous_baseline"], 90)
            self.assertEqual(summary["n5_delta_only"]["ActionEligible"], 21)
            self.assertEqual(summary["n5_delta_only"]["ActionExecuted"], 0)
            self.assertEqual(summary["n5_delta_only"]["stock_ActionEligible"], 0)
            self.assertEqual(summary["n5_delta_only"]["stock_ActionExecuted"], 0)
            self.assertEqual(summary["n5_delta_only"]["b2_hint_final_proof_rows"], 0)
            self.assertEqual(len(n5_delta_rows), 21)
            self.assertEqual({row["asset_kind"] for row in n5_delta_rows}, {"index", "board"})
            self.assertEqual({row["event_type"] for row in n5_delta_rows}, {"ActionEligible"})
            self.assertEqual(
                {row["source_condition_run_id"] for row in n5_delta_rows},
                {"condition_layer_20260625_source_20260625_for_20260626_v1"},
            )
            self.assertTrue(all(row["source_mode"] == "replay" for row in n5_delta_rows))
            self.assertTrue(all(value is False for value in summary["safety_flags"].values() if isinstance(value, bool)))

            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            self.assertIn("n4_delta_attribution", workbook.sheetnames)
            self.assertIn("n5_delta_only_messages", workbook.sheetnames)
            self.assertIn("excluded_stock_replay_risk", workbook.sheetnames)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("validation_mode", "asset_unit_fix_delta_v1"), lineage_rows)
            self.assertIn(("asset_unit_fix_delta.corrected_only", 111), lineage_rows)

    def test_asset_unit_fix_delta_n4_payload_preserves_passthrough_fields(self) -> None:
        row = {
            "delta_classification": "index_board_unit_fix_new_signal",
            "for_trade_date": "20260626",
            "asset_kind": "board",
            "identity_key": "board:TDX:881002",
            "direction": "buy",
            "signal_type": "B_BUY",
            "condition_key": "BUY:Q,M,W,D",
            "original_condition_key": "BUY:Q,M,W,D",
            "trigger_period": "D",
            "trigger_mark_candidate": "normal",
            "trigger_time": "2026-06-26T14:47:00+08:00",
            "trigger_price": "10.00",
            "source_condition_run_id": "condition_layer_20260625_source_20260625_for_20260626_v1",
            "source_metric_run_id": "realtime_action_confirmation_metric_20260626_until_1447__asset_unit_fix_v1",
            "source_n3p_live_target_run_id": "realtime_action_confirmation_metric_20260626_until_1447__asset_unit_fix_v1",
            "source_mode": "live_current_1m",
            "c1_dependency": False,
            "n5_entry_allowed": True,
        }

        message = _asset_unit_fix_delta_n4_message(
            row,
            trade_date="2026-06-26",
            n4_run_id="local_replay_20260626_field_mapping__n4_asset_unit_fix_delta_v1",
            ordinal=1,
        )
        outbox_row = _n4_message_to_outbox_row(
            message,
            n4_run_id="local_replay_20260626_field_mapping__n4_asset_unit_fix_delta_v1",
        )
        payload = outbox_row["payload_json"]

        self.assertEqual(payload["source_condition_run_id"], row["source_condition_run_id"])
        self.assertEqual(payload["source_metric_run_id"], row["source_metric_run_id"])
        self.assertEqual(payload["source_n3p_live_target_run_id"], row["source_n3p_live_target_run_id"])
        self.assertEqual(payload["source_mode"], row["source_mode"])
        self.assertFalse(payload["c1_dependency"])
        self.assertTrue(payload["n5_entry_allowed"])
        self.assertEqual(payload["original_condition_key"], row["original_condition_key"])

    def test_asset_unit_fix_delta_missing_source_condition_run_id_fails_closed(self) -> None:
        with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_DELTA_SOURCE_MISMATCH"):
            _build_asset_unit_fix_n5_delta_only_messages(
                trade_date="2026-06-26",
                replay_run_id="local_replay_20260626_missing_source_condition",
                replay_config={},
                delta_rows=[
                    {
                        "delta_classification": "index_board_unit_fix_new_signal",
                        "for_trade_date": "20260626",
                        "asset_kind": "index",
                        "identity_key": "index:SH:000001",
                        "signal_type": "B_BUY",
                        "condition_key": "BUY:D",
                        "trigger_period": "D",
                        "trigger_time": "2026-06-26T14:47:00+08:00",
                    }
                ],
            )

    def test_invalid_validation_mode_fails_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_INVALID_VALIDATION_MODE"):
                create_local_replay_job(
                    replay_root=Path(tmp),
                    trade_date="2026-06-26",
                    validation_mode="unknown_validation_mode_v9",
                )

    def test_replay_profiling_writes_phase_artifact_when_enabled(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_explicit_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="09:31",
                enable_profiling=True,
            )

            artifact_dir = Path(job["artifact_dir"])
            profile = json.loads((artifact_dir / "replay_profile.json").read_text(encoding="utf-8"))
            self.assertEqual(profile["job_id"], job["job_id"])
            self.assertEqual(profile["trade_date"], "2026-06-26")
            self.assertEqual(profile["status"], "completed")
            self.assertEqual(profile["replay_engine_version"], "canonical_plan_v1")
            self.assertTrue((artifact_dir / "replay_profile.md").exists())
            self.assertEqual(profile["safety_flags"]["database_write"], False)
            self.assertEqual(
                set(profile["phases"].keys()),
                {
                    "source_bundle_json_load",
                    "source_bundle_normalization_validation",
                    "minute_loop_initialization",
                    "per_minute_n3p_plan_only",
                    "per_minute_b2_plan_only",
                    "per_minute_n4_ordinary_matcher_lifecycle",
                    "per_minute_n4_hint_matcher_lifecycle",
                    "per_minute_n5_actioneligible_planner",
                    "per_minute_n5_actionexecuted_evaluator",
                    "artifact_jsonl_serialization",
                    "excel_generation",
                },
            )
            self.assertIn(
                profile["bottleneck_classification"],
                {
                    "JSON_LOAD_BOTTLENECK",
                    "NORMALIZATION_BOTTLENECK",
                    "N3P_PLAN_BOTTLENECK",
                    "B2_PLAN_BOTTLENECK",
                    "N4_MATCHER_BOTTLENECK",
                    "N5_PLANNER_BOTTLENECK",
                    "EXCEL_SERIALIZATION_BOTTLENECK",
                    "ARTIFACT_JSONL_BOTTLENECK",
                    "UNKNOWN_TIMEOUT",
                },
            )

    def test_n3p_replay_cache_reuses_repeated_active_target_minute(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_source_bundle(root)

            with patch(
                "ashare_v3.web.n6_replay.build_n3p_plan_only_replay_artifact",
                wraps=build_n3p_plan_only_replay_artifact,
            ) as wrapped_n3p:
                job = create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    start_hhmm="09:58",
                    end_hhmm="10:02",
                    asset_scope="index_board_only",
                    source_bundle_key="20260626_index_board_only",
                    enable_profiling=True,
                )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            profile = json.loads((artifact_dir / "replay_profile.json").read_text(encoding="utf-8"))
            n4_messages = read_replay_messages(artifact_dir, "n4")
            n5_messages = read_replay_messages(artifact_dir, "n5")

            self.assertEqual(wrapped_n3p.call_count, 1)
            self.assertEqual(summary["n3p_cache_stats"]["n3p_cache_misses"], 1)
            self.assertEqual(summary["n3p_cache_stats"]["n3p_cache_hits"], 2)
            self.assertEqual(summary["n3p_cache_stats"]["n3p_cache_saved_calls"], 2)
            self.assertEqual(summary["n3p_cache_stats"]["n3p_cache_key_count"], 1)
            self.assertEqual(summary["n3p_cache_stats"]["empty_minute_fast_path_count"], 2)
            self.assertEqual(profile["metadata"]["n3p_cache_stats"], summary["n3p_cache_stats"])
            self.assertEqual(
                profile["phases"]["per_minute_n3p_plan_only"]["metrics"]["n3p_cache_hit"],
                2,
            )
            self.assertEqual(
                profile["phases"]["per_minute_n3p_plan_only"]["metrics"]["empty_minute_fast_path"],
                2,
            )
            self.assertNotIn("stock", {row.get("asset_kind") for row in n4_messages})
            self.assertNotIn("stock", {row.get("asset_kind") for row in n5_messages})
            self.assertFalse(any("stock:SZ:002668" in json.dumps(row) for row in [*n4_messages, *n5_messages]))

    def test_asset_scope_index_board_only_filters_out_stock_and_records_lineage(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_path = self._seed_mixed_asset_scope_bundle(root)
            original_text = source_path.read_text(encoding="utf-8")

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
                asset_scope="index_board_only",
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            config = json.loads((artifact_dir / "replay_config.json").read_text(encoding="utf-8"))
            status = json.loads((artifact_dir / "replay_status.json").read_text(encoding="utf-8"))
            n4_messages = read_replay_messages(artifact_dir, "n4")
            n5_messages = read_replay_messages(artifact_dir, "n5")
            self.assertEqual(summary["asset_scope"], "index_board_only")
            self.assertTrue(summary["asset_scope_filter_applied"])
            self.assertEqual(summary["asset_scope_allowed_asset_kinds"], ["index", "board"])
            self.assertEqual(summary["asset_scope_source_counts_before"]["candidates"]["stock"], 1)
            self.assertEqual(summary["asset_scope_source_counts_after"]["candidates"]["stock"], 0)
            self.assertEqual(summary["asset_scope_source_counts_after"]["context"]["stock"], 0)
            self.assertEqual(summary["asset_scope_source_counts_after"]["b2_snapshot"]["stock"], 0)
            self.assertGreater(summary["asset_scope_source_counts_after"]["source_records"]["index"], 0)
            self.assertGreater(summary["asset_scope_source_counts_after"]["source_records"]["board"], 0)
            self.assertEqual(config["asset_scope"], "index_board_only")
            self.assertEqual(status["asset_scope"], "index_board_only")
            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("asset_scope", "index_board_only"), lineage_rows)
            self.assertIn(("asset_scope_filter_applied", True), lineage_rows)
            self.assertIn(("asset_scope_source_counts_after.candidates.stock", 0), lineage_rows)
            self.assertNotIn("stock", {row.get("asset_kind") for row in n4_messages})
            self.assertNotIn("stock", {row.get("asset_kind") for row in n5_messages})
            self.assertFalse(
                any("manual_state_changed_002668" in str(row.get("event_id") or "") for row in n4_messages)
            )
            self.assertEqual(source_path.read_text(encoding="utf-8"), original_text)

    def test_full_day_shadow_v1_index_board_only_writes_shadow_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only_full_day",
                validation_mode="full_day_shadow_v1",
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            timeline = read_replay_timeline(artifact_dir)
            n4_messages = read_replay_messages(artifact_dir, "n4")
            n5_messages = read_replay_messages(artifact_dir, "n5")
            n4_shadow_rows = [
                json.loads(line)
                for line in (artifact_dir / "n4_shadow_state_transitions.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            n4_shadow_evaluations = [
                json.loads(line)
                for line in (artifact_dir / "n4_shadow_evaluations.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            n5_shadow_rows = [
                json.loads(line)
                for line in (artifact_dir / "n5_shadow_action_windows.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            shadow_report = json.loads((artifact_dir / "shadow_validation_report.json").read_text(encoding="utf-8"))

            self.assertEqual(summary["validation_mode"], "full_day_shadow_v1")
            self.assertTrue(summary["shadow_mode"])
            self.assertEqual(summary["asset_scope"], "index_board_only")
            self.assertEqual(summary["shadow"]["stock_n4_messages"], 0)
            self.assertEqual(summary["shadow"]["stock_n5_messages"], 0)
            self.assertEqual(summary["shadow"]["trigger_state_changed_action_entries"], 0)
            self.assertEqual(summary["shadow"]["ActionEligible"], summary["shadow"]["TriggerMatched"])
            self.assertEqual(shadow_report["stock_n4_messages"], 0)
            self.assertEqual(shadow_report["stock_n5_messages"], 0)
            self.assertEqual(len(timeline), 240)
            self.assertNotIn("12:00", {row["minute"] for row in timeline})
            self.assertNotIn("12:30", {row["minute"] for row in timeline})
            self.assertNotIn("stock", {row.get("asset_kind") for row in n4_messages})
            self.assertNotIn("stock", {row.get("asset_kind") for row in n5_messages})
            self.assertFalse(any(row.get("source_trigger_event_type") == "TriggerStateChanged" for row in n5_messages))
            self.assertGreaterEqual(summary["shadow"]["noop_suppressed"], 0)
            self.assertGreaterEqual(len(n4_shadow_rows), len(n4_messages))
            self.assertEqual(
                len(n4_shadow_evaluations),
                sum(int(row["n4_noop_suppressed"] or 0) for row in timeline) + len(n4_messages),
            )
            self.assertIn("09:31", {row["minute"] for row in n4_shadow_evaluations})
            self.assertIn("NoOp", {row["event_type"] for row in n4_shadow_evaluations})
            first_minute_evaluations = [row for row in n4_shadow_evaluations if row["minute"] == "09:31"]
            self.assertTrue(first_minute_evaluations)
            first_minute_metric_labels = {
                row["metric_time_label"]
                for row in first_minute_evaluations
                if row.get("metric_time_label")
            }
            self.assertTrue(first_minute_metric_labels)
            self.assertEqual(
                first_minute_metric_labels,
                {"2026-06-26 09:31"},
            )
            self.assertGreaterEqual(len(n5_shadow_rows), len(n5_messages))

            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            self.assertIn("n4_shadow_evaluations", workbook.sheetnames)
            self.assertIn("n4_shadow_state_transitions", workbook.sheetnames)
            self.assertIn("n5_shadow_action_windows", workbook.sheetnames)
            self.assertIn("shadow_quality_blockers", workbook.sheetnames)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("validation_mode", "full_day_shadow_v1"), lineage_rows)
            self.assertIn(("shadow_mode", True), lineage_rows)

    def test_full_day_shadow_v1_evaluates_scope_from_open_not_candidate_snapshot_minute(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source_path = self._seed_index_board_full_day_source_bundle(root)
            bundle = json.loads(source_path.read_text(encoding="utf-8"))
            for candidate in bundle["candidates"]:
                candidate["minute_label"] = "2026-06-26 14:47"
                candidate["observed_at"] = "2026-06-26 14:48:00+08:00"
            source_path.write_text(json.dumps(bundle, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")

            with patch(
                "ashare_v3.web.n6_replay.build_n3p_plan_only_replay_artifact",
                wraps=build_n3p_plan_only_replay_artifact,
            ) as wrapped_n3p:
                job = create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    start_hhmm="09:31",
                    end_hhmm="09:32",
                    asset_scope="index_board_only",
                    source_bundle_key="20260626_index_board_only_full_day",
                    validation_mode="full_day_shadow_v1",
                )

            timeline = read_replay_timeline(Path(job["artifact_dir"]))

            self.assertEqual(wrapped_n3p.call_count, 2)
            self.assertEqual([row["minute"] for row in timeline], ["09:31", "09:32"])
            self.assertTrue(all(row["n3_rows"] > 0 for row in timeline))

    def test_full_day_shadow_v1_routes_hint_candidates_by_original_condition_key(self) -> None:
        candidates = [
            {
                "asset_kind": "board",
                "identity_key": "board:TDX:881026",
                "condition_key": "LIVE_CURRENT_1M:B_BUY",
                "original_condition_key": "BUY_HINT",
                "condition_keys": ["BUY_HINT"],
                "minute_label": "2026-06-26 14:47",
            },
            {
                "asset_kind": "board",
                "identity_key": "board:TDX:881002",
                "condition_key": "LIVE_CURRENT_1M:B_BUY",
                "original_condition_key": "BUY:Q,M,W,D",
                "condition_keys": ["BUY:Q,M,W,D"],
                "minute_label": "2026-06-26 14:47",
            },
        ]

        hints = _active_candidates_for_minute(
            candidates,
            "09:31",
            hint_only=True,
            full_day_shadow=True,
            trade_date="2026-06-26",
        )
        ordinary = _active_candidates_for_minute(
            candidates,
            "09:31",
            hint_only=False,
            full_day_shadow=True,
            trade_date="2026-06-26",
        )

        self.assertEqual([row["identity_key"] for row in hints], ["board:TDX:881026"])
        self.assertEqual([row["identity_key"] for row in ordinary], ["board:TDX:881002"])

    def test_full_day_shadow_v1_builds_generic_closed_confirmation_metrics(self) -> None:
        metric_rows = [
            {
                "asset_kind": "board",
                "identity_key": "board:TDX:881002",
                "projection_run_id": "local_replay_20260626_test__n3p",
                "action_confirmation_metric_id": 123,
                "metric_ready": True,
                "metric_quality_status": "passed",
                "metric_minute_label": "09:31",
                "metric_time": "2026-06-26T09:31:00+08:00",
                "metric_time_label": "2026-06-26 09:31",
                "raw_json": {
                    "signal_type": "B_BUY",
                    "condition_key": "LIVE_CURRENT_1M:B_BUY",
                    "original_condition_key": "BUY:D",
                    "condition_keys": ["BUY:D"],
                },
            }
        ]

        closed = _build_closed_confirmation_metric_rows_for_minute(
            metric_rows=metric_rows,
            trade_date="2026-06-26",
            current_minute="09:32",
        )

        self.assertEqual(len(closed), 1)
        self.assertEqual(closed[0]["identity_key"], "board:TDX:881002")
        self.assertTrue(closed[0]["is_closed_1m"])
        self.assertEqual(closed[0]["metric_time_label"], "2026-06-26 09:31")

    def test_full_day_shadow_v1_b2_input_uses_c1_minute_snapshot(self) -> None:
        source_bundle = {
            "source_records": {
                "board:TDX:881002": [
                    {
                        "asset_kind": "board",
                        "identity_key": "board:TDX:881002",
                        "code": "881002",
                        "display_code": "881002",
                        "datetime": "2026-06-26 09:31:00+08:00",
                        "bar_time": "2026-06-26T09:31:00+08:00",
                        "open": 10,
                        "high": 11,
                        "low": 9,
                        "close": 10.5,
                        "volume": 100,
                        "amount": 1000,
                    },
                    {
                        "asset_kind": "board",
                        "identity_key": "board:TDX:881002",
                        "code": "881002",
                        "display_code": "881002",
                        "datetime": "2026-06-26 09:32:00+08:00",
                        "bar_time": "2026-06-26T09:32:00+08:00",
                        "open": 10.5,
                        "high": 12,
                        "low": 10,
                        "close": 11.5,
                        "volume": 120,
                        "amount": 1200,
                    },
                ]
            },
            "b2_input": {
                "snapshot_rows": [],
                "live_current_rows_by_asset": {"board": []},
                "previous_day_rows_by_asset": {"board": []},
            },
        }

        b2_input = _build_canonical_b2_input(
            source_bundle,
            minute="09:31",
            trade_date="2026-06-26",
            shadow_mode=True,
        )

        self.assertEqual(len(b2_input["snapshot_rows"]), 1)
        self.assertEqual(b2_input["snapshot_rows"][0]["identity_key"], "board:TDX:881002")
        self.assertEqual(b2_input["snapshot_rows"][0]["snapshot_time"].isoformat(), "2026-06-26T09:31:00+08:00")
        self.assertEqual(len(b2_input["live_current_rows_by_asset"]["board"]), 1)

    def test_full_day_shadow_v1_does_not_filter_closed_metrics_by_current_minute_run_id(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)
            captured_run_ids: list[str | None] = []

            def fake_executed_report(**kwargs):
                captured_run_ids.append(kwargs.get("confirmation_metric_run_id"))
                return {
                    "side_effect_guard": {"db_written": False},
                    "action_executed_plans": [],
                }

            with patch(
                "ashare_v3.web.n6_replay.provisional_action_executed_dry_run.build_provisional_action_executed_dry_run_report",
                side_effect=fake_executed_report,
            ):
                create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    start_hhmm="09:31",
                    end_hhmm="09:31",
                    asset_scope="index_board_only",
                    source_bundle_key="20260626_index_board_only_full_day",
                    validation_mode="full_day_shadow_v1",
                )

        self.assertTrue(captured_run_ids)
        self.assertEqual(set(captured_run_ids), {None})

    def test_full_day_shadow_v1_prefilter_audit_writes_prefilter_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="09:32",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only_full_day",
                validation_mode="full_day_shadow_v1",
                n3p_strategy="prefilter_audit",
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            status = json.loads((artifact_dir / "replay_status.json").read_text(encoding="utf-8"))
            audit_rows = [
                json.loads(line)
                for line in (artifact_dir / "n4_lightweight_prefilter_audit.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            demand_rows = [
                json.loads(line)
                for line in (artifact_dir / "n3p_demand_plan.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]

            self.assertEqual(summary["n3p_strategy"], "prefilter_audit")
            self.assertEqual(status["n3p_strategy"], "prefilter_audit")
            self.assertGreater(len(audit_rows), 0)
            self.assertEqual(len(demand_rows), 2)
            self.assertEqual(summary["n3p_prefilter"]["false_negative_count"], 0)
            self.assertEqual(summary["n3p_prefilter"]["strategy"], "prefilter_audit")
            self.assertIn("prefilter_keep", {row["decision"] for row in audit_rows})

            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            self.assertIn("n4_prefilter_audit", workbook.sheetnames)
            self.assertIn("n3p_demand_plan", workbook.sheetnames)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("n3p_strategy", "prefilter_audit"), lineage_rows)

    def test_full_day_shadow_v1_prefilter_respects_any_period_thresholds(self) -> None:
        buy_context = {
            "condition_key": "BUY:Q,M,W,D",
            "direction": "buy",
            "period_trigger_baseline_json": {
                "periods": {
                    "Q": {"previous_entity_high": "1048.56"},
                    "M": {"previous_entity_high": "1000.00"},
                    "W": {"previous_entity_high": "980.00"},
                    "D": {"previous_entity_high": "950.00"},
                }
            },
        }
        buy_decision, buy_reason, buy_threshold = _prefilter_decision_for_context(
            context=buy_context,
            source_row={"high": "970.24", "low": "960.07", "close": "960.29"},
            previous_live_state=False,
            open_action_window=False,
        )
        self.assertEqual(buy_decision, "prefilter_keep")
        self.assertEqual(buy_reason, "buy_price_candidate")
        self.assertEqual(str(buy_threshold), "950.00")

        sell_context = {
            "condition_key": "SELL:Y,M,D",
            "direction": "sell",
            "period_trigger_baseline_json": {
                "periods": {
                    "Y": {"previous_entity_low": "401.12"},
                    "M": {"previous_entity_low": "403.00"},
                    "D": {"previous_entity_low": "404.00"},
                }
            },
        }
        sell_decision, sell_reason, sell_threshold = _prefilter_decision_for_context(
            context=sell_context,
            source_row={"high": "404.92", "low": "402.34", "close": "402.34"},
            previous_live_state=False,
            open_action_window=False,
        )
        self.assertEqual(sell_decision, "prefilter_keep")
        self.assertEqual(sell_reason, "sell_price_candidate")
        self.assertEqual(str(sell_threshold), "404.00")

    def test_full_day_shadow_v1_prefilter_prune_matches_audit_fixture_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            audit = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="09:33",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only_full_day",
                validation_mode="full_day_shadow_v1",
                n3p_strategy="prefilter_audit",
            )
            prune = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="09:33",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only_full_day",
                validation_mode="full_day_shadow_v1",
                n3p_strategy="prefilter_prune",
            )

            audit_dir = Path(audit["artifact_dir"])
            prune_dir = Path(prune["artifact_dir"])
            audit_summary = json.loads((audit_dir / "replay_summary.json").read_text(encoding="utf-8"))
            prune_summary = json.loads((prune_dir / "replay_summary.json").read_text(encoding="utf-8"))

            self.assertEqual(audit_summary["n4"], prune_summary["n4"])
            self.assertEqual(audit_summary["n5"], prune_summary["n5"])
            self.assertEqual(audit_summary["shadow"], prune_summary["shadow"])
            self.assertEqual(prune_summary["n3p_prefilter"]["false_negative_count"], 0)
            self.assertGreaterEqual(prune_summary["n3p_prefilter"]["n3p_calls_saved"], 0)

    def test_full_day_shadow_v1_prefilter_audit_fails_closed_on_false_negative(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            with patch("ashare_v3.web.n6_replay._build_n4_lightweight_prefilter_rows", return_value=[{
                "minute": "09:31",
                "decision": "prefilter_drop",
                "reason": "test_drop",
                "asset_kind": "index",
                "identity_key": "index:SH:000300",
                "condition_key": "BUY:D",
                "signal_type": "B_BUY",
            }]), patch("ashare_v3.web.n6_replay._annotate_prefilter_rows_with_canonical_events", return_value=[{
                "minute": "09:31",
                "decision": "prefilter_drop",
                "reason": "test_drop",
                "asset_kind": "index",
                "identity_key": "index:SH:000300",
                "condition_key": "BUY:D",
                "signal_type": "B_BUY",
                "canonical_event_type": "TriggerMatched",
                "false_negative": True,
            }]):
                with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_PREFILTER_FALSE_NEGATIVE"):
                    create_local_replay_job(
                        replay_root=root,
                        trade_date="2026-06-26",
                        start_hhmm="09:31",
                        end_hhmm="09:31",
                        asset_scope="index_board_only",
                        source_bundle_key="20260626_index_board_only_full_day",
                        validation_mode="full_day_shadow_v1",
                        n3p_strategy="prefilter_audit",
                    )

    def test_n3p_reduction_mode_rejects_invalid_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_INVALID_N3P_REDUCTION_MODE"):
                create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    start_hhmm="09:31",
                    end_hhmm="09:31",
                    asset_scope="index_board_only",
                    source_bundle_key="20260626_index_board_only_full_day",
                    validation_mode="full_day_shadow_v1",
                    n3p_strategy="prefilter_prune",
                    n3p_reduction_mode="unknown",
                )

    def test_n3p_negative_cache_rejects_invalid_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_INVALID_N3P_NEGATIVE_CACHE"):
                create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    start_hhmm="09:31",
                    end_hhmm="09:31",
                    asset_scope="index_board_only",
                    source_bundle_key="20260626_index_board_only_full_day",
                    validation_mode="full_day_shadow_v1",
                    n3p_strategy="prefilter_prune",
                    n3p_reduction_mode="active_state_fast_path",
                    n3p_negative_cache="unknown",
                )

    def test_n3p_negative_cache_decision_fails_open_unless_safe_proof_matches(self) -> None:
        cache = _N3PNegativeProofCache(
            enabled=True,
            replay_config={
                "historical_source_hash": "hash-a",
                "source_bundle_key": "20260626_index_board_only_full_day",
                "asset_scope": "index_board_only",
                "replay_engine_version": "canonical_plan_v1",
            },
        )
        candidate = {
            "asset_kind": "board",
            "identity_key": "board:TDX:881001",
            "signal_type": "S_SELL",
            "condition_key": "SELL:D",
            "minute_label": "2026-06-26 09:31",
        }
        proof = {
            "proof_version": "n3p_plan_only_proof_summary_v1",
            "asset_kind": "board",
            "identity_key": "board:TDX:881001",
            "stable_trigger_key": "20260626|board|board:TDX:881001|S_SELL|SELL:D|D|normal|09:31",
            "source_input_fingerprint": "source-a",
            "context_fingerprint": "context-a",
            "safe_negative_cacheable": True,
        }

        missing = cache.decision(candidate=candidate, proof_context={})
        self.assertFalse(missing["skip_full_n3p"])
        self.assertEqual(missing["decision"], "fail_open_missing_proof")

        bad_version = dict(proof)
        bad_version["proof_version"] = "unknown"
        cache.store([bad_version])
        unknown = cache.decision(candidate=candidate, proof_context={"source_input_fingerprint": "source-a", "context_fingerprint": "context-a"})
        self.assertFalse(unknown["skip_full_n3p"])
        self.assertEqual(unknown["decision"], "fail_open_unknown_proof_version")

        unsafe = dict(proof)
        unsafe["safe_negative_cacheable"] = False
        cache.store([unsafe])
        not_safe = cache.decision(candidate=candidate, proof_context={"source_input_fingerprint": "source-a", "context_fingerprint": "context-a"})
        self.assertFalse(not_safe["skip_full_n3p"])
        self.assertEqual(not_safe["decision"], "fail_open_not_safe_cacheable")

        cache.store([proof])
        changed = cache.decision(candidate=candidate, proof_context={"source_input_fingerprint": "source-b", "context_fingerprint": "context-a"})
        self.assertFalse(changed["skip_full_n3p"])
        self.assertEqual(changed["decision"], "fail_open_fingerprint_changed")

        matched = cache.decision(candidate=candidate, proof_context={"source_input_fingerprint": "source-a", "context_fingerprint": "context-a"})
        self.assertTrue(matched["skip_full_n3p"])
        self.assertEqual(matched["decision"], "negative_cache_hit")

    def test_n3p_negative_cache_matches_original_condition_key_and_requested_periods(self) -> None:
        cache = _N3PNegativeProofCache(
            enabled=True,
            replay_config={
                "historical_source_hash": "hash-a",
                "source_bundle_key": "20260626_index_board_only_full_day",
                "asset_scope": "index_board_only",
                "replay_engine_version": "canonical_plan_v1",
            },
        )
        candidate = {
            "for_trade_date": "20260626",
            "asset_kind": "board",
            "identity_key": "board:TDX:881034",
            "signal_type": "B_BUY",
            "condition_key": "LIVE_CURRENT_1M:B_BUY",
            "original_condition_key": "BUY:D,W,M,Q",
            "condition_keys": ["BUY:D,W,M,Q"],
            "minute_label": "2026-06-26 09:32",
            "source_condition_pool_id": 211486,
            "source_minute_target_scope_id": 199999,
        }
        proof_context = _n3p_negative_cache_proof_context(candidate)
        proof = {
            "proof_version": "n3p_plan_only_proof_summary_v1",
            "asset_kind": "board",
            "identity_key": "board:TDX:881034",
            "stable_trigger_key": "20260626|board|board:TDX:881034|B_BUY|BUY:D,W,M,Q|Q|normal|09:31",
            "stable_trigger_family_key": "20260626|board|board:TDX:881034|B_BUY|BUY:D,W,M,Q|Q|normal",
            "requested_periods": ["D", "W", "M", "Q"],
            "original_condition_key": "BUY:D,W,M,Q",
            "source_input_fingerprint": proof_context["source_input_fingerprint"],
            "context_fingerprint": proof_context["context_fingerprint"],
            "safe_negative_cacheable": True,
            "safe_negative_cacheable_reason": "amount_chain_failed_for_required_period",
            "next_recompute_condition": "source_or_context_or_amount_chain_boundary_changed",
        }

        cache.store([proof])
        matched = cache.decision(candidate=candidate, proof_context=proof_context)

        self.assertTrue(matched["skip_full_n3p"])
        self.assertEqual(matched["decision"], "negative_cache_hit")
        self.assertEqual(matched["stable_trigger_family_key"], proof["stable_trigger_family_key"])
        self.assertEqual(matched["requested_periods"], ["D", "W", "M", "Q"])
        self.assertEqual(matched["safe_negative_cacheable_reason"], "amount_chain_failed_for_required_period")

    def test_active_state_fast_path_requires_shadow_prefilter_prune(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_N3P_REDUCTION_REQUIRES_SHADOW_PREFILTER_PRUNE"):
                create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    start_hhmm="09:31",
                    end_hhmm="09:31",
                    asset_scope="index_board_only",
                    source_bundle_key="20260626_index_board_only_full_day",
                    validation_mode="full_day_shadow_v1",
                    n3p_strategy="prefilter_audit",
                    n3p_reduction_mode="active_state_fast_path",
                )

    def test_full_day_shadow_active_state_fast_path_writes_reduction_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="09:33",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only_full_day",
                validation_mode="full_day_shadow_v1",
                n3p_strategy="prefilter_prune",
                n3p_reduction_mode="active_state_fast_path",
                enable_profiling=True,
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            status = json.loads((artifact_dir / "replay_status.json").read_text(encoding="utf-8"))
            profile = json.loads((artifact_dir / "replay_profile.json").read_text(encoding="utf-8"))
            reduction_rows = [
                json.loads(line)
                for line in (artifact_dir / "n3p_active_state_reduction.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]

            self.assertEqual(summary["n3p_reduction_mode"], "active_state_fast_path")
            self.assertEqual(status["n3p_reduction_mode"], "active_state_fast_path")
            self.assertIn("n3p_active_state_reduction", summary)
            self.assertIn("shadow_state_fast_path_count", summary["n3p_active_state_reduction"])
            self.assertIn("n3p_calls_saved", summary["n3p_active_state_reduction"])
            self.assertGreaterEqual(len(reduction_rows), 0)
            self.assertEqual(profile["metadata"]["n3p_reduction_mode"], "active_state_fast_path")

            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            self.assertIn("n3p_active_state_reduction", workbook.sheetnames)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("n3p_reduction_mode", "active_state_fast_path"), lineage_rows)

    def test_full_day_shadow_negative_cache_writes_proof_artifacts_and_counters(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="09:33",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only_full_day",
                validation_mode="full_day_shadow_v1",
                n3p_strategy="prefilter_prune",
                n3p_reduction_mode="active_state_fast_path",
                n3p_negative_cache="enabled",
                enable_profiling=True,
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            profile = json.loads((artifact_dir / "replay_profile.json").read_text(encoding="utf-8"))
            proof_rows = [
                json.loads(line)
                for line in (artifact_dir / "n3p_plan_only_proof_summary.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            decision_rows = [
                json.loads(line)
                for line in (artifact_dir / "n3p_negative_cache_decisions.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]

            self.assertEqual(summary["n3p_negative_cache"]["mode"], "enabled")
            self.assertIn("n3p_negative_cache_hits", summary["n3p_negative_cache"])
            self.assertIn("n3p_negative_cache_false_negative_count", summary["n3p_negative_cache"])
            self.assertEqual(summary["n3p_negative_cache"]["n3p_negative_cache_false_negative_count"], 0)
            self.assertGreater(len(proof_rows), 0)
            self.assertGreater(len(decision_rows), 0)
            self.assertEqual(profile["metadata"]["n3p_negative_cache"]["mode"], "enabled")

            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            self.assertIn("n3p_proof_summary", workbook.sheetnames)
            self.assertIn("n3p_negative_cache_decisions", workbook.sheetnames)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("n3p_negative_cache.mode", "enabled"), lineage_rows)

    def test_canonical_fixture_b2_hint_rows_are_ready_for_n4_matching(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="10:15",
                end_hhmm="10:20",
                replay_engine_version="canonical_plan_v1",
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            n4_messages = read_replay_messages(artifact_dir, "n4")
            n5_messages = read_replay_messages(artifact_dir, "n5")
            hint_messages = [row for row in n4_messages if row.get("source") == "hint"]
            hint_eligible = [
                row
                for row in n5_messages
                if row.get("event_type") == "ActionEligible"
                and row.get("condition_key") in {"BUY_HINT", "SELL_HINT"}
            ]

            self.assertEqual(summary["n4"]["hint_TriggerMatched"], 2)
            self.assertEqual({row["condition_key"] for row in hint_messages}, {"BUY_HINT", "SELL_HINT"})
            self.assertEqual({row["trigger_mark_candidate"] for row in hint_messages}, {"30m_volume", "30m_shrink"})
            self.assertEqual({row["projection_30m_flag"] for row in hint_messages}, {True})
            self.assertEqual(len(hint_eligible), 2)

    def test_full_day_shadow_writes_n5_evaluator_demand_artifacts_and_counters(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            job = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="09:33",
                asset_scope="index_board_only",
                source_bundle_key="20260626_index_board_only_full_day",
                validation_mode="full_day_shadow_v1",
                n3p_strategy="prefilter_prune",
                n3p_reduction_mode="active_state_fast_path",
                n3p_negative_cache="enabled",
                enable_profiling=True,
            )

            artifact_dir = Path(job["artifact_dir"])
            summary = json.loads((artifact_dir / "replay_summary.json").read_text(encoding="utf-8"))
            profile = json.loads((artifact_dir / "replay_profile.json").read_text(encoding="utf-8"))
            demand_rows = [
                json.loads(line)
                for line in (artifact_dir / "n5_evaluator_demand_plan.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            index_rows = [
                json.loads(line)
                for line in (artifact_dir / "n5_confirmation_metric_index_stats.jsonl").read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]

            self.assertEqual(len(demand_rows), 3)
            self.assertEqual(len(index_rows), 3)
            self.assertIn("n5_evaluator_demand", summary)
            self.assertIn("n5_evaluator_rows_saved", summary["n5_evaluator_demand"])
            self.assertIn("n5_evaluator_fail_open_count", summary["n5_evaluator_demand"])
            self.assertIn("n5_evaluator_demand", profile["metadata"])

            workbook = load_workbook(artifact_dir / "n3_n5_full_day_replay.xlsx", read_only=True, data_only=True)
            self.assertIn("n5_evaluator_demand_plan", workbook.sheetnames)
            self.assertIn("n5_confirmation_metric_index_stats", workbook.sheetnames)
            lineage_rows = list(workbook["lineage_and_safety"].iter_rows(values_only=True))
            self.assertIn(("n5_evaluator_demand.n5_evaluator_fail_open_count", 0), lineage_rows)

    def test_n5_evaluator_demand_inputs_close_executed_window_and_filter_metrics(self) -> None:
        eligible_rows = [
            {
                "event_id": "eligible-executed",
                "asset_kind": "board",
                "identity_key": "board:TDX:881001",
                "payload_json": {
                    "signal_type": "B_BUY",
                    "condition_key": "BUY:Y,Q,M,W,D",
                    "action_mark": "normal",
                    "selected_metric_time": "2026-06-26 09:31",
                },
            },
            {
                "event_id": "eligible-open",
                "asset_kind": "index",
                "identity_key": "index:SH:000001",
                "payload_json": {
                    "signal_type": "S_SELL",
                    "condition_key": "SELL:Y,Q,M,W,D",
                    "action_mark": "normal",
                    "selected_metric_time": "2026-06-26 09:32",
                },
            },
        ]
        closed_metric_rows = [
            {
                "asset_kind": "board",
                "identity_key": "board:TDX:881001",
                "signal_type": "B_BUY",
                "metric_minute_label": "09:31",
                "is_closed_1m": True,
                "metric_ready": True,
            },
            {
                "asset_kind": "index",
                "identity_key": "index:SH:000001",
                "signal_type": "S_SELL",
                "metric_minute_label": "09:32",
                "is_closed_1m": True,
                "metric_ready": True,
            },
            {
                "asset_kind": "index",
                "identity_key": "index:SH:000300",
                "signal_type": "B_BUY",
                "metric_minute_label": "09:32",
                "is_closed_1m": True,
                "metric_ready": True,
            },
        ]

        demand = _build_n5_evaluator_demand_inputs(
            minute="09:32",
            tracked_eligible_rows=eligible_rows,
            closed_metric_rows=closed_metric_rows,
            executed_eligible_event_ids={"eligible-executed"},
            terminal_eligible_event_ids=set(),
        )

        self.assertEqual([row["event_id"] for row in demand["eligible_rows"]], ["eligible-open"])
        self.assertEqual(len(demand["confirmation_metric_rows"]), 1)
        self.assertEqual(demand["confirmation_metric_rows"][0]["identity_key"], "index:SH:000001")
        self.assertEqual(demand["plan_row"]["open_eligible_count"], 1)
        self.assertEqual(demand["plan_row"]["closed_window_count"], 1)
        self.assertEqual(demand["plan_row"]["n5_evaluator_rows_saved"], 3)
        self.assertEqual(demand["plan_row"]["fail_open"], False)

    def test_n5_evaluator_demand_inputs_fail_open_on_unindexable_eligible(self) -> None:
        eligible_rows = [
            {
                "event_id": "eligible-open",
                "asset_kind": "index",
                "identity_key": "index:SH:000001",
                "payload_json": {
                    "signal_type": "S_SELL",
                    "condition_key": "SELL:Y,Q,M,W,D",
                },
            },
        ]
        closed_metric_rows = [
            {
                "asset_kind": "index",
                "identity_key": "index:SH:000001",
                "signal_type": "S_SELL",
                "metric_minute_label": "09:32",
                "is_closed_1m": True,
                "metric_ready": True,
            },
            {
                "asset_kind": "board",
                "identity_key": "board:TDX:881001",
                "signal_type": "B_BUY",
                "metric_minute_label": "09:32",
                "is_closed_1m": True,
                "metric_ready": True,
            },
        ]

        demand = _build_n5_evaluator_demand_inputs(
            minute="09:32",
            tracked_eligible_rows=eligible_rows,
            closed_metric_rows=closed_metric_rows,
            executed_eligible_event_ids=set(),
            terminal_eligible_event_ids=set(),
        )

        self.assertEqual(len(demand["eligible_rows"]), 1)
        self.assertEqual(len(demand["confirmation_metric_rows"]), 2)
        self.assertEqual(demand["plan_row"]["fail_open"], True)
        self.assertEqual(demand["plan_row"]["n5_evaluator_fail_open_count"], 1)

    def test_duplicate_action_executed_is_suppressed_by_stable_action_key(self) -> None:
        messages = [
            {
                "event_type": "ActionExecuted",
                "asset_kind": "board",
                "identity_key": "board:TDX:881127",
                "signal_type": "B_BUY",
                "condition_key": "BUY:Y,Q,M,W,D",
                "action_mark": "normal",
                "event_id": "first",
            },
            {
                "event_type": "ActionExecuted",
                "asset_kind": "board",
                "identity_key": "board:TDX:881127",
                "signal_type": "B_BUY",
                "condition_key": "BUY:Y,Q,M,W,D",
                "action_mark": "normal",
                "event_id": "second",
            },
            {"event_type": "ActionEligible", "identity_key": "board:TDX:881127"},
        ]
        kept, audit_rows, suppressed = _suppress_duplicate_action_executed_messages(
            messages,
            executed_action_keys=set(),
            minute="09:34",
        )

        self.assertEqual(suppressed, 1)
        self.assertEqual([row.get("event_id") for row in kept if row.get("event_type") == "ActionExecuted"], ["first"])
        self.assertEqual(audit_rows[0]["reduction_reason"], "duplicate_execution_suppressed")
        self.assertEqual(audit_rows[0]["action_key"], "board|board:TDX:881127|B_BUY|BUY:Y,Q,M,W,D|normal")

    def test_active_state_fast_path_skips_only_retained_live_state_rows(self) -> None:
        rows = [
            {
                "minute": "09:33",
                "decision": "prefilter_keep",
                "reason": "previous_live_state",
                "asset_kind": "board",
                "identity_key": "board:TDX:881127",
                "signal_type": "B_BUY",
                "condition_key": "BUY:Y,Q,M,W,D",
                "active_reduction_eligible": True,
            },
            {
                "minute": "09:33",
                "decision": "prefilter_keep",
                "reason": "previous_live_state",
                "asset_kind": "board",
                "identity_key": "board:TDX:881128",
                "signal_type": "B_BUY",
                "condition_key": "BUY:D",
                "active_reduction_eligible": False,
                "active_reduction_reason": "possible_state_change",
            },
        ]

        skip_keys, audit_rows = _active_state_reduction_skip_keys(
            rows,
            minute="09:33",
            reduction_mode="active_state_fast_path",
        )

        self.assertEqual(skip_keys, {"board|board:TDX:881127|B_BUY|BUY:Y,Q,M,W,D"})
        self.assertEqual(len(audit_rows), 1)
        self.assertEqual(audit_rows[0]["reduction_reason"], "live_state_retained")
        self.assertTrue(audit_rows[0]["n3p_skipped"])

    def test_full_day_shadow_v1_requires_index_board_only_scope(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_full_day_source_bundle(root)

            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SHADOW_REQUIRES_INDEX_BOARD_ONLY"):
                create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    asset_scope="all",
                    source_bundle_key="20260626_index_board_only_full_day",
                    validation_mode="full_day_shadow_v1",
                )

    def test_full_day_shadow_v1_rejects_partial_index_board_source_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_index_board_source_bundle(root)

            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_C1_SOURCE_INCOMPLETE"):
                create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    asset_scope="index_board_only",
                    source_bundle_key="20260626_index_board_only",
                    validation_mode="full_day_shadow_v1",
                )

    def test_asset_scope_specific_scopes_filter_expected_asset_kinds(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_mixed_asset_scope_bundle(root)

            index_job = create_local_replay_job(replay_root=root, trade_date="2026-06-26", asset_scope="index_only")
            index_summary = read_replay_job(root, index_job["job_id"])["summary"]
            self.assertEqual(index_summary["asset_scope_source_counts_after"]["candidates"]["index"], 1)
            self.assertEqual(index_summary["asset_scope_source_counts_after"]["candidates"]["stock"], 0)
            self.assertEqual(index_summary["asset_scope_source_counts_after"]["candidates"]["board"], 0)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_mixed_asset_scope_bundle(root)
            board_job = create_local_replay_job(replay_root=root, trade_date="2026-06-26", asset_scope="board_only")
            board_summary = read_replay_job(root, board_job["job_id"])["summary"]
            self.assertEqual(board_summary["asset_scope_source_counts_after"]["candidates"]["board"], 1)
            self.assertEqual(board_summary["asset_scope_source_counts_after"]["candidates"]["stock"], 0)
            self.assertEqual(board_summary["asset_scope_source_counts_after"]["candidates"]["index"], 0)

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._seed_mixed_asset_scope_bundle(root)
            stock_job = create_local_replay_job(replay_root=root, trade_date="2026-06-26", asset_scope="stock_only")
            stock_summary = read_replay_job(root, stock_job["job_id"])["summary"]
            self.assertEqual(stock_summary["asset_scope_source_counts_after"]["candidates"]["stock"], 1)
            self.assertEqual(stock_summary["asset_scope_source_counts_after"]["candidates"]["index"], 0)
            self.assertEqual(stock_summary["asset_scope_source_counts_after"]["candidates"]["board"], 0)

    def test_asset_scope_invalid_and_empty_fail_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_INVALID_ASSET_SCOPE"):
                create_local_replay_job(
                    replay_root=Path(tmp),
                    trade_date="2026-06-26",
                    asset_scope="unknown_scope_v9",
                )

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            target = root / "_sources" / "20260626" / "source_bundle.json"
            target.parent.mkdir(parents=True, exist_ok=True)
            bundle = _canonical_fixture_source_bundle(job_id="local_replay_20260626_explicit_seed")
            target.write_text(json.dumps(bundle, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SCOPE_EMPTY"):
                create_local_replay_job(
                    replay_root=root,
                    trade_date="2026-06-26",
                    asset_scope="index_only",
                )

    def test_buy_sell_hint_enter_n5_eligible_but_not_final_proof(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            job = create_local_replay_job(
                replay_root=Path(tmp),
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
            )

            n5_messages = read_replay_messages(Path(job["artifact_dir"]), "n5")
            eligible = [row for row in n5_messages if row["event_type"] == "ActionEligible"]
            executed = [row for row in n5_messages if row["event_type"] == "ActionExecuted"]

            self.assertEqual({row["condition_key"] for row in eligible if row["condition_key"].endswith("_HINT")}, {"BUY_HINT", "SELL_HINT"})
            self.assertEqual({row["signal_type"] for row in eligible}, {"B_BUY", "S_SELL"})
            self.assertEqual(len(executed), 0)
            self.assertTrue(all(row.get("final_proof_source") != "B2" for row in executed))
            self.assertEqual(eligible[0]["minute"], "10:00")

    def test_canonical_plan_v1_job_generates_n4_and_n5_messages_without_state_change_execution(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            job = create_local_replay_job(
                replay_root=Path(tmp),
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
            )

            summary = read_replay_job(Path(tmp), job["job_id"])["summary"]
            n4_messages = read_replay_messages(Path(job["artifact_dir"]), "n4")
            n5_messages = read_replay_messages(Path(job["artifact_dir"]), "n5")

            self.assertEqual(summary["replay_engine_version"], "canonical_plan_v1")
            self.assertEqual(summary["source_policy"], "historical_minute_local_replay")
            self.assertEqual(summary["n4"]["ordinary_TriggerMatched"], 1)
            self.assertEqual(summary["n4"]["hint_TriggerMatched"], 2)
            self.assertEqual(summary["n4"]["ordinary_TriggerStateChanged"], 1)
            self.assertEqual(summary["n5"]["ActionEligible"], 3)
            self.assertEqual(summary["n5"]["ActionExecuted"], 0)
            self.assertTrue(all(row.get("source_mode") == "replay" for row in n4_messages))
            self.assertTrue(all(row.get("source_mode") == "replay" for row in n5_messages))
            self.assertEqual(
                {row["event_type"] for row in n5_messages},
                {"ActionEligible"},
            )
            self.assertEqual(
                {row["event_type"] for row in n5_messages if row.get("condition_key") in {"BUY_HINT", "SELL_HINT"}},
                {"ActionEligible"},
            )
            self.assertEqual(
                {row["event_type"] for row in n4_messages if row.get("event_type") == "TriggerStateChanged"},
                {"TriggerStateChanged"},
            )
            self.assertEqual(
                [row for row in n5_messages if row.get("source_trigger_event_type") == "TriggerStateChanged"],
                [],
            )
            self.assertTrue(all(value is False for value in summary["plan_only_side_effects"].values()))
            ordinary_matches = [row for row in n4_messages if row.get("source") == "ordinary" and row.get("event_type") == "TriggerMatched"]
            self.assertEqual(len(ordinary_matches), 1)
            self.assertEqual(ordinary_matches[0]["minute"], "10:00")

    def test_fixture_v1_engine_remains_available_without_canonical_trace(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            job = create_local_replay_job(
                replay_root=Path(tmp),
                trade_date="2026-06-26",
                replay_engine_version="fixture_v1",
            )

            summary = read_replay_job(Path(tmp), job["job_id"])["summary"]
            self.assertEqual(job["replay_engine_version"], "fixture_v1")
            self.assertEqual(summary["replay_engine_version"], "fixture_v1")
            self.assertEqual(summary["source_policy"], "fixture harness, not canonical planner proof")
            self.assertEqual(summary["n4"]["ordinary_TriggerMatched"], 1)
            self.assertEqual(summary["n4"]["hint_TriggerMatched"], 2)
            self.assertEqual(summary["n5"]["ActionEligible"], 3)
            self.assertEqual(summary["n5"]["ActionExecuted"], 1)
            self.assertEqual(summary["canonical_planner_trace"], {})

    def test_unknown_replay_engine_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SIDE_EFFECT_RISK"):
                create_local_replay_job(
                    replay_root=Path(tmp),
                    trade_date="2026-06-26",
                    replay_engine_version="unknown_engine_v9",
                )

    def test_replay_job_rejects_production_like_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SIDE_EFFECT_RISK"):
                create_local_replay_job(
                    replay_root=Path(tmp),
                    trade_date="2026-06-26",
                    start_hhmm="09:31",
                    end_hhmm="15:00",
                    job_id="trigger_provisional_ordinary_20260626_until_1447",
                )

    def test_read_replay_job_and_dates_are_artifact_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            created = create_local_replay_job(
                replay_root=root,
                trade_date="2026-06-26",
                start_hhmm="09:31",
                end_hhmm="15:00",
            )

            self.assertEqual(list_replay_dates(root), ["2026-06-26"])
            loaded = read_replay_job(root, created["job_id"])
            self.assertEqual(loaded["job_id"], created["job_id"])
            self.assertEqual(loaded["summary"]["n4"]["hint_TriggerMatched"], 2)
            self.assertEqual(loaded["summary"]["replay_engine_version"], "canonical_plan_v1")
            self.assertEqual(loaded["summary"]["n3"]["source_mode"], "replay")
            timeline = read_replay_timeline(Path(created["artifact_dir"]))
            self.assertEqual(len(timeline), 240)
            ten_oclock = [row for row in timeline if row["minute"] == "10:00"][0]
            ten_oh_one = [row for row in timeline if row["minute"] == "10:01"][0]
            thirteen_thirty = [row for row in timeline if row["minute"] == "13:30"][0]
            self.assertNotIn("12:00", {row["minute"] for row in timeline})
            self.assertNotIn("12:30", {row["minute"] for row in timeline})
            self.assertEqual(ten_oclock["n4_ordinary"], 1)
            self.assertEqual(ten_oclock["n5_eligible"], 1)
            self.assertEqual(ten_oclock["n5_executed"], 0)
            self.assertEqual(ten_oh_one["n5_executed"], 0)
            self.assertEqual(thirteen_thirty["n4_ordinary"], 1)

    def test_n3p_plan_only_wrapper_matches_n4_context_by_original_condition_key(self) -> None:
        payload = live_688596_payload(include_n4_context=True)
        payload["candidates"][0]["condition_key"] = "LIVE_CURRENT_1M:B_BUY"
        payload["candidates"][0]["original_condition_key"] = "BUY:M,W,D"
        source_bundle = {
            **payload,
            "replay_config": replay_config(),
        }

        report = build_n3p_plan_only_replay_artifact(source_bundle=source_bundle)

        stock_row = report["metric_rows_by_asset"]["stock"][0]
        source = stock_row["trace_json"]["higher_period_context_source"]
        self.assertEqual(source["higher_period_context_match_strategy"], "asset_kind+identity_key+original_condition_key")
        self.assertEqual(source["original_condition_key"], "BUY:M,W,D")
        self.assertEqual(stock_row["raw_json"]["original_condition_key"], "BUY:M,W,D")
        self.assertEqual(report["quality_summary"]["blocked_reasons"], [])

    def test_n3p_plan_only_wrapper_matches_n4_context_by_condition_keys_array(self) -> None:
        payload = live_688596_payload(include_n4_context=True)
        payload["candidates"][0]["condition_key"] = "LIVE_CURRENT_1M:B_BUY"
        payload["candidates"][0]["condition_keys"] = ["BUY:M,W,D"]
        source_bundle = {
            **payload,
            "replay_config": replay_config(),
        }

        report = build_n3p_plan_only_replay_artifact(source_bundle=source_bundle)

        stock_row = report["metric_rows_by_asset"]["stock"][0]
        self.assertEqual(
            stock_row["trace_json"]["higher_period_context_source"]["higher_period_context_match_strategy"],
            "asset_kind+identity_key+condition_keys",
        )

    def test_n3p_plan_only_wrapper_injects_688596_amount_chain_seed(self) -> None:
        source_bundle = {
            **live_688596_payload(include_n4_context=True),
            "replay_config": replay_config(),
        }

        report = build_n3p_plan_only_replay_artifact(source_bundle=source_bundle)

        stock_row = report["metric_rows_by_asset"]["stock"][0]
        self.assertTrue(stock_row["trace_json"]["trigger_amount_chain_pass"]["D"])
        self.assertTrue(stock_row["trace_json"]["trigger_amount_chain_pass"]["W"])
        self.assertTrue(stock_row["trace_json"]["trigger_amount_chain_pass"]["M"])

    def test_n3p_plan_only_wrapper_sparse_no_trade_does_not_generate_fake_ready_metric(self) -> None:
        payload = live_688596_payload(include_n4_context=True)
        payload["candidates"] = payload["candidates"][:1]
        payload["candidates"][0]["code"] = "600346"
        payload["candidates"][0]["identity_key"] = "stock:SH:600346"
        payload["candidates"][0]["display_code"] = "600346"
        payload["candidates"][0]["name"] = "恒力石化"
        payload["source_records"] = {"600346": []}
        payload["live_current_sparse_no_trade_exceptions"] = [
            {
                "asset_kind": "stock",
                "identity_key": "stock:SH:600346",
                "code": "600346",
                "display_code": "600346",
                "name": "恒力石化",
                "reason": "adapter_sparse_no_trade",
                "latest_row_minute": "2026-06-26 09:31",
                "expected_target_minute": "2026-06-26 13:55",
                "latest_row": {"minute_label": "2026-06-26 09:31"},
            }
        ]
        source_bundle = {
            **payload,
            "replay_config": replay_config(),
        }

        report = build_n3p_plan_only_replay_artifact(source_bundle=source_bundle)

        self.assertEqual(report["quality_summary"]["metric_ready_count"], 0)
        self.assertEqual(report["quality_summary"]["metric_not_ready_count"], 1)
        self.assertFalse(report["metric_rows_by_asset"]["stock"][0]["metric_ready"])
        self.assertNotEqual(report["metric_rows_by_asset"]["stock"][0]["metric_time_label"], "2026-06-26 13:55")
        self.assertEqual(report["adapter_report"]["live_current_sparse_no_trade_exception_count"], 1)

    def test_b2_plan_only_wrapper_keeps_live_current_lineage_but_marks_source_mode_replay(self) -> None:
        source_bundle = {
            "replay_config": replay_config(),
        }

        report = build_b2_plan_only_replay_artifact(
            source_bundle=source_bundle,
            b1_snapshot_rows=[sample_snapshot()],
            live_current_minute_rows_by_asset={"stock": sample_current_bars()},
            previous_day_minute_rows_by_asset={"stock": sample_previous_bars()},
        )

        row = report["projection_rows_by_asset"]["stock"][0]
        self.assertEqual(row["source_fact_ids"]["canonical_source_fact_kind"], "live_current_1m")
        self.assertEqual(row["source_fact_ids"]["source_mode"], "replay")
        self.assertEqual(row["raw_json"]["source_mode"], "replay")
        self.assertEqual(row["raw_json"]["source_live_minute_kind"], "live_current_1m")
        self.assertEqual(row["source_fact_ids"]["source_live_minute_run_id"], LIVE_CURRENT_1M_SOURCE_RUN_ID)

    def test_b2_plan_only_wrapper_buy_sell_hint_rows_remain_consumable_for_n4(self) -> None:
        stock_snapshot = sample_snapshot()
        board_snapshot = dict(sample_snapshot())
        board_snapshot.update(
            {
                "snapshot_id": 2,
                "identity_key": "board:TDX:881002",
                "exchange": "TDX",
                "code": "881002",
                "display_code": "881002",
                "name": "煤炭开采",
            }
        )
        source_bundle = {
            "replay_config": replay_config(),
            "candidates": [
                {"asset_kind": "stock", "identity_key": "stock:SH:600000", "condition_key": "BUY_HINT", "signal_type": "B_BUY"},
                {"asset_kind": "board", "identity_key": "board:TDX:881002", "condition_key": "SELL_HINT", "signal_type": "S_SELL"},
            ],
        }

        report = build_b2_plan_only_replay_artifact(
            source_bundle=source_bundle,
            b1_snapshot_rows=[stock_snapshot, board_snapshot],
            live_current_minute_rows_by_asset={"stock": sample_current_bars(), "board": sample_current_bars()},
            previous_day_minute_rows_by_asset={"stock": sample_previous_bars(), "board": sample_previous_bars()},
        )

        rows = report["projection_rows"]
        self.assertEqual({row["raw_json"]["original_condition_key"] for row in rows}, {"BUY_HINT", "SELL_HINT"})
        self.assertEqual({row["raw_json"]["signal_type"] for row in rows}, {"B_BUY", "S_SELL"})

    def test_plan_only_wrappers_do_not_call_db_write_or_outbox_checkpoint_paths(self) -> None:
        source_bundle = {
            **live_688596_payload(include_n4_context=True),
            "replay_config": replay_config(),
        }

        with patch("ashare_v3.web.n6_replay.writer.write_rows_to_db", side_effect=AssertionError("write_rows_to_db_called")):
            n3p_report = build_n3p_plan_only_replay_artifact(source_bundle=source_bundle)
        self.assertEqual(n3p_report["side_effects"]["database_written"], False)

        with patch(
            "ashare_v3.web.n6_replay.projection_execute.write_projection_execute_transaction",
            side_effect=AssertionError("write_projection_execute_transaction_called"),
        ), patch(
            "ashare_v3.web.n6_replay.projection_execute.audited_n3_market_execute_connect",
            side_effect=AssertionError("db_connect_called"),
        ):
            b2_report = build_b2_plan_only_replay_artifact(
                source_bundle={"replay_config": replay_config()},
                b1_snapshot_rows=[sample_snapshot()],
                live_current_minute_rows_by_asset={"stock": sample_current_bars()},
                previous_day_minute_rows_by_asset={"stock": sample_previous_bars()},
            )
        self.assertFalse(b2_report["side_effects"]["outbox_inbox_checkpoint_consumed_or_updated"])

    def test_plan_only_wrappers_reject_production_run_id(self) -> None:
        with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SIDE_EFFECT_RISK"):
            build_n3p_plan_only_replay_artifact(
                source_bundle={
                    **live_688596_payload(include_n4_context=True),
                    "replay_config": replay_config(replay_run_id="realtime_projection_metric_20260626_prod"),
                }
            )

    def test_plan_only_wrappers_accept_local_replay_run_id(self) -> None:
        report = build_n3p_plan_only_replay_artifact(
            source_bundle={
                **live_688596_payload(include_n4_context=True),
                "replay_config": replay_config(replay_run_id="local_replay_20260626_093100_deadbeef"),
            }
        )
        self.assertEqual(report["replay_run_id"], "local_replay_20260626_093100_deadbeef")

    def test_plan_only_wrappers_fail_closed_on_side_effect_intent(self) -> None:
        with self.assertRaisesRegex(N6ReplayBlocked, "BLOCKED_REPLAY_SIDE_EFFECT_RISK"):
            build_n3p_plan_only_replay_artifact(
                source_bundle={
                    **live_688596_payload(include_n4_context=True),
                    "replay_config": replay_config(execute=True),
                }
            )


def replay_config(
    *,
    replay_run_id: str = "local_replay_20260626_093100_deadbeef",
    execute: bool = False,
) -> dict[str, object]:
    return {
        "replay_run_id": replay_run_id,
        "job_id": replay_run_id,
        "trade_date": "20260626",
        "source_trade_date": "20260625",
        "prev_trade_date": "20260625",
        "for_trade_date": "20260626",
        "until_hhmm": "1500",
        "source_mode": "replay",
        "execute": execute,
        "source_snapshot_run_id": SOURCE_SNAPSHOT_RUN_ID,
        "source_previous_day_minute_run_id": SOURCE_PREVIOUS_DAY_MINUTE_RUN_ID,
        "source_live_minute_run_id": LIVE_CURRENT_1M_SOURCE_RUN_ID,
        "source_condition_run_id": "condition_layer_20260625_source_20260625_for_20260626_v1",
        "source_subscription_run_id": "market_data_subscription_20260626_condition_layer_20260625_source_20260625_for_20260626_v1",
        "target_absence_counts": clean_target_counts(),
        "calculation_config": sample_calculation_config(),
        "latest_closed_minute": "2026-06-26T11:05:00+08:00",
        "fact_only_snapshot_trace_policy": {
            "allow_missing_snapshot_event_id": True,
            "required_trace_fields": ["snapshot_id", "subscription_id", "pull_plan_id", "source_adapter"],
        },
    }
